#!/usr/bin/env python3
"""Validate a derived GTM analyst workbook without changing the audit."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from gtm_privacy import privacy_findings
from gtm_privacy_scan import scan_xlsx
from gtm_workbook_readability import (
    MANIFEST_KIND,
    MANIFEST_SCHEMA_VERSION,
    PLACEHOLDER_RE,
    UNSUPPORTED_CLAIMS,
    artifact_paths,
    build_model,
    default_manifest_path,
    input_hash_manifest,
    load_inputs,
    projected_delta_text,
    sha256_file,
    validate_manifest_path,
    workbook_sheet_hashes,
)

# Keep the delivery contract local to the gate.  A builder-side sheet-order or
# navigation regression must not redefine what the validator accepts.
HUMAN_SHEETS = [
    "A1 Overview",
    "A2 Actions",
    "A3 Decisions",
    "A4 Audit Register",
    "A5 Custom HTML",
]
ORIGINAL_SHEETS = [
    "01 Summary",
    "02 Cleanup Plan",
    "03 Operational Review",
    "04 Configuration Review",
    "05 Architecture Review",
    "06 Custom Code Review",
    "07 Reconciled Operations",
    "08 Source & Gates",
]
SECTION_MARKER = "—"

LINK_RE = re.compile(
    r"^#(?:(?:'(?P<quoted>(?:[^']|'')+)')|(?P<plain>[^!]+))!"
    r"(?P<coordinate>[A-Z]{1,3}[1-9][0-9]*)$"
)


def expected_audit_rows(model: dict[str, Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    for section in model["audit_sections"]:
        rows.append(
            [
                SECTION_MARKER,
                "",
                "",
                f"{section['label']} ({len(section['rows'])})",
                "",
                "",
            ]
        )
        rows.extend(row["values"] for row in section["rows"])
    return rows


def expected_action_rows(model: dict[str, Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    if model["common_validation"]:
        rows.append(
            [
                model["labels"]["shared_validation"],
                "",
                "",
                "",
                "",
                "",
                "",
                model["common_validation"],
            ]
        )
    rows.extend(row["values"] for row in model["action_rows"])
    return rows


def expected_decision_rows(model: dict[str, Any]) -> list[list[str]]:
    return [row["values"] for row in model["decision_rows"]]


def expected_custom_rows(model: dict[str, Any]) -> list[list[str]]:
    return [row["values"] for row in model["custom_rows"]]


def cell_value(cell: Any) -> str:
    return "" if cell.value is None else str(cell.value)


def table_errors(
    sheet: Any,
    headers: list[str],
    expected_rows: list[list[str]],
) -> list[str]:
    errors: list[str] = []
    if sheet.max_column != len(headers):
        errors.append(
            f"{sheet.title}: expected {len(headers)} columns, found {sheet.max_column}"
        )
    actual_headers = [
        cell_value(sheet.cell(1, column))
        for column in range(1, max(sheet.max_column, len(headers)) + 1)
    ]
    if actual_headers != headers:
        errors.append(
            f"{sheet.title}: header mismatch; expected={headers}, actual={actual_headers}"
        )
    actual_count = max(0, sheet.max_row - 1)
    if actual_count != len(expected_rows):
        errors.append(
            f"{sheet.title}: expected {len(expected_rows)} data rows, found {actual_count}"
        )
    for row_index, expected in enumerate(expected_rows, start=2):
        if row_index > sheet.max_row:
            break
        actual = [
            cell_value(sheet.cell(row_index, column))
            for column in range(1, len(headers) + 1)
        ]
        if actual != expected:
            errors.append(
                f"{sheet.title}!{row_index}: row differs from authoritative projection; "
                f"expected={expected}, actual={actual}"
            )
            if len(errors) >= 25:
                errors.append(f"{sheet.title}: additional row differences suppressed")
                break
    return errors


def action_note_errors(sheet: Any, model: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    start_row = 3 if model["common_validation"] else 2
    for offset, row in enumerate(model["action_rows"]):
        row_number = start_row + offset
        expected = str(row["notes"].get(5) or "")
        comment = sheet.cell(row_number, 6).comment
        actual = str(comment.text or "") if comment else ""
        if actual != expected:
            errors.append(
                f"A2 Actions!{row_number}: exact structured-action note differs "
                "from the authoritative operation"
            )
    return errors


def independent_action_utility_errors(sheet: Any) -> list[str]:
    """Check the human contract without trusting the builder's projected rows."""

    errors: list[str] = []
    generic_markers = (
        "reduces maintenance risk without changing unrelated",
        "preserves affected measurement families",
        "improves maintainability",
        "reduces maintenance risk",
        "keeps the container clean",
    )
    machine_problem_markers = (
        "non-canonical unicode form",
        "source-proven stale or broken configuration",
    )
    generic_action_markers = (
        "apply the structured exact change below",
        "apply the approved target state",
    )
    for row_number in range(2, sheet.max_row + 1):
        operation_cell = cell_value(sheet.cell(row_number, 1))
        if "OP-" not in operation_cell:
            continue
        problem = cell_value(sheet.cell(row_number, 4)).strip()
        consequence = cell_value(sheet.cell(row_number, 5)).strip()
        exact_change = cell_value(sheet.cell(row_number, 6)).strip()
        verification = cell_value(sheet.cell(row_number, 8)).strip()
        for label, value in (
            ("literal problem", problem),
            ("consequence", consequence),
            ("exact change", exact_change),
            ("static verification and rollback", verification),
        ):
            if len(value.split()) < 5:
                errors.append(
                    f"A2 Actions!{row_number}: {label} is not standalone and readable"
                )
        if problem.casefold() == consequence.casefold():
            errors.append(
                f"A2 Actions!{row_number}: problem and consequence repeat the same text"
            )
        if any(marker in consequence.casefold() for marker in generic_markers):
            errors.append(
                f"A2 Actions!{row_number}: consequence uses generic impact boilerplate"
            )
        if any(marker in problem.casefold() for marker in machine_problem_markers):
            errors.append(
                f"A2 Actions!{row_number}: problem exposes machine wording instead of "
                "the literal GTM behavior"
            )
        if any(marker in exact_change.casefold() for marker in generic_action_markers):
            errors.append(
                f"A2 Actions!{row_number}: exact change uses an instruction placeholder"
            )
        if "static readback:" not in verification.casefold():
            errors.append(
                f"A2 Actions!{row_number}: static readback is not explicit"
            )
        if not sheet.cell(row_number, 6).comment:
            errors.append(
                f"A2 Actions!{row_number}: exact structured mutation note is missing"
            )
    return errors


def independent_decision_utility_errors(sheet: Any) -> list[str]:
    """Reject owner rows that preserve IDs but do not explain the decision's use."""

    errors: list[str] = []
    obsolete_unlock = (
        "selects the exact keep, repair, replacement, remap, or removal target"
    )
    for row_number in range(2, sheet.max_row + 1):
        identifier = cell_value(sheet.cell(row_number, 1)).strip()
        if not re.search(r"\bD-\d+\b", identifier):
            continue
        question = cell_value(sheet.cell(row_number, 2)).strip()
        recommendation = cell_value(sheet.cell(row_number, 3)).strip()
        objects = cell_value(sheet.cell(row_number, 4)).strip()
        unlock = cell_value(sheet.cell(row_number, 6)).strip()
        for label, value in (
            ("owner question", question),
            ("recommended next step", recommendation),
            ("affected scope", objects),
            ("why the answer is needed", unlock),
        ):
            if len(value.split()) < 4:
                errors.append(
                    f"A3 Decisions!{row_number}: {label} is not standalone and readable"
                )
        if obsolete_unlock in unlock.casefold():
            errors.append(
                f"A3 Decisions!{row_number}: decision consequence uses abstract action taxonomy"
            )
    return errors


def input_binding_errors(
    manifest: dict[str, Any],
    inputs: dict[str, Any],
    analyst_workbook: Path,
) -> list[str]:
    errors: list[str] = []
    if manifest.get("kind") != MANIFEST_KIND:
        errors.append("Transformation manifest kind is invalid")
    if manifest.get("schema_version") != MANIFEST_SCHEMA_VERSION:
        errors.append("Transformation manifest schema version is invalid")
    if manifest.get("source_sha256") != inputs["source_sha256"]:
        errors.append("Transformation manifest source SHA-256 is stale")
    expected_inputs = input_hash_manifest(inputs["paths"])
    if manifest.get("inputs") != expected_inputs:
        errors.append("One or more readability input hashes changed after the build")
    canonical = manifest.get("canonical_workbook") or {}
    if canonical.get("sha256") != sha256_file(inputs["paths"]["canonical_workbook"]):
        errors.append("Canonical workbook hash changed after the build")
    analyst = manifest.get("analyst_workbook") or {}
    if analyst.get("name") != analyst_workbook.name:
        errors.append("Transformation manifest names another analyst workbook")
    if analyst.get("sha256") != sha256_file(analyst_workbook):
        errors.append("Analyst workbook hash does not match the transformation manifest")
    if analyst.get("human_sheets") != HUMAN_SHEETS:
        errors.append("Transformation manifest has the wrong human sheet contract")
    if analyst.get("columns") != inputs["model"]["headers"]:
        errors.append("Transformation manifest has stale human-column definitions")
    return errors


def original_preservation_errors(
    manifest: dict[str, Any],
    inputs: dict[str, Any],
    analyst_workbook: Path,
) -> list[str]:
    errors: list[str] = []
    expected = inputs["standard_sheet_hashes"]
    declared = (manifest.get("canonical_workbook") or {}).get("sheets")
    if declared != expected:
        errors.append("Canonical sheet hashes in the manifest are stale")
    transformed = workbook_sheet_hashes(analyst_workbook, ORIGINAL_SHEETS)
    if set(transformed) != set(ORIGINAL_SHEETS):
        errors.append("The analyst workbook omits one or more original sheets")
        return errors
    for name in ORIGINAL_SHEETS:
        if transformed.get(name) != expected.get(name):
            errors.append(f"Original sheet changed during transformation: {name}")
    return errors


def workbook_structure_errors(workbook: Any, model: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    expected_order = HUMAN_SHEETS + ORIGINAL_SHEETS
    if workbook.sheetnames != expected_order:
        errors.append(
            f"Workbook sheet order mismatch; expected={expected_order}, "
            f"actual={workbook.sheetnames}"
        )
    for name in HUMAN_SHEETS:
        if name not in workbook.sheetnames:
            errors.append(f"Missing human sheet: {name}")
        elif workbook[name].sheet_state != "visible":
            errors.append(f"Human sheet must be visible: {name}")
    for name in ORIGINAL_SHEETS:
        if name not in workbook.sheetnames:
            errors.append(f"Missing canonical technical sheet: {name}")
        elif workbook[name].sheet_state != "hidden":
            errors.append(
                f"Canonical technical sheet must be retained but hidden by default: {name}"
            )
    if errors:
        return errors
    errors.extend(
        table_errors(
            workbook["A4 Audit Register"],
            model["headers"]["audit"],
            expected_audit_rows(model),
        )
    )
    errors.extend(
        table_errors(
            workbook["A2 Actions"],
            model["headers"]["actions"],
            expected_action_rows(model),
        )
    )
    errors.extend(action_note_errors(workbook["A2 Actions"], model))
    errors.extend(independent_action_utility_errors(workbook["A2 Actions"]))
    errors.extend(
        table_errors(
            workbook["A3 Decisions"],
            model["headers"]["decisions"],
            expected_decision_rows(model),
        )
    )
    errors.extend(independent_decision_utility_errors(workbook["A3 Decisions"]))
    errors.extend(
        table_errors(
            workbook["A5 Custom HTML"],
            model["headers"]["html"],
            expected_custom_rows(model),
        )
    )
    overview_values = {
        cell_value(cell)
        for row in workbook["A1 Overview"].iter_rows()
        for cell in row
        if cell.value is not None
    }
    counts = model["counts"]
    labels = model["labels"]
    priority_text = ", ".join(
        labels["counts"]["priority"].format(
            priority=labels["priority_labels"][priority],
            count=count,
        )
        for priority, count in counts["priority"].items()
        if count
    ) or labels["counts"]["no_actions"]
    required_overview = {
        str(counts["audit_records"]),
        str(counts["operations"]),
        str(counts["custom_html_tags"]),
        labels["counts"]["owner"].format(
            sources=counts["owner_source_records"],
            topics=counts["decision_topics"],
        ),
        labels["counts"]["retained"].format(
            retained=counts["retained"],
            exceptions=counts["documented_exceptions"],
        ),
        priority_text,
        projected_delta_text(
            model["projected_object_counts"],
            labels["counts"]["no_delta"],
        ),
        model["first_actions"],
        model["measurement_summary"],
        labels["counts"]["reconciliation"].format(
            findings=counts["audit_records"],
            operations=counts["operations"],
            retained=counts["retained"] + counts["documented_exceptions"],
            decisions=counts["decision_topics"],
        ),
        labels["counts"]["approval_scope"].format(
            bulk=counts["bulk_operations"],
            individual=counts["individual_operations"],
            activation=counts["activation_operations"],
        ),
        labels["counts"]["change_scope"].format(
            maintenance=counts["maintenance_operations"],
            behavior=counts["behavior_operations"],
        ),
        labels["counts"]["remaining"].format(
            remaining=counts["remaining_records"],
            decisions=counts["decision_topics"],
            limits=counts["evidence_limits"],
        ),
    }
    for value in sorted(required_overview):
        if value not in overview_values:
            errors.append(f"A1 Overview does not expose required value: {value}")
    return errors


def parse_link(value: str) -> tuple[str, str] | None:
    match = LINK_RE.fullmatch(value)
    if not match:
        return None
    sheet_name = match.group("quoted") or match.group("plain") or ""
    sheet_name = sheet_name.replace("''", "'")
    return sheet_name, match.group("coordinate")


def hyperlink_errors(
    workbook: Any,
    manifest: dict[str, Any],
) -> list[str]:
    from openpyxl.utils.cell import coordinate_to_tuple

    errors: list[str] = []
    actual: set[tuple[str, str]] = set()
    allowed_directions = {
        ("A4 Audit Register", "A2 Actions"),
        ("A4 Audit Register", "A3 Decisions"),
        ("A4 Audit Register", "A5 Custom HTML"),
        ("A2 Actions", "A4 Audit Register"),
        ("A3 Decisions", "A4 Audit Register"),
        ("A3 Decisions", "A5 Custom HTML"),
        ("A5 Custom HTML", "A4 Audit Register"),
        ("A5 Custom HTML", "A3 Decisions"),
    }
    for sheet_name in HUMAN_SHEETS:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.hyperlink:
                    continue
                target = str(
                    cell.hyperlink.target or cell.hyperlink.location or ""
                )
                parsed = parse_link(target)
                if not parsed:
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: invalid internal hyperlink {target!r}"
                    )
                    continue
                target_sheet, coordinate = parsed
                if target_sheet not in HUMAN_SHEETS:
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: link targets non-human or hidden "
                        f"sheet {target_sheet!r}"
                    )
                    continue
                if target_sheet not in workbook.sheetnames:
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: unknown target sheet {target_sheet!r}"
                    )
                    continue
                target_row, target_column = coordinate_to_tuple(coordinate)
                target_tab = workbook[target_sheet]
                if (
                    target_row > target_tab.max_row
                    or target_column > target_tab.max_column
                ):
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: target is outside used cells"
                    )
                    continue
                target_cell = target_tab.cell(target_row, target_column)
                if target_cell.value in {None, ""}:
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: target cell is empty"
                    )
                    continue
                if (sheet_name, target_sheet) not in allowed_directions:
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: invalid analyst-navigation "
                        f"direction to {target_sheet!r}"
                    )
                actual.add(
                    (
                        f"{sheet_name}!{cell.coordinate}",
                        f"{target_sheet}!{coordinate}",
                    )
                )
    expected = {
        (str(row.get("source") or ""), str(row.get("target") or ""))
        for row in manifest.get("links") or []
    }
    if actual != expected:
        missing = sorted(expected - actual)
        unknown = sorted(actual - expected)
        errors.append(
            f"Visible-sheet hyperlink map differs; missing={missing[:10]}, "
            f"unknown={unknown[:10]}"
        )
    return errors


def readability_errors(workbook: Any) -> list[str]:
    errors: list[str] = []
    for sheet_name in HUMAN_SHEETS:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: formula cells are forbidden"
                    )
                if cell.value is None:
                    continue
                text = str(cell.value)
                if PLACEHOLDER_RE.search(text):
                    errors.append(
                        f"{sheet_name}!{cell.coordinate}: unresolved placeholder"
                    )
                lower = text.casefold()
                for claim in UNSUPPORTED_CLAIMS:
                    if claim.casefold() in lower:
                        errors.append(
                            f"{sheet_name}!{cell.coordinate}: unsupported absolute claim "
                            f"{claim!r}"
                        )
    return errors


def comment_privacy_errors(workbook: Any) -> list[str]:
    errors: list[str] = []
    for sheet_name in HUMAN_SHEETS:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.comment:
                    continue
                for finding in privacy_findings(cell.comment.text):
                    errors.append(
                        f"{sheet_name}!{cell.coordinate} comment: {finding}"
                    )
    return errors


def classify_gates(errors: dict[str, list[str]]) -> dict[str, str]:
    return {
        name: ("pass" if not values else "fail")
        for name, values in errors.items()
    }


def update_manifest(
    path: Path,
    manifest: dict[str, Any],
    errors: dict[str, list[str]],
) -> dict[str, Any]:
    flattened = [
        f"{gate}: {message}"
        for gate, values in errors.items()
        for message in values
    ]
    manifest["status"] = "pass" if not flattened else "fail"
    manifest["gates"] = classify_gates(errors)
    manifest["errors"] = flattened
    path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def validate(
    package_dir: Path,
    operations_path: Path,
    canonical_workbook: Path,
    analyst_workbook: Path,
    *,
    future_state_path: Path | None = None,
    completion_gate_path: Path | None = None,
    decision_topics_path: Path | None = None,
    manifest_path: Path | None = None,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to validate XLSX output") from exc

    manifest_path = manifest_path or default_manifest_path(analyst_workbook)
    if not analyst_workbook.is_file():
        raise FileNotFoundError(f"Missing analyst workbook: {analyst_workbook}")
    paths = artifact_paths(
        package_dir,
        operations_path,
        canonical_workbook,
        future_state_path,
        completion_gate_path,
        decision_topics_path,
    )
    validate_manifest_path(manifest_path, analyst_workbook, paths)
    if not manifest_path.is_file():
        raise FileNotFoundError(f"Missing transformation manifest: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    if (
        manifest.get("kind") != MANIFEST_KIND
        or manifest.get("schema_version") != MANIFEST_SCHEMA_VERSION
    ):
        raise ValueError(
            "Refusing to update a file that is not this workbook's "
            "transformation manifest"
        )
    inputs = load_inputs(paths)
    model = build_model(inputs, str(manifest.get("language") or "en"))
    inputs["model"] = model

    errors: dict[str, list[str]] = {
        "input_binding": input_binding_errors(manifest, inputs, analyst_workbook),
        "original_preservation": [],
        "workbook_integrity": [],
        "audit_coverage": [],
        "action_coverage_and_direction": [],
        "decision_coverage": [],
        "custom_html_coverage": [],
        "visible_links": [],
        "readability": [],
        "privacy": [],
    }

    try:
        errors["original_preservation"].extend(
            original_preservation_errors(manifest, inputs, analyst_workbook)
        )
    except Exception as exc:
        errors["original_preservation"].append(
            f"Could not compare original sheets: {exc}"
        )
        errors["workbook_integrity"].append(
            f"Workbook open failed during preservation check: {exc}"
        )

    try:
        workbook = load_workbook(analyst_workbook, read_only=False, data_only=False)
    except Exception as exc:
        errors["workbook_integrity"].append(f"Workbook open failed: {exc}")
    else:
        try:
            structure = workbook_structure_errors(workbook, model)
            errors["workbook_integrity"].extend(
                message
                for message in structure
                if message.startswith("Workbook")
                or message.startswith("Missing")
                or message.startswith("Human")
                or message.startswith("A1")
            )
            errors["audit_coverage"].extend(
                message
                for message in structure
                if message.startswith("A4 Audit Register")
            )
            errors["action_coverage_and_direction"].extend(
                message
                for message in structure
                if message.startswith("A2 Actions")
            )
            errors["decision_coverage"].extend(
                message
                for message in structure
                if message.startswith("A3 Decisions")
            )
            errors["custom_html_coverage"].extend(
                message
                for message in structure
                if message.startswith("A5 Custom HTML")
            )
            errors["visible_links"].extend(hyperlink_errors(workbook, manifest))
            errors["readability"].extend(readability_errors(workbook))
            errors["privacy"].extend(comment_privacy_errors(workbook))
        finally:
            workbook.close()

        try:
            # Reopening in read-only mode catches damaged package structures that a
            # normal load may not expose until worksheet iteration.
            reopened = load_workbook(
                analyst_workbook,
                read_only=True,
                data_only=False,
            )
            try:
                for sheet in reopened.worksheets:
                    for _row in sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row, 2),
                    ):
                        pass
            finally:
                reopened.close()
        except Exception as exc:  # pragma: no cover - parser-specific failures
            errors["workbook_integrity"].append(f"Workbook reopen failed: {exc}")

        try:
            errors["privacy"].extend(scan_xlsx(analyst_workbook, all_sheets=True))
        except Exception as exc:  # pragma: no cover - parser-specific failures
            errors["privacy"].append(f"Workbook privacy scan failed: {exc}")

    return update_manifest(manifest_path, manifest, errors)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("package_dir", type=Path)
    parser.add_argument("operations", type=Path)
    parser.add_argument("canonical_workbook", type=Path)
    parser.add_argument("analyst_workbook", type=Path)
    parser.add_argument("--future-state", type=Path)
    parser.add_argument("--completion-gate", type=Path)
    parser.add_argument("--decision-topics", type=Path)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()
    try:
        result = validate(
            args.package_dir,
            args.operations,
            args.canonical_workbook,
            args.analyst_workbook,
            future_state_path=args.future_state,
            completion_gate_path=args.completion_gate,
            decision_topics_path=args.decision_topics,
            manifest_path=args.manifest,
        )
    except (FileNotFoundError, RuntimeError, ValueError, json.JSONDecodeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(
        json.dumps(
            {
                "status": result["status"],
                "workbook": str(args.analyst_workbook),
                "manifest": str(args.manifest or default_manifest_path(args.analyst_workbook)),
                "gates": result.get("gates"),
                "errors": result.get("errors"),
            },
            ensure_ascii=False,
            indent=2 if args.pretty else None,
        )
    )
    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
