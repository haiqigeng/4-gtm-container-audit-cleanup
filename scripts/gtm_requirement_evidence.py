#!/usr/bin/env python3
"""Normalize an approved tracking plan into source-labelled audit evidence."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import unicodedata
from itertools import chain
from pathlib import Path
from typing import Any

from gtm_lib import as_list, stable_hash, write_json

HEADER_ALIASES = {
    "event_name": {
        "event",
        "eventname",
        "nomevenement",
        "evenement",
        "nomdelevenement",
        "ga4event",
    },
    "object_name": {
        "tag",
        "tagname",
        "balise",
        "nombalise",
        "nomdelabalise",
        "gtmobject",
    },
    "destination": {
        "destination",
        "measurementid",
        "pixelid",
        "accountid",
        "identifiantdestination",
    },
    "requirement": {
        "requirement",
        "description",
        "businessrequirement",
        "besoin",
        "regle",
        "specification",
    },
    "status": {"status", "statut", "approval", "validation"},
}
RECOGNIZED_HEADERS = set().union(*HEADER_ALIASES.values())


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalized_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "", text.casefold())


def scalar_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    rendered = " ".join(str(value).split())
    return rendered[:2000]


def recognized_value(fields: dict[str, Any], semantic_field: str) -> str:
    aliases = HEADER_ALIASES[semantic_field]
    for header, value in fields.items():
        if normalized_header(header) in aliases and scalar_text(value):
            return scalar_text(value)
    return ""


def normalized_requirement_rows(
    rows: list[tuple[str, int, dict[str, Any]]], source_sha256: str
) -> list[dict[str, Any]]:
    result = []
    for sheet, row_number, fields in rows:
        visible = {
            str(header): scalar_text(value)
            for header, value in fields.items()
            if str(header).strip() and scalar_text(value)
        }
        if not visible:
            continue
        identity = {
            "source_sha256": source_sha256,
            "sheet": sheet,
            "row_number": row_number,
            "fields": visible,
        }
        result.append(
            {
                "requirement_id": f"REQ-{stable_hash(identity, 12).upper()}",
                "source_sheet": sheet,
                "source_row": row_number,
                "event_name": recognized_value(visible, "event_name"),
                "object_name": recognized_value(visible, "object_name"),
                "destination": recognized_value(visible, "destination"),
                "requirement": recognized_value(visible, "requirement"),
                "status": recognized_value(visible, "status"),
                "source_fields": visible,
                "source_row_sha256": stable_hash(identity, 64),
            }
        )
    ids = [row["requirement_id"] for row in result]
    if len(ids) != len(set(ids)):
        raise ValueError("approved requirement rows do not have unique identities")
    return result


def json_rows(path: Path) -> list[tuple[str, int, dict[str, Any]]]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(payload, list):
        values = payload
    elif isinstance(payload, dict):
        values = next(
            (
                value
                for key in ("requirements", "rows", "events")
                if isinstance((value := payload.get(key)), list)
            ),
            [],
        )
    else:
        raise ValueError("approved requirement JSON must be an object or array")
    return [
        ("JSON", index, row)
        for index, row in enumerate(values, start=1)
        if isinstance(row, dict)
    ]


def csv_rows(path: Path) -> list[tuple[str, int, dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [
            ("CSV", index, dict(row))
            for index, row in enumerate(reader, start=2)
        ]


def xlsx_rows(path: Path) -> list[tuple[str, int, dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read approved XLSX requirements") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    result: list[tuple[str, int, dict[str, Any]]] = []
    try:
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            buffered = []
            for row_number in range(1, 21):
                try:
                    row = next(values)
                except StopIteration:
                    break
                rendered = [scalar_text(value) for value in row]
                buffered.append((row_number, row, rendered))
            candidates = [
                (
                    sum(
                        normalized_header(value) in RECOGNIZED_HEADERS
                        for value in rendered
                        if value
                    ),
                    row_number,
                    row,
                    rendered,
                )
                for row_number, row, rendered in buffered
                if any(rendered)
            ]
            if not candidates:
                continue
            recognized = [candidate for candidate in candidates if candidate[0] > 0]
            _score, header_row, _header_values, rendered_headers = max(
                recognized or candidates,
                key=lambda candidate: (candidate[0], -candidate[1]),
            )
            headers: list[str] = []
            seen_headers: dict[str, int] = {}
            for index, value in enumerate(rendered_headers, start=1):
                base = value or f"Column {index}"
                seen_headers[base] = seen_headers.get(base, 0) + 1
                suffix = f" [{seen_headers[base]}]" if seen_headers[base] > 1 else ""
                headers.append(base + suffix)
            remaining_rows = chain(
                (
                    (row_number, row)
                    for row_number, row, _rendered in buffered
                    if row_number > header_row
                ),
                enumerate(values, start=21),
            )
            for row_number, row in remaining_rows:
                fields = {
                    header: row[index] if index < len(row) else None
                    for index, header in enumerate(headers)
                }
                result.append((sheet.title, row_number, fields))
    finally:
        workbook.close()
    return result


def build_requirement_evidence(path: Path) -> dict[str, Any]:
    suffix = path.suffix.casefold()
    if suffix == ".json":
        raw_rows = json_rows(path)
    elif suffix == ".csv":
        raw_rows = csv_rows(path)
    elif suffix in {".xlsx", ".xlsm"}:
        raw_rows = xlsx_rows(path)
    else:
        raise ValueError("approved requirements must be JSON, CSV, XLSX, or XLSM")
    source_hash = sha256_file(path)
    rows = normalized_requirement_rows(raw_rows, source_hash)
    return {
        "kind": "gtm_approved_requirement_evidence",
        "schema_version": 1,
        "source_file": path.name,
        "source_sha256": source_hash,
        "approval_status": "analyst_supplied_as_approved",
        "evidence_role": (
            "Approved requirement context only. It may support legacy, replacement, "
            "and necessity decisions but is never container evidence and never changes "
            "the tracking plan."
        ),
        "counts": {
            "rows": len(rows),
            "rows_with_event_name": sum(bool(row["event_name"]) for row in rows),
            "rows_with_object_name": sum(bool(row["object_name"]) for row in rows),
            "rows_with_destination": sum(bool(row["destination"]) for row in rows),
        },
        "requirements": rows,
    }


def object_requirement_links(
    obj: dict[str, Any], object_name: str, evidence: dict[str, Any] | None
) -> list[dict[str, Any]]:
    """Create exact-match candidates only; do not infer semantic equivalence."""

    if not evidence:
        return []
    scalar_values: set[str] = set()

    def visit(value: Any) -> None:
        if isinstance(value, dict):
            for nested in value.values():
                visit(nested)
        elif isinstance(value, list):
            for nested in value:
                visit(nested)
        elif value is not None:
            rendered = scalar_text(value).casefold()
            if rendered:
                scalar_values.add(rendered)

    visit(obj)
    normalized_name = scalar_text(object_name).casefold()
    links = []
    for requirement in as_list(evidence.get("requirements")):
        matches = []
        requirement_name = scalar_text(requirement.get("object_name")).casefold()
        event_name = scalar_text(requirement.get("event_name")).casefold()
        destination = scalar_text(requirement.get("destination")).casefold()
        if requirement_name and requirement_name == normalized_name:
            matches.append("exact_object_name")
        if event_name and event_name in scalar_values:
            matches.append("exact_event_value")
        if destination and destination in scalar_values:
            matches.append("exact_destination_value")
        if matches:
            links.append(
                {
                    "requirement_id": str(requirement.get("requirement_id") or ""),
                    "match_types": matches,
                    "source_sheet": str(requirement.get("source_sheet") or ""),
                    "source_row": requirement.get("source_row"),
                    "event_name": str(requirement.get("event_name") or ""),
                    "object_name": str(requirement.get("object_name") or ""),
                    "destination": str(requirement.get("destination") or ""),
                    "requirement": str(requirement.get("requirement") or ""),
                    "source_row_sha256": str(
                        requirement.get("source_row_sha256") or ""
                    ),
                    "interpretation_boundary": (
                        "Exact text/value match only; replacement, necessity, and semantic "
                        "equivalence still require independent analyst judgment."
                    ),
                }
            )
    return links


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("requirements", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()
    try:
        payload = build_requirement_evidence(args.requirements)
        write_json(args.output, payload, args.pretty)
        print(
            json.dumps(
                {"status": "pass", "output": str(args.output), **payload["counts"]}
            )
        )
        return 0
    except (OSError, ValueError, RuntimeError, json.JSONDecodeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
