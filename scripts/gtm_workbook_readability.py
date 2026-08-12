#!/usr/bin/env python3
"""Add lossless analyst-facing views to a validated GTM cleanup workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from gtm_human_rows import (
    operation_action_text,
    operation_problem_text,
    static_verification_text,
)
from gtm_lib import as_list, load_json
from gtm_privacy import redact_text, spreadsheet_safe_text

ORIGINAL_SHEETS = [
    "01 Summary",
    "02 Cleanup Plan",
    "03 Operational Review",
    "04 Configuration Review",
    "05 Architecture Review",
    "06 Custom Code Review",
    "07 Reconciled Operations",
    "08 Source & Gates",
]
HUMAN_SHEETS = [
    "A1 Overview",
    "A2 Actions",
    "A3 Decisions",
    "A4 Audit Register",
    "A5 Custom HTML",
]
MANIFEST_KIND = "gtm_workbook_readability_manifest"
MANIFEST_SCHEMA_VERSION = 1

HEADER_FILL = "17365D"
HEADER_FONT = "FFFFFF"
SUBHEADER_FILL = "DCE6F1"
ALT_FILL = "F6F8FB"
GRID_COLOR = "D7DEE8"
LINK_COLOR = "0563C1"
WARNING_FILL = "FFF2CC"
HUMAN_TAB_COLOR = "1F4E78"
PRIORITY_FILLS = {
    "Critical": ("8B0000", "FFFFFF"),
    "High": ("C00000", "FFFFFF"),
    "Medium": ("FFD966", "3F2D00"),
    "Low": ("C6E0B4", "244000"),
}
PRIORITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
DISPOSITION_ORDER = {
    "cleanup_operation": 0,
    "owner_decision_needed": 1,
    "keep": 2,
    "documented_exception": 3,
    "container_evidence_limit": 4,
    "not_applicable": 5,
}
SECTION_MARKER = "—"

LOCALES: dict[str, dict[str, Any]] = {
    "en": {
        "headers": {
            "audit": [
                "ID",
                "Area",
                "Objects",
                "Finding",
                "Outcome / waiting for",
                "Priority",
            ],
            "actions": [
                "Order + OP ID",
                "Priority",
                "Objects",
                "Literal problem",
                "Consequence if unchanged",
                "Exact change",
                "Preconditions / approval",
                "Static verification + rollback",
            ],
            "decisions": [
                "Decision",
                "Question",
                "Recommendation",
                "Affected items",
                "Measurement families",
                "What the answer unlocks",
            ],
            "html": [
                "Tag",
                "State / execution context",
                "Functional role",
                "Technical health",
                "Replacement / simplification candidate",
                "Simplest safe target",
                "Exact action / decision",
            ],
        },
        "title": "GTM container audit — analyst workbook",
        "status": (
            "Static audit and projected-container gates passed. "
            "No GTM change has been executed."
        ),
        "overview": {
            "audit_records": "Audit records",
            "operations": "Atomic actions",
            "owner": "Owner decisions",
            "html": "Custom HTML tags",
            "retained": "Retained / exceptions",
            "priorities": "Action priorities",
            "deltas": "Projected object deltas",
            "first_actions": "First cleanup actions",
            "measurement": "Measurement target state",
            "boundary": "Evidence boundary",
            "next": "Next analyst step",
            "navigation": "How to use this workbook",
            "reconciliation": "How the audit reconciles",
            "approval_scope": "Approval scope",
            "change_scope": "Operation impact",
            "remaining": "Remaining findings",
        },
        "counts": {
            "owner": "{sources} source / {topics} topics",
            "retained": "{retained} retained; {exceptions} documented exceptions",
            "priority": "{priority} {count}",
            "no_actions": "No proposed action",
            "no_delta": "No count change",
            "no_families": "No confirmed measurement family in the operation packet",
            "measurement": (
                "{total} families: {changed} changed; {retained} retained; "
                "{reviewed} reviewed without operation; {owner} owner-blocked; "
                "{boundary} evidence-limited"
            ),
            "reconciliation": (
                "{findings} reconciled findings produce {operations} atomic operations; "
                "{retained} retained/exception records and {decisions} owner topics remain visible."
            ),
            "approval_scope": "{bulk} bulk-eligible; {individual} individual; {activation} activation-sensitive",
            "change_scope": "{maintenance} maintenance-only; {behavior} behavior-changing",
            "remaining": "{remaining} records remain without an operation ({decisions} owner topics; {limits} evidence limits)",
        },
        "priority_labels": {
            "Critical": "Critical",
            "High": "High",
            "Medium": "Medium",
            "Low": "Low",
        },
        "next_step": (
            "Review every A2 action and A3 owner question before authorising any GTM change."
        ),
        "navigation": (
            "A1–A5 are the analyst views. Sheets 01–08 are the unchanged canonical "
            "audit record. Unhide a technical sheet and filter by the stable ID or "
            "object key when deeper proof is needed."
        ),
        "boundary": (
            "Container-only static evidence. No GTM Preview, live dataLayer value, "
            "network request, CMP state, vendor acceptance, or unseen server-container "
            "behaviour was verified."
        ),
        "sections": {
            "cleanup_operation": "Actions",
            "owner_decision_needed": "Owner decisions",
            "keep": "Retained — no action",
            "documented_exception": "Documented exceptions — no action",
            "container_evidence_limit": "Evidence limitations",
            "not_applicable": "Not applicable — no action",
        },
        "outcomes": {
            "cleanup_operation": "Action {ids} — waiting for analyst approval",
            "owner_decision_needed": "Decision {topic} — waiting for owner answer",
            "keep": "Retained — no action",
            "documented_exception": "Documented exception — no action",
            "container_evidence_limit": "Evidence limitation",
            "not_applicable": "Not applicable — no action",
        },
        "shared_validation": "Shared validation",
        "source_items": "{count} source items — expand",
        "states": {
            "active": "Active",
            "paused": "Paused",
            "deprecated": "Named deprecated",
        },
        "custom": {
            "long": "Long Custom HTML ({length} characters).",
            "direct": "Already reads a configured dataLayer/GTM source: {sources}.",
            "candidate": (
                "Potential dataLayer candidate: {sources}. "
                "The key match does not prove live equivalence."
            ),
            "legacy": (
                "Uses legacy browser/DOM value acquisition; no source-proven "
                "dataLayer replacement was identified."
            ),
            "producer": (
                "Produces or writes dataLayer data; treat it as a producer, not a "
                "replacement consumer."
            ),
            "loader": "Vendor/script loader; no native replacement is proven by the export.",
            "none": "No source-proven native or dataLayer replacement was identified.",
            "conflict": "Cleanup conflict: {objects} are scheduled for deletion by {operations}.",
            "decisions": "Related decision: {topics}.",
            "health": "{status}; selected disposition: {disposition}. {findings}",
            "operation_target": "Approved target is defined by {operations}.",
            "owner_target": "Keep unchanged until {topics} selects the exact target.",
            "keep_target": "Keep the exported implementation; no static code mutation is proposed.",
            "candidate_native": "Native/template candidate: {candidates}.",
            "candidate_duplicate": "Consolidation candidate with identical code: {tags}.",
        },
    },
    "fr-FR": {
        "headers": {
            "audit": [
                "ID",
                "Domaine",
                "Objets",
                "Constat",
                "Résultat / attente",
                "Priorité",
            ],
            "actions": [
                "Ordre + ID OP",
                "Priorité",
                "Objets",
                "Problème concret",
                "Conséquence sans correction",
                "Modification exacte",
                "Prérequis / approbation",
                "Vérification statique + retour arrière",
            ],
            "decisions": [
                "Décision",
                "Question",
                "Recommandation",
                "Éléments concernés",
                "Familles de mesure",
                "Ce que la réponse débloque",
            ],
            "html": [
                "Tag",
                "État / contexte d’exécution",
                "Rôle fonctionnel",
                "Santé technique",
                "Candidat de remplacement / simplification",
                "Cible sûre la plus simple",
                "Action / décision exacte",
            ],
        },
        "title": "Audit du conteneur GTM — classeur analyste",
        "status": (
            "Les contrôles statiques de l’audit et du conteneur projeté sont validés. "
            "Aucune modification GTM n’a été exécutée."
        ),
        "overview": {
            "audit_records": "Éléments audités",
            "operations": "Actions atomiques",
            "owner": "Décisions propriétaire",
            "html": "Tags Custom HTML",
            "retained": "Conservés / exceptions",
            "priorities": "Priorités des actions",
            "deltas": "Évolution projetée des objets",
            "first_actions": "Premières actions de nettoyage",
            "measurement": "État cible des mesures",
            "boundary": "Limite de preuve",
            "next": "Prochaine étape analyste",
            "navigation": "Utilisation du classeur",
            "reconciliation": "Réconciliation de l’audit",
            "approval_scope": "Périmètre d’approbation",
            "change_scope": "Impact des opérations",
            "remaining": "Constats restants",
        },
        "counts": {
            "owner": "{sources} sources / {topics} sujets",
            "retained": "{retained} conservés ; {exceptions} exceptions documentées",
            "priority": "{priority} {count}",
            "no_actions": "Aucune action proposée",
            "no_delta": "Aucune variation du nombre d’objets",
            "no_families": "Aucune famille de mesure confirmée dans le plan d’opérations",
            "measurement": (
                "{total} familles : {changed} modifiées ; {retained} conservées ; "
                "{reviewed} examinées sans opération ; {owner} bloquées par décision ; "
                "{boundary} limitées par la preuve"
            ),
            "reconciliation": (
                "{findings} constats réconciliés produisent {operations} opérations atomiques ; "
                "{retained} lignes conservées/exceptions et {decisions} sujets propriétaire restent visibles."
            ),
            "approval_scope": "{bulk} regroupables ; {individual} individuels ; {activation} sensibles à l’activation",
            "change_scope": "{maintenance} maintenance uniquement ; {behavior} changement de comportement",
            "remaining": "{remaining} lignes restent sans opération ({decisions} sujets propriétaire ; {limits} limites de preuve)",
        },
        "priority_labels": {
            "Critical": "Critique",
            "High": "Haute",
            "Medium": "Moyenne",
            "Low": "Faible",
        },
        "next_step": (
            "Examiner chaque action A2 et chaque question A3 avant d’autoriser "
            "une modification GTM."
        ),
        "navigation": (
            "A1–A5 sont les vues analyste. Les onglets 01–08 conservent l’audit "
            "canonique inchangé. Afficher un onglet technique puis filtrer par ID "
            "stable ou clé d’objet pour consulter la preuve détaillée."
        ),
        "boundary": (
            "Preuve statique limitée au conteneur. Aucun Preview GTM, valeur dataLayer "
            "live, appel réseau, état CMP, retour fournisseur ou comportement d’un "
            "conteneur serveur non visible n’a été vérifié."
        ),
        "sections": {
            "cleanup_operation": "Actions",
            "owner_decision_needed": "Décisions propriétaire",
            "keep": "Conservés — aucune action",
            "documented_exception": "Exceptions documentées — aucune action",
            "container_evidence_limit": "Limites de preuve",
            "not_applicable": "Non applicable — aucune action",
        },
        "outcomes": {
            "cleanup_operation": "Action {ids} — en attente d’approbation analyste",
            "owner_decision_needed": "Décision {topic} — en attente de réponse",
            "keep": "Conservé — aucune action",
            "documented_exception": "Exception documentée — aucune action",
            "container_evidence_limit": "Limite de preuve",
            "not_applicable": "Non applicable — aucune action",
        },
        "shared_validation": "Validation commune",
        "source_items": "{count} éléments sources — développer",
        "states": {
            "active": "Actif",
            "paused": "En pause",
            "deprecated": "Nommé obsolète",
        },
        "custom": {
            "long": "Custom HTML long ({length} caractères).",
            "direct": "Utilise déjà une source dataLayer/GTM configurée : {sources}.",
            "candidate": (
                "Source dataLayer potentielle : {sources}. "
                "La correspondance de clé ne prouve pas l’équivalence live."
            ),
            "legacy": (
                "Utilise une acquisition de valeur legacy via navigateur/DOM ; "
                "aucun remplacement dataLayer n’est prouvé par la source."
            ),
            "producer": (
                "Produit ou écrit des données dans le dataLayer ; le traiter comme "
                "producteur et non comme consommateur à remplacer."
            ),
            "loader": (
                "Chargeur fournisseur/script ; aucun remplacement natif n’est prouvé "
                "par l’export."
            ),
            "none": "Aucun remplacement natif ou dataLayer n’est prouvé par la source.",
            "conflict": (
                "Conflit de nettoyage : {objects} doivent être supprimés par {operations}."
            ),
            "decisions": "Décision associée : {topics}.",
            "health": "{status} ; disposition retenue : {disposition}. {findings}",
            "operation_target": "La cible approuvée est définie par {operations}.",
            "owner_target": "Conserver sans modification jusqu’à ce que {topics} choisisse la cible exacte.",
            "keep_target": "Conserver l’implémentation exportée ; aucune mutation statique du code n’est proposée.",
            "candidate_native": "Candidat natif/template : {candidates}.",
            "candidate_duplicate": "Candidat à la consolidation avec code identique : {tags}.",
        },
    },
}

OBJECT_ID_FIELDS = {
    "tag": "tagId",
    "trigger": "triggerId",
    "variable": "variableId",
    "folder": "folderId",
    "customTemplate": "templateId",
    "client": "clientId",
    "zone": "zoneId",
    "gtagConfig": "gtagConfigId",
    "transformation": "transformationId",
}
ACTION_MUTATION_FIELDS = (
    "canonical_object_key",
    "creations",
    "additions",
    "changes",
    "remaps",
    "renames",
    "deletions",
)
MAX_ACTION_NOTE_TEXT = 30_000
IDENTIFIER_STOP_KEYS = {
    "content",
    "event",
    "page",
    "status",
    "user",
    "value",
    "data",
    "result",
    "resource",
}
PLACEHOLDER_RE = re.compile(
    r"(?i)\b(?:TO" r"DO|TBD|LOREM IPSUM)\b|<placeholder>"
)
UNSUPPORTED_CLAIMS = (
    "zero measurement loss",
    "all integrations preserved",
    "guaranteed live behaviour",
    "guaranteed live behavior",
    "legal consent compliance confirmed",
    "aucune perte de mesure",
    "toutes les intégrations sont préservées",
    "comportement live garanti",
    "conformité juridique du consentement confirmée",
)


def normalize_space(value: Any) -> str:
    return " ".join(str(value or "").split())


def safe_text(value: Any) -> str:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return spreadsheet_safe_text(redact_text("" if value is None else str(value)))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def json_hash_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    return value


def workbook_sheet_hashes(path: Path, names: list[str] | None = None) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to inspect XLSX output") from exc

    workbook = load_workbook(path, read_only=False, data_only=False)
    selected = names or workbook.sheetnames
    result: dict[str, Any] = {}
    try:
        for name in selected:
            if name not in workbook.sheetnames:
                continue
            sheet = workbook[name]
            digest = hashlib.sha256()
            identity = {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
            }
            digest.update(
                json.dumps(identity, ensure_ascii=False, sort_keys=True).encode("utf-8")
            )
            for row in sheet.iter_rows():
                for cell in row:
                    comment = getattr(cell, "comment", None)
                    hyperlink = getattr(cell, "hyperlink", None)
                    if cell.value is None and comment is None and hyperlink is None:
                        continue
                    record = [
                        cell.coordinate,
                        cell.data_type,
                        json_hash_value(cell.value),
                        (
                            {
                                "author": str(comment.author or ""),
                                "text": str(comment.text or ""),
                            }
                            if comment is not None
                            else None
                        ),
                        (
                            {
                                "target": str(hyperlink.target or ""),
                                "location": str(hyperlink.location or ""),
                                "display": str(hyperlink.display or ""),
                                "tooltip": str(hyperlink.tooltip or ""),
                            }
                            if hyperlink is not None
                            else None
                        ),
                    ]
                    digest.update(
                        json.dumps(
                            record,
                            ensure_ascii=False,
                            sort_keys=True,
                            default=str,
                        ).encode("utf-8")
                    )
            result[name] = {**identity, "content_sha256": digest.hexdigest()}
    finally:
        workbook.close()
    return result


def default_manifest_path(output: Path) -> Path:
    return output.with_name(f"{output.stem}.manifest.json")


def validate_manifest_path(
    manifest_path: Path,
    analyst_workbook: Path,
    input_paths: dict[str, Path],
) -> None:
    manifest_resolved = manifest_path.resolve()
    analyst_resolved = analyst_workbook.resolve()
    if manifest_resolved.parent != analyst_resolved.parent:
        raise ValueError(
            "The transformation manifest must be stored beside the analyst workbook"
        )
    protected = {analyst_resolved}
    protected.update(path.resolve() for path in input_paths.values())
    if manifest_resolved in protected:
        raise ValueError(
            "The transformation manifest path cannot overwrite a workbook or audit input"
        )
    if manifest_path.exists():
        protected_paths = [analyst_workbook, *input_paths.values()]
        if any(
            path.exists() and manifest_path.samefile(path)
            for path in protected_paths
        ):
            raise ValueError(
                "The transformation manifest cannot alias a workbook or audit input"
            )


def artifact_paths(
    package_dir: Path,
    operations_path: Path,
    standard_workbook: Path,
    future_state_path: Path | None,
    completion_gate_path: Path | None,
    decision_topics_path: Path | None,
) -> dict[str, Path]:
    paths = {
        "audit_package_manifest": package_dir / "audit_package_manifest.json",
        "context": package_dir / "context.json",
        "source_model": package_dir / "source_model.json",
        "operational_review": package_dir / "operational_review.json",
        "configuration_review": package_dir / "configuration_review.json",
        "architecture_review": package_dir / "architecture_review.json",
        "technical_code_findings": package_dir / "technical_code_findings.json",
        "reconciled_operations": operations_path,
        "future_state_gate": future_state_path or package_dir / "future_state_gate.json",
        "completion_gate": completion_gate_path or package_dir / "completion_gate.json",
        "canonical_workbook": standard_workbook,
    }
    if decision_topics_path is not None:
        paths["decision_topics"] = decision_topics_path
    missing = [f"{role}: {path}" for role, path in paths.items() if not path.is_file()]
    if missing:
        raise FileNotFoundError("Missing readability input(s): " + "; ".join(missing))
    return paths


def load_inputs(paths: dict[str, Path]) -> dict[str, Any]:
    payloads: dict[str, Any] = {}
    for role, path in paths.items():
        if path.suffix.lower() == ".json":
            payloads[role] = load_json(path)

    manifest = payloads["audit_package_manifest"]
    operations = payloads["reconciled_operations"]
    source_sha = str(manifest.get("source_sha256") or "")
    if not source_sha:
        raise ValueError("audit_package_manifest.json has no source_sha256")
    for role in (
        "context",
        "source_model",
        "operational_review",
        "configuration_review",
        "architecture_review",
        "technical_code_findings",
        "reconciled_operations",
        "future_state_gate",
    ):
        candidate = str(payloads[role].get("source_sha256") or "")
        if not candidate:
            raise ValueError(f"{role} has no source SHA-256")
        if candidate != source_sha:
            raise ValueError(f"{role} does not match the locked source SHA-256")
    completion = payloads["completion_gate"]
    if completion.get("kind") != "gtm_three_run_completion_gate":
        raise ValueError("completion_gate has an invalid kind")
    expected_runs = [
        "operational_sanitation",
        "configuration_correctness",
        "business_architecture",
    ]
    if as_list(completion.get("three_required_runs")) != expected_runs:
        raise ValueError("completion_gate does not attest the three required runs")
    if completion.get("completion_mode") != "audit_and_cleanup_plan":
        raise ValueError("completion_gate has the wrong completion mode")
    completion_future = completion.get("future_state")
    if not isinstance(completion_future, dict):
        raise ValueError("completion_gate has no bound future-state result")
    if completion_future.get("status") != "pass":
        raise ValueError("completion_gate future-state result is not pass")
    nested_completion_source = str(
        completion_future.get("source_sha256") or ""
    )
    if not nested_completion_source:
        raise ValueError("completion_gate future-state result has no source SHA-256")
    if nested_completion_source != source_sha:
        raise ValueError(
            "completion_gate future-state result does not match the locked source"
        )
    completion_source = str(completion.get("source_sha256") or "")
    if completion_source and completion_source != source_sha:
        raise ValueError(
            "completion_gate does not match the locked source SHA-256"
        )

    if str(operations.get("plan_status") or "") != "complete":
        raise ValueError("The reconciled operation plan is not complete")
    if str(payloads["future_state_gate"].get("status") or "") != "pass":
        raise ValueError("The future-state gate is not pass")
    if str(payloads["completion_gate"].get("status") or "") != "pass":
        raise ValueError("The three-run completion gate is not pass")
    if str(manifest.get("status") or "") not in {
        "ready_for_semantic_review",
        "complete",
        "pass",
    }:
        raise ValueError("The audit package manifest is not a completed/usable package")

    standard_hashes = workbook_sheet_hashes(paths["canonical_workbook"])
    if list(standard_hashes) != ORIGINAL_SHEETS:
        raise ValueError(
            "The canonical workbook must contain exactly the eight canonical sheets "
            "before readability transformation"
        )
    payloads["paths"] = paths
    payloads["source_sha256"] = source_sha
    payloads["standard_sheet_hashes"] = standard_hashes
    return payloads


def object_label(object_key: str, catalog: dict[str, Any]) -> str:
    key = safe_text(object_key)
    name = safe_text((catalog.get(object_key) or {}).get("object_name"))
    return f"{key} — {name}" if name and name != key else key


def friendly_value(value: Any, limit: int = 120) -> str:
    text = safe_text(value)
    if len(text) <= limit:
        return text
    return (
        text[: max(1, limit - 1)].rstrip()
        + "… [full value in cell note]"
    )


def field_label(json_path: Any) -> str:
    text = str(json_path or "")
    if not text:
        return "configured field"
    parts = re.findall(r"[A-Za-z_][A-Za-z0-9_]*|\[\d+\]", text)
    if not parts:
        return "configured field"
    return ".".join(parts[-3:])


def creation_object_key(creation: dict[str, Any]) -> str:
    layer = str(creation.get("layer") or "object")
    payload = creation.get("object") or {}
    field = OBJECT_ID_FIELDS.get(layer)
    identifier = payload.get(field) if field else None
    identifier = identifier or payload.get("name") or "new"
    return f"{layer}:{identifier}"


def creation_label(creation: dict[str, Any], catalog: dict[str, Any]) -> str:
    key = creation_object_key(creation)
    payload = creation.get("object") or {}
    name = safe_text(payload.get("name"))
    if name:
        return f"{key} — {name}"
    return object_label(key, catalog)


def structured_action_note(operation: dict[str, Any]) -> str:
    payload = {
        field: operation.get(field)
        for field in ACTION_MUTATION_FIELDS
    }
    note = safe_text(
        "Authoritative structured mutation for "
        f"{operation.get('operation_id') or operation.get('operation_key')}:\n"
        + json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
        )
    )
    if len(note) > MAX_ACTION_NOTE_TEXT:
        raise ValueError(
            f"{operation.get('operation_id') or operation.get('operation_key')} "
            "structured mutation is too long for a lossless analyst-workbook note"
        )
    return note


def deterministic_action_text(operation: dict[str, Any], catalog: dict[str, Any]) -> str:
    clauses: list[str] = []
    canonical = str(operation.get("canonical_object_key") or "")
    remaps = as_list(operation.get("remaps"))
    deletions = as_list(operation.get("deletions"))
    if canonical and (remaps or deletions):
        clauses.append(f"Keep {object_label(canonical, catalog)}")

    for creation in as_list(operation.get("creations")):
        clauses.append(f"Create {creation_label(creation, catalog)}")

    for addition in as_list(operation.get("additions")):
        key = str(addition.get("object_key") or "")
        mode = str(addition.get("mode") or "add")
        value = friendly_value(addition.get("value"))
        clauses.append(
            f"{mode.capitalize()} {value} in {field_label(addition.get('json_path'))} "
            f"on {object_label(key, catalog)}"
        )

    for change in as_list(operation.get("changes")):
        key = str(change.get("object_key") or "")
        clauses.append(
            f"On {object_label(key, catalog)}, change "
            f"{field_label(change.get('json_path'))} from "
            f"{friendly_value(change.get('before'))} to "
            f"{friendly_value(change.get('after'))}"
        )

    for remap in remaps:
        source = str(remap.get("from_object_key") or "")
        target = str(remap.get("to_object_key") or "")
        consumers = [
            object_label(str(value), catalog)
            for value in as_list(remap.get("consumer_object_keys"))
        ]
        consumer_text = ", ".join(consumers) or "the listed consumers"
        clauses.append(
            f"Repoint {consumer_text} from {object_label(source, catalog)} "
            f"to {object_label(target, catalog)}"
        )

    for rename in as_list(operation.get("renames")):
        key = str(rename.get("object_key") or "")
        clauses.append(
            f"Rename {object_label(key, catalog)} from "
            f"{friendly_value(rename.get('before'))} to "
            f"{friendly_value(rename.get('after'))}"
        )

    for deletion in deletions:
        key = str(deletion.get("object_key") or "")
        if key.startswith("builtInVariable:"):
            clauses.append(f"Disable/deselect {object_label(key, catalog)}")
        else:
            clauses.append(f"Delete {object_label(key, catalog)}")

    if not clauses:
        raise ValueError(
            f"{operation.get('operation_id') or operation.get('operation_key')} "
            "has no structured mutation to render"
        )
    return safe_text("; ".join(clauses) + ".")


def normalized_source_reference(value: Any, ledger_ids: set[str]) -> str:
    text = str(value or "")
    if text in ledger_ids:
        return text
    text = text.split(":operation:", 1)[0]
    return text if text in ledger_ids else ""


def operation_source_ids(operation: dict[str, Any], ledger_ids: set[str]) -> list[str]:
    values = {
        normalized_source_reference(value, ledger_ids)
        for value in as_list(operation.get("source_references"))
    }
    return sorted(value for value in values if value)


GENERIC_IMPACT_MARKERS = (
    "reduces maintenance risk without changing unrelated",
    "preserves affected measurement families",
    "improves maintainability",
    "reduces maintenance risk",
    "keeps the container clean",
    "see the evidence package",
)
GENERIC_IMPACT_RE = re.compile(
    r"\bpreserves?\s+\d+\s+affected measurement families\b|"
    r"\bsee (?:the )?evidence package\b",
    re.I,
)


def problem_specific_consequence(problem: str, object_text: str) -> str:
    subject = object_text or "the affected object"
    lowered = problem.casefold()
    rules = (
        (
            ("settimeout", "without an exported attempt"),
            f"{subject} can keep polling for the entire page lifetime after the expected dependency never becomes available.",
        ),
        (
            ("origin", "substring"),
            f"{subject} can accept a message from an unrelated origin whose URL merely contains the trusted text.",
        ),
        (
            ("postmessage", "datalayer", "payload"),
            f"{subject} can copy an unexpected cross-window payload into dataLayer and trigger downstream tags with unapproved fields.",
        ),
        (
            ("consent", "initialization"),
            f"{subject} can establish the default consent state after other tags have already evaluated their consent route.",
        ),
        (
            ("cookie duration", "day count"),
            f"{subject} retains the cookie for a different number of days than the setter and its callers declare.",
        ),
        (
            ("cookie", "secure", "samesite"),
            f"{subject} writes a cookie without the exported transport and cross-site attributes expected by the approved cookie policy.",
        ),
        (
            ("setdefaultvalue", ".includes"),
            f"{subject} can throw when the dataLayer value is absent, so the intended consent or routing value is not returned.",
        ),
        (
            ("literal string 'undefined'",),
            f"{subject} sends the text 'undefined' to its consumers instead of an explicit missing-value fallback.",
        ),
        (
            ("promises an hour", "date.now"),
            f"{subject} gives consumers an epoch-millisecond timestamp where its name promises an hour value.",
        ),
    )
    for markers, consequence in rules:
        if all(marker in lowered for marker in markers):
            return consequence
    return ""


def visible_consequence(operation: dict[str, Any], problem: str) -> str:
    """Prefer a literal consequence over reusable cleanup boilerplate."""

    supplied = safe_text(operation.get("why_it_matters"))
    lowered = supplied.casefold()
    if supplied and not any(
        marker in lowered for marker in GENERIC_IMPACT_MARKERS
    ) and not GENERIC_IMPACT_RE.search(supplied):
        return supplied
    object_text, _note = compact_objects(operation.get("affected_objects"), 220)
    literal = problem_specific_consequence(problem, object_text)
    if literal:
        return safe_text(literal)
    if as_list(operation.get("remaps")):
        return safe_text(
            f"Without the remap, {object_text or 'the listed consumers'} continue to "
            "depend on the legacy or incorrect object described in the problem."
        )
    if as_list(operation.get("changes")) or as_list(operation.get("additions")):
        return safe_text(
            f"Without this correction, {object_text or 'the affected object'} keeps "
            f"the wrong or incomplete configured behavior: {problem}"
        )
    if as_list(operation.get("deletions")):
        return safe_text(
            f"Leaving {object_text or 'the listed object'} in the container preserves "
            "a redundant, obsolete, or misleading configuration that can be reused or "
            "edited by mistake."
        )
    if as_list(operation.get("renames")):
        return safe_text(
            f"Without the rename, {object_text or 'the retained object'} remains easy "
            "to misidentify during maintenance even though its configured behavior is unchanged."
        )
    return safe_text(
        f"If left unchanged, the container keeps the exact problem described here: {problem}"
    )


def approval_and_preconditions(operation: dict[str, Any]) -> str:
    values = [
        safe_text(operation.get("preconditions")),
        safe_text(operation.get("blocker")),
    ]
    safety = operation.get("execution_safety") or {}
    approval = safety.get("approval") or {}
    scope = safe_text(approval.get("scope"))
    reasons = ", ".join(
        safe_text(value) for value in as_list(approval.get("reasons")) if safe_text(value)
    )
    if scope:
        values.append(f"Approval scope: {scope}" + (f" ({reasons})" if reasons else ""))
    decommission = safety.get("decommission") or {}
    if decommission.get("required"):
        values.append(
            "Quarantine first; deletion needs a separate post-observation approval."
        )
    return safe_text(" ".join(value for value in values if value))


def verification_and_rollback(operation: dict[str, Any]) -> str:
    verification = static_verification_text(operation)
    rollback = safe_text(operation.get("rollback"))
    return safe_text(
        f"Static readback: {verification}."
        + (f" Rollback: {rollback}" if rollback else "")
    )


def compact_finding(record: dict[str, Any], limit: int = 430) -> tuple[str, str]:
    full = safe_text(record.get("summary") or record.get("title") or record.get("decision_id"))
    decision_id = re.escape(str(record.get("decision_id") or ""))
    compact = re.sub(rf"^{decision_id}\s+records\s+", "", full, flags=re.I)
    compact = re.split(r"\s+Affected source object\(s\):", compact, maxsplit=1)[0]
    compact = normalize_space(compact)
    if len(compact) <= limit:
        return compact, full if compact != full else ""
    boundary = compact.rfind(". ", 0, limit)
    if boundary < limit // 2:
        boundary = compact.rfind(" ", 0, limit)
    boundary = boundary if boundary > 0 else limit
    return compact[:boundary].rstrip(" .;") + "…", full


def compact_objects(value: Any, limit: int = 360) -> tuple[str, str]:
    full = safe_text(value)
    if len(full) <= limit:
        return full, ""
    items = [item.strip() for item in full.split(";") if item.strip()]
    if len(items) > 3:
        visible = "; ".join(items[:3]) + f"; +{len(items) - 3} more — full scope in note"
        return safe_text(visible), full
    boundary = full.rfind(" ", 0, limit)
    boundary = boundary if boundary > limit // 2 else limit
    return full[:boundary].rstrip(" ;") + "… — full scope in note", full


def normalize_topic_text(value: Any) -> str:
    return normalize_space(value).casefold()


def decision_answer_class(record: dict[str, Any]) -> str:
    text = normalize_topic_text(
        " ".join(
            str(record.get(field) or "")
            for field in (
                "problem_type",
                "owner_question",
                "recommended_action",
            )
        )
    )
    classes = (
        ("consent", ("consent", "cmp", "storage")),
        ("folder_taxonomy", ("folder", "taxonomy", "classification")),
        ("naming", ("rename", "naming", "name convention")),
        ("retirement", ("remove", "delete", "retire", "decommission")),
        ("retention", ("keep", "retain", "exception")),
        ("canonical_route", ("canonical", "replace", "route", "source", "mapping")),
        ("ownership", ("owner", "ownership")),
        ("lifecycle", ("paused", "lifecycle", "rollback")),
    )
    return next(
        (name for name, markers in classes if any(marker in text for marker in markers)),
        "exact",
    )


def decision_group_key(record: dict[str, Any]) -> tuple[str, ...]:
    object_keys = tuple(
        sorted(
            str(value)
            for value in as_list(record.get("source_object_keys"))
            if str(value)
        )
    )
    answer_class = decision_answer_class(record)
    problem_type = normalize_topic_text(record.get("problem_type"))
    if object_keys and answer_class != "exact":
        return ("semantic", problem_type, answer_class, *object_keys)
    return (
        "exact",
        normalize_topic_text(record.get("owner_question")),
        normalize_topic_text(record.get("recommended_action")),
    )


def default_decision_topics(owner_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for record in owner_records:
        grouped[decision_group_key(record)].append(record)
    safe_groups: list[list[dict[str, Any]]] = []
    for key, records in grouped.items():
        source_lenses = [
            str(record.get("decision_id") or "").partition("-")[0]
            for record in records
        ]
        if key[0] != "semantic" or len(source_lenses) == len(set(source_lenses)):
            safe_groups.append(records)
            continue
        exact_groups: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            exact_groups[
                (
                    normalize_topic_text(record.get("owner_question")),
                    normalize_topic_text(record.get("recommended_action")),
                )
            ].append(record)
        safe_groups.extend(exact_groups.values())
    ordered = sorted(
        safe_groups,
        key=lambda rows: min(str(row.get("decision_id") or "") for row in rows),
    )
    topics = []
    for index, records in enumerate(ordered, start=1):
        first = records[0]
        topics.append(
            {
                "topic_id": f"D-{index:02d}",
                "title": safe_text(
                    first.get("problem_type") or first.get("title") or "Owner decision"
                ),
                "question": safe_text(first.get("owner_question")),
                "recommendation": safe_text(first.get("recommended_action")),
                "source_ids": sorted(str(row.get("decision_id")) for row in records),
            }
        )
    return topics


def decision_topics(
    owner_records: list[dict[str, Any]],
    editorial: dict[str, Any] | None,
    source_sha256: str,
) -> list[dict[str, Any]]:
    if editorial is None:
        return default_decision_topics(owner_records)
    if editorial.get("kind") not in {
        "gtm_readability_decision_topics",
        "gtm_workbook_readability_editorial",
    }:
        raise ValueError("Unsupported decision-topic artifact kind")
    artifact_sha = str(editorial.get("source_sha256") or "")
    if not artifact_sha:
        raise ValueError("Decision-topic artifact has no source SHA-256")
    if artifact_sha != source_sha256:
        raise ValueError("Decision-topic artifact does not match the locked source")

    owner_by_id = {str(row.get("decision_id")): row for row in owner_records}
    expected = set(owner_by_id)
    seen: set[str] = set()
    topics: list[dict[str, Any]] = []
    topic_ids: set[str] = set()
    for index, topic in enumerate(as_list(editorial.get("topics")), start=1):
        topic_id = str(topic.get("topic_id") or f"D-{index:02d}")
        if topic_id in topic_ids:
            raise ValueError(f"Duplicate decision topic ID: {topic_id}")
        topic_ids.add(topic_id)
        source_ids = [str(value) for value in as_list(topic.get("source_ids"))]
        if not source_ids or any(not value for value in source_ids):
            raise ValueError(f"{topic_id} has no source decision")
        internal_duplicates = sorted(
            {
                source_id
                for source_id in source_ids
                if source_ids.count(source_id) > 1
            }
        )
        unknown = sorted(set(source_ids) - expected)
        duplicate = sorted(set(source_ids) & seen)
        if unknown or duplicate or internal_duplicates:
            raise ValueError(
                f"{topic_id} has invalid source mapping; unknown={unknown}, "
                f"duplicate={sorted(set(duplicate + internal_duplicates))}"
            )
        seen.update(source_ids)
        first = owner_by_id[source_ids[0]]
        source_question = safe_text(first.get("owner_question"))
        source_recommendation = safe_text(first.get("recommended_action"))
        question = safe_text(topic.get("question") or source_question)
        recommendation = safe_text(
            topic.get("recommendation") or source_recommendation
        )
        if len(source_ids) == 1 and (
            question != source_question
            or recommendation != source_recommendation
        ):
            raise ValueError(
                f"{topic_id} cannot rewrite a single source question or recommendation"
            )
        if not question or not recommendation:
            raise ValueError(f"{topic_id} needs a question and recommendation")
        topics.append(
            {
                "topic_id": topic_id,
                "title": safe_text(
                    topic.get("title")
                    or first.get("problem_type")
                    or first.get("title")
                    or "Owner decision"
                ),
                "question": question,
                "recommendation": recommendation,
                "source_ids": source_ids,
            }
        )
    missing = sorted(expected - seen)
    if missing:
        raise ValueError(f"Decision-topic artifact omits owner decisions: {missing}")
    return topics


def normalized_identifier(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"^(?:_?pc_?|om_?|dlv?_?|gtm_?)+", "", text)
    return re.sub(r"[^a-z0-9]+", "", text)


def legacy_source_keys(technical: dict[str, Any], configuration: dict[str, Any]) -> set[str]:
    values: set[str] = set()

    def add_literal(value: Any) -> None:
        text = str(value or "").strip()
        if not text:
            return
        values.add(text)
        # Cookie comparisons commonly expose "name=value" while a dataLayer
        # variable exposes only the semantic key. Keep both as unproven
        # candidates so the analyst sees the possible modernization path.
        if "=" in text:
            name = text.split("=", 1)[0].strip()
            if name:
                values.add(name)

    for field in ("localStorage_use", "sessionStorage_use"):
        for item in as_list(technical.get(field)):
            add_literal(str(item).split(":", 1)[-1])
    for expression in as_list(technical.get("return_expressions")):
        text = str((expression or {}).get("expression") or "")
        for literal in re.findall(r"[\"']([^\"']{3,80})[\"']", text):
            add_literal(literal)
    for fact in as_list(configuration.get("source_facts")):
        if str(fact.get("json_path") or "").endswith(".parameter[0].value"):
            preview = str(fact.get("value_preview") or "")
            for literal in re.findall(r"[\"']([^\"']{3,80})[\"']", preview):
                add_literal(literal)
    return {value for value in values if value}


def code_haystack(technical: dict[str, Any], configuration: dict[str, Any]) -> str:
    values: list[str] = [
        str(technical.get("technical_current_behavior") or ""),
        str(technical.get("technical_plain_language_summary") or ""),
    ]
    values.extend(str(item) for item in as_list(technical.get("referenced_gtm_variables")))
    for expression in as_list(technical.get("return_expressions")):
        values.append(str((expression or {}).get("expression") or ""))
    for fact in as_list(configuration.get("source_facts")):
        values.append(str(fact.get("value_preview") or ""))
    for block in as_list(configuration.get("code_behavior_blocks")):
        for field in ("purpose", "inputs", "outputs", "side_effects"):
            values.append(str(block.get(field) or ""))
    return "\n".join(values).casefold()


def data_layer_candidates(
    technical: dict[str, Any],
    configuration: dict[str, Any],
    variable_sources: list[dict[str, Any]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    haystack = code_haystack(technical, configuration)
    referenced_names = {
        str(value or "").strip().strip("{}").strip().casefold()
        for value in as_list(technical.get("referenced_gtm_variables"))
        if str(value or "").strip()
    }
    legacy_keys = {
        normalized_identifier(value)
        for value in legacy_source_keys(technical, configuration)
    }
    legacy_keys = {
        value
        for value in legacy_keys
        if len(value) >= 6 and value not in IDENTIFIER_STOP_KEYS
    }
    direct: list[dict[str, str]] = []
    candidates: list[dict[str, str]] = []
    seen_direct: set[str] = set()
    seen_candidate: set[str] = set()
    for variable in variable_sources:
        path = str(variable.get("data_layer_path") or "")
        if not path:
            continue
        key = f"variable:{variable.get('object_id')}"
        record = {
            "object_key": key,
            "object_name": str(variable.get("object_name") or ""),
            "data_layer_path": path,
        }
        normalized = normalized_identifier(path)
        object_name = str(variable.get("object_name") or "").strip().casefold()
        exact_gtm_reference = bool(object_name and object_name in referenced_names)
        explicit_path_read = bool(
            technical.get("dataLayer_reads")
            and len(path) >= 5
            and re.search(
                rf"(?<![A-Za-z0-9_]){re.escape(path.casefold())}(?![A-Za-z0-9_])",
                haystack,
            )
        )
        if (exact_gtm_reference or explicit_path_read) and key not in seen_direct:
            direct.append(record)
            seen_direct.add(key)
        if normalized in legacy_keys and key not in seen_candidate:
            candidates.append(record)
            seen_candidate.add(key)
    return direct, candidates


def deletion_operation_map(operations: list[dict[str, Any]]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = defaultdict(list)
    for operation in operations:
        operation_id = str(operation.get("operation_id") or "")
        for deletion in as_list(operation.get("deletions")):
            key = str(deletion.get("object_key") or "")
            if operation_id and key:
                mapping[key].append(operation_id)
    return {key: sorted(set(values)) for key, values in mapping.items()}


def custom_html_role(technical: dict[str, Any]) -> str:
    length = int(technical.get("code_length") or 0)
    parts = [f"{length:,} characters of custom JavaScript"]
    if technical.get("dataLayer_reads"):
        parts.append("reads dataLayer")
    if technical.get("dataLayer_pushes_or_writes"):
        parts.append("writes dataLayer")
    if technical.get("localStorage_use") or technical.get("sessionStorage_use"):
        parts.append("uses browser storage")
    if technical.get("cookies_read_written"):
        parts.append("uses cookies")
    if technical.get("dom_reads_writes"):
        parts.append("reads or changes the DOM")
    if technical.get("event_listeners"):
        parts.append("registers listeners")
    if technical.get("external_scripts_loaded"):
        parts.append("loads an external script")
    if technical.get("network_calls"):
        parts.append("performs a network call")
    return safe_text("; ".join(parts) + ".")


def custom_execution_context(
    technical: dict[str, Any], configuration: dict[str, Any], labels: dict[str, Any]
) -> str:
    state = (
        labels["states"]["paused"]
        if configuration.get("paused")
        else labels["states"]["active"]
    )
    context: list[str] = [state]
    if "deprecated" in str(technical.get("object_name") or "").casefold():
        context.append(labels["states"]["deprecated"])
    execution = normalize_space(configuration.get("execution_logic"))
    if execution:
        context.append("Configured execution: " + execution[:360])
    listener = technical.get("listener_lifecycle") or {}
    if listener.get("registration_count"):
        context.append(
            "Listener lifecycle: "
            + (
                "guarded"
                if listener.get("has_stable_registration_guard")
                else "registration guard not exported"
            )
        )
    if listener.get("window_load_listener"):
        context.append(
            "load timing has readyState branch"
            if listener.get("ready_state_branch")
            else "load timing has no readyState branch"
        )
    timers = technical.get("timer_lifecycle") or {}
    if timers.get("set_interval"):
        context.append(
            "interval lifecycle includes clearInterval"
            if timers.get("clear_interval")
            else "interval lifecycle has no exported clearInterval"
        )
    if timers.get("set_timeout"):
        context.append(
            "one-shot timer with exported cancellation"
            if timers.get("clear_timeout")
            else "one-shot timer; no cancellation path exported"
        )
    observer = technical.get("observer_lifecycle") or {}
    if observer.get("mutation_observer"):
        context.append(
            "MutationObserver has disconnect lifecycle"
            if observer.get("disconnect")
            else "MutationObserver has no exported disconnect lifecycle"
        )
    return safe_text(". ".join(context) + ".")


def custom_technical_health(
    technical: dict[str, Any], labels: dict[str, Any]
) -> str:
    findings = [
        str(value)
        for field in (
            "technical_code_security_findings",
            "technical_code_health_findings",
            "technical_code_optimization_findings",
        )
        for value in as_list(technical.get(field))
        if str(value)
    ]
    visible = " ".join(findings[:4])
    if len(findings) > 4:
        visible += f" +{len(findings) - 4} additional source-locked findings."
    if not visible:
        visible = "No static technical defect detected in the exported code."
    return safe_text(
        labels["custom"]["health"].format(
            status=technical.get("technical_code_health_status")
            or "static review complete",
            disposition=technical.get("technical_disposition")
            or technical.get("technical_action_candidate")
            or "keep",
            findings=visible,
        )
    )


def operation_object_keys(operation: dict[str, Any]) -> set[str]:
    keys = {
        str(value)
        for field in ("affected_object_keys", "source_object_keys")
        for value in as_list(operation.get(field))
        if str(value)
    }
    for field in ("changes", "additions", "deletions", "renames"):
        keys.update(
            str(item.get("object_key") or "")
            for item in as_list(operation.get(field))
            if isinstance(item, dict) and str(item.get("object_key") or "")
        )
    for remap in as_list(operation.get("remaps")):
        if not isinstance(remap, dict):
            continue
        keys.update(
            str(value)
            for value in (
                remap.get("from_object_key"),
                remap.get("to_object_key"),
                *as_list(remap.get("consumer_object_keys")),
            )
            if str(value)
        )
    return keys


def operation_is_maintenance_only(operation: dict[str, Any]) -> bool:
    if as_list(operation.get("creations")) or as_list(operation.get("remaps")):
        return False
    deletions = [
        str(item.get("object_key") or "")
        for item in as_list(operation.get("deletions"))
        if isinstance(item, dict) and str(item.get("object_key") or "")
    ]
    if deletions:
        reachability = str(
            (operation.get("priority_basis") or {}).get("active_reachability") or ""
        )
        metadata_deletions = all(
            key.startswith(("folder:", "builtInVariable:")) for key in deletions
        )
        if not metadata_deletions and reachability not in {
            "inactive_or_unreferenced",
            "metadata_only",
            "paused_only",
        }:
            return False
    behavior_markers = re.compile(
        r"firingTriggerId|blockingTriggerId|setupTag|teardownTag|parameter|consent|"
        r"schedule|paused|type$",
        re.I,
    )
    return not any(
        behavior_markers.search(str(item.get("json_path") or ""))
        for field in ("changes", "additions")
        for item in as_list(operation.get(field))
        if isinstance(item, dict)
    )


def locale(language: str) -> dict[str, Any]:
    if language not in LOCALES:
        raise ValueError(f"Unsupported report language: {language}")
    return LOCALES[language]


def input_hash_manifest(paths: dict[str, Path]) -> dict[str, Any]:
    return {
        role: {"name": path.name, "sha256": sha256_file(path)}
        for role, path in sorted(paths.items())
    }


def build_model(inputs: dict[str, Any], language: str) -> dict[str, Any]:
    labels = locale(language)
    operations_payload = inputs["reconciled_operations"]
    operations = as_list(operations_payload.get("operations"))
    ledger = as_list(operations_payload.get("decision_ledger"))
    ledger_ids = {str(row.get("decision_id") or "") for row in ledger}
    if "" in ledger_ids or len(ledger_ids) != len(ledger):
        raise ValueError("Decision ledger IDs must be nonblank and unique")
    unknown_dispositions = sorted(
        {
            str(row.get("disposition") or "")
            for row in ledger
            if str(row.get("disposition") or "") not in DISPOSITION_ORDER
        }
    )
    if unknown_dispositions:
        raise ValueError(
            "Decision ledger contains unsupported dispositions: "
            + ", ".join(unknown_dispositions)
        )
    operation_by_id = {
        str(row.get("operation_id") or ""): row for row in operations
    }
    if "" in operation_by_id or len(operation_by_id) != len(operations):
        raise ValueError("Operation IDs must be nonblank and unique")
    catalog = operations_payload.get("object_catalog") or {}
    preservation_families = as_list(
        (operations_payload.get("measurement_preservation") or {}).get("families")
    )
    family_label_by_id = {
        str(family.get("family_id") or ""): safe_text(
            family.get("family_label") or family.get("family_id")
        )
        for family in preservation_families
        if str(family.get("family_id") or "")
    }

    owner_records = [
        row for row in ledger if row.get("disposition") == "owner_decision_needed"
    ]
    topics = decision_topics(
        owner_records,
        inputs.get("decision_topics"),
        inputs["source_sha256"],
    )
    topic_by_source = {
        source_id: topic["topic_id"]
        for topic in topics
        for source_id in topic["source_ids"]
    }

    operation_sources = {
        operation_id: operation_source_ids(operation, ledger_ids)
        for operation_id, operation in operation_by_id.items()
    }
    audit_sections: list[dict[str, Any]] = []
    audit_rows_by_id: dict[str, dict[str, Any]] = {}
    for disposition in DISPOSITION_ORDER:
        records = [row for row in ledger if row.get("disposition") == disposition]
        if not records:
            continue
        records.sort(
            key=lambda row: (
                min(
                    (
                        PRIORITY_ORDER.get(
                            str(operation_by_id[operation_id].get("priority") or ""),
                            9,
                        )
                        for operation_id in as_list(row.get("compiled_operation_ids"))
                        if operation_id in operation_by_id
                    ),
                    default=9,
                ),
                str(row.get("area") or ""),
                str(row.get("decision_id") or ""),
            )
        )
        section_rows = []
        for record in records:
            decision_id = str(record.get("decision_id"))
            operation_ids = [
                str(value)
                for value in as_list(record.get("compiled_operation_ids"))
                if str(value) in operation_by_id
            ]
            if disposition == "cleanup_operation":
                if not operation_ids:
                    raise ValueError(f"{decision_id} has cleanup disposition without operation")
                outcome = labels["outcomes"][disposition].format(
                    ids=", ".join(operation_ids)
                )
            elif disposition == "owner_decision_needed":
                topic_id = topic_by_source.get(decision_id)
                if not topic_id:
                    raise ValueError(f"{decision_id} has no decision topic")
                outcome = labels["outcomes"][disposition].format(topic=topic_id)
            else:
                outcome = labels["outcomes"][disposition]

            priority_values = [
                str(operation_by_id[operation_id].get("priority") or "")
                for operation_id in operation_ids
            ]
            priority = (
                min(
                    priority_values,
                    key=lambda value: PRIORITY_ORDER.get(value, 9),
                )
                if priority_values
                else ""
            )
            finding, finding_note = compact_finding(record)
            objects, objects_note = compact_objects(record.get("affected_objects"))
            values = [
                decision_id,
                safe_text(record.get("area") or record.get("problem_type")),
                objects,
                finding,
                safe_text(outcome),
                safe_text(priority),
            ]
            row = {
                "kind": "audit",
                "id": decision_id,
                "disposition": disposition,
                "operation_ids": operation_ids,
                "topic_id": topic_by_source.get(decision_id, ""),
                "source_object_keys": [
                    str(value) for value in as_list(record.get("source_object_keys"))
                ],
                "values": values,
                "notes": {2: objects_note, 3: finding_note},
                "record": record,
            }
            section_rows.append(row)
            audit_rows_by_id[decision_id] = row
        audit_sections.append(
            {
                "disposition": disposition,
                "label": labels["sections"][disposition],
                "collapsed": disposition
                in {
                    "keep",
                    "documented_exception",
                    "container_evidence_limit",
                    "not_applicable",
                },
                "rows": section_rows,
            }
        )

    action_rows = []
    common_validation = ""
    for operation in sorted(
        operations,
        key=lambda row: (
            int(row.get("execution_order") or 0),
            str(row.get("operation_id") or ""),
        ),
    ):
        operation_id = str(operation.get("operation_id"))
        objects, objects_note = compact_objects(operation.get("affected_objects"))
        problem = operation_problem_text(operation, catalog)
        if len(normalize_space(problem)) < 24 or normalize_space(problem).startswith("Impact:"):
            problem = safe_text(
                f"{objects or operation_id} is covered by this source-evidenced cleanup "
                "operation; the exact structured mutation below defines the correction."
            )
        consequence = visible_consequence(operation, problem)
        exact_change = operation_action_text(operation, catalog)
        if len(normalize_space(exact_change)) < 12:
            exact_change = deterministic_action_text(operation, catalog)
        preconditions = approval_and_preconditions(operation)
        validation = verification_and_rollback(operation)
        action_note = structured_action_note(operation)
        action_rows.append(
            {
                "kind": "action",
                "id": operation_id,
                "source_ids": operation_sources[operation_id],
                "values": [
                    safe_text(f"{operation.get('execution_order')} · {operation_id}"),
                    safe_text(operation.get("priority")),
                    objects,
                    problem,
                    consequence,
                    exact_change,
                    preconditions,
                    validation,
                ],
                "notes": {2: objects_note, 5: action_note},
                "operation": operation,
            }
        )

    owner_by_id = {
        str(record.get("decision_id")): record for record in owner_records
    }

    def decision_family_labels(records: list[dict[str, Any]]) -> list[str]:
        source_keys = {
            str(value)
            for record in records
            for value in as_list(record.get("source_object_keys"))
            if str(value)
        }
        operation_ids = {
            str(value)
            for record in records
            for value in as_list(record.get("compiled_operation_ids"))
            if str(value)
        }
        family_ids = sorted(
            {
                str(family.get("family_id") or "")
                for family in preservation_families
                if str(family.get("family_id") or "")
                and (
                    source_keys
                    & {
                        str(value)
                        for value in as_list(family.get("source_object_keys"))
                        if str(value)
                    }
                    or operation_ids
                    & {
                        str(value)
                        for value in as_list(family.get("related_operation_ids"))
                        if str(value)
                    }
                )
            }
        )
        return [
            (
                f"{family_label_by_id[family_id]} ({family_id})"
                if family_label_by_id.get(family_id) not in {"", family_id}
                else family_id
            )
            for family_id in family_ids
        ]

    def decision_unlock(topic: dict[str, Any], records: list[dict[str, Any]]) -> str:
        object_keys = {
            str(value)
            for record in records
            for value in as_list(record.get("source_object_keys"))
            if str(value)
        }
        if not object_keys:
            object_keys = {
                item.strip()
                for record in records
                for item in str(record.get("affected_objects") or "").split(";")
                if item.strip()
            }
        scope = ", ".join(sorted(object_keys)) or "the named source condition"
        source_ids = ", ".join(
            sorted(str(record.get("decision_id") or "") for record in records)
        )
        return safe_text(
            f"Why this answer is needed: GTM cannot safely change {scope} until the "
            f"owner answers {source_ids}. Once answered, follow this concrete next "
            f"step only for those objects: {topic['recommendation']}"
        )

    decision_rows = []
    for topic in topics:
        source_ids = list(topic["source_ids"])
        source_records = [owner_by_id[source_id] for source_id in source_ids]
        if len(source_ids) == 1:
            source = source_records[0]
            objects, objects_note = compact_objects(source.get("affected_objects"))
            family_labels = decision_family_labels(source_records)
            decision_rows.append(
                {
                    "kind": "decision_single",
                    "topic_id": topic["topic_id"],
                    "source_ids": source_ids,
                    "source_object_keys": [
                        str(value) for value in as_list(source.get("source_object_keys"))
                    ],
                    "values": [
                        safe_text(
                            f"{topic['topic_id']} — {topic['title']} ({source_ids[0]})"
                        ),
                        topic["question"],
                        topic["recommendation"],
                        objects,
                        ", ".join(family_labels) or "None linked by container evidence",
                        decision_unlock(topic, source_records),
                    ],
                    "notes": {3: objects_note},
                }
            )
            continue
        decision_rows.append(
            {
                "kind": "decision_parent",
                "topic_id": topic["topic_id"],
                "source_ids": source_ids,
                "values": [
                    safe_text(f"{topic['topic_id']} — {topic['title']}"),
                    topic["question"],
                    topic["recommendation"],
                    labels["source_items"].format(count=len(source_ids)),
                    ", ".join(decision_family_labels(source_records))
                    or "None linked by container evidence",
                    decision_unlock(topic, source_records),
                ],
                "notes": {},
            }
        )
        for source in source_records:
            source_id = str(source.get("decision_id"))
            objects, objects_note = compact_objects(source.get("affected_objects"))
            source_question = safe_text(source.get("owner_question"))
            source_recommendation = safe_text(source.get("recommended_action"))
            decision_rows.append(
                {
                    "kind": "decision_child",
                    "topic_id": topic["topic_id"],
                    "source_ids": [source_id],
                    "source_object_keys": [
                        str(value) for value in as_list(source.get("source_object_keys"))
                    ],
                    "values": [
                        safe_text(f"↳ {source_id}"),
                        source_question if source_question != topic["question"] else "",
                        (
                            source_recommendation
                            if source_recommendation != topic["recommendation"]
                            else ""
                        ),
                        objects,
                        ", ".join(decision_family_labels([source]))
                        or "None linked by container evidence",
                        decision_unlock(topic, [source]),
                    ],
                    "notes": {
                        0: safe_text(source.get("summary")),
                        3: objects_note,
                    },
                }
            )

    technical_rows = [
        row
        for row in as_list(inputs["technical_code_findings"].get("rows"))
        if row.get("layer") == "tag" and row.get("type") == "html"
    ]
    configuration_by_key = {
        str(row.get("object_key")): row
        for row in as_list(inputs["configuration_review"].get("rows"))
    }
    missing_configuration = sorted(
        f"tag:{row.get('object_id')}"
        for row in technical_rows
        if f"tag:{row.get('object_id')}" not in configuration_by_key
    )
    if missing_configuration:
        raise ValueError(
            "Custom HTML tags are missing configuration-review records: "
            + ", ".join(missing_configuration)
        )
    variable_sources = as_list(inputs["source_model"].get("variable_sources"))
    delete_map = deletion_operation_map(operations)
    operation_ids_by_object: dict[str, list[str]] = defaultdict(list)
    for operation in operations:
        operation_id = str(operation.get("operation_id") or "")
        for key in operation_object_keys(operation):
            if operation_id:
                operation_ids_by_object[key].append(operation_id)
    identical_code_tags: dict[str, list[str]] = defaultdict(list)
    for technical in technical_rows:
        code_hash = str(technical.get("code_hash") or "")
        if code_hash:
            identical_code_tags[code_hash].append(
                f"tag:{technical.get('object_id')} — {technical.get('object_name')}"
            )
    owner_topics_by_object: dict[str, set[str]] = defaultdict(set)
    for source_id, topic_id in topic_by_source.items():
        for key in as_list(owner_by_id[source_id].get("source_object_keys")):
            owner_topics_by_object[str(key)].add(topic_id)

    custom_rows = []
    custom_conflicts: dict[str, Any] = {}
    for technical in sorted(technical_rows, key=lambda row: str(row.get("object_id") or "")):
        tag_key = f"tag:{technical.get('object_id')}"
        configuration = configuration_by_key.get(tag_key) or {}
        direct, candidates = data_layer_candidates(
            technical,
            configuration,
            variable_sources,
        )
        replacement_parts = []
        code_length = int(technical.get("code_length") or 0)
        if code_length >= 4000:
            replacement_parts.append(
                labels["custom"]["long"].format(length=f"{code_length:,}")
            )
        legacy_acquisition = bool(
            technical.get("cookies_read_written")
            or technical.get("localStorage_use")
            or technical.get("sessionStorage_use")
            or technical.get("dom_selector_reads")
        )
        if direct:
            replacement_parts.append(
                labels["custom"]["direct"].format(
                    sources=", ".join(
                        f"{row['object_key']} ({safe_text(row['data_layer_path'])})"
                        for row in direct
                    )
                )
            )
        if candidates:
            candidate_labels = [
                f"{row['object_key']} ({safe_text(row['data_layer_path'])})"
                for row in candidates
            ]
            replacement_parts.append(
                labels["custom"]["candidate"].format(
                    sources=", ".join(candidate_labels)
                )
            )
        elif legacy_acquisition:
            replacement_parts.append(labels["custom"]["legacy"])
        if not direct and not candidates and not legacy_acquisition:
            if technical.get("dataLayer_pushes_or_writes"):
                replacement_parts.append(labels["custom"]["producer"])
                replacement_parts.append(
                    "Site-side data production candidate only when the application can own "
                    "the same event, fields, type, timing, consent state, and consumers."
                )
            elif technical.get("external_scripts_loaded"):
                replacement_parts.append(labels["custom"]["loader"])
            else:
                replacement_parts.append(labels["custom"]["none"])

        native_candidates = []
        if technical.get("manual_gtag_calls"):
            native_candidates.append("native Google tag/event tag")
        if technical.get("external_scripts_loaded"):
            native_candidates.append("maintained vendor template, if installed and equivalent")
        if any(
            "small helper variable" in str(value).casefold()
            for value in as_list(technical.get("technical_code_optimization_findings"))
        ):
            native_candidates.append("built-in variable, lookup table, or regex table")
        if native_candidates:
            replacement_parts.append(
                labels["custom"]["candidate_native"].format(
                    candidates=", ".join(native_candidates)
                )
            )
        code_peers = identical_code_tags.get(str(technical.get("code_hash") or ""), [])
        if len(code_peers) > 1:
            replacement_parts.append(
                labels["custom"]["candidate_duplicate"].format(
                    tags=", ".join(code_peers)
                )
            )

        conflicts = {
            row["object_key"]: delete_map[row["object_key"]]
            for row in candidates
            if row["object_key"] in delete_map
        }
        if conflicts:
            conflict_objects = ", ".join(sorted(conflicts))
            conflict_operations = ", ".join(
                sorted({op_id for values in conflicts.values() for op_id in values})
            )
            replacement_parts.append(
                labels["custom"]["conflict"].format(
                    objects=conflict_objects,
                    operations=conflict_operations,
                )
            )
            custom_conflicts[tag_key] = conflicts
        related_topics = sorted(owner_topics_by_object.get(tag_key, set()))
        tag_operations = sorted(set(operation_ids_by_object.get(tag_key, [])))
        if tag_operations:
            simplest_target = labels["custom"]["operation_target"].format(
                operations=", ".join(tag_operations)
            )
        elif related_topics:
            simplest_target = labels["custom"]["owner_target"].format(
                topics=", ".join(related_topics)
            )
        elif technical.get("technical_disposition") == "keep":
            simplest_target = labels["custom"]["keep_target"]
        else:
            simplest_target = safe_text(
                technical.get("technical_expected_clean_state")
                or "Retain the smallest source-proven implementation after exact equivalence review."
            )
        exact_parts = []
        if tag_operations:
            exact_parts.append(
                "Planned operation(s): "
                + "; ".join(
                    f"{operation_id} — "
                    + safe_text(
                        operation_by_id[operation_id].get("title")
                        or operation_by_id[operation_id].get("exact_proposed_action")
                    )
                    for operation_id in tag_operations
                    if operation_id in operation_by_id
                )
            )
        technical_action = safe_text(technical.get("technical_exact_proposed_action"))
        if technical_action:
            exact_parts.append(technical_action)
        if related_topics:
            exact_parts.append(
                labels["custom"]["decisions"].format(
                    topics=", ".join(related_topics)
                )
            )
        if not exact_parts:
            exact_parts.append(labels["custom"]["keep_target"])
        custom_rows.append(
            {
                "kind": "custom_html",
                "id": tag_key,
                "configuration_id": str(configuration.get("review_id") or ""),
                "related_topics": related_topics,
                "conflicts": conflicts,
                "values": [
                    safe_text(f"{tag_key} — {technical.get('object_name')}"),
                    custom_execution_context(technical, configuration, labels),
                    custom_html_role(technical),
                    custom_technical_health(technical, labels),
                    safe_text(" ".join(replacement_parts)),
                    safe_text(simplest_target),
                    safe_text(" ".join(exact_parts)),
                ],
                "notes": {
                    3: safe_text(
                        {
                            "health": technical.get("technical_code_health_findings") or [],
                            "security": technical.get("technical_code_security_findings") or [],
                            "optimization": technical.get("technical_code_optimization_findings") or [],
                            "cookie_writes": technical.get("cookie_writes") or [],
                            "listener_lifecycle": technical.get("listener_lifecycle") or {},
                            "timer_lifecycle": technical.get("timer_lifecycle") or {},
                            "observer_lifecycle": technical.get("observer_lifecycle") or {},
                        }
                    )
                },
            }
        )

    disposition_counts = defaultdict(int)
    for row in ledger:
        disposition_counts[str(row.get("disposition") or "")] += 1
    priority_counts = {
        priority: sum(1 for row in operations if row.get("priority") == priority)
        for priority in PRIORITY_ORDER
    }
    priority_first_actions = sorted(
        action_rows,
        key=lambda row: (
            PRIORITY_ORDER.get(str(row["operation"].get("priority") or ""), 9),
            int(row["operation"].get("execution_order") or 0),
            row["id"],
        ),
    )
    first_actions = " | ".join(
        safe_text(
            f"{row['id']} [{row['operation'].get('priority') or ''}]: "
            f"{row['operation'].get('title') or row['values'][5]}"
        )[:260]
        for row in priority_first_actions[:3]
    ) or labels["counts"]["no_actions"]
    bulk_operations = sum(
        1
        for operation in operations
        if (
            ((operation.get("execution_safety") or {}).get("approval") or {}).get(
                "scope"
            )
            == "bulk_eligible_exact_low_risk_bundle"
        )
    )
    simulated_activation = (
        inputs["future_state_gate"].get("configured_activation_risk") or {}
    )
    if isinstance(simulated_activation, dict) and "flag" in simulated_activation:
        confirmed_candidates = {
            str(value)
            for value in as_list(simulated_activation.get("candidate_operation_ids"))
            if str(value)
        }
        activation_operations = (
            sum(
                1
                for operation in operations
                if str(operation.get("operation_id") or "") in confirmed_candidates
            )
            if confirmed_candidates or not simulated_activation.get("flag")
            else sum(
                1
                for operation in operations
                if bool(
                    (
                        (operation.get("execution_safety") or {}).get(
                            "configured_activation_risk"
                        )
                        or {}
                    ).get("flag")
                )
            )
        )
    else:
        activation_operations = sum(
            1
            for operation in operations
            if bool(
                (
                    (operation.get("execution_safety") or {}).get(
                        "configured_activation_risk"
                    )
                    or {}
                ).get("flag")
            )
        )
    maintenance_operations = sum(operation_is_maintenance_only(row) for row in operations)
    behavior_operations = len(operations) - maintenance_operations
    remaining_records = len(ledger) - disposition_counts["cleanup_operation"]
    preservation_counts = defaultdict(int)
    for family in preservation_families:
        preservation_counts[str(family.get("preservation_status") or "")] += 1
    measurement_summary = (
        labels["counts"]["measurement"].format(
            total=len(preservation_families),
            changed=preservation_counts["planned_change"],
            retained=preservation_counts["retained_unchanged"],
            reviewed=preservation_counts["reviewed_no_operation"],
            owner=preservation_counts["owner_confirmation_required"],
            boundary=preservation_counts["container_evidence_boundary"],
        )
        if preservation_families
        else labels["counts"]["no_families"]
    )
    model = {
        "language": language,
        "headers": labels["headers"],
        "labels": labels,
        "audit_sections": audit_sections,
        "audit_rows_by_id": audit_rows_by_id,
        "action_rows": action_rows,
        "common_validation": common_validation,
        "decision_rows": decision_rows,
        "decision_topics": topics,
        "topic_by_source": topic_by_source,
        "custom_rows": custom_rows,
        "custom_conflicts": custom_conflicts,
        "counts": {
            "audit_records": len(ledger),
            "operations": len(operations),
            "owner_source_records": len(owner_records),
            "decision_topics": len(topics),
            "custom_html_tags": len(custom_rows),
            "retained": disposition_counts["keep"],
            "documented_exceptions": disposition_counts["documented_exception"],
            "evidence_limits": disposition_counts["container_evidence_limit"],
            "priority": priority_counts,
            "bulk_operations": bulk_operations,
            "individual_operations": len(operations) - bulk_operations,
            "activation_operations": activation_operations,
            "maintenance_operations": maintenance_operations,
            "behavior_operations": behavior_operations,
            "remaining_records": remaining_records,
        },
        "operation_by_id": operation_by_id,
        "ledger_by_id": {
            str(row.get("decision_id")): row for row in ledger
        },
        "projected_object_counts": operations_payload.get("projected_object_counts") or {},
        "first_actions": first_actions,
        "measurement_summary": measurement_summary,
    }
    return model


def apply_header_style(sheet: Any, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=GRID_COLOR)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, safe_text(header))
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 32
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.outlinePr.summaryBelow = False


def estimate_row_height(values: list[Any], widths: list[int]) -> float:
    lines = 1
    for index, value in enumerate(values):
        text = str(value or "")
        width = max(10, widths[min(index, len(widths) - 1)])
        cell_lines = sum(
            max(1, math.ceil(len(line) / max(10, int(width * 1.3))))
            for line in text.splitlines() or [""]
        )
        lines = max(lines, cell_lines)
    return float(min(120, max(30, lines * 15)))


def style_data_row(
    sheet: Any,
    row_number: int,
    values: list[Any],
    widths: list[int],
    *,
    alternate: bool,
    priority_column: int | None = None,
    warning: bool = False,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    hair = Side(style="hair", color=GRID_COLOR)
    for cell in sheet[row_number]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=hair)
        if alternate:
            cell.fill = PatternFill("solid", fgColor=ALT_FILL)
    if priority_column is not None:
        priority = str(sheet.cell(row_number, priority_column).value or "")
        if priority in PRIORITY_FILLS:
            fill, font = PRIORITY_FILLS[priority]
            cell = sheet.cell(row_number, priority_column)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(color=font, bold=True)
    if warning:
        for cell in sheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=WARNING_FILL)
    sheet.row_dimensions[row_number].height = estimate_row_height(values, widths)


def style_section_row(sheet: Any, row_number: int, widths: list[int]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    bottom = Side(style="thin", color=GRID_COLOR)
    for cell in sheet[row_number]:
        cell.fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
        cell.font = Font(color=HEADER_FILL, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=bottom)
    sheet.row_dimensions[row_number].height = estimate_row_height(
        [cell.value for cell in sheet[row_number]], widths
    )


def add_note(cell: Any, text: str) -> None:
    if not text:
        return
    from openpyxl.comments import Comment

    note = safe_text(text)
    if not note:
        return
    if cell.comment and cell.comment.text:
        note = f"{cell.comment.text}\n\n{note}"
    cell.comment = Comment(note, "GTM audit readability")


def add_internal_link(cell: Any, sheet_name: str, coordinate: str) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import quote_sheetname

    cell.hyperlink = f"#{quote_sheetname(sheet_name)}!{coordinate}"
    cell.font = Font(color=LINK_COLOR, underline="single")


def append_values(sheet: Any, values: list[Any]) -> int:
    sheet.append([safe_text(value) for value in values])
    return sheet.max_row


def set_widths(sheet: Any, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.auto_filter.ref = sheet.dimensions


def write_audit_sheet(sheet: Any, model: dict[str, Any]) -> dict[str, int]:
    headers = model["headers"]["audit"]
    widths = [18, 24, 42, 58, 38, 14]
    apply_header_style(sheet, headers)
    row_by_id: dict[str, int] = {}
    data_index = 0
    for section in model["audit_sections"]:
        rows = section["rows"]
        section_row = append_values(
            sheet,
            [
                SECTION_MARKER,
                "",
                "",
                f"{section['label']} ({len(rows)})",
                "",
                "",
            ],
        )
        style_section_row(sheet, section_row, widths)
        first_child = sheet.max_row + 1
        for row in rows:
            row_number = append_values(sheet, row["values"])
            row_by_id[row["id"]] = row_number
            data_index += 1
            style_data_row(
                sheet,
                row_number,
                row["values"],
                widths,
                alternate=data_index % 2 == 0,
                priority_column=6,
            )
            for column_index, note in row["notes"].items():
                add_note(sheet.cell(row_number, column_index + 1), note)
            source_run = safe_text(row["record"].get("source_run"))
            source_keys = ", ".join(row["source_object_keys"])
            add_note(
                sheet.cell(row_number, 1),
                (
                    f"Source run: {source_run}. "
                    f"Technical lookup keys: {source_keys or row['id']}."
                ),
            )
        last_child = sheet.max_row
        if first_child <= last_child:
            for row_number in range(first_child, last_child + 1):
                dimension = sheet.row_dimensions[row_number]
                dimension.outlineLevel = 1
                dimension.hidden = bool(section["collapsed"])
            sheet.row_dimensions[section_row].collapsed = bool(section["collapsed"])
    set_widths(sheet, widths)
    return row_by_id


def write_action_sheet(sheet: Any, model: dict[str, Any]) -> dict[str, int]:
    headers = model["headers"]["actions"]
    widths = [22, 14, 38, 54, 48, 58, 38, 58]
    apply_header_style(sheet, headers)
    if model["common_validation"]:
        shared_row = append_values(
            sheet,
            [
                model["labels"]["shared_validation"],
                "",
                "",
                "",
                "",
                "",
                "",
                model["common_validation"],
            ],
        )
        style_section_row(sheet, shared_row, widths)
        add_note(
            sheet.cell(1, 8),
            "Blank operation cells use the shared validation stated in the first data row.",
        )
    row_by_id: dict[str, int] = {}
    for index, row in enumerate(model["action_rows"], start=1):
        row_number = append_values(sheet, row["values"])
        row_by_id[row["id"]] = row_number
        style_data_row(
            sheet,
            row_number,
            row["values"],
            widths,
            alternate=index % 2 == 0,
            priority_column=2,
        )
        for column_index, note in row["notes"].items():
            add_note(sheet.cell(row_number, column_index + 1), note)
        if row["source_ids"]:
            add_note(
                sheet.cell(row_number, 1),
                "Source audit IDs: " + ", ".join(row["source_ids"]),
            )
    set_widths(sheet, widths)
    return row_by_id


def write_decision_sheet(
    sheet: Any, model: dict[str, Any]
) -> tuple[dict[str, int], dict[str, int]]:
    headers = model["headers"]["decisions"]
    widths = [34, 54, 52, 42, 24, 58]
    apply_header_style(sheet, headers)
    topic_rows: dict[str, int] = {}
    source_rows: dict[str, int] = {}
    pending_parent: tuple[int, str] | None = None
    child_start = 0
    for alternate, row in enumerate(model["decision_rows"], start=1):
        if pending_parent and row["kind"] != "decision_child":
            parent_row, _topic = pending_parent
            for number in range(child_start, sheet.max_row + 1):
                dimension = sheet.row_dimensions[number]
                dimension.outlineLevel = 1
                dimension.hidden = True
            sheet.row_dimensions[parent_row].collapsed = True
            pending_parent = None
        row_number = append_values(sheet, row["values"])
        if row["kind"] == "decision_parent":
            topic_rows[row["topic_id"]] = row_number
            style_section_row(sheet, row_number, widths)
            pending_parent = (row_number, row["topic_id"])
            child_start = row_number + 1
        else:
            if row["kind"] == "decision_single":
                topic_rows[row["topic_id"]] = row_number
            for source_id in row["source_ids"]:
                source_rows[source_id] = row_number
            style_data_row(
                sheet,
                row_number,
                row["values"],
                widths,
                alternate=alternate % 2 == 0,
            )
        for column_index, note in row["notes"].items():
            add_note(sheet.cell(row_number, column_index + 1), note)
    if pending_parent:
        parent_row, _topic = pending_parent
        for number in range(child_start, sheet.max_row + 1):
            dimension = sheet.row_dimensions[number]
            dimension.outlineLevel = 1
            dimension.hidden = True
        sheet.row_dimensions[parent_row].collapsed = True
    set_widths(sheet, widths)
    return topic_rows, source_rows


def write_custom_html_sheet(sheet: Any, model: dict[str, Any]) -> dict[str, int]:
    headers = model["headers"]["html"]
    widths = [36, 42, 46, 64, 70, 54, 72]
    apply_header_style(sheet, headers)
    add_note(
        sheet.cell(1, 5),
        (
            "A dataLayer, native-tag, template, consolidation, or site-side candidate is "
            "static evidence only. Before replacement, compare exact value, type, format, "
            "timing, fallback, route, consent state, trigger use, and every downstream consumer."
        ),
    )
    row_by_id: dict[str, int] = {}
    for index, row in enumerate(model["custom_rows"], start=1):
        row_number = append_values(sheet, row["values"])
        row_by_id[row["id"]] = row_number
        style_data_row(
            sheet,
            row_number,
            row["values"],
            widths,
            alternate=index % 2 == 0,
            warning=bool(row["conflicts"]),
        )
        for column_index, note in row.get("notes", {}).items():
            add_note(sheet.cell(row_number, column_index + 1), note)
    set_widths(sheet, widths)
    return row_by_id


def projected_delta_text(counts: dict[str, Any], no_change: str) -> str:
    values = []
    for layer, row in sorted(counts.items()):
        delta = int((row or {}).get("delta") or 0)
        if delta:
            values.append(f"{layer} {delta:+d}")
    return ", ".join(values) or no_change


def container_label(inputs: dict[str, Any]) -> str:
    source_file = str(inputs["audit_package_manifest"].get("source_file") or "")
    match = re.search(r"\bGTM-[A-Z0-9]+\b", source_file, flags=re.I)
    return match.group(0).upper() if match else Path(source_file).name


def write_overview_sheet(sheet: Any, model: dict[str, Any], inputs: dict[str, Any]) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    labels = model["labels"]
    counts = model["counts"]
    context = inputs["context"].get("context") or {}
    website = safe_text(context.get("website_url"))
    project_context = " · ".join(
        value for value in (website, container_label(inputs)) if value
    )
    sheet.merge_cells("A1:F1")
    sheet["A1"] = labels["title"]
    sheet["A1"].font = Font(size=20, bold=True, color=HEADER_FONT)
    sheet["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 36

    sheet.merge_cells("A2:F2")
    sheet["A2"] = safe_text(project_context)
    sheet["A2"].font = Font(size=11, color="44546A")
    sheet["A2"].alignment = Alignment(vertical="center")

    sheet.merge_cells("A3:F3")
    sheet["A3"] = labels["status"]
    sheet["A3"].fill = PatternFill("solid", fgColor="E2F0D9")
    sheet["A3"].font = Font(bold=True, color="375623")
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 32

    metrics = [
        (
            labels["overview"]["audit_records"],
            str(counts["audit_records"]),
        ),
        (
            labels["overview"]["operations"],
            str(counts["operations"]),
        ),
        (
            labels["overview"]["owner"],
            labels["counts"]["owner"].format(
                sources=counts["owner_source_records"],
                topics=counts["decision_topics"],
            ),
        ),
        (
            labels["overview"]["html"],
            str(counts["custom_html_tags"]),
        ),
    ]
    for index, (title, value) in enumerate(metrics):
        start = index % 2 * 3 + 1
        row = 5 + index // 2 * 2
        sheet.merge_cells(
            start_row=row,
            start_column=start,
            end_row=row,
            end_column=start + 1,
        )
        sheet.merge_cells(
            start_row=row + 1,
            start_column=start,
            end_row=row + 1,
            end_column=start + 1,
        )
        title_cell = sheet.cell(row, start, title)
        value_cell = sheet.cell(row + 1, start, value)
        title_cell.fill = PatternFill("solid", fgColor=SUBHEADER_FILL)
        title_cell.font = Font(bold=True, color=HEADER_FILL)
        value_cell.font = Font(size=16, bold=True, color=HEADER_FILL)
        for cell in (title_cell, value_cell):
            cell.alignment = Alignment(horizontal="center", vertical="center")

    priority_text = ", ".join(
        labels["counts"]["priority"].format(
            priority=labels["priority_labels"][priority],
            count=count,
        )
        for priority, count in counts["priority"].items()
        if count
    ) or labels["counts"]["no_actions"]
    detail_rows = [
        (
            labels["overview"]["reconciliation"],
            labels["counts"]["reconciliation"].format(
                findings=counts["audit_records"],
                operations=counts["operations"],
                retained=counts["retained"] + counts["documented_exceptions"],
                decisions=counts["decision_topics"],
            ),
        ),
        (
            labels["overview"]["approval_scope"],
            labels["counts"]["approval_scope"].format(
                bulk=counts["bulk_operations"],
                individual=counts["individual_operations"],
                activation=counts["activation_operations"],
            ),
        ),
        (
            labels["overview"]["change_scope"],
            labels["counts"]["change_scope"].format(
                maintenance=counts["maintenance_operations"],
                behavior=counts["behavior_operations"],
            ),
        ),
        (
            labels["overview"]["remaining"],
            labels["counts"]["remaining"].format(
                remaining=counts["remaining_records"],
                decisions=counts["decision_topics"],
                limits=counts["evidence_limits"],
            ),
        ),
        (
            labels["overview"]["retained"],
            labels["counts"]["retained"].format(
                retained=counts["retained"],
                exceptions=counts["documented_exceptions"],
            ),
        ),
        (labels["overview"]["priorities"], priority_text),
        (
            labels["overview"]["deltas"],
            projected_delta_text(
                model["projected_object_counts"],
                labels["counts"]["no_delta"],
            ),
        ),
        (labels["overview"]["first_actions"], model["first_actions"]),
        (labels["overview"]["measurement"], model["measurement_summary"]),
        (labels["overview"]["boundary"], labels["boundary"]),
        (labels["overview"]["next"], labels["next_step"]),
        (labels["overview"]["navigation"], labels["navigation"]),
    ]
    start_row = 10
    thin = Side(style="hair", color=GRID_COLOR)
    for offset, (title, value) in enumerate(detail_rows):
        row = start_row + offset
        sheet.cell(row, 1, title)
        sheet.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=6,
        )
        sheet.cell(row, 2, safe_text(value))
        sheet.cell(row, 1).font = Font(bold=True, color=HEADER_FILL)
        for cell in sheet[row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=thin)
        sheet.row_dimensions[row].height = estimate_row_height(
            [title, value], [24, 90]
        )
    for column, width in enumerate([24, 25, 4, 24, 25, 4], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"


def apply_cross_links(
    workbook: Any,
    model: dict[str, Any],
    audit_rows: dict[str, int],
    action_rows: dict[str, int],
    topic_rows: dict[str, int],
    decision_source_rows: dict[str, int],
    custom_rows: dict[str, int],
) -> list[dict[str, str]]:
    links: list[dict[str, str]] = []
    audit_sheet = workbook["A4 Audit Register"]
    action_sheet = workbook["A2 Actions"]
    decision_sheet = workbook["A3 Decisions"]
    custom_sheet = workbook["A5 Custom HTML"]

    for section in model["audit_sections"]:
        for row in section["rows"]:
            source_row = audit_rows[row["id"]]
            target_sheet = ""
            target_row = 0
            if row["operation_ids"]:
                target_sheet = "A2 Actions"
                target_row = action_rows[row["operation_ids"][0]]
            elif row["topic_id"]:
                target_sheet = "A3 Decisions"
                target_row = topic_rows[row["topic_id"]]
            else:
                html_keys = [
                    key for key in row["source_object_keys"] if key in custom_rows
                ]
                if html_keys:
                    target_sheet = "A5 Custom HTML"
                    target_row = custom_rows[html_keys[0]]
            if target_sheet:
                cell = audit_sheet.cell(source_row, 5)
                add_internal_link(cell, target_sheet, f"A{target_row}")
                links.append(
                    {
                        "source": f"A4 Audit Register!{cell.coordinate}",
                        "target": f"{target_sheet}!A{target_row}",
                    }
                )

    for row in model["action_rows"]:
        if not row["source_ids"]:
            continue
        source_row = action_rows[row["id"]]
        target_id = next(
            (value for value in row["source_ids"] if value in audit_rows),
            "",
        )
        if target_id:
            cell = action_sheet.cell(source_row, 1)
            add_internal_link(cell, "A4 Audit Register", f"A{audit_rows[target_id]}")
            links.append(
                {
                    "source": f"A2 Actions!{cell.coordinate}",
                    "target": f"A4 Audit Register!A{audit_rows[target_id]}",
                }
            )

    for source_id, source_row in decision_source_rows.items():
        if source_id not in audit_rows:
            continue
        cell = decision_sheet.cell(source_row, 1)
        add_internal_link(cell, "A4 Audit Register", f"A{audit_rows[source_id]}")
        links.append(
            {
                "source": f"A3 Decisions!{cell.coordinate}",
                "target": f"A4 Audit Register!A{audit_rows[source_id]}",
            }
        )

    for row in model["decision_rows"]:
        if row["kind"] not in {"decision_single", "decision_child"}:
            continue
        source_id = row["source_ids"][0]
        source_row = decision_source_rows.get(source_id)
        if not source_row:
            continue
        html_keys = [
            key for key in row.get("source_object_keys", []) if key in custom_rows
        ]
        if html_keys:
            cell = decision_sheet.cell(source_row, 4)
            add_internal_link(cell, "A5 Custom HTML", f"A{custom_rows[html_keys[0]]}")
            links.append(
                {
                    "source": f"A3 Decisions!{cell.coordinate}",
                    "target": f"A5 Custom HTML!A{custom_rows[html_keys[0]]}",
                }
            )
            if len(html_keys) > 1:
                add_note(cell, "Related Custom HTML tags: " + ", ".join(html_keys))

    for row in model["custom_rows"]:
        source_row = custom_rows[row["id"]]
        config_id = row["configuration_id"]
        if config_id and config_id in audit_rows:
            cell = custom_sheet.cell(source_row, 1)
            add_internal_link(cell, "A4 Audit Register", f"A{audit_rows[config_id]}")
            links.append(
                {
                    "source": f"A5 Custom HTML!{cell.coordinate}",
                    "target": f"A4 Audit Register!A{audit_rows[config_id]}",
                }
            )
        if row["related_topics"]:
            topic_id = row["related_topics"][0]
            if topic_id in topic_rows:
                cell = custom_sheet.cell(source_row, 7)
                add_internal_link(cell, "A3 Decisions", f"A{topic_rows[topic_id]}")
                links.append(
                    {
                        "source": f"A5 Custom HTML!{cell.coordinate}",
                        "target": f"A3 Decisions!A{topic_rows[topic_id]}",
                    }
                )
                if len(row["related_topics"]) > 1:
                    add_note(
                        cell,
                        "Related decision topics: " + ", ".join(row["related_topics"]),
                    )
    return links


def create_human_sheets(workbook: Any) -> dict[str, Any]:
    for name in HUMAN_SHEETS:
        if name in workbook.sheetnames:
            raise ValueError(f"The canonical workbook already contains {name}")
    sheets = {}
    for index, name in enumerate(HUMAN_SHEETS):
        sheet = workbook.create_sheet(name, index)
        sheet.sheet_properties.tabColor = HUMAN_TAB_COLOR
        sheets[name] = sheet
    return sheets


def build_readability_workbook(
    inputs: dict[str, Any],
    model: dict[str, Any],
    output: Path,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build XLSX output") from exc

    if output.exists():
        raise FileExistsError(f"Output already exists: {output}")
    source_workbook = inputs["paths"]["canonical_workbook"]
    if output.resolve() == source_workbook.resolve():
        raise ValueError("The analyst workbook output must differ from the canonical workbook")

    workbook = load_workbook(source_workbook)
    sheets = create_human_sheets(workbook)
    write_overview_sheet(sheets["A1 Overview"], model, inputs)
    action_rows = write_action_sheet(sheets["A2 Actions"], model)
    topic_rows, decision_source_rows = write_decision_sheet(
        sheets["A3 Decisions"], model
    )
    audit_rows = write_audit_sheet(sheets["A4 Audit Register"], model)
    custom_rows = write_custom_html_sheet(sheets["A5 Custom HTML"], model)
    links = apply_cross_links(
        workbook,
        model,
        audit_rows,
        action_rows,
        topic_rows,
        decision_source_rows,
        custom_rows,
    )
    for name in ORIGINAL_SHEETS:
        workbook[name].sheet_state = "hidden"
    workbook.active = 0
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()

    # Reopen immediately so a corrupt derived file never receives a build manifest.
    check = load_workbook(output, read_only=True, data_only=False)
    try:
        if check.sheetnames[: len(HUMAN_SHEETS)] != HUMAN_SHEETS:
            raise ValueError("Saved workbook does not contain the required human tab order")
    finally:
        check.close()
    return {
        "audit_rows": audit_rows,
        "action_rows": action_rows,
        "topic_rows": topic_rows,
        "decision_source_rows": decision_source_rows,
        "custom_rows": custom_rows,
        "links": links,
    }


def transformation_manifest(
    inputs: dict[str, Any],
    model: dict[str, Any],
    output: Path,
    row_index: dict[str, Any],
) -> dict[str, Any]:
    source_manifest = inputs["audit_package_manifest"]
    return {
        "kind": MANIFEST_KIND,
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "status": "built",
        "language": model["language"],
        "source_file": Path(str(source_manifest.get("source_file") or "")).name,
        "source_sha256": inputs["source_sha256"],
        "skill_version": (
            (source_manifest.get("skill_runtime_identity") or {}).get("project_version")
        ),
        "canonical_workbook": {
            "name": inputs["paths"]["canonical_workbook"].name,
            "sha256": sha256_file(inputs["paths"]["canonical_workbook"]),
            "sheets": inputs["standard_sheet_hashes"],
        },
        "analyst_workbook": {
            "name": output.name,
            "sha256": sha256_file(output),
            "human_sheets": HUMAN_SHEETS,
            "columns": model["headers"],
        },
        "inputs": input_hash_manifest(inputs["paths"]),
        "coverage": model["counts"],
        "decision_topics": [
            {
                "topic_id": topic["topic_id"],
                "source_ids": topic["source_ids"],
            }
            for topic in model["decision_topics"]
        ],
        "custom_html_cleanup_conflicts": model["custom_conflicts"],
        "links": row_index["links"],
        "gates": {},
        "errors": [],
        "scope_boundary": (
            "Derived analyst workbook only; no audit verdict, operation, GTM object, "
            "version, workspace, or publication was changed."
        ),
        "fallback": {
            "name": inputs["paths"]["canonical_workbook"].name,
            "deliver_when_readability_gate_fails": True,
        },
    }


def write_manifest(path: Path, payload: dict[str, Any]) -> None:
    if path.exists():
        raise FileExistsError(f"Manifest already exists: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def build(
    package_dir: Path,
    operations_path: Path,
    standard_workbook: Path,
    output: Path,
    *,
    future_state_path: Path | None = None,
    completion_gate_path: Path | None = None,
    decision_topics_path: Path | None = None,
    manifest_path: Path | None = None,
    language: str = "en",
) -> tuple[Path, Path, dict[str, Any]]:
    paths = artifact_paths(
        package_dir,
        operations_path,
        standard_workbook,
        future_state_path,
        completion_gate_path,
        decision_topics_path,
    )
    manifest_path = manifest_path or default_manifest_path(output)
    validate_manifest_path(manifest_path, output, paths)
    if manifest_path.exists():
        raise FileExistsError(f"Manifest already exists: {manifest_path}")
    inputs = load_inputs(paths)
    model = build_model(inputs, language)
    row_index = build_readability_workbook(inputs, model, output)
    manifest = transformation_manifest(inputs, model, output, row_index)
    write_manifest(manifest_path, manifest)
    return output, manifest_path, manifest


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("package_dir", type=Path)
    parser.add_argument("operations", type=Path)
    parser.add_argument("canonical_workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--future-state", type=Path)
    parser.add_argument("--completion-gate", type=Path)
    parser.add_argument("--decision-topics", type=Path)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--language", choices=sorted(LOCALES), default="en")
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()
    try:
        output, manifest_path, manifest = build(
            args.package_dir,
            args.operations,
            args.canonical_workbook,
            args.output,
            future_state_path=args.future_state,
            completion_gate_path=args.completion_gate,
            decision_topics_path=args.decision_topics,
            manifest_path=args.manifest,
            language=args.language,
        )
    except (FileNotFoundError, FileExistsError, RuntimeError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    result = {
        "output": str(output),
        "manifest": str(manifest_path),
        "status": manifest["status"],
        "coverage": manifest["coverage"],
    }
    print(
        json.dumps(
            result,
            ensure_ascii=False,
            indent=2 if args.pretty else None,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
