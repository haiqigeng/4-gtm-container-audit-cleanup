#!/usr/bin/env python3
"""Validate the compact GTM cleanup-plan workbook structure and content."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from gtm_lib import as_list
from gtm_taxonomy import (
    CLEANUP_PLAN_COLUMNS,
    GENERAL_CATEGORY_BY_PROBLEM_TYPE,
    GENERAL_PROBLEM_CATEGORIES,
    general_problem_category,
)
from gtm_workbook import load_xlsx_workbook, normalize_header
from gtm_workbook_build import (
    CANONICAL_SHEETS,
    MAX_ROW_HEIGHT,
    estimated_cell_lines,
)

PLACEHOLDER_PATTERNS = (
    re.compile(r"\b(?:tbd|to" r"do|lorem ipsum)\b", re.I),
    re.compile(r"\b(?:configuration|code) (?:reviewed|inspected)\b", re.I),
    re.compile(r"\b(?:review|check) (?:later|in gtm)\b", re.I),
)
VISIBLE_INTERNAL_DUMP_PATTERNS = (
    re.compile(r"\b[a-f0-9]{32,}\b", re.I),
    re.compile(r"\b(?:line_hash|config_hash|object_identity|contract_sha256)\b", re.I),
    re.compile(r"\$\.containerVersion\b"),
    re.compile(r"\{\s*\"[A-Za-z0-9_]+\"\s*:"),
)


def duplicate_columns(sheet_name: str, rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return []
    headers = list(rows[0])
    columns = {
        header: tuple(str(row.get(header, "") or "").strip() for row in rows) for header in headers
    }
    errors = []
    for index, left in enumerate(headers):
        if not any(columns[left]):
            continue
        for right in headers[index + 1 :]:
            if columns[left] == columns[right]:
                errors.append(f"{sheet_name}: duplicate column content in {left!r} and {right!r}")
    return errors


def placeholder_errors(sheet_name: str, rows: list[dict[str, Any]]) -> list[str]:
    errors = []
    for row_number, row in enumerate(rows, start=2):
        for header, value in row.items():
            text = str(value or "")
            if any(pattern.search(text) for pattern in PLACEHOLDER_PATTERNS):
                errors.append(
                    f"{sheet_name}!{header} row {row_number}: generic or deferred wording"
                )
    return errors


def workbook_structure_errors(
    workbook_rows: dict[str, list[dict[str, Any]]]
) -> list[str]:
    errors: list[str] = []
    if list(workbook_rows) != CANONICAL_SHEETS:
        errors.append("workbook tabs do not match the canonical eight-tab order")
    if len(workbook_rows) > 8:
        errors.append("workbook has more than eight tabs")
    for name, rows in workbook_rows.items():
        headers = list(rows[0]) if rows else []
        column_limit = 7 if name == "02 Cleanup Plan" else 6
        if len(headers) > column_limit:
            errors.append(f"{name}: more than {column_limit} columns")
        if name == "02 Cleanup Plan":
            expected_headers = [
                normalize_header(value) for value in CLEANUP_PLAN_COLUMNS
            ]
            if headers != expected_headers:
                errors.append(
                    "02 Cleanup Plan columns do not match the canonical "
                    "seven-column order"
                )
        errors.extend(duplicate_columns(name, rows))
        errors.extend(placeholder_errors(name, rows))
    if "Change Log" in workbook_rows:
        errors.append("cleanup plan must not contain a change-log tab")
    return errors


def rendered_cell_errors(sheet: Any) -> list[str]:
    errors: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "[truncated]" in cell.value.lower():
                errors.append(f"{sheet.title}!{cell.coordinate}: proof text was truncated")
            if cell.data_type == "f":
                errors.append(f"{sheet.title}!{cell.coordinate}: formulas are not allowed")
            elif (
                isinstance(cell.value, str)
                and cell.value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r", "\n"))
                and not cell.value.startswith("'")
            ):
                errors.append(
                    f"{sheet.title}!{cell.coordinate}: formula-like text is not escaped"
                )
    return errors


def rendered_sheet_errors(sheet: Any) -> list[str]:
    errors: list[str] = []
    column_limit = 7 if sheet.title == "02 Cleanup Plan" else 6
    if sheet.max_column > column_limit:
        errors.append(
            f"{sheet.title}: rendered workbook exceeds {column_limit} columns"
        )
    if sheet.title == "02 Cleanup Plan":
        expected_filter = f"A1:G{sheet.max_row}"
        if sheet.auto_filter.ref != expected_filter:
            errors.append(
                "02 Cleanup Plan filter does not span all seven canonical columns"
            )
        if sheet.freeze_panes != "A2":
            errors.append("02 Cleanup Plan header row is not frozen")
    if any((dimension.width or 0) > 92 for dimension in sheet.column_dimensions.values()):
        errors.append(f"{sheet.title}: column width exceeds 92")
    if any(
        (dimension.height or 0) > MAX_ROW_HEIGHT
        for dimension in sheet.row_dimensions.values()
    ):
        errors.append(f"{sheet.title}: row height exceeds {MAX_ROW_HEIGHT}")
    for row_number in range(2, sheet.max_row + 1):
        height = float(sheet.row_dimensions[row_number].height or 15)
        required_lines = 1
        for cell in sheet[row_number]:
            width = float(
                sheet.column_dimensions[cell.column_letter].width or 10
            )
            required_lines = max(
                required_lines, estimated_cell_lines(cell.value, int(width))
            )
        required_height = max(36, required_lines * 15)
        if required_height > height + 1:
            errors.append(
                f"{sheet.title} row {row_number}: visible text needs "
                f"{required_height}pt but row height is {height}pt"
            )
    errors.extend(rendered_cell_errors(sheet))
    return errors


def rendered_workbook_errors(workbook_path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return [f"openpyxl is required for workbook state validation: {exc}"]
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    errors: list[str] = []
    visible = [sheet.title for sheet in workbook if sheet.sheet_state == "visible"]
    if visible != CANONICAL_SHEETS[:2]:
        errors.append("only Summary and Cleanup Plan may be visible")
    for sheet in workbook:
        errors.extend(rendered_sheet_errors(sheet))
    workbook.close()
    return errors


def operations_alignment_errors(
    workbook_rows: dict[str, list[dict[str, Any]]], operations: dict[str, Any]
) -> list[str]:
    errors: list[str] = []
    owner_count = sum(
        1
        for row in as_list(operations.get("decision_ledger"))
        if row.get("disposition") == "owner_decision_needed"
    )
    evidence_limit_count = sum(
        1
        for row in as_list(operations.get("decision_ledger"))
        if row.get("disposition") == "container_evidence_limit"
    )
    scope_row_required = bool(evidence_limit_count)
    cleanup_rows = workbook_rows.get("02 Cleanup Plan", [])
    summary_values = {
        str(row.get("decision") or ""): (
            "" if row.get("value") is None else str(row.get("value"))
        )
        for row in workbook_rows.get("01 Summary", [])
    }
    overall_status = summary_values.get("Overall status", "").lower()
    action_completeness = operations.get("action_completeness") or {}
    incomplete = action_completeness.get("status") != "pass"
    rendered_ids = [str(row.get("id") or "") for row in cleanup_rows]

    for row_number, row in enumerate(cleanup_rows, start=2):
        identifier = str(row.get("id") or "")
        # Workbook headers are normalized by gtm_workbook.normalize_header.
        category = str(row.get("general_problem_category") or "").strip()
        area_and_type = str(row.get("area_problem_type") or "").strip()
        if category not in GENERAL_PROBLEM_CATEGORIES:
            errors.append(
                f"Cleanup Plan row {row_number} has an unsupported general "
                f"problem category: {category!r}"
            )
        matched_problem_type = next(
            (
                problem_type
                for problem_type in sorted(
                    GENERAL_CATEGORY_BY_PROBLEM_TYPE,
                    key=len,
                    reverse=True,
                )
                if area_and_type.endswith(f" / {problem_type}")
            ),
            "",
        )
        if not matched_problem_type:
            errors.append(
                f"Cleanup Plan row {row_number} has an unsupported Area / "
                f"problem type value: {area_and_type!r}"
            )
        elif category != general_problem_category(matched_problem_type):
            errors.append(
                f"Cleanup Plan row {row_number} general category does not match "
                f"{matched_problem_type!r}"
            )
        affected = str(row.get("affected_object_s") or "").strip()
        if identifier and not affected:
            errors.append(
                f"Cleanup Plan row {row_number} has an empty affected-object scope"
            )
        for header, value in row.items():
            text = str(value or "")
            if any(pattern.search(text) for pattern in VISIBLE_INTERNAL_DUMP_PATTERNS):
                errors.append(
                    f"Cleanup Plan row {row_number} {header!r} exposes internal "
                    "hash, token, JSON, or source-path proof"
                )

    expected_summary_counts = {
        "Proposed operations": len(as_list(operations.get("operations"))),
        "Action-completeness errors": len(
            as_list(action_completeness.get("errors"))
        ),
        "Owner decisions": owner_count,
        "Container evidence limits": evidence_limit_count,
    }
    for label, expected_count in expected_summary_counts.items():
        value = summary_values.get(label, "")
        if not re.fullmatch(r"\d+", value.strip()) or int(value) != expected_count:
            errors.append(
                f"Summary {label!r} count does not match the operations packet"
            )

    if incomplete:
        if rendered_ids != ["BLOCKED-001"]:
            errors.append(
                "action-incomplete output must show only one BLOCKED-001 draft row"
            )
        if "incomplete cleanup plan" not in overall_status:
            errors.append("Summary status does not expose incomplete cleanup actions")
        errors.append("cleanup plan action completeness is not pass")
        if set((operations.get("run_statuses") or {}).values()) != {"complete"}:
            errors.append("operations do not record three complete review runs")
        return errors

    expected_operation_ids = {
        str(row.get("operation_id") or "")
        for row in as_list(operations.get("operations"))
        if row.get("operation_id")
    }
    rendered_operation_ids = [
        value
        for row in cleanup_rows
        for value in re.findall(r"\bOP-\d{4}\b", str(row.get("id") or ""))
    ]
    if set(rendered_operation_ids) != expected_operation_ids:
        missing = sorted(expected_operation_ids - set(rendered_operation_ids))
        unknown = sorted(set(rendered_operation_ids) - expected_operation_ids)
        if missing:
            errors.append(
                "Cleanup Plan omits operation IDs: " + ", ".join(missing)
            )
        if unknown:
            errors.append(
                "Cleanup Plan contains unknown operation IDs: " + ", ".join(unknown)
            )
    duplicates = sorted(
        operation_id
        for operation_id in set(rendered_operation_ids)
        if rendered_operation_ids.count(operation_id) > 1
    )
    if duplicates:
        errors.append(
            "Cleanup Plan repeats operation IDs across visible rows: "
            + ", ".join(duplicates)
        )

    owner_groups: dict[tuple[str, str, str, str], list[dict[str, Any]]] = {}
    for row in as_list(operations.get("decision_ledger")):
        if row.get("disposition") != "owner_decision_needed":
            continue
        key = (
            str(row.get("area") or "Governance / ownership"),
            str(row.get("problem_type") or "Unclear business purpose"),
            " ".join(str(row.get("owner_question") or "").split()),
            " ".join(str(row.get("recommended_action") or "").split()),
        )
        owner_groups.setdefault(key, []).append(row)
    rendered_owner_rows = [
        row for row in cleanup_rows if row.get("status") == "Owner confirmation"
    ]
    if len(rendered_owner_rows) != len(owner_groups):
        errors.append(
            "Cleanup Plan must show each distinct owner question exactly once"
        )
    for (_area, _problem_type, question, recommendation), source_rows in owner_groups.items():
        decision_ids = [
            str(row.get("decision_id") or "") for row in source_rows if row.get("decision_id")
        ]
        representative_id = decision_ids[0] if decision_ids else ""
        matches = []
        for row in rendered_owner_rows:
            action_text = str(row.get("action_priority_qa") or "")
            # Long source-bound questions can enumerate dozens of objects.
            # The visible workbook may show a concise wording when it still
            # carries a precise source decision reference plus both labels;
            # the full immutable text remains in the reconciled packet and
            # hidden review sheet. This keeps the plan readable without
            # losing the link required to approve the exact decision.
            compact_reference_match = bool(
                representative_id
                and re.search(
                    rf"(?<![A-Za-z0-9_-]){re.escape(representative_id)}(?![A-Za-z0-9_-])",
                    action_text,
                )
                and "Question:" in action_text
                and "Recommendation:" in action_text
            )
            # Decision IDs take precedence: different source decisions may
            # legitimately have the same owner wording while retaining
            # separate taxonomy, evidence, and approval scope.
            full_text_match = bool(
                not representative_id
                and question
                and question in action_text
                and recommendation in action_text
            )
            if compact_reference_match or full_text_match:
                matches.append(row)
        if len(matches) != 1:
            errors.append(
                "Cleanup Plan omits or repeats a distinct owner question/recommendation"
            )
    if scope_row_required and rendered_ids.count("SCOPE-001") != 1:
        errors.append("Cleanup Plan must contain one consolidated evidence-boundary row")
    if not scope_row_required and "SCOPE-001" in rendered_ids:
        errors.append(
            "Cleanup Plan contains an evidence-boundary row without source decisions "
            "or runtime handoffs"
        )
    single_owner_ids = {
        str(rows[0].get("decision_id") or "")
        for rows in owner_groups.values()
        if len(rows) == 1
    }
    allowed_standalone_ids = single_owner_ids | (
        {"SCOPE-001"} if scope_row_required else set()
    )
    for identifier in rendered_ids:
        if (
            re.search(r"\bOP-\d{4}\b", identifier)
            or identifier in allowed_standalone_ids
            or re.fullmatch(r"OWNER-\d{3}", identifier)
        ):
            continue
        errors.append(f"Cleanup Plan contains an unlinked visible row ID: {identifier!r}")
    if owner_count and "owner decisions required" not in overall_status:
        errors.append("Summary status does not expose unresolved owner decisions")
    if set((operations.get("run_statuses") or {}).values()) != {"complete"}:
        errors.append("operations do not record three complete review runs")
    return errors


def validate_workbook(
    workbook_path: Path, operations_path: Path | None = None
) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    if not workbook_path.is_file():
        return [f"workbook does not exist: {workbook_path}"], warnings
    workbook_rows = load_xlsx_workbook(workbook_path)
    errors.extend(workbook_structure_errors(workbook_rows))
    errors.extend(rendered_workbook_errors(workbook_path))

    if operations_path:
        operations = json.loads(operations_path.read_text(encoding="utf-8"))
        errors.extend(operations_alignment_errors(workbook_rows, operations))
    return errors, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--operations", type=Path)
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()
    errors, warnings = validate_workbook(args.workbook, args.operations)
    report = {
        "kind": "gtm_cleanup_workbook_gate",
        "status": "pass" if not errors else "fail",
        "workbook": str(args.workbook),
        "errors": errors,
        "warnings": warnings,
    }
    print(json.dumps(report, indent=2 if args.pretty else None))
    for warning in warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    for error in errors:
        print(f"ERROR: {error}", file=sys.stderr)
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
