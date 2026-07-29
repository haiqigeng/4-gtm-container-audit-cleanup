#!/usr/bin/env python3
"""Build the compact stakeholder GTM cleanup-plan workbook."""

from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any

from gtm_lib import as_list, load_json
from gtm_privacy import redact_text, spreadsheet_safe_text
from gtm_taxonomy import CLEANUP_PLAN_COLUMNS, GENERAL_PROBLEM_CATEGORIES

CANONICAL_SHEETS = [
    "01 Summary",
    "02 Cleanup Plan",
    "03 Operational Review",
    "04 Configuration Review",
    "05 Architecture Review",
    "06 Custom Code Review",
    "07 Reconciled Operations",
    "08 Source & Gates",
]

HEADER_FILL = "16324F"
HEADER_FONT = "FFFFFF"
GRID_COLOR = "C8D2DC"
MAX_CELL_TEXT = 24000
MAX_RENDERED_LINES = 12
MAX_ROW_HEIGHT = MAX_RENDERED_LINES * 15


def clean_text(value: Any) -> str:
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    else:
        # Counts of zero and explicit false states are evidence, not blanks.
        text = "" if value is None else str(value)
    text = redact_text(text)
    return spreadsheet_safe_text(text)


def join_text(values: list[Any]) -> str:
    return "; ".join(
        clean_text(value)
        for value in values
        if value is not None and str(value).strip()
    )


def code_display_text(value: Any) -> str:
    """Render code behavior without exposing internal segment hash identities."""
    text = clean_text(value)
    text = re.sub(
        r"(?i)(?:api[_-]?key|access[_-]?token|refresh[_-]?token|token|secret|"
        r"password|authorization|email|phone|user[_-]?id|client[_-]?id)\s*[:=]\s*"
        r"<redacted(?:-[^>]+)?>",
        "redacted_field",
        text,
    )
    return re.sub(r"(?i)\b[0-9a-f]{16}\b", "segment", text)


def cell_chunks(value: Any, max_chars: int = MAX_CELL_TEXT) -> list[str]:
    """Return lossless, formula-safe chunks that fit comfortably in an XLSX cell."""
    text = clean_text(value)
    max_chars = max(1, min(MAX_CELL_TEXT, max_chars))
    if len(text) <= max_chars:
        return [text]
    return [
        spreadsheet_safe_text(text[start : start + max_chars])
        for start in range(0, len(text), max_chars)
    ]


def table_widths(header_count: int) -> list[int]:
    if header_count == len(CLEANUP_PLAN_COLUMNS):
        return [16, 28, 28, 42, 54, 54, 54]
    return [16, 28, 42, 54, 54, 54] if header_count >= 6 else [28, 92]


def rendered_cell_capacity(width: int) -> int:
    # Keep one line of slack for uneven explicit line breaks and the leading
    # apostrophe used to escape formula-like spreadsheet text.
    return max(180, int(max(12, width * 1.25) * (MAX_RENDERED_LINES - 1)))


def estimated_cell_lines(value: Any, width: int) -> int:
    text = str(value or "")
    return max(
        1,
        sum(
            max(1, math.ceil(len(line) / max(12, width * 1.25)))
            for line in text.split("\n")
        ),
    )


def expanded_table_rows(
    rows: list[dict[str, Any]], headers: list[str], split_long_cells: bool
) -> list[list[str]]:
    expanded: list[list[str]] = []
    widths = table_widths(len(headers))
    for row_number, row in enumerate(rows, start=2):
        chunks_by_column = []
        for column_number, header in enumerate(headers):
            width = widths[min(column_number, len(widths) - 1)]
            capacity = rendered_cell_capacity(width)
            text = clean_text(row.get(header, ""))
            if not split_long_cells and len(text) > capacity:
                raise ValueError(
                    f"visible workbook row {row_number} column {header!r} needs "
                    f"{estimated_cell_lines(text, width)} rendered lines; summarize "
                    "the user-facing row instead of clipping it"
                )
            chunks_by_column.append(
                cell_chunks(text, capacity if split_long_cells else MAX_CELL_TEXT)
            )
        part_count = max(len(chunks) for chunks in chunks_by_column)
        for part in range(part_count):
            expanded.append(
                [
                    chunks[part] if part < len(chunks) else ""
                    for chunks in chunks_by_column
                ]
            )
    return expanded


def decision_text(row: dict[str, Any]) -> str:
    operation = row.get("operation") or {}
    operations = as_list(row.get("operations"))
    action = operation.get("exact_proposed_action") if operation else ""
    if not action and operations:
        action = join_text([item.get("exact_proposed_action") for item in operations])
    return join_text(
        [
            row.get("disposition"),
            row.get("owner_question"),
            row.get("recommended_action"),
            action,
            row.get("confidence"),
        ]
    )


def operational_rows(review: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for finding in as_list(review.get("findings")):
        rows.append(
            {
                "Finding": clean_text(finding.get("finding_id")),
                "Type / module": clean_text(
                    f"{finding.get('finding_type')} / {finding.get('module_name')}"
                ),
                "Affected objects": join_text(
                    [
                        f"{object_id} - {name}"
                        for object_id, name in zip(
                            as_list(finding.get("object_ids")),
                            as_list(finding.get("object_names")),
                            strict=False,
                        )
                    ]
                ),
                "Evidence": clean_text(finding.get("deterministic_evidence")),
                "Decision": join_text([finding.get("disposition"), finding.get("rationale")]),
                "Action / status": join_text(
                    [
                        finding.get("exact_proposed_action"),
                        finding.get("priority"),
                        finding.get("execution_readiness"),
                        finding.get("owner_question"),
                        finding.get("recommended_action"),
                    ]
                ),
            }
        )
    return rows


def configuration_rows(review: dict[str, Any]) -> list[dict[str, str]]:
    """Render one decision-oriented proof row per object.

    The exact branch, trace, contract, and code-line coverage remains in the
    JSON evidence package. Repeating every passing obligation in XLSX made the
    analyst deliverable larger without improving discovery or decisions.
    """

    rows: list[dict[str, str]] = []
    for item in as_list(review.get("rows")):
        object_label = clean_text(
            f"{item.get('object_key')} - {item.get('object_name')} [{item.get('object_type')}]"
        )
        defects = [
            f"{defect.get('defect_id')}: {defect.get('statement')}"
            for defect in as_list(item.get("defects"))
        ]
        contract_exceptions = [
            f"{check.get('contract_topic')}: {check.get('verdict')}"
            for check in as_list(item.get("contract_checks"))
            if check.get("verdict") in {"Non-compliant", "Unproven"}
        ]
        logic_exceptions = [
            f"{check.get('check_key')}: {check.get('verdict')}"
            for check in as_list(item.get("logic_cross_checks"))
            if check.get("verdict") in {"Issue", "Unclear"}
        ]
        rows.append(
            {
                "Object": object_label,
                "Configured behavior": join_text(
                    [
                        item.get("purpose"),
                        item.get("execution_logic"),
                        item.get("inputs_and_terminal_sources"),
                        item.get("configured_output_or_side_effect"),
                        item.get("consumer_contract"),
                        item.get("consent_and_sequence"),
                    ]
                ),
                "Correctness": join_text(
                    [item.get("correctness_verdict"), item.get("correctness_basis")]
                ),
                "Key evidence": join_text(as_list(item.get("evidence_anchors"))),
                "Defects / limits": join_text(
                    [*defects, *contract_exceptions, *logic_exceptions]
                ),
                "Decision / action": decision_text(item),
            }
        )
    return rows


def architecture_rows(review: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in as_list(review.get("families")):
        rows.append(
            {
                "Family / comparison": clean_text(
                    f"{item.get('family_id')} - {item.get('family_label')}"
                ),
                "Members / chain": join_text(as_list(item.get("chain_object_keys"))),
                "Business and execution logic": join_text(
                    [
                        item.get("business_action"),
                        item.get("family_purpose"),
                        item.get("execution_path_summary"),
                        item.get("payload_coherence"),
                        item.get("consent_and_sequence_coherence"),
                    ]
                ),
                "Relationship verdict": join_text(
                    [item.get("relationship_verdict"), item.get("analyst_rationale")]
                ),
                "Necessity / target state": join_text(
                    [item.get("necessity_and_ownership"), item.get("target_architecture")]
                ),
                "Decision": decision_text(item),
            }
        )
    for item in as_list(review.get("comparisons")):
        rows.append(
            {
                "Family / comparison": clean_text(item.get("comparison_id")),
                "Members / chain": clean_text(item.get("candidate_object_keys", [])),
                "Business and execution logic": join_text(
                    [item.get("candidate_basis"), item.get("architecture_effect")]
                ),
                "Relationship verdict": join_text(
                    [item.get("relationship_verdict"), item.get("analyst_rationale")]
                ),
                "Necessity / target state": join_text(
                    [
                        assessment.get("necessity")
                        for assessment in as_list(item.get("member_assessments"))
                    ]
                ),
                "Decision": decision_text(item),
            }
        )
    return rows


def code_rows(configuration: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for item in as_list(configuration.get("rows")):
        if not item.get("required_code_line_hashes"):
            continue
        object_label = clean_text(f"{item.get('object_key')} - {item.get('object_name')}")
        blocks = as_list(item.get("code_behavior_blocks"))
        findings = as_list(item.get("technical_finding_reviews"))
        rows.append(
            {
                "Object": object_label,
                "Coverage": clean_text(
                    f"{len(item.get('required_code_line_hashes', []))} executable lines / "
                    f"{len(blocks)} behavior blocks"
                ),
                "Behavior": code_display_text(
                    join_text([block.get("purpose") for block in blocks])
                ),
                "Inputs / outputs / side effects": code_display_text(
                    join_text(
                        [
                            value
                            for block in blocks
                            for value in (
                                block.get("inputs"),
                                block.get("outputs"),
                                block.get("side_effects"),
                            )
                        ]
                    )
                ),
                "Code findings": code_display_text(
                    join_text(
                        [
                            f"{finding.get('finding_key')}: {finding.get('verdict')} - "
                            f"{finding.get('rationale')}"
                            for finding in findings
                        ]
                    )
                ),
                "Decision": code_display_text(decision_text(item)),
            }
        )
    return rows


def operation_rows(payload: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for operation in as_list(payload.get("operations")):
        rows.append(
            {
                "Operation": clean_text(
                    f"{operation.get('operation_id')} / {operation.get('resolution_status')}"
                ),
                "Area / problem": clean_text(
                    f"{operation.get('area')} / {operation.get('problem_type')}"
                ),
                "Affected objects": clean_text(operation.get("affected_objects")),
                "Reason / target state": join_text(
                    [
                        operation.get("problem"),
                        operation.get("why_it_matters"),
                        operation.get("expected_clean_state"),
                        (
                            "Affected measurement families: "
                            + ", ".join(
                                str(value)
                                for value in as_list(
                                    operation.get("affected_measurement_family_ids")
                                )
                            )
                            if as_list(operation.get("affected_measurement_family_ids"))
                            else ""
                        ),
                        operation.get("retained_behavior"),
                    ]
                ),
                "Exact mutation": clean_text(
                    {
                        "action": operation.get("exact_proposed_action"),
                        "creations": operation.get("creations", []),
                        "additions": operation.get("additions", []),
                        "changes": operation.get("changes", []),
                        "remaps": operation.get("remaps", []),
                        "deletions": operation.get("deletions", []),
                        "renames": operation.get("renames", []),
                    }
                ),
                "Priority / QA / rollback": join_text(
                    [
                        operation.get("priority"),
                        operation.get("execution_readiness"),
                        operation.get("qa_steps"),
                        operation.get("rollback"),
                    ]
                ),
            }
        )
    return rows


def source_rows(
    manifest: dict[str, Any],
    source: dict[str, Any],
    operations: dict[str, Any],
) -> list[dict[str, str]]:
    return [
        {"Check": "Source file", "Result": clean_text(manifest.get("source_file"))},
        {"Check": "Source SHA-256", "Result": clean_text(manifest.get("source_sha256"))},
        {
            "Check": "Shared deterministic facts",
            "Result": clean_text(manifest.get("shared_facts_sha256")),
        },
        {
            "Check": "Audit context",
            "Result": clean_text(manifest.get("context_sha256")),
        },
        {
            "Check": "Independent run input contracts",
            "Result": clean_text(
                {
                    key: (value or {}).get("contract_sha256")
                    for key, value in (
                        manifest.get("run_input_contracts") or {}
                    ).items()
                }
            ),
        },
        {
            "Check": "Source model coverage",
            "Result": clean_text(source.get("coverage_gate")),
        },
        {
            "Check": "Operational sanitation run",
            "Result": clean_text(
                (operations.get("run_statuses") or {}).get("operational_sanitation")
            ),
        },
        {
            "Check": "Configuration correctness run",
            "Result": clean_text(
                (operations.get("run_statuses") or {}).get("configuration_correctness")
            ),
        },
        {
            "Check": "Business architecture run",
            "Result": clean_text(
                (operations.get("run_statuses") or {}).get("business_architecture")
            ),
        },
        {"Check": "Execution route", "Result": clean_text(operations.get("route"))},
        {"Check": "Cleanup plan", "Result": clean_text(operations.get("plan_status"))},
        {
            "Check": "Measurement preservation",
            "Result": clean_text(
                (operations.get("measurement_preservation") or {}).get("status")
            ),
        },
        {
            "Check": "Decision ledger records",
            "Result": clean_text(len(as_list(operations.get("decision_ledger")))),
        },
    ]


def add_table(
    sheet: Any,
    rows: list[dict[str, Any]],
    headers: list[str] | None = None,
    *,
    split_long_cells: bool = False,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    headers = headers or (list(rows[0]) if rows else ["Status"])
    if not rows:
        rows = [{headers[0]: "No rows"}]
    sheet.append(headers)
    for values in expanded_table_rows(rows, headers, split_long_cells):
        sheet.append(values)
    thin = Side(style="thin", color=GRID_COLOR)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 30
    widths = table_widths(len(headers))
    for row_number in range(2, sheet.max_row + 1):
        estimated_lines = 1
        for column_number, cell in enumerate(sheet[row_number], start=1):
            value = str(cell.value or "")
            width = widths[min(column_number - 1, len(widths) - 1)]
            estimated_lines = max(
                estimated_lines, estimated_cell_lines(value, width)
            )
        if estimated_lines > MAX_RENDERED_LINES:
            raise ValueError(
                f"{sheet.title} row {row_number} still needs {estimated_lines} "
                "rendered lines after chunking"
            )
        sheet.row_dimensions[row_number].height = max(36, estimated_lines * 15)
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = PatternFill("solid", fgColor="F5F8FA")
        for cell in sheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color=GRID_COLOR))
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = widths[
            min(index - 1, len(widths) - 1)
        ]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def build_workbook(
    manifest: dict[str, Any],
    source: dict[str, Any],
    operational: dict[str, Any],
    configuration: dict[str, Any],
    architecture: dict[str, Any],
    operations: dict[str, Any],
    human_rows: dict[str, Any],
    output: Path,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build XLSX output") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in CANONICAL_SHEETS}
    operation_count = len(as_list(operations.get("operations")))
    ledger = as_list(operations.get("decision_ledger"))
    owner_decisions = sum(
        1
        for row in ledger
        if row.get("disposition") == "owner_decision_needed"
    )
    evidence_limits = sum(
        1
        for row in ledger
        if row.get("disposition") == "container_evidence_limit"
    )
    retained_decisions = sum(1 for row in ledger if row.get("disposition") == "keep")
    documented_exceptions = sum(
        1 for row in ledger if row.get("disposition") == "documented_exception"
    )
    preservation = operations.get("measurement_preservation") or {}
    preservation_counts = preservation.get("counts") or {}
    preservation_families = as_list(preservation.get("families"))
    target_organization = operations.get("target_organization") or {}
    target_naming = target_organization.get("naming") or {}
    target_folders = target_organization.get("folders") or {}
    target_paused = target_organization.get("paused_lifecycle") or {}
    # The detailed target state stays in the evidence tabs and JSON packet.
    # Keep the visible summary scannable even for containers with dozens of
    # measurement families; embedding each family’s full route description here
    # produces clipped/over-height stakeholder rows rather than useful
    # information.
    target_state_summary = join_text(
        [
            f"{family.get('family_id')} {family.get('family_label')} "
            f"[{family.get('preservation_status')}]"
            for family in preservation_families[:12]
        ]
    )
    if len(preservation_families) > 12:
        target_state_summary += (
            f"; +{len(preservation_families) - 12} more reviewed measurement families"
        )
    retained_families = [
        str(row.get("title") or row.get("decision_id") or "")
        for row in ledger
        if row.get("source_run") == "business_architecture"
        and str(row.get("decision_id") or "").startswith("FAM-")
        and row.get("disposition") == "keep"
    ]
    retained_family_summary = join_text(retained_families[:6])
    if len(retained_families) > 6:
        retained_family_summary += f"; +{len(retained_families) - 6} more retained families"
    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    highest_impact = sorted(
        as_list(operations.get("operations")),
        key=lambda row: (
            priority_order.get(str(row.get("priority") or ""), 4),
            str(row.get("operation_id") or ""),
        ),
    )[:3]
    highest_impact_summary = join_text(
        [
            f"[{row.get('priority')}] {row.get('operation_id')}: "
            f"{' '.join(str(row.get('title') or row.get('problem') or '').split())[:110]}"
            for row in highest_impact
        ]
    )
    action_errors = as_list((operations.get("action_completeness") or {}).get("errors"))
    if operations.get("plan_status") != "complete":
        overall_status = (
            "Incomplete cleanup plan; exact actions and owner decisions required"
            if owner_decisions
            else "Incomplete cleanup plan; exact actions required"
        )
    elif owner_decisions and operation_count:
        overall_status = (
            "Cleanup operations ready for scoped approval; owner decisions required for "
            "affected objects"
        )
    elif owner_decisions:
        overall_status = "Owner decisions required for affected retained objects"
    elif evidence_limits:
        overall_status = "Plan ready with documented container evidence limits"
    elif operation_count:
        overall_status = "Ready for human approval"
    else:
        overall_status = "Complete audit and cleanup plan; no cleanup operation is justified"
    next_step = (
        "Resolve every listed action-completeness error and rebuild the cleanup plan."
        if operations.get("plan_status") != "complete"
        else (
            "Approve, reject, or amend proposed operations by scope; resolve each owner "
            "question before changing the objects it covers."
            if operation_count
            else "Resolve the listed owner questions for the affected retained objects."
        )
        if owner_decisions
        else "Approve, reject, or amend the proposed operations before any GTM mutation."
    )
    priority_counts = {
        priority: sum(
            1
            for operation in as_list(operations.get("operations"))
            if operation.get("priority") == priority
        )
        for priority in ("Critical", "High", "Medium", "Low")
    }
    priority_summary = ", ".join(
        f"{priority} {count}"
        for priority, count in priority_counts.items()
        if count
    ) or "no proposed operation"
    projected_deltas = [
        f"{layer} {counts.get('delta', 0):+d}"
        for layer, counts in sorted(
            (operations.get("projected_object_counts") or {}).items()
        )
        if counts.get("delta")
    ]
    clean_modules = sum(
        1
        for row in as_list(operational.get("module_results"))
        if row.get("module_status") == "zero_findings"
    )
    operational_synopsis = (
        f"{operation_count} proposed operation(s) ({priority_summary}); "
        f"{owner_decisions} owner decision(s); {clean_modules} sanitation module(s) "
        f"confirmed clean; {len(preservation_families)} measurement-family target "
        f"state(s); projected object delta: "
        f"{', '.join(projected_deltas) if projected_deltas else 'none'}. "
        f"Next: {next_step}"
    )
    summary = [
        {"Decision": "Overall status", "Value": overall_status},
        {"Decision": "Operational synopsis", "Value": operational_synopsis},
        {"Decision": "Source", "Value": manifest.get("source_file")},
        {"Decision": "Objects reviewed", "Value": len(as_list(configuration.get("rows")))},
        {
            "Decision": "Operational findings reviewed",
            "Value": len(as_list(operational.get("findings"))),
        },
        {
            "Decision": "Business families reviewed",
            "Value": len(as_list(architecture.get("families"))),
        },
        {
            "Decision": "Retained / no-change decisions",
            "Value": retained_decisions,
        },
        {"Decision": "Documented owner exceptions", "Value": documented_exceptions},
        {
            "Decision": "Retained business-family architecture",
            "Value": retained_family_summary or "No retained family decision recorded",
        },
        {
            "Decision": "Measurement-family preservation",
            "Value": join_text(
                [
                    f"{status.replace('_', ' ')}: {count}"
                    for status, count in preservation_counts.items()
                    if count
                ]
            )
            or "No source-confirmed business family was generated",
        },
        {
            "Decision": "Target-state architecture",
            "Value": target_state_summary or "No target-state family summary recorded",
        },
        {
            "Decision": "Target organization",
            "Value": join_text(
                [
                    target_organization.get("status"),
                    target_organization.get("scope"),
                ]
            )
            or "No organization change justified",
        },
        {
            "Decision": "Naming convention and exact renames",
            "Value": (
                f"Policy: {target_naming.get('selected_policy')}. "
                f"Patterns: {join_text(as_list(target_naming.get('target_patterns'))) or 'none'}. "
                f"Exact renames: {len(as_list(target_naming.get('exact_renames')))}. "
                f"Open policy decisions: "
                f"{len(as_list(target_naming.get('confirmation_decision_ids')))}."
            ),
        },
        {
            "Decision": "Folder target and paused lifecycle",
            "Value": (
                f"Exact folder actions: "
                f"{len(as_list(target_folders.get('exact_actions')))}; "
                f"open folder decisions: "
                f"{len(as_list(target_folders.get('unresolved_decision_ids')))}; "
                f"paused tags proposed for retirement: "
                f"{len(as_list(target_paused.get('proposed_retirement_keys')))}; "
                f"paused tags retained/pending: "
                f"{len(as_list(target_paused.get('retained_pending_or_necessary_keys')))}."
            ),
        },
        {
            "Decision": "Preservation evidence boundary",
            "Value": preservation.get("scope")
            or "Container-visible configuration only; runtime certification is separate.",
        },
        {
            "Decision": "Runtime verification",
            "Value": (
                "Out of scope for this container-only audit; no Preview, browser, "
                "network, CMP, vendor, or server-side test was run or planned."
            ),
        },
        {
            "Decision": "Highest-impact proposed actions",
            "Value": highest_impact_summary or "No cleanup operation proposed",
        },
        {"Decision": "Proposed operations", "Value": operation_count},
        {"Decision": "Action-completeness errors", "Value": len(action_errors)},
        {"Decision": "Owner decisions", "Value": owner_decisions},
        {"Decision": "Container evidence limits", "Value": evidence_limits},
        {
            "Decision": "Projected object counts",
            "Value": join_text(
                [
                    f"{layer}: {counts.get('before', 0)} -> {counts.get('after', 0)} "
                    f"({counts.get('delta', 0):+d})"
                    for layer, counts in sorted(
                        (operations.get("projected_object_counts") or {}).items()
                    )
                ]
            )
            or "No count-changing operation proposed",
        },
        {"Decision": "Execution route", "Value": operations.get("route")},
        {"Decision": "Cleanup plan status", "Value": operations.get("plan_status")},
        {
            "Decision": "Next step",
            "Value": next_step,
        },
        {
            "Decision": "Filterable problem taxonomy",
            "Value": (
                "Filter Cleanup Plan by General problem category, or filter Affected "
                "object(s) with tag:, trigger:, variable:, builtInVariable:, folder:, "
                "or customTemplate:. Categories: "
                + ", ".join(sorted(GENERAL_PROBLEM_CATEGORIES))
                + "."
            ),
        },
    ]
    add_table(sheets["01 Summary"], summary, ["Decision", "Value"])
    add_table(
        sheets["02 Cleanup Plan"],
        as_list(human_rows.get("rows")),
        list(CLEANUP_PLAN_COLUMNS),
    )
    from openpyxl.comments import Comment

    sheets["02 Cleanup Plan"]["C1"].comment = Comment(
        "Broad filter. The exact audit lens and problem type remain in the next column.",
        "GTM cleanup skill",
    )
    sheets["02 Cleanup Plan"]["E1"].comment = Comment(
        "Use Text Filters > Contains with tag:, trigger:, variable:, "
        "builtInVariable:, folder:, or customTemplate: to filter by GTM layer.",
        "GTM cleanup skill",
    )
    add_table(
        sheets["03 Operational Review"],
        operational_rows(operational),
        split_long_cells=True,
    )
    add_table(
        sheets["04 Configuration Review"],
        configuration_rows(configuration),
        split_long_cells=True,
    )
    add_table(
        sheets["05 Architecture Review"],
        architecture_rows(architecture),
        split_long_cells=True,
    )
    add_table(
        sheets["06 Custom Code Review"],
        code_rows(configuration),
        split_long_cells=True,
    )
    add_table(
        sheets["07 Reconciled Operations"],
        operation_rows(operations),
        split_long_cells=True,
    )
    add_table(
        sheets["08 Source & Gates"],
        source_rows(manifest, source, operations),
        split_long_cells=True,
    )
    for name in CANONICAL_SHEETS[2:]:
        sheets[name].sheet_state = "hidden"
    workbook.active = 0
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("package_dir", type=Path)
    parser.add_argument("operations", type=Path)
    parser.add_argument("human_rows", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    build_workbook(
        load_json(args.package_dir / "audit_package_manifest.json"),
        load_json(args.package_dir / "source_model.json"),
        load_json(args.package_dir / "operational_review.json"),
        load_json(args.package_dir / "configuration_review.json"),
        load_json(args.package_dir / "architecture_review.json"),
        load_json(args.operations),
        load_json(args.human_rows),
        args.output,
    )
    print(json.dumps({"output": str(args.output), "tabs": CANONICAL_SHEETS}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
