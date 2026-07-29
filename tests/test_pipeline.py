from __future__ import annotations

import copy
import json
import re
import subprocess
import sys
import tempfile
import tomllib
import unittest
import urllib.error
from pathlib import Path
from unittest.mock import MagicMock, patch

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from tests.review_helpers import complete_review_attestation  # noqa: E402

from build_skill_package import build as build_skill_bundle  # noqa: E402
from check_release import (  # noqa: E402
    check_production_test_imports,
    check_project_version,
    check_release_tag,
    check_repository_layout,
    git_ls_files,
)
from gtm_architecture_review import (  # noqa: E402
    scaffold_review as scaffold_architecture,
)
from gtm_architecture_review import (  # noqa: E402
    validate_review as validate_architecture,
)
from gtm_audit_gate_check import validate_workbook  # noqa: E402
from gtm_audit_package_build import build_package  # noqa: E402
from gtm_baseline_audit import audit_export  # noqa: E402
from gtm_change_log_build import build_change_log  # noqa: E402
from gtm_configuration_review import (  # noqa: E402
    scaffold_review as scaffold_configuration,
)
from gtm_configuration_review import (  # noqa: E402
    validate_review as validate_configuration,
)
from gtm_consent_model import tag_consent_route  # noqa: E402
from gtm_context_model import build_context_model  # noqa: E402
from gtm_custom_code_extract import extract_export  # noqa: E402
from gtm_diff_operations import operations as diff_operations  # noqa: E402
from gtm_future_state_check import apply_operations, check_future_state  # noqa: E402
from gtm_human_rows import build_rows  # noqa: E402
from gtm_lib import container_version  # noqa: E402
from gtm_operation_compile import (  # noqa: E402
    action_completeness_report,
    compile_operations,
    operation_priority_basis,
    reconcile_ledger_resolutions,
    runtime_neutral_operational_deletions,
    runtime_qa_handoff,
    source_object_catalog,
)
from gtm_operational_review import (  # noqa: E402  # noqa: E402
    MANDATORY_OPERATIONAL_MODULES,
    mandatory_module_errors,
    validate_deterministic_repair,
)
from gtm_operational_review import (  # noqa: E402
    scaffold_review as scaffold_operational,
)
from gtm_operational_review import (  # noqa: E402
    validate_review as validate_operational,
)
from gtm_privacy import (  # noqa: E402
    privacy_findings,
    redact_text,
    sanitize_url,
    spreadsheet_safe_text,
)
from gtm_privacy_scan import scan_xlsx  # noqa: E402
from gtm_relationships import (  # noqa: E402
    object_records,
    relationship_candidates,
)
from gtm_relationships import (  # noqa: E402
    scan_export as scan_relationships,
)
from gtm_review_common import (  # noqa: E402
    object_consumer_map,
    object_keys,
    object_name_map,
    object_source_path_map,
    validate_operation_set,
    validate_structured_actions,
)
from gtm_review_shards import (  # noqa: E402
    merge_review,
    review_requires_sharding,
    split_review,
)
from gtm_shared_facts import build_shared_facts  # noqa: E402
from gtm_source_model import build_model  # noqa: E402
from gtm_taxonomy import (  # noqa: E402
    CLEANUP_PLAN_COLUMNS,
    GENERAL_CATEGORY_BY_PROBLEM_TYPE,
    GENERAL_PROBLEM_CATEGORIES,
    PROBLEM_TYPES,
    general_problem_category,
)
from gtm_three_run_gate import run_gate  # noqa: E402
from gtm_validate_artifact import missing_references  # noqa: E402
from gtm_validate_artifact import validate as validate_artifact  # noqa: E402
from gtm_vendor_registry import (  # noqa: E402
    load_registry,
    official_url_error,
    validate_registry,
    vendor_record,
)
from gtm_workbook_build import (  # noqa: E402
    CANONICAL_SHEETS,
    MAX_CELL_TEXT,
    MAX_ROW_HEIGHT,
    add_table,
    build_workbook,
)


def condition(operator: str, left: str, right: str) -> dict:
    return {
        "type": operator,
        "parameter": [
            {"type": "TEMPLATE", "key": "arg0", "value": left},
            {"type": "TEMPLATE", "key": "arg1", "value": right},
        ],
    }


def sample_export() -> dict:
    return {
        "exportFormatVersion": 2,
        "containerVersion": {
            "accountId": "100",
            "containerId": "200",
            "containerVersionId": "1",
            "container": {"publicId": "GTM-TEST", "usageContext": ["WEB"]},
            "tag": [
                {
                    "tagId": "1",
                    "name": "GA4 - Purchase - All",
                    "type": "gaawe",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "eventName", "value": "purchase"},
                        {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                        {
                            "type": "MAP",
                            "key": "eventParameters",
                            "map": [
                                {
                                    "type": "TEMPLATE",
                                    "key": "value",
                                    "value": "{{DLV - Value}}",
                                },
                                {
                                    "type": "TEMPLATE",
                                    "key": "items",
                                    "value": "{{DLV - Items}}",
                                },
                                {
                                    "type": "TEMPLATE",
                                    "key": "transaction_id",
                                    "value": "{{DLV - Transaction ID}}",
                                },
                                {
                                    "type": "TEMPLATE",
                                    "key": "currency",
                                    "value": "EUR",
                                },
                            ],
                        },
                    ],
                    "firingTriggerId": ["10"],
                    "blockingTriggerId": ["13"],
                    "setupTag": [{"tagName": "Utility - Consent Defaults"}],
                },
                {
                    "tagId": "2",
                    "name": "Meta - Purchase - All",
                    "type": "html",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": (
                                "<script>\n"
                                "var items = {{DLV - Items}} || [];\n"
                                "fbq('track', 'Purchase', {contents: items});\n"
                                "</script>"
                            ),
                        }
                    ],
                    "firingTriggerId": ["12"],
                },
                {
                    "tagId": "3",
                    "name": "Utility - Consent Defaults",
                    "type": "html",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": "<script>\nwindow.consentDefault = 'denied';\n</script>",
                        }
                    ],
                    "parentFolderId": "101",
                },
                {
                    "tagId": "4",
                    "name": "Paused - Helper Consumer",
                    "type": "html",
                    "paused": True,
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": "<script>void({{CJS - Paused Only}});</script>",
                        }
                    ],
                    "firingTriggerId": ["10"],
                },
            ],
            "trigger": [
                {
                    "triggerId": "10",
                    "name": "Purchase",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "purchase")],
                },
                {
                    "triggerId": "11",
                    "name": "Purchase copy",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "purchase")],
                },
                {
                    "triggerId": "12",
                    "name": "TG - Purchase only",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "10"}],
                        }
                    ],
                },
                {
                    "triggerId": "13",
                    "name": "Block - Page view",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "page_view")],
                },
                {
                    "triggerId": "14",
                    "name": "Click - Invalid regex",
                    "type": "LINK_CLICK",
                    "filter": [condition("MATCH_REGEX", "{{Click URL}}", "(")],
                },
                {
                    "triggerId": "15",
                    "name": "Funnel question 1",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "funnel_question_1")],
                },
                {
                    "triggerId": "16",
                    "name": "Funnel step impression Q1",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [
                        condition("EQUALS", "{{_event}}", "funnel_step_impression")
                    ],
                },
            ],
            "variable": [
                {
                    "variableId": "20",
                    "name": "DLV - Items",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {"type": "TEMPLATE", "key": "name", "value": "ecommerce.items"},
                    ],
                },
                {
                    "variableId": "21",
                    "name": "DLV - Items copy",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {"type": "TEMPLATE", "key": "name", "value": "ecommerce.items"},
                    ],
                },
                {
                    "variableId": "22",
                    "name": "CJS - Page URL Mirror",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function() {\n  return {{Page URL}};\n}",
                        }
                    ],
                },
                {
                    "variableId": "23",
                    "name": "CJS - Paused Only",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function() {\n  return 'paused-value';\n}",
                        }
                    ],
                },
                {
                    "variableId": "24",
                    "name": "DLV - Value",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {"type": "TEMPLATE", "key": "name", "value": "ecommerce.value"},
                    ],
                },
                {
                    "variableId": "25",
                    "name": "DLV - Transaction ID",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {
                            "type": "TEMPLATE",
                            "key": "name",
                            "value": "ecommerce.transaction_id",
                        },
                    ],
                },
            ],
            "folder": [
                {"folderId": "100", "name": "Unused folder"},
                {"folderId": "101", "name": "Utilities"},
            ],
            "builtInVariable": [
                {"name": "Page URL", "type": "PAGE_URL"},
                {"name": "Click URL", "type": "CLICK_URL"},
            ],
        },
    }


def fixed_slot_formula_export() -> dict:
    data = sample_export()
    variables = data["containerVersion"]["variable"]
    for variable_id, index in (("30", 1), ("31", 2), ("32", 3)):
        variables.append(
            {
                "variableId": variable_id,
                "name": f"DLV - Product Price {index}",
                "type": "v",
                "parameter": [
                    {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                    {
                        "type": "TEMPLATE",
                        "key": "name",
                        "value": f"ecommerce.product_price_{index}",
                    },
                ],
            }
        )
    variables.append(
        {
            "variableId": "33",
            "name": "CJS - Total Price",
            "type": "jsm",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "javascript",
                    "value": (
                        "function() {\n"
                        "  return Number({{DLV - Product Price 1}} || 0) + "
                        "Number({{DLV - Product Price 2}} || 0) + "
                        "Number({{DLV - Product Price 3}} || 0);\n"
                        "}"
                    ),
                }
            ],
        }
    )
    event_parameters = data["containerVersion"]["tag"][0]["parameter"][2]["map"]
    event_parameters[0]["value"] = "{{CJS - Total Price}}"
    return data


def object_specific_text(row: dict, subject: str, field: str = "") -> str:
    field_terms = [
        str(value)
        for value in (row.get("field_evidence_requirements") or {}).get(field, [])[:3]
    ]
    if not field_terms:
        field_terms = [row["object_name"] or row["object_key"], row["object_type"]]
    field_terms.extend(
        str(value)
        for value in row.get("specificity_tokens", [])[:2]
        if str(value) not in field_terms
    )
    evidence = " and ".join(field_terms)
    return (
        f"{row['object_name'] or row['object_key']} is a {row['object_type']} that {subject}; "
        f"the exported configuration specifically names {evidence}."
    )


def concrete_purpose_subject(row: dict) -> str:
    return {
        "tag": "sends, loads, routes, or records the configured measurement action",
        "trigger": "matches and activates the configured event or condition scope",
        "variable": "reads, calculates, maps, or returns its configured value",
        "zone": "restricts, scopes, allows, and governs the configured child container",
        "customTemplate": "defines and executes the exported template behavior",
        "client": "claims, parses, and routes the configured request",
        "gtagConfig": "configures, routes, sets, and governs the Google tag behavior",
        "transformation": "transforms, allows, or redacts the configured fields",
    }.get(row.get("layer"), "implements its concrete exported action")


def branch_role(path: str) -> str:
    lowered = path.lower()
    if "consent" in lowered or "storage" in lowered:
        return "Consent"
    if "filter" in lowered or "condition" in lowered or "operator" in lowered:
        return "Condition"
    if any(value in lowered for value in ("firingtriggerid", "blockingtriggerid", "triggerids")):
        return "Routing"
    if any(
        value in lowered
        for value in (
            "setuptag",
            "teardowntag",
            "tagfiringoption",
            "schedulestartms",
            "scheduleendms",
        )
    ):
        return "Execution control"
    if "childcontainer" in lowered or "typerestriction" in lowered:
        return "Condition"
    return "Input"


def contract_topic_anchors(
    topic: dict, fact_by_path: dict, row: dict, contract_anchor: str
) -> list[str]:
    events = [str(value).lower() for value in topic.get("configured_event_values", [])]
    anchors = [
        path
        for path, fact in fact_by_path.items()
        if path in row["required_logic_anchors"]
        and any(
            event in str(fact.get("value_preview") or "").lower() for event in events
        )
    ]
    return anchors or [contract_anchor]


def behavior_signal_text(fact: dict) -> str:
    signals = []
    for signal in fact.get("required_behavior_signals", []):
        required_terms = [
            str(group[0])
            for group in signal.get("required_term_groups", [])
            if isinstance(group, list) and group
        ]
        signals.append(
            f"{str(signal.get('signal') or 'source behavior').replace('_', ' ')} "
            f"uses {' '.join(required_terms)}"
        )
    return "; ".join(signals)


def complete_configuration(
    export_path: Path, shared_facts: dict | None = None
) -> dict:
    review = scaffold_configuration(export_path, shared_facts=shared_facts)
    for row in review["rows"]:
        row.update(
            {
                "review_status": "complete",
                "purpose": object_specific_text(
                    row, concrete_purpose_subject(row), "purpose"
                ),
                "execution_logic": object_specific_text(
                    row, "runs under the named trigger, condition, or call route", "execution_logic"
                ),
                "inputs_and_terminal_sources": object_specific_text(
                    row,
                    "reads the listed GTM references and terminal configuration",
                    "inputs_and_terminal_sources",
                ),
                "configured_output_or_side_effect": object_specific_text(
                    row,
                    "produces the named event, return value, or browser side effect",
                    "configured_output_or_side_effect",
                ),
                "consumer_contract": object_specific_text(
                    row,
                    "supplies the listed consumer objects with that configured value",
                    "consumer_contract",
                ),
                "consent_and_sequence": object_specific_text(
                    row,
                    "uses the listed blocking, consent, and sequencing controls",
                    "consent_and_sequence",
                ),
                "correctness_verdict": "Correct",
                "correctness_basis": object_specific_text(
                    row,
                    "has matching inputs, route, output, and consumers in this source",
                    "correctness_basis",
                ),
                "defects": [],
                "contract_checks": [],
                "code_behavior_blocks": [],
                "technical_facts_assessment": "",
                "technical_finding_reviews": [],
                "configuration_branch_reviews": [
                    {
                        "json_path": branch["json_path"],
                        "value_hash": branch["value_hash"],
                        "logic_role": branch_role(branch["json_path"]),
                        "interpretation": (
                            f"At {branch['json_path']}, value {branch.get('value_preview')} is the "
                            f"{branch_role(branch['json_path']).lower()} branch for "
                            f"{row['object_name'] or row['object_key']}."
                        ),
                        "configured_effect": (
                            f"The {branch_role(branch['json_path']).lower()} setting "
                            f"{branch.get('value_preview')} at {branch['json_path']} is read, "
                            "matched, routed, or applied when that configured branch executes for "
                            f"{row['object_name'] or row['object_key']}."
                        ),
                        "correctness": "Correct",
                    }
                    for branch in row["required_branch_reviews"]
                ],
                "evidence_anchors": list(row["required_logic_anchors"]),
                "consumer_evidence_keys": [
                    item["consumer_key"] for item in row["export_consumers"]
                ],
                "reference_traces": [
                    {
                        "reference": item["reference"],
                        "object_chain": item["required_object_keys"],
                        "evidence_anchors": item["required_evidence_anchors"],
                        "terminal_states": item["terminal_states"],
                        "terminal_source": (
                            f"Reference {item['reference']} terminates in the source states "
                            f"{', '.join(item['terminal_states'])} after the listed variable chain."
                        ),
                        "node_reviews": [
                            {
                                "object_key": node["object_key"],
                                "object_name": node["object_name"],
                                "object_type": node["object_type"],
                                "config_hash": node["config_hash"],
                                "source_json_path": node["source_json_path"],
                                "referenced_variables": node["referenced_variables"],
                                "configured_parameters": node["configured_parameters"],
                                "semantic_role": node["semantic_role"],
                                "evidence_anchors": node["required_evidence_anchors"],
                                "configured_function": (
                                    f"{node['object_name']} ({node['object_type']}) reads "
                                    f"{' and '.join(node['specificity_tokens'][:3])} from its parameters."
                                ),
                                "configured_output": (
                                    f"{node['object_name']} ({node['object_type']}) returns the value "
                                    f"selected by {' and '.join(node['specificity_tokens'][:3])}."
                                ),
                                "output_type_and_shape": (
                                    f"{node['object_name']} keeps the {node['object_type']} output "
                                    f"shape associated with {' and '.join(node['specificity_tokens'][:2])}."
                                ),
                                "availability_and_fallback": (
                                    f"{node['object_name']} makes {' and '.join(node['specificity_tokens'][:2])} "
                                    "available only where its listed source exists, with no extra fallback."
                                ),
                                "consumer_compatibility": (
                                    f"{node['object_name']} supplies its {node['object_type']} and "
                                    f"{' and '.join(node['specificity_tokens'][:2])} value to "
                                    f"{row['object_name'] or row['object_key']}."
                                ),
                            }
                            for node in item["required_nodes"]
                        ],
                        "edge_reviews": [
                            {
                                **edge,
                                "dependency_meaning": (
                                    f"{edge['from_object_key']} reads {edge['reference']} from "
                                    f"{edge['to_object_key']} before returning its configured result."
                                ),
                            }
                            for edge in item["required_edges"]
                        ],
                        "terminal_reviews": [
                            {
                                **terminal,
                                "terminal_meaning": (
                                    f"{terminal['reference']} resolves to {terminal['configured_source']} "
                                    f"as the final {terminal['state']} source."
                                ),
                                "consumer_compatibility": (
                                    f"The {terminal['reference']} {terminal['state']} source supplies "
                                    f"the configured value to {row['object_name'] or row['object_key']}."
                                ),
                            }
                            for terminal in item["terminal_requirements"]
                        ],
                    }
                    for item in row["reference_trace_requirements"]
                ],
                "logic_cross_checks": [
                    {
                        "check_key": check["check_key"],
                        "verdict": "Aligned",
                        "conclusion": (
                            f"For {row['object_name'] or row['object_key']}, "
                            f"{check['question']} The exported facts "
                            f"{' and '.join(check['required_terms'][:2])} remain aligned in this "
                            "controlled fixture configuration."
                        ),
                        "evidence_anchors": list(check["allowed_evidence_anchors"][:2]),
                    }
                    for check in row["required_logic_cross_checks"]
                ],
                "disposition": "keep",
                "owner_question": "",
                "operation": {},
                "confidence": "High",
                "evidence_citations": {
                    field: list((row.get("field_evidence_paths") or {}).get(field, []))[
                        : 2 if field == "correctness_basis" else 1
                    ]
                    for field in (
                        "purpose",
                        "execution_logic",
                        "inputs_and_terminal_sources",
                        "configured_output_or_side_effect",
                        "consumer_contract",
                        "consent_and_sequence",
                        "correctness_basis",
                    )
                },
            }
        )
        if row["required_contract_topics"]:
            fact_by_path = {
                fact["json_path"]: fact for fact in row.get("source_facts", [])
            }
            contract_anchor = row["required_logic_anchors"][0]

            row["contract_checks"] = [
                {
                    "contract_topic": topic["topic_key"],
                    "contract_field": (f"{topic['vendor']} {topic['topic']} exported contract"),
                    "configured_value": (
                        f"At {contract_topic_anchors(topic, fact_by_path, row, contract_anchor)[0]}, "
                        "the exported value "
                        f"{fact_by_path[contract_topic_anchors(topic, fact_by_path, row, contract_anchor)[0]].get('value_preview')} configures "
                        f"{topic['topic']} for {row['object_name']}; vendor-specific events are "
                        f"{', '.join(topic.get('configured_event_values', [])) or 'not exported'}"
                    ),
                    "expected_rule": (
                        f"The official {topic['vendor']} documentation defines the required "
                        f"{topic['topic']} behavior and value types"
                    ),
                    "source": (
                        topic["official_doc_candidates"][0]
                        if topic["official_doc_candidates"]
                        else ""
                    ),
                    "identified_vendor": topic["vendor"],
                    "official_source_basis": (
                        f"The cited page is the registered official {topic['vendor']} reference, "
                        "or no authoritative vendor identity and source is visible for this "
                        "unclassified integration."
                    ),
                    "research_status": (
                        "Searched current official vendor documentation using the exported "
                        "integration hostname and code identifiers, but no authoritative vendor "
                        "identity or official source could be established."
                        if topic.get("research_required")
                        and not topic["official_doc_candidates"]
                        else "Official source is registered for this detected vendor."
                    ),
                    "verdict": (
                        "Non-compliant"
                        if topic.get("deterministic_contract_state")
                        == "known_noncompliant"
                        else "Unproven"
                        if topic.get("deterministic_contract_state")
                        == "unproven_from_container"
                        or (
                            topic.get("research_required")
                            and not topic["official_doc_candidates"]
                        )
                        else "Compliant"
                    ),
                    "evidence_anchors": contract_topic_anchors(
                        topic, fact_by_path, row, contract_anchor
                    ),
                }
                for topic in row["required_contract_topics"]
            ]
        if row["required_code_line_hashes"]:
            previews = " ".join(
                str(item.get("line_preview") or "") for item in row["code_line_facts"]
            )
            markers = [
                token
                for token in re.findall(r"[A-Za-z_$][A-Za-z0-9_.$:/-]{3,}", previews)
                if token.lower() not in {"function", "return", "const", "false", "true", "script"}
            ]
            marker = markers[0] if markers else row["object_name"]
            segment_marker_text = " and ".join(dict.fromkeys(markers)) or marker
            required_behavior_text = "; ".join(
                text
                for fact in row["code_line_facts"]
                if (text := behavior_signal_text(fact))
            )
            row["code_behavior_blocks"] = [
                {
                    "line_hashes": list(row["required_code_line_hashes"]),
                    "start_line": min(item["line_number"] for item in row["code_line_facts"]),
                    "end_line": max(item["line_number"] for item in row["code_line_facts"]),
                    "purpose": f"The {marker} block implements the exact exported helper behavior.",
                    "inputs": f"The {marker} block reads only the variables and literals visible here.",
                    "outputs": f"The {marker} block returns or sends the output shown by these lines.",
                    "side_effects": f"The {marker} block has the browser effects identified in static facts.",
                    "health_assessment": (
                        f"The {marker} implementation is coherent for this controlled fixture; "
                        f"every source segment is identified by {segment_marker_text}. "
                        f"Source-visible behavior: {required_behavior_text or 'no additional static signal'}."
                    ),
                }
            ]
            row["technical_facts_assessment"] = (
                f"The {marker} code facts, parser result, side effects, and line behavior are "
                "accounted for in this container-only assessment."
            )
            fallback_segment_reviews = []
            for fact in row["code_line_facts"]:
                fact_markers = [
                    token
                    for token in re.findall(
                        r"[A-Za-z_$][A-Za-z0-9_.$:/-]{3,}",
                        str(fact.get("line_preview") or ""),
                    )
                    if token.lower()
                    not in {
                        "function",
                        "return",
                        "const",
                        "false",
                        "true",
                        "undefined",
                        "script",
                    }
                ][:4]
                fact_marker_text = (
                    " and ".join(fact_markers) or "the exported syntax boundary"
                )
                fact_behavior_text = behavior_signal_text(fact)
                fallback_segment_reviews.append(
                    {
                        "line_hash": fact["line_hash"],
                        "behavior": (
                            "Mandatory line-by-line review of this segment identifies "
                            f"{fact_marker_text} and maps its inputs, output, side effects, "
                            "and execution behavior. Source-visible behavior: "
                            f"{fact_behavior_text or 'no additional static signal'}."
                        ),
                    }
                )
            row["technical_finding_reviews"] = [
                {
                    "finding_key": item["finding_key"],
                    "source_statement": item["statement"],
                    "verdict": (
                        "Documented exception"
                        if item["category"] in {"parser", "health", "security"}
                        else "False positive"
                    ),
                    "rationale": (
                        (
                            f"The {marker} mandatory line-by-line code blocks cover every "
                            "exported line; the parser boundary is recorded without claiming "
                            "AST coverage."
                        )
                        if item["category"] == "parser"
                        else (
                            f"The {marker} controlled fixture documents the exact source signal "
                            f"{item['statement']} with identifiers {segment_marker_text}; its "
                            "exported behavior remains explicit and accepted for this test."
                        )
                    ),
                    "proposed_action": "",
                    "exception_basis": (
                        f"This controlled fixture accepts the source-proven risk "
                        f"{item['statement']} because its behavior is required and retained "
                        "under an explicit test constraint."
                        if item["category"] in {"health", "security"}
                        else ""
                    ),
                    "owner_question": "",
                    "fallback_line_hashes": (
                        list(row["required_code_line_hashes"])
                        if item["category"] == "parser"
                        else []
                    ),
                    "parser_boundary": (
                        f"The exact parser status "
                        f"{row['technical_code_facts'].get('javascript_parser')} leaves syntax "
                        "coverage incomplete for GTM substitutions or template wrapper code."
                        if item["category"] == "parser"
                        else ""
                    ),
                    "manual_review_method": (
                        f"A mandatory line-by-line review follows source identifiers "
                        f"{' and '.join(markers[:4]) or marker}, maps every segment hash to its "
                        "inputs, output, side effects, and execution branch, and claims no AST proof."
                        if item["category"] == "parser"
                        else ""
                    ),
                    "fallback_segment_reviews": (
                        copy.deepcopy(fallback_segment_reviews)
                        if item["category"] == "parser"
                        else []
                    ),
                }
                for item in row["required_technical_findings"]
            ]
        elif row["required_technical_findings"]:
            row["technical_finding_reviews"] = [
                {
                    "finding_key": item["finding_key"],
                    "source_statement": item["statement"],
                    "verdict": "Documented exception",
                    "rationale": (
                        f"The exported {row['object_name'] or row['object_key']} metadata "
                        f"proves the bounded source limitation {item['statement']} and exposes "
                        "no executable segment that could receive a fabricated code review."
                    ),
                    "proposed_action": "",
                    "exception_basis": (
                        f"The controlled fixture retains the exact source boundary "
                        f"{item['statement']} while requiring executable template source "
                        "before correctness certification."
                    ),
                    "owner_question": "",
                    "fallback_line_hashes": [],
                    "parser_boundary": "",
                    "manual_review_method": "",
                    "fallback_segment_reviews": [],
                }
                for item in row["required_technical_findings"]
            ]
        issue_obligations = [
            item
            for item in row["required_configuration_obligations"]
            if item["required_outcome"] == "Issue"
        ]
        unclear_obligations = [
            item
            for item in row["required_configuration_obligations"]
            if item["required_outcome"] == "Unclear"
        ]
        required_topic_by_key = {
            item["topic_key"]: item for item in row["required_contract_topics"]
        }
        issue_contract_topics = {
            topic
            for item in issue_obligations
            for topic in item.get("affected_contract_topics", [])
        }
        unclear_contract_topics = {
            topic
            for item in unclear_obligations
            for topic in item.get("affected_contract_topics", [])
        }
        for check in row["contract_checks"]:
            topic = required_topic_by_key[check["contract_topic"]]["topic"]
            if topic in issue_contract_topics:
                check["verdict"] = "Non-compliant"
            elif topic in unclear_contract_topics and check["verdict"] == "Compliant":
                check["verdict"] = "Unproven"
        issue_paths = {
            anchor
            for item in issue_obligations
            for anchor in item["evidence_anchors"]
        }
        unclear_paths = {
            anchor
            for item in unclear_obligations
            for anchor in item["evidence_anchors"]
        }
        noncompliant_checks = [
            check for check in row["contract_checks"] if check["verdict"] == "Non-compliant"
        ]
        unproven_checks = [
            check for check in row["contract_checks"] if check["verdict"] == "Unproven"
        ]
        issue_paths.update(
            anchor
            for check in noncompliant_checks
            for anchor in check["evidence_anchors"]
        )
        unclear_paths.update(
            anchor
            for check in unproven_checks
            for anchor in check["evidence_anchors"]
            if anchor not in issue_paths
        )
        for branch in row["configuration_branch_reviews"]:
            if branch["json_path"] in issue_paths:
                branch["correctness"] = "Issue"
            elif branch["json_path"] in unclear_paths:
                branch["correctness"] = "Unclear"
        issue_check_keys = {
            key
            for item in issue_obligations
            for key in item["affected_logic_checks"]
        }
        unclear_check_keys = {
            key
            for item in unclear_obligations
            for key in item["affected_logic_checks"]
        }
        if noncompliant_checks:
            issue_check_keys.add("vendor_contract_alignment")
        elif unproven_checks:
            unclear_check_keys.add("vendor_contract_alignment")
        for check in row["logic_cross_checks"]:
            requirement = next(
                item
                for item in row["required_logic_cross_checks"]
                if item["check_key"] == check["check_key"]
            )
            if check["check_key"] in issue_check_keys:
                check["verdict"] = "Issue"
                issue_evidence = issue_paths | {
                    anchor
                    for contract in noncompliant_checks
                    for anchor in contract["evidence_anchors"]
                }
                allowed = set(requirement["allowed_evidence_anchors"])
                check["evidence_anchors"] = (
                    sorted(issue_evidence & allowed)[:2]
                    or sorted(allowed)[:1]
                )
                check["conclusion"] = (
                    f"For {row['object_name'] or row['object_key']}, the exported facts "
                    f"{' and '.join(requirement['required_terms'][:2])} "
                    "contain a deterministic configuration defect and are not aligned."
                )
            elif check["check_key"] in unclear_check_keys:
                check["verdict"] = "Unclear"
                check["conclusion"] = (
                    f"For {row['object_name'] or row['object_key']}, the exported facts "
                    f"{' and '.join(requirement['required_terms'][:2])} "
                    "leave a specific container-only contract unproven."
                )
            relevant_obligations = [
                item
                for item in row["required_configuration_obligations"]
                if check["check_key"] in item["affected_logic_checks"]
            ]
            if relevant_obligations:
                check["conclusion"] += " " + " ".join(
                    f"Obligation {item['obligation_key']}: {item['statement']}"
                    for item in relevant_obligations
                )
        row["defects"] = [
            {
                "defect_id": f"AUTO-{index:03d}",
                "statement": item["statement"],
                "configured_effect": (
                    f"The exported state for {item['obligation_key']} makes the configured "
                    "execution, payload, or dependency behavior invalid."
                ),
                "expected_behavior": (
                    "The source must expose a valid, resolvable, officially supported "
                    "configuration before this object can be certified."
                ),
                "evidence_anchors": list(item["evidence_anchors"]),
                "code_line_hashes": [],
                "technical_finding_keys": [],
            }
            for index, item in enumerate(issue_obligations, start=1)
        ]
        defect_index = len(row["defects"])
        for check in noncompliant_checks:
            defect_index += 1
            row["defects"].append(
                {
                    "defect_id": f"AUTO-{defect_index:03d}",
                    "statement": (
                        f"Official contract topic {check['contract_topic']} is non-compliant."
                    ),
                    "configured_effect": (
                        f"The exported value in {check['configured_value']} does not satisfy "
                        "the cited vendor contract."
                    ),
                    "expected_behavior": (
                        f"Use the officially supported names, required fields, and value types "
                        f"for {check['contract_topic']}."
                    ),
                    "evidence_anchors": list(check["evidence_anchors"]),
                    "code_line_hashes": [],
                    "technical_finding_keys": [],
                }
            )
        required_topic_by_key = {
            item["topic_key"]: item for item in row["required_contract_topics"]
        }
        runtime_unproven_checks = [
            check
            for check in unproven_checks
            if required_topic_by_key.get(check["contract_topic"], {}).get(
                "deterministic_contract_state"
            )
            == "unproven_from_container"
        ]
        blocking_unproven_checks = [
            check for check in unproven_checks if check not in runtime_unproven_checks
        ]
        technical_limits = [
            str(value)
            for value in (row.get("technical_code_facts") or {}).get(
                "container_evidence_limits", []
            )
            if str(value)
            and not str(value).startswith("No material external behavior")
            and "does not expose executable behavior" not in str(value).lower()
        ]
        runtime_required = bool(runtime_unproven_checks or technical_limits)
        row["external_evidence_status"] = (
            "runtime_handoff_required" if runtime_required else "none"
        )
        row["external_evidence_summary"] = (
            f"{row['object_key']} ({row['object_name']}) has container-visible "
            f"configuration that can be judged separately, while runtime topics "
            f"{[check['contract_topic'] for check in runtime_unproven_checks]!r} "
            f"and external effects {technical_limits!r} remain unproven."
            if runtime_required
            else ""
        )
        row["external_evidence_next_action"] = (
            f"In GTM Preview, test {row['object_key']} on its affected event route; "
            "capture resolved inputs, consent state, browser requests, and the "
            "destination response required by the listed contract."
            if runtime_required
            else ""
        )
        unresolved = bool(unclear_obligations or blocking_unproven_checks)
        if row["defects"]:
            first_defect = row["defects"][0]
            first_anchor = next(
                (
                    str(value)
                    for value in first_defect.get("evidence_anchors", [])
                    if str(value)
                ),
                row["required_logic_anchors"][0],
            )
            row["correctness_verdict"] = "Issue"
            row["correctness_basis"] = object_specific_text(
                row,
                "contains the listed deterministic defects and cannot be certified as correct",
                "correctness_basis",
            )
            row["disposition"] = "owner_decision_needed"
            row["owner_question"] = (
                f"For {row['object_key']}, which valid replacement value or route should "
                f"correct defect {first_defect['defect_id']} at {first_anchor}?"
            )
            row["recommended_action"] = (
                f"For {row['object_key']}, repair defect {first_defect['defect_id']} at "
                f"{first_anchor} by replacing the invalid exported configuration with the "
                "owner-approved valid value; retain the object only after reference and "
                "measurement-preservation QA."
            )
            row["confidence"] = "High"
        elif unresolved:
            unresolved_contracts = [
                check["contract_topic"]
                for check in [*blocking_unproven_checks, *runtime_unproven_checks]
            ]
            unresolved_scope = (
                ", ".join(unresolved_contracts[:2])
                or row["required_logic_anchors"][0]
            )
            row["correctness_verdict"] = "Owner decision needed"
            row["correctness_basis"] = object_specific_text(
                row,
                "leaves the listed runtime, external, or owner-controlled contract unproven",
                "correctness_basis",
            )
            row["disposition"] = "owner_decision_needed"
            row["owner_question"] = (
                f"For {row['object_key']}, what runtime or owner evidence establishes "
                f"the required {unresolved_scope} contract?"
            )
            row["recommended_action"] = (
                f"Test {row['object_key']} against {unresolved_scope}; retain its current "
                "configuration only if the captured route and destination evidence satisfies "
                "that contract, otherwise prepare the field-level correction."
            )
            row["confidence"] = "Medium"
        for technical_review in row["technical_finding_reviews"]:
            if "no reviewable executable behavior" in technical_review[
                "source_statement"
            ].lower():
                technical_review["verdict"] = "Owner decision needed"
                technical_review["rationale"] = (
                    "The exported template metadata does not expose executable behavior, so an "
                    "owner must provide the original template source or approve removal."
                )
                technical_review["owner_question"] = (
                    "Which owner can provide the custom-template executable source and permissions, "
                    "or approve removal of the opaque implementation?"
                )
    review["run_status"] = "complete"
    review["completion_attestation"] = complete_review_attestation(
        review, decision_authoring_method="independent_test_fixture_review"
    )
    return review


def complete_operational(export_path: Path) -> dict:
    review = scaffold_operational(export_path)
    for row in review["findings"]:
        base = {
            "review_status": "complete",
            "rationale": (
                f"The source evidence {' and '.join(row['rationale_evidence_terms'][:2])} "
                "supports the recorded controlled-fixture resolution."
            ),
        }
        repair = row.get("deterministic_repair") or {}
        exact_repair = str(repair.get("status") or "").startswith("unique_")
        if row.get("deterministic_action_candidate") == "delete_candidate":
            source_keys = [
                str(value)
                for value in row.get("shared_fact_object_keys", [])
                if str(value)
            ]
            row.update(
                {
                    **base,
                    "disposition": "cleanup_operation",
                    "operation_key": (
                        "delete-" + str(row["finding_id"]).lower().replace("_", "-")
                    ),
                    "title": f"Remove source-proven unused {row['finding_type']}",
                    "area": "GTM hygiene",
                    "problem_type": "Unused object",
                    "problem": row["deterministic_evidence"],
                    "why_it_matters": (
                        "An export-unreachable object adds maintenance ambiguity without "
                        "supporting the active measurement graph."
                    ),
                    "expected_clean_state": (
                        "The unused object is absent and every retained reference still resolves."
                    ),
                    "exact_proposed_action": (
                        "Delete " + ", ".join(source_keys) + " after the final dependency readback."
                    ),
                    "canonical_object_key": "",
                    "canonical_selection_rationale": "",
                    "creations": [],
                    "additions": [],
                    "changes": [],
                    "remaps": [],
                    "deletions": [
                        {
                            "object_key": key,
                            "reason": (
                                "The complete source execution graph has no active "
                                "consumer for this object."
                            ),
                        }
                        for key in source_keys
                    ],
                    "renames": [],
                    "preconditions": (
                        "Re-read the workspace and confirm the object remains outside "
                        "every active consumer and sequence."
                    ),
                    "qa_steps": (
                        "Re-export and confirm the object is absent, references resolve, "
                        "and active object counts match the projected delta."
                    ),
                    "rollback": (
                        "Restore the exact object from the source export if a dependency "
                        "appears during readback."
                    ),
                    "priority": "Low",
                    "confidence": "High",
                    "execution_readiness": "approval_required",
                    "owner_question": "",
                    "recommended_action": "",
                    "challenge_review": {},
                }
            )
        elif exact_repair:
            repair_kind = str(repair.get("repair_kind") or "")
            canonical_key = (
                str(repair.get("target_object_key") or "")
                if repair_kind
                in {"variable_reference", "setupTag", "teardownTag"}
                else ""
            )
            problem_type = (
                "Naming inconsistency"
                if repair_kind == "object_name"
                else "Over-firing"
                if repair_kind == "blockingTriggerId"
                else "Broken reference"
            )
            row.update(
                {
                    **base,
                    "disposition": "cleanup_operation",
                    "operation_key": (
                        "repair-" + str(row["finding_id"]).lower().replace("_", "-")
                    ),
                    "title": (
                        f"Apply source-proven {repair_kind} repair"
                    ),
                    "area": (
                        "Event firing logic"
                        if repair_kind == "blockingTriggerId"
                        else "GTM hygiene"
                        if repair_kind == "object_name"
                        else "Tracking plan / dataLayer"
                    ),
                    "problem_type": problem_type,
                    "problem": row["deterministic_evidence"],
                    "why_it_matters": (
                        "The exact source mismatch leaves stale metadata, a broken "
                        "reference, or an ineffective execution control in the container."
                    ),
                    "expected_clean_state": (
                        "Every listed source field equals the generated canonical value "
                        "while unrelated configuration remains unchanged."
                    ),
                    "exact_proposed_action": row["default_action"],
                    "canonical_object_key": canonical_key,
                    "canonical_selection_rationale": (
                        f"{canonical_key} is the only container-visible target with a "
                        "matching active consumer route, configuration, and source name."
                        if canonical_key
                        else ""
                    ),
                    "creations": [],
                    "additions": [],
                    "changes": copy.deepcopy(repair.get("changes", [])),
                    "remaps": [],
                    "deletions": copy.deepcopy(repair.get("deletions", [])),
                    "renames": copy.deepcopy(repair.get("renames", [])),
                    "preconditions": (
                        "Re-read every listed object and confirm its before value still "
                        "matches the source export."
                    ),
                    "qa_steps": (
                        "Re-export and verify the exact changed field or name, all "
                        "references, and the affected tag firing route."
                    ),
                    "rollback": (
                        "Restore each exact before value or name recorded in this "
                        "operation if post-change QA fails."
                    ),
                    "priority": "Medium",
                    "confidence": "High",
                    "execution_readiness": "approval_required",
                    "owner_question": "",
                    "recommended_action": "",
                    "challenge_review": {},
                }
            )
        else:
            source_scope = [
                str(value)
                for value in (
                    row.get("shared_fact_object_keys")
                    or row.get("object_identities")
                    or row.get("object_ids")
                    or []
                )
                if str(value)
            ]
            scope_text = ", ".join(source_scope[:3]) or row["finding_id"]
            row.update(
                {
                    **base,
                    "disposition": "owner_decision_needed",
                    "owner_question": (
                        f"For {scope_text}, which exported route or approved business purpose "
                        f"requires retention of this {row['finding_type']} finding?"
                    ),
                    "recommended_action": (
                        f"For {scope_text}, apply the source recommendation "
                        f"{row['default_action']} Retain the current state only when the owner "
                        "documents the cited route or business distinction."
                    ),
                }
            )
    review["run_status"] = "complete"
    review["completion_attestation"] = complete_review_attestation(
        review, decision_authoring_method="independent_test_fixture_review"
    )
    return review


def member_assessments(
    item: dict,
    keys_field: str,
    anchors_field: str,
    paused_field: str,
    terms_field: str,
) -> list:
    distinguishing_map = (
        item.get("member_distinguishing_terms")
        or item.get("candidate_distinguishing_terms")
        or {}
    )
    return [
        {
            "object_key": key,
            "configured_role": (
                f"{key} uses {' and '.join(item[terms_field][key][:2])} to perform its "
                "specific role in this measurement chain."
            ),
            "necessity": (
                f"{key} remains necessary while {item[terms_field][key][0]} has a distinct "
                "consumer, route, or terminal source."
            ),
            "distinguishing_configuration": (
                f"{key} is distinguished by "
                f"{' and '.join((distinguishing_map.get(key) or item[terms_field][key])[:2])} "
                "in its route, payload, or dependency configuration; its own source facts include "
                f"{' and '.join(item[terms_field][key][:2])}."
            ),
            "status": "paused" if item[paused_field].get(key, False) else "active",
            "evidence_anchors": item[anchors_field][key][:1],
        }
        for key in item[keys_field]
    ]


def architecture_text(row: dict, field: str, statement: str) -> str:
    terms = list((row.get("field_evidence_requirements") or {}).get(field, [])[:3])
    for token in row.get("chain_specificity_tokens", [])[:3]:
        if token not in terms:
            terms.append(token)
    return f"{statement}; the source specifically includes {' and '.join(terms)}."


def architecture_caution_text(cautions: list[dict]) -> str:
    keys = {str(item.get("caution_key") or "") for item in cautions}
    statements = []
    if "deduplication_alignment_unproven" in keys:
        statements.append(
            "Runtime deduplication through event ID or transaction ID remains unproven from "
            "the visible container"
        )
    if "consent_alignment_unproven_or_conflicting" in keys:
        statements.append(
            "end-to-end browser and server consent alignment remains unproven and unresolved"
        )
    return (" " + "; ".join(statements) + ".") if statements else ""


def unsafe_owner_question(comparison_types: set[str], candidate_keys: list[str]) -> str:
    identities = " and ".join(candidate_keys[:2])
    prefix = f"For {identities}, "
    if "browser_server_consent_deduplication_review" in comparison_types:
        return prefix + (
            "Which owner approves the browser and server routes, and what evidence resolves "
            "their consent forwarding and event-ID deduplication contract?"
        )
    if "shared_zone_child_container" in comparison_types:
        return prefix + (
            "Which Zone owns the shared child container, and what non-overlapping boundary "
            "scope justifies retaining both Zone routes?"
        )
    if "cyclic_trigger_group_dependency" in comparison_types:
        return prefix + (
            "Which trigger-group dependency should be removed to break the exported cycle, "
            "and which trigger route is canonical?"
        )
    return prefix + (
        "Which tag route is canonical, and what trigger or consent distinction justifies "
        "retaining the same payload on the other route?"
    )


def exact_duplicate_operation_from_comparison(row: dict) -> dict:
    canonical = str(row["recommended_canonical_object_key"])
    duplicates = [
        str(key) for key in row["candidate_object_keys"] if str(key) != canonical
    ]
    remaps = [
        {
            "from_object_key": duplicate,
            "to_object_key": canonical,
            "consumer_object_keys": list(
                (row.get("candidate_consumer_keys") or {}).get(duplicate, [])
            ),
        }
        for duplicate in duplicates
        if (row.get("candidate_consumer_keys") or {}).get(duplicate)
    ]
    return {
        "operation_key": f"ARCH-{row['comparison_id']}-CONSOLIDATE",
        "title": f"Consolidate exact duplicate {row['comparison_id']}",
        "area": "GTM hygiene",
        "problem_type": "Exact duplicate",
        "problem": (
            f"{', '.join(duplicates)} duplicates {canonical} with identical "
            "source-visible configuration."
        ),
        "why_it_matters": (
            "Keeping identical copies adds ownership ambiguity and future drift risk."
        ),
        "expected_clean_state": (
            f"{canonical} is the sole retained configuration and serves every "
            "former consumer."
        ),
        "exact_proposed_action": (
            f"Keep {canonical}, remap every listed consumer, and delete "
            f"{', '.join(duplicates)}."
        ),
        "canonical_object_key": canonical,
        "canonical_selection_rationale": str(
            row.get("recommended_canonical_basis") or ""
        ),
        "creations": [],
        "additions": [],
        "changes": [],
        "remaps": remaps,
        "deletions": [
            {
                "object_key": duplicate,
                "reason": f"{duplicate} exactly duplicates {canonical}.",
            }
            for duplicate in duplicates
        ],
        "renames": [],
        "preconditions": (
            "Re-read every candidate and consumer and confirm the exact source "
            "configuration and consumer sets are unchanged."
        ),
        "qa_steps": (
            "Re-export; confirm all consumers resolve to the canonical object, "
            "duplicates are absent, and no reference is missing."
        ),
        "rollback": (
            "Restore deleted duplicates and reverse only the listed remaps from the "
            "locked export."
        ),
        "priority": "Medium",
        "confidence": "High",
        "execution_readiness": "approval_required",
        "challenge_review": {},
    }


def complete_architecture(
    export_path: Path, shared_facts: dict | None = None
) -> dict:
    review = scaffold_architecture(export_path, shared_facts=shared_facts)
    non_retention_types = {
        "same_tag_payload_different_route",
        "shared_zone_child_container",
        "cyclic_trigger_group_dependency",
        "browser_server_consent_deduplication_review",
    }
    for row in review["families"]:
        basis = row["family_label"] or row["family_key"]
        member_keys = set(row["member_object_keys"])
        unsafe_family_comparisons = [
            candidate
            for candidate in review["comparisons"]
            if set(candidate["candidate_object_keys"]) <= member_keys
            and set(candidate["comparison_types"]) & non_retention_types
        ]
        unresolved_member_relationship = bool(unsafe_family_comparisons)
        missing_positive_distinction = any(
            not row.get("member_distinguishing_terms", {}).get(key)
            for key in row["member_object_keys"]
        )
        unresolved_member_relationship = (
            unresolved_member_relationship or missing_positive_distinction
        )
        family_comparison_types = {
            comparison_type
            for candidate in unsafe_family_comparisons
            for comparison_type in candidate["comparison_types"]
            if comparison_type in non_retention_types
        }
        family_cautions = [
            caution
            for candidate in unsafe_family_comparisons
            for caution in candidate.get("required_caution_states", [])
        ]
        row.update(
            {
                "review_status": "complete",
                "business_action": architecture_text(
                    row, "business_action", f"The {basis} family records one named business action"
                ),
                "family_purpose": architecture_text(
                    row, "family_purpose", f"The {basis} family serves one destination outcome"
                ),
                "member_assessments": member_assessments(
                    row,
                    "member_object_keys",
                    "available_member_evidence_anchors",
                    "member_paused_status",
                    "member_evidence_terms",
                ),
                "chain_assessments": member_assessments(
                    row,
                    "chain_object_keys",
                    "available_chain_evidence_anchors",
                    "chain_paused_status",
                    "chain_evidence_terms",
                ),
                "execution_path_summary": architecture_text(
                    row, "execution_path_summary", f"The {basis} route connects its tag and dependencies"
                ),
                "payload_coherence": architecture_text(
                    row, "payload_coherence", f"The {basis} payload matches its event and destination"
                ),
                "consent_and_sequence_coherence": architecture_text(
                    row,
                    "consent_and_sequence_coherence",
                    f"The {basis} route uses its listed consent and sequence controls",
                ),
                "necessity_and_ownership": architecture_text(
                    row,
                    "necessity_and_ownership",
                    f"The {basis} members retain distinct chain ownership",
                ),
                "relationship_verdict": (
                    "Owner decision needed"
                    if unresolved_member_relationship
                    else "Complementary"
                ),
                "analyst_rationale": architecture_text(
                    row,
                    "analyst_rationale",
                    f"The {basis} members {', '.join(row['member_object_keys'])} "
                    "use the cited event, destination, trigger, and dependency facts "
                    "as separate chain roles",
                )
                + architecture_caution_text(family_cautions),
                "target_architecture": architecture_text(
                    row,
                    "target_architecture",
                    f"Keep the {basis} chain minimal while preserving these distinct roles",
                ),
                "disposition": (
                    "owner_decision_needed" if unresolved_member_relationship else "keep"
                ),
                "owner_question": (
                    unsafe_owner_question(
                        family_comparison_types, row["member_object_keys"]
                    )
                    if unresolved_member_relationship
                    else ""
                ),
                "recommended_action": (
                    (
                        f"Resolve the unsafe "
                        f"{', '.join(sorted(family_comparison_types))} relationship "
                        "to one evidence-supported route while preserving required "
                        "measurement."
                        if family_comparison_types
                        else f"For {', '.join(row['member_object_keys'])}, obtain a "
                        "positive event, destination, trigger, or dependency distinction; "
                        "otherwise consolidate to one owner-approved route."
                    )
                    if unresolved_member_relationship
                    else ""
                ),
                "operations": [],
                "confidence": "High",
            }
        )
    for row in review["comparisons"]:
        exact = "exact_configuration" in row.get("comparison_types", [])
        recommended_canonical = str(
            row.get("recommended_canonical_object_key") or ""
        )
        owner_required = exact or bool(
            set(row.get("comparison_types", [])) & non_retention_types
        )
        owner_required = owner_required or any(
            not row.get("candidate_distinguishing_terms", {}).get(key)
            for key in row["candidate_object_keys"]
        )
        comparison_types = set(row.get("comparison_types", []))
        caution_text = architecture_caution_text(
            row.get("required_caution_states", [])
        )
        row.update(
            {
                "review_status": "complete",
                "member_assessments": member_assessments(
                    row,
                    "candidate_object_keys",
                    "available_member_evidence_anchors",
                    "candidate_paused_status",
                    "candidate_evidence_terms",
                ),
                "relationship_verdict": (
                    "Owner decision needed" if owner_required else "Intentional variant"
                ),
                "analyst_rationale": architecture_text(
                    row,
                    "analyst_rationale",
                    f"{row['comparison_id']} compares "
                    f"{', '.join(row['candidate_object_keys'])}; their cited event, "
                    "destination, trigger, and dependency facts assign different "
                    "execution roles",
                )
                + caution_text,
                "architecture_effect": architecture_text(
                    row,
                    "architecture_effect",
                    f"{row['comparison_id']} retains "
                    f"{', '.join(row['candidate_object_keys'])} separately only where "
                    "their positive source terms define different route or payload scope",
                )
                + caution_text,
                "disposition": "owner_decision_needed" if owner_required else "keep",
                "owner_question": (
                    unsafe_owner_question(
                        comparison_types, row["candidate_object_keys"]
                    )
                    if owner_required
                    else ""
                ),
                "recommended_action": (
                    (
                        f"Use {recommended_canonical} as the default canonical object, "
                        "remap every visible consumer, and delete only the proven duplicate "
                        "after approval."
                        if exact
                        else f"Resolve {row['comparison_id']} to the simplest "
                        "evidence-supported route and retain variants only where "
                        "source-visible behavior proves they are necessary."
                    )
                    if owner_required
                    else ""
                ),
                "canonical_selection_rationale": (
                    str(row.get("recommended_canonical_basis") or "")
                    if exact
                    else ""
                ),
                "operations": [],
                "confidence": "High",
            }
        )
        if exact:
            operation = exact_duplicate_operation_from_comparison(row)
            row.update(
                {
                    "relationship_verdict": "Exact duplicate",
                    "analyst_rationale": architecture_text(
                        row,
                        "analyst_rationale",
                        f"{row['comparison_id']} has identical source-visible "
                        f"configuration across {', '.join(row['candidate_object_keys'])} "
                        "without a distinct route, payload, consent, or consumer behavior",
                    ),
                    "architecture_effect": architecture_text(
                        row,
                        "architecture_effect",
                        f"Retain {recommended_canonical} as the source-ranked canonical "
                        "object and remove the exact duplicate after complete remap",
                    ),
                    "disposition": "cleanup_operation",
                    "owner_question": "",
                    "recommended_action": "",
                    "operations": [operation],
                }
            )
    source_records = object_records(container_version(json.loads(export_path.read_text(encoding="utf-8"))))
    source_record_by_key = {
        record["object_key"]: record
        for layer_records in source_records.values()
        for record in layer_records
    }
    reviewed_keys = sorted(
        record["object_key"]
        for layer_records in source_records.values()
        for record in layer_records
    )
    attestation = review["open_discovery_attestation"]
    source_terms = [
        term
        for family in review["families"]
        for terms in family["chain_evidence_terms"].values()
        for term in terms
    ]
    source_terms = list(dict.fromkeys([*source_terms, *reviewed_keys]))
    while len(source_terms) < 3:
        source_terms.append(f"source-object-{len(source_terms) + 1}")

    def method_source_terms(method_row: dict) -> list[str]:
        keys = method_row["candidate_object_keys"] or reviewed_keys
        values = [
            value
            for key in keys
            for value in (
                source_record_by_key[key]["object_name"],
                source_record_by_key[key]["object_key"],
            )
        ]
        return list(dict.fromkeys(values))[:6]
    attestation.update(
        {
            "review_status": "complete",
            "reviewed_object_keys": reviewed_keys,
            "discovered_comparison_ids": [],
            "zero_discovery_rationale": (
                f"Every object, including {source_terms[0]}, {source_terms[1]}, and "
                f"{source_terms[2]}, was compared by "
                + ", ".join(
                    row["method"].replace("_", " ")
                    for row in review["discovery_method_coverage"]
                )
                + " without finding another source-grounded relationship."
            ),
            "method_reviews": [
                {
                    **row,
                    "review_status": "complete",
                    "reviewed_comparison_ids": list(row["comparison_ids"]),
                    "reviewed_object_keys": reviewed_keys,
                    "additional_discovery_ids": [],
                    "conclusion": (
                        f"The {row['method'].replace('_', ' ')} scan reviewed "
                        f"{len(reviewed_keys)} source objects including "
                        f"{' and '.join(method_source_terms(row))}, and candidates "
                        f"{', '.join(row['comparison_ids']) or 'with no generated comparison'}; "
                        "no additional source-grounded relationship was found in this fixture."
                    ),
                }
                for row in review["discovery_method_coverage"]
            ],
        }
    )
    review["run_status"] = "complete"
    review["completion_attestation"] = complete_review_attestation(
        review, decision_authoring_method="independent_test_fixture_review"
    )
    return review


def value_discovery_row(export_path: Path) -> dict:
    records = object_records(
        container_version(json.loads(export_path.read_text(encoding="utf-8")))
    )
    by_key = {
        record["object_key"]: record
        for layer_records in records.values()
        for record in layer_records
    }
    keys = ["variable:22", "variable:24"]
    return {
        "comparison_id": "DISC-VALUE-001",
        "comparison_origin": "analyst_discovered",
        "discovery_methods": ["terminal_source_formula_and_output_overlap"],
        "comparison_types": ["shared_terminal_source"],
        "candidate_object_keys": keys,
        "candidate_basis": [
            "Recursive chain review found two differently named value helpers that require "
            "consumer-level comparison outside deterministic similarity groups."
        ],
        "review_status": "complete",
        "member_assessments": [
            {
                "object_key": key,
                "configured_role": (
                    f"{by_key[key]['object_name']} ({by_key[key]['object_type']}) returns "
                    f"{' and '.join(by_key[key]['specificity_tokens'][:2])} to its consumers."
                ),
                "necessity": (
                    f"{by_key[key]['object_name']} remains necessary while "
                    f"{by_key[key]['specificity_tokens'][0]} serves a distinct consumer."
                ),
                "distinguishing_configuration": (
                    f"{by_key[key]['object_name']} differs as a {by_key[key]['object_type']} "
                    f"through {' and '.join(by_key[key]['specificity_tokens'][:2])}."
                ),
                "status": "paused" if by_key[key]["paused"] else "active",
                "evidence_anchors": by_key[key]["evidence_anchors"][:1],
            }
            for key in keys
        ],
        "relationship_verdict": "Intentional variant",
        "analyst_rationale": (
            "CJS - Page URL Mirror mirrors Page URL while DLV - Value reads "
            "ecommerce.value, so shared value wording does not establish duplication."
        ),
        "architecture_effect": (
            "Keep CJS - Page URL Mirror and DLV - Value as separate variable roles while "
            "Page URL and ecommerce.value consumers remain distinct."
        ),
        "disposition": "keep",
        "owner_question": "",
        "operations": [],
        "confidence": "High",
    }


def duplicate_variable_operation() -> dict:
    return {
        "operation_key": "consolidate-ecommerce-items-dlv",
        "title": "Consolidate duplicate ecommerce items variables",
        "area": "GTM hygiene",
        "problem_type": "Exact duplicate",
        "problem": "Two data-layer variables read the same ecommerce.items source with identical settings.",
        "why_it_matters": "Maintaining both variables creates needless ambiguity and duplicate ownership work.",
        "expected_clean_state": "One canonical ecommerce.items variable serves every existing consumer.",
        "exact_proposed_action": "Keep variable 20 and delete unused duplicate variable 21.",
        "canonical_object_key": "variable:20",
        "canonical_selection_rationale": (
            "Select variable:20 as canonical because it is active, has the "
            "container-visible tag:1 consumer, and variable:21 has no consumer."
        ),
        "changes": [],
        "remaps": [],
        "deletions": [
            {
                "object_key": "variable:21",
                "reason": "Variable 21 duplicates variable 20 and has no active consumer.",
            }
        ],
        "renames": [],
        "preconditions": "Confirm variable 21 still has no export-visible consumer before execution.",
        "qa_steps": "Re-export the workspace and confirm every ecommerce.items reference resolves to variable 20.",
        "rollback": "Restore variable 21 from the original container export if a missed dependency appears.",
        "priority": "Medium",
        "confidence": "High",
        "execution_readiness": "approval_required",
        "challenge_review": {},
    }


def align_duplicate_operation(operational: dict, architecture: dict) -> None:
    operation = duplicate_variable_operation()
    for finding in operational["findings"]:
        if finding.get("module_name") not in {
            "duplicate_variable_logic",
            "duplicate_variable_paths",
        }:
            continue
        if set(finding.get("object_ids", [])) != {"20", "21"}:
            continue
        finding.update(copy.deepcopy(operation))
        finding["disposition"] = "cleanup_operation"
        finding["rationale"] = (
            "Variables 20 and 21 have identical data-layer settings, and variable 21 has no "
            "consumer, so one canonical variable is sufficient."
        )
    for comparison in architecture["comparisons"]:
        if set(comparison.get("candidate_object_keys", [])) != {"variable:20", "variable:21"}:
            continue
        comparison.update(
            {
                "relationship_verdict": "Exact duplicate",
                "analyst_rationale": (
                    "DLV - Items and DLV - Items Copy read ecommerce.items with identical type "
                    "and settings, while only DLV - Items has an active consumer."
                ),
                "architecture_effect": (
                    "Keeping DLV - Items and DLV - Items Copy adds maintenance work because both "
                    "use ecommerce.items without a distinct route, payload, or consumer contract."
                ),
                "disposition": "cleanup_operation",
                "operations": [copy.deepcopy(operation)],
            }
        )


class PipelineTests(unittest.TestCase):
    maxDiff = None

    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.root = Path(self.tempdir.name)
        self.export_path = self.root / "container.json"
        self.export_path.write_text(json.dumps(sample_export()), encoding="utf-8")

    def completed_reviews(self) -> tuple[dict, dict, dict]:
        return (
            complete_operational(self.export_path),
            complete_configuration(self.export_path),
            complete_architecture(self.export_path),
        )

    def write_review(self, name: str, payload: dict) -> Path:
        path = self.root / name
        path.write_text(json.dumps(payload), encoding="utf-8")
        return path

    def test_system_and_builtin_references_are_not_missing(self) -> None:
        report = missing_references(container_version(sample_export()))
        self.assertEqual([], report["undefinedVariableReferences"])
        configuration = scaffold_configuration(self.export_path)
        mirror = next(row for row in configuration["rows"] if row["object_id"] == "22")
        trace = next(
            item
            for item in mirror["reference_trace_requirements"]
            if item["reference"] == "Page URL"
        )
        self.assertEqual(["built_in"], trace["terminal_states"])

    def test_canonical_remap_does_not_turn_consumers_into_architecture_conflicts(self) -> None:
        from gtm_architecture_review import operation_behavior_keys
        from gtm_operation_compile import behavior_impact_keys

        operation = {
            "remaps": [
                {
                    "from_object_key": "trigger:10",
                    "to_object_key": "trigger:11",
                    "consumer_object_keys": ["tag:20", "tag:21"],
                }
            ],
            "creations": [],
            "additions": [],
            "changes": [],
            "deletions": [],
        }
        expected = {"trigger:10", "trigger:11"}
        self.assertEqual(expected, operation_behavior_keys(operation))
        self.assertEqual(expected, behavior_impact_keys(operation))

    def test_folder_operations_are_runtime_and_architecture_neutral(self) -> None:
        from gtm_architecture_review import operation_behavior_keys
        from gtm_operation_compile import behavior_impact_keys

        operation = {
            "creations": [
                {
                    "layer": "folder",
                    "object": {"folderId": "20", "name": "Media"},
                    "reason": "Create a bounded operational folder.",
                }
            ],
            "additions": [],
            "changes": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].parentFolderId",
                    "before": "10",
                    "after": "20",
                }
            ],
            "remaps": [
                {
                    "from_object_key": "folder:10",
                    "to_object_key": "folder:20",
                    "consumer_object_keys": ["tag:1"],
                }
            ],
            "deletions": [
                {"object_key": "folder:10", "reason": "Replaced by role folder."}
            ],
        }
        self.assertEqual(set(), operation_behavior_keys(operation))
        self.assertEqual(set(), behavior_impact_keys(operation))

    def test_architecture_exact_duplicate_can_remove_an_entire_inactive_set(self) -> None:
        from gtm_architecture_review import validate_decision

        operation = duplicate_variable_operation()
        operation["canonical_object_key"] = ""
        operation["remaps"] = []
        operation["deletions"] = [
            {"object_key": "variable:20", "reason": "Inactive duplicate removed."},
            {"object_key": "variable:21", "reason": "Inactive duplicate removed."},
        ]
        row = {
            "relationship_verdict": "Exact duplicate",
            "disposition": "cleanup_operation",
            "confidence": "High",
            "owner_question": "",
            "recommended_action": "",
            "operations": [operation],
        }
        errors = validate_decision(
            row,
            {"variable:20", "variable:21"},
            "inactive exact duplicate set",
            {"variable:20": set(), "variable:21": set()},
            ["variable:20", "variable:21"],
        )
        self.assertFalse(any("lacks canonical object" in error for error in errors), errors)

    def test_non_destructive_fix_does_not_create_fake_relationship_conflicts(self) -> None:
        from gtm_operation_compile import (
            comparison_reconciliation_errors,
            family_reconciliation_errors,
        )

        comparisons = [
            {
                "comparison_id": "REL-KEEP",
                "candidate_object_keys": ["tag:1", "tag:2"],
                "relationship_verdict": "Intentional variant",
                "disposition": "keep",
            }
        ]
        families = [
            {
                "family_id": "FAM-KEEP",
                "member_object_keys": ["tag:1"],
                "chain_object_keys": ["tag:1", "trigger:10"],
                "relationship_verdict": "Complementary",
                "disposition": "keep",
            }
        ]
        self.assertEqual(
            [],
            comparison_reconciliation_errors(
                "fix-tag-consent", set(), {"tag:1"}, comparisons
            ),
        )
        self.assertEqual(
            [],
            family_reconciliation_errors(
                "consolidate-trigger", {"trigger:10"}, {"trigger:10"}, families
            ),
        )

    def test_distinguishing_terms_do_not_expose_opaque_behavior_signatures(self) -> None:
        from gtm_architecture_review import (
            distinguishing_terms_for_keys,
            validate_retention_distinctions,
        )

        terms = distinguishing_terms_for_keys(
            ["tag:1", "tag:2"],
            {
                "tag:1": {
                    "behavior_signatures": {
                        "configuration": "same",
                        "execution_route": "route-a",
                    }
                },
                "tag:2": {
                    "behavior_signatures": {
                        "configuration": "same",
                        "execution_route": "route-b",
                    }
                },
            },
        )
        self.assertEqual([], terms["tag:1"])
        self.assertEqual([], terms["tag:2"])
        errors = validate_retention_distinctions(
            {
                "relationship_verdict": "Intentional variant",
                "member_assessments": [
                    {
                        "object_key": "tag:1",
                        "distinguishing_configuration": (
                            "execution route signature aabbccdd"
                        ),
                    },
                    {
                        "object_key": "tag:2",
                        "distinguishing_configuration": (
                            "execution route signature 11223344"
                        ),
                    },
                ],
            },
            ["tag:1", "tag:2"],
            {
                "tag:1": ["execution route signature aabbccdd"],
                "tag:2": ["execution route signature 11223344"],
            },
            {
                "tag:1": {"execution_route": "aabbccdd"},
                "tag:2": {"execution_route": "11223344"},
            },
            "opaque retention",
        )
        self.assertTrue(any("opaque signature" in error for error in errors), errors)

    def test_distinguishing_terms_preserve_source_visible_baseline_scope(self) -> None:
        from gtm_architecture_review import (
            configured_parameter_terms,
            custom_code_return_terms,
            distinguishing_terms_for_keys,
        )

        terms = distinguishing_terms_for_keys(
            ["trigger:1", "trigger:2", "tag:3", "tag:4"],
            {
                "trigger:1": {
                    "object_type": "CUSTOM_EVENT",
                    "trigger_conditions": ["equals|{{_event}}|consent_ready|"],
                },
                "trigger:2": {
                    "object_type": "CUSTOM_EVENT",
                    "trigger_conditions": [
                        "equals|{{_event}}|consent_ready|",
                        "equals|{{Page Path}}|/funnel|",
                    ],
                },
                "tag:3": {"object_type": "vendor_template_a"},
                "tag:4": {"object_type": "vendor_template_b"},
            },
        )
        self.assertIn(
            "baseline condition set _event equals consent_ready", terms["trigger:1"]
        )
        self.assertTrue(
            any("page path equals /funnel" in term for term in terms["trigger:2"])
        )
        self.assertIn("object type vendor_template_a", terms["tag:3"])
        self.assertIn("object type vendor_template_b", terms["tag:4"])
        self.assertEqual(
            ["parameter conversionid 123456"],
            configured_parameter_terms(
                {
                    "source_leaf_facts": [
                        {
                            "json_path": "$.containerVersion.tag[0].parameter[2].key",
                            "value_preview": "conversionId",
                        },
                        {
                            "json_path": "$.containerVersion.tag[0].parameter[2].value",
                            "value_preview": "123456",
                        },
                        {
                            "json_path": "$.containerVersion.tag[0].parameter[3].map[0].key",
                            "value_preview": "key",
                        },
                    ]
                }
            ),
        )
        self.assertIn(
            "custom return location",
            custom_code_return_terms(
                {
                    "custom_code_facts": {
                        "return_expressions": [
                            {"expression": "dl || window.location.href"}
                        ]
                    }
                }
            ),
        )

    def test_missing_reference_check_ignores_custom_template_test_examples(self) -> None:
        from gtm_validate_artifact import missing_references

        template_data = """___TERMS_OF_SERVICE___
accepted
___SANDBOXED_JS_FOR_WEB_TEMPLATE___
data.gtmOnSuccess();
___TESTS___
scenarios:
- name: test-only variable
  code: \"return '{{Event}}';\"
"""
        report = missing_references(
            {
                "tag": [],
                "trigger": [],
                "variable": [],
                "builtInVariable": [],
                "folder": [],
                "customTemplate": [
                    {
                        "templateId": "1",
                        "name": "Template with test reference",
                        "templateData": template_data,
                    }
                ],
            }
        )
        self.assertEqual([], report["undefinedVariableReferences"])

    def test_gallery_template_type_is_reachable_and_not_flagged_unused(self) -> None:
        from gtm_architecture_review import dependency_graph

        data = sample_export()
        data["containerVersion"].setdefault("customTemplate", []).append(
            {
                "templateId": "90",
                "accountId": "100",
                "name": "Gallery-backed template",
                "galleryReference": {"galleryTemplateId": "GALLERY90"},
                "templateData": "___SANDBOXED_JS_FOR_WEB_TEMPLATE___\ndata.gtmOnSuccess();",
            }
        )
        data["containerVersion"]["tag"].append(
            {
                "tagId": "91",
                "name": "Tag using gallery template",
                "type": "cvt_GALLERY90",
                "firingTriggerId": ["10"],
            }
        )
        self.export_path.write_text(json.dumps(data), encoding="utf-8")

        scan = audit_export(self.export_path)
        self.assertFalse(
            any(
                row["module_name"] == "unused_custom_templates"
                and "90" in row["object_ids"]
                for row in scan["findings"]
            )
        )
        self.assertIn(
            "tag:91", object_consumer_map(self.export_path)["customTemplate:90"]
        )
        self.assertNotIn(
            "90", missing_references(data["containerVersion"])["missingCustomTemplateReferences"]
        )

        graph, _ = dependency_graph(
            data["containerVersion"], object_records(data["containerVersion"])
        )
        self.assertTrue(
            any(
                edge["to_object_key"] == "customTemplate:90"
                for edge in graph["tag:91"]
            )
        )

    def test_future_state_review_candidate_is_visible_but_nonblocking(self) -> None:
        from gtm_future_state_check import blocking_new_operational_findings

        rows = [
            {"finding_id": "CANDIDATE", "finding_class": "review_candidate"},
            {"finding_id": "DEFECT", "finding_class": "deterministic_defect"},
        ]
        self.assertEqual(
            ["DEFECT"],
            [
                row["finding_id"]
                for row in blocking_new_operational_findings(rows)
            ],
        )

    def test_projected_quality_review_returns_report(self) -> None:
        from gtm_future_state_check import projected_quality_review

        report, errors = projected_quality_review(
            self.export_path,
            sample_export(),
            {"plan_status": "partial", "operations": []},
        )
        self.assertIsInstance(report, dict)
        self.assertIsInstance(errors, list)
        self.assertEqual(errors, report["errors"])

    def test_container_only_contract_has_no_d4_or_runtime_gate(self) -> None:
        paths = [ROOT / "SKILL.md", ROOT / "README.md"]
        paths.extend((ROOT / "scripts").glob("*.py"))
        paths.extend((ROOT / "references").rglob("*.md"))
        for path in paths:
            content = path.read_text(encoding="utf-8").lower()
            self.assertNotIn("d4_required", content, str(path))
            self.assertNotIn("runtime_qa_required", content, str(path))
        removed_runtime_reference = (
            ROOT / "references" / "02-commands" / ("runtime-qa-" + "templates.md")
        )
        self.assertFalse(removed_runtime_reference.exists())

    def test_operational_scan_catches_basic_cleanup_failures(self) -> None:
        findings = [
            row
            for row in audit_export(self.export_path)["findings"]
            if row["finding_type"] != "zero_findings"
        ]
        types = {row["finding_type"] for row in findings}
        self.assertIn("duplicate_configuration", types)
        self.assertIn("duplicate_variable_path", types)
        self.assertIn("single_member_trigger_group", types)
        self.assertIn("invalid_trigger_regex", types)
        self.assertIn("ineffective_blocking_trigger", types)
        self.assertIn("variable_mirrors_builtin", types)
        unused_builtins = {
            object_id
            for row in findings
            if row["finding_type"] == "unused_built_in_variable"
            for object_id in row["object_ids"]
        }
        self.assertEqual({"Page URL", "Click URL"}, unused_builtins)
        paused = [row for row in findings if row["module_name"] == "used_only_by_paused_tags"]
        self.assertTrue(any("23" in row["object_ids"] for row in paused))
        triggerless = [
            row for row in findings if row["module_name"] == "tags_without_firing_triggers"
        ]
        self.assertFalse(any("3" in row["object_ids"] for row in triggerless))
        allowed_resolutions = {
            "cleanup_operation",
            "documented_exception",
            "owner_decision_needed",
            "keep",
        }
        self.assertTrue(
            all(
                {
                    value.strip()
                    for value in row["required_resolution"].split("|")
                    if value.strip()
                }
                <= allowed_resolutions
                for row in findings
            )
        )

    def test_fixed_slot_business_formula_cannot_pass_as_generic_false_positive(self) -> None:
        path = self.root / "fixed-slot-formula.json"
        path.write_text(json.dumps(fixed_slot_formula_export()), encoding="utf-8")
        findings = [
            row for row in audit_export(path)["findings"] if row["finding_type"] != "zero_findings"
        ]
        formula = next(
            row for row in findings if row["finding_type"] == "fixed_slot_business_formula"
        )
        self.assertEqual(["33"], formula["object_ids"])
        technical = extract_export(path)
        technical_row = next(row for row in technical["rows"] if row["object_id"] == "33")
        self.assertTrue(technical_row["fixed_slot_aggregation"])
        self.assertEqual([1, 2, 3], technical_row["fixed_slot_groups"][0]["indexes"])

        review = complete_configuration(path)
        total = next(row for row in review["rows"] if row["object_key"] == "variable:33")
        fixed_review = next(
            row
            for row in total["technical_finding_reviews"]
            if "fixed numbered value slots" in row["source_statement"].lower()
        )
        fixed_review["verdict"] = "False positive"
        fixed_review["rationale"] = "The code was inspected and appears acceptable."
        errors, _ = validate_configuration(
            path,
            self.write_review("fixed-slot-generic-dismissal.json", review),
        )
        self.assertTrue(any("fixed-slot business formula" in error for error in errors))

    def test_consent_purposes_with_same_logic_are_compared(self) -> None:
        data = sample_export()
        shared_config = [
            {"type": "TEMPLATE", "key": "name", "value": "OnetrustActiveGroups"},
        ]
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "40",
                    "name": "CJS - analytics_storage",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function(){ return {{OnetrustActiveGroups}}.indexOf(',2,') > -1; }",
                        },
                        *shared_config,
                    ],
                },
                {
                    "variableId": "41",
                    "name": "CJS - ad_storage",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function(){ return {{OnetrustActiveGroups}}.indexOf(',2,') > -1; }",
                        },
                        *shared_config,
                    ],
                },
            ]
        )
        path = self.root / "consent-shared-logic.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        findings = audit_export(path)["findings"]
        consent = [
            row
            for row in findings
            if row["finding_type"] == "different_consent_purposes_share_logic"
        ]
        self.assertEqual(1, len(consent))
        self.assertEqual({"40", "41"}, set(consent[0]["object_ids"]))

    def test_server_forwarded_consent_is_distinct_from_missing_client_control(self) -> None:
        variable = {
            "variableId": "50",
            "name": "DLV - Server consent state",
            "type": "v",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "name",
                    "value": "consent_state",
                }
            ],
        }
        tag = {
            "tagId": "51",
            "name": "Google tag - Server transport",
            "type": "googtag",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "transport_url",
                    "value": "https://collect.example.test",
                },
                {
                    "type": "TEMPLATE",
                    "key": "server_consent",
                    "value": "{{DLV - Server consent state}}",
                },
            ],
        }
        route = tag_consent_route(
            tag,
            "$.containerVersion.tag[0]",
            variables=[variable],
        )
        self.assertEqual("server_forwarding_candidate", route["effective_control_status"])
        self.assertEqual(["collect.example.test"], route["server_routing_hosts"])
        self.assertEqual(
            ["DLV - Server consent state"],
            route["server_consent_forwarding_variables"],
        )
        self.assertTrue(route["forwarded_cmp_signal_visible"])
        self.assertEqual("not_visible_in_web_export", route["server_enforcement_visibility"])

        data = sample_export()
        data["containerVersion"]["variable"].append(variable)
        media_transport = {
            **tag,
            "tagId": "53",
            "name": "Meta - Lead server transporter",
            "type": "html",
            "firingTriggerId": ["10"],
        }
        data["containerVersion"]["tag"].append(media_transport)
        export = self.root / "server-forwarded-consent.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        findings = audit_export(export)["findings"]
        self.assertFalse(
            any(
                row["finding_type"] == "media_consent_route_requires_review"
                and "53" in row["object_ids"]
                for row in findings
            )
        )

    def test_server_route_without_forwarded_consent_remains_unproven(self) -> None:
        tag = {
            "tagId": "52",
            "name": "Google tag - Incomplete server transport",
            "type": "googtag",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "transport_url",
                    "value": "https://collect.example.test",
                }
            ],
        }
        route = tag_consent_route(tag)
        self.assertEqual("server_contract_unproven", route["effective_control_status"])
        self.assertEqual([], route["server_consent_forwarding_evidence"])

    def test_manual_consent_enum_only_enables_additional_checks_when_needed(self) -> None:
        expectations = {
            "notSet": ("NOT_SET", False, "native_consent_capability"),
            "notNeeded": ("NOT_NEEDED", False, "native_consent_capability"),
            "needed": ("NEEDED", True, "explicit_export_control"),
            "unexpected": ("UNEXPECTED", False, "unrecognized_consent_status"),
        }
        for raw_status, expected in expectations.items():
            with self.subTest(raw_status=raw_status):
                route = tag_consent_route(
                    {
                        "tagId": "consent",
                        "name": "Google tag consent enum",
                        "type": "googtag",
                        "consentSettings": {"consentStatus": raw_status},
                    }
                )
                self.assertEqual(expected[0], route["consent_status"])
                self.assertEqual(expected[1], route["additional_consent_checks_visible"])
                self.assertEqual(expected[2], route["effective_control_status"])

    def test_consent_control_requires_payload_and_preserves_every_vendor(self) -> None:
        route = tag_consent_route(
            {
                "tagId": "consent-boundary",
                "name": "Consent bootstrap",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "eventName",
                        "value": "consent_ready",
                    },
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            "<script>fbq('track','Purchase');"
                            "ttq.track('CompletePayment');</script>"
                        ),
                    },
                ],
                "blockingTriggerId": ["99"],
            }
        )
        self.assertEqual({"Meta", "TikTok"}, set(route["detected_vendors"]))
        self.assertEqual("blocker_control_candidate", route["effective_control_status"])
        self.assertFalse(route["forwarded_cmp_signal_visible"])
        self.assertEqual([], route["forwarded_consent_purposes"])
        self.assertEqual([], route["server_consent_forwarding_evidence"])

    def test_context_inference_uses_reachable_behavior_and_separates_gateway(self) -> None:
        data = sample_export()
        data["containerVersion"]["trigger"].append(
            {
                "triggerId": "99",
                "name": "Lead form submit orphan",
                "type": "CUSTOM_EVENT",
                "customEventFilter": [
                    condition("EQUALS", "{{_event}}", "generate_lead")
                ],
            }
        )
        data["containerVersion"]["tag"][0]["parameter"].append(
            {
                "type": "TEMPLATE",
                "key": "server_container_url",
                "value": "https://collect.example.test",
            }
        )
        export = self.root / "reachable-context.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        context = build_context_model(export)["context"]
        self.assertEqual("ecommerce", context["business_model"])
        self.assertNotIn("lead_or_quote", context["business_signals"])
        self.assertEqual(["collect.example.test"], context["server_routing_hosts"])
        self.assertEqual(
            "not_visible_in_container_export",
            context["google_tag_gateway"]["status"],
        )

    def test_cross_object_dependencies_and_empty_objects_remain_reviewable(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["setupTag"].append(
            {"tagName": "GA4 - Purchase - All"}
        )
        data["containerVersion"]["tag"][0]["teardownTag"] = [{}]
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["10", "999"]},
                "typeRestriction": {},
            }
        ]
        export = self.root / "cross-object-dependencies.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = scaffold_configuration(export)
        purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        self.assertTrue(
            any(
                fact["json_path"].endswith(".teardownTag[0]")
                and fact["value_preview"] == "{}"
                for fact in purchase["source_facts"]
            )
        )
        setup = next(
            trace
            for trace in purchase["execution_dependency_traces"]
            if trace["relation"] == "setupTag"
            and trace["reference"] == "GA4 - Purchase - All"
        )
        self.assertEqual("tag:1", setup["targets"][0]["object_key"])
        self.assertEqual("cycle", setup["resolution_state"])
        self.assertTrue(
            any(
                fact["json_path"].startswith("$.containerVersion.trigger[0]")
                for fact in purchase["execution_dependency_facts"]
            )
        )
        self.assertTrue(
            any(
                fact["json_path"].startswith("$.containerVersion.tag[2].parameter")
                and "consentDefault" in str(fact.get("value_preview") or "")
                for fact in purchase["execution_dependency_facts"]
            )
        )
        zone = next(row for row in review["rows"] if row["object_key"] == "zone:70")
        missing = next(
            trace
            for trace in zone["execution_dependency_traces"]
            if trace["reference"] == "999"
        )
        self.assertEqual("missing", missing["resolution_state"])
        architecture = scaffold_architecture(export)
        zone_family = next(
            family
            for family in architecture["families"]
            if "zone:70" in family["member_object_keys"]
        )
        self.assertTrue(
            any(
                edge.get("target_reference") == "999"
                and edge.get("resolution_state") == "missing"
                for edge in zone_family["chain_edges"]
            )
        )
        transaction = next(
            row for row in review["rows"] if row["object_key"] == "variable:25"
        )
        self.assertTrue(transaction["consumer_dependency_facts"])

    def test_custom_code_facts_do_not_infer_loader_dom_or_return_shapes(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].append(
            {
                "variableId": "90",
                "name": "CJS - Delegated URL",
                "type": "jsm",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "javascript",
                        "value": "function(){ return {{Click URL}}; }",
                    }
                ],
            }
        )
        data["containerVersion"]["tag"].append(
            {
                "tagId": "91",
                "name": "Dynamic vendor loader",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            "<script>var s=document.createElement('script');"
                            "s.src='https://cdn.example.test/sdk.js';"
                            "document.head.appendChild(s);window.vendorReady=true;</script>"
                        ),
                    }
                ],
                "firingTriggerId": ["10"],
            }
        )
        data["containerVersion"]["customTemplate"] = [
            {
                "templateId": "tpl-90",
                "name": "Opaque template",
                "templateData": "function(){ return 1; }",
            }
        ]
        export = self.root / "custom-code-representation.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        with patch.dict(sys.modules, {"esprima": None}):
            technical = extract_export(export)
        delegated = next(row for row in technical["rows"] if row["object_id"] == "90")
        self.assertEqual(
            "gtm_variable_reference_type_unresolved", delegated["returned_value_type"]
        )
        self.assertTrue(delegated["parser_input_normalized"])
        loader = next(row for row in technical["rows"] if row["object_id"] == "91")
        self.assertTrue(loader["dom_mutations"])
        self.assertFalse(loader["dom_selector_reads"])
        self.assertFalse(
            any(
                "more than one script" in finding.lower()
                for finding in loader["technical_code_optimization_findings"]
            )
        )
        self.assertIn("window/global write", loader["technical_current_behavior"])
        template = next(
            row for row in technical["rows"] if row["object_id"] == "tpl-90"
        )
        self.assertEqual("owner_decision_needed", template["technical_action_candidate"])

    def test_operational_scan_covers_paused_groups_and_template_duplicates(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][3].pop("firingTriggerId")
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "30",
                    "name": "TG - Empty",
                    "type": "TRIGGER_GROUP",
                    "parameter": [],
                },
                {
                    "triggerId": "31",
                    "name": "TG - Duplicate members",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [
                                {"type": "TEMPLATE", "value": "10"},
                                {"type": "TEMPLATE", "value": "10"},
                            ],
                        }
                    ],
                },
                {
                    "triggerId": "32",
                    "name": "TG - Cycle A",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "33"}],
                        }
                    ],
                },
                {
                    "triggerId": "33",
                    "name": "TG - Cycle B",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "32"}],
                        }
                    ],
                },
            ]
        )
        data["containerVersion"]["customTemplate"] = [
            {"templateId": "tpl-1", "name": "Template one", "templateData": "same"},
            {"templateId": "tpl-2", "name": "Template two", "templateData": "same"},
        ]
        path = self.root / "operational-adversary.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        findings = [
            row for row in audit_export(path)["findings"] if row["finding_type"] != "zero_findings"
        ]
        types = {row["finding_type"] for row in findings}
        self.assertIn("paused_objects_for_lifecycle_review", types)
        self.assertIn("empty_trigger_group", types)
        self.assertIn("duplicate_trigger_group_members", types)
        self.assertIn("cyclic_trigger_groups", types)
        self.assertTrue(
            any(
                row["module_name"] == "duplicate_custom_template_configurations" for row in findings
            )
        )
        triggerless = [
            row for row in findings if row["module_name"] == "tags_without_firing_triggers"
        ]
        self.assertFalse(any("4" in row["object_ids"] for row in triggerless))

    def test_operational_scan_resolves_reachability_sequences_and_contradictions(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "80",
                    "name": "Orphan A",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "{{Orphan B}}"}
                    ],
                },
                {
                    "variableId": "81",
                    "name": "Orphan B",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "{{Orphan A}}"}
                    ],
                },
            ]
        )
        data["containerVersion"]["tag"].extend(
            [
                {
                    "tagId": "80",
                    "name": "Sequence A",
                    "type": "html",
                    "setupTag": [{"tagName": "Sequence B"}],
                    "scheduleStartMs": "200",
                    "scheduleEndMs": "100",
                },
                {
                    "tagId": "81",
                    "name": "Sequence B",
                    "type": "html",
                    "teardownTag": [{"tagName": "Sequence A"}],
                },
            ]
        )
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "80",
                    "name": "CE - Impossible",
                    "type": "CUSTOM_EVENT",
                    "filter": [
                        condition("EQUALS", "{{_event}}", "purchase"),
                        condition("NOT_EQUALS", "{{_event}}", "purchase"),
                    ],
                },
                {
                    "triggerId": "90",
                    "name": "TG - Parent",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "91"}],
                        }
                    ],
                },
                {
                    "triggerId": "91",
                    "name": "TG - Child",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "10"}],
                        }
                    ],
                },
            ]
        )
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "95",
                "name": "Impossible Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {
                    "condition": [
                        condition("EQUALS", "market", "fr"),
                        condition("NOT_EQUALS", "market", "fr"),
                    ]
                },
            }
        ]
        path = self.root / "reachability-and-sequence.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        scan = audit_export(path)
        findings = [
            row for row in scan["findings"] if row["finding_type"] != "zero_findings"
        ]
        finding_types = {row["finding_type"] for row in findings}
        self.assertIn("cyclic_tag_sequence", finding_types)
        self.assertIn("invalid_tag_schedule_order", finding_types)
        self.assertIn("contradictory_trigger_conditions", finding_types)
        self.assertIn("contradictory_zone_boundary_conditions", finding_types)
        self.assertIn("nested_trigger_groups", finding_types)
        unused_variable_ids = {
            object_id
            for row in findings
            if row["module_name"] == "unused_variables"
            for object_id in row["object_ids"]
        }
        self.assertTrue({"80", "81"}.issubset(unused_variable_ids))
        unreachable_tags = {
            object_id
            for row in findings
            if row["module_name"] == "tags_without_firing_triggers"
            for object_id in row["object_ids"]
        }
        self.assertTrue({"80", "81"}.issubset(unreachable_tags))

    def test_operational_scan_fails_visible_on_malformed_nested_controls(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["setupTag"] = {
            "tagName": "Utility - Consent Defaults"
        }
        data["containerVersion"]["tag"][0]["teardownTag"] = ["bad-entry", {}]
        data["containerVersion"]["tag"][0]["consentSettings"] = []
        data["containerVersion"]["tag"][1]["consentSettings"] = {
            "consentStatus": "futureMaybe"
        }
        group = next(
            row for row in data["containerVersion"]["trigger"] if row["triggerId"] == "12"
        )
        group["parameter"][0]["list"] = ["bad-member", {"value": ""}]
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "95",
                "name": "Malformed Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"condition": "not-an-array"},
                "typeRestriction": {
                    "enable": True,
                    "whitelistedTypeId": "html",
                },
            }
        ]
        path = self.root / "malformed-nested-controls.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        scan = audit_export(path)
        self.assertEqual("complete", scan["run_status"])
        finding_types = {
            row["finding_type"]
            for row in scan["findings"]
            if row["finding_type"] != "zero_findings"
        }
        self.assertTrue(
            {
                "invalid_tag_sequence_shape",
                "invalid_tag_sequence_entry",
                "tag_sequence_target_missing_name",
                "invalid_trigger_group_member_structure",
                "invalid_zone_boundary_field_shape",
                "invalid_zone_type_allowlist_shape",
                "invalid_consent_settings_shape",
                "unrecognized_manual_consent_status",
            }.issubset(finding_types)
        )
        package = self.root / "malformed-nested-package"
        manifest = build_package(path, package, pretty=True)
        self.assertEqual("pass", manifest["status"])
        self.assertTrue((package / "operational_review.json").exists())
        self.assertTrue((package / "configuration_review.json").exists())
        self.assertTrue((package / "architecture_review.json").exists())

    def test_ineffective_blocker_requires_every_firing_route_to_be_exact(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["firingTriggerId"].append("14")
        path = self.root / "mixed-firing-routes.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        findings = audit_export(path)["findings"]
        self.assertFalse(
            any(
                row["finding_type"] == "ineffective_blocking_trigger"
                and "1" in row["object_ids"]
                for row in findings
            )
        )

    def test_relationship_run_detects_scope_and_status_without_names_only(self) -> None:
        rows = relationship_candidates(container_version(sample_export()))
        q1 = [
            row
            for row in rows
            if "shared_business_scope" in row["comparison_types"]
            and {"trigger:15", "trigger:16"}.issubset(row["candidate_object_keys"])
        ]
        self.assertEqual(1, len(q1))
        duplicate = next(
            row
            for row in rows
            if {"variable:20", "variable:21"} == set(row["candidate_object_keys"])
        )
        self.assertIn("exact_configuration", duplicate["comparison_types"])
        self.assertEqual(False, duplicate["candidate_paused_status"]["variable:20"])

    def test_duplicate_logic_ignores_export_metadata_and_folder_placement(self) -> None:
        data = sample_export()
        variables = {
            row["variableId"]: row for row in data["containerVersion"]["variable"]
        }
        variables["20"].update(
            {
                "workspaceId": "7",
                "tagManagerUrl": "https://tagmanager.google.com/variable/20",
                "notes": "Original implementation",
                "parentFolderId": "100",
            }
        )
        variables["21"].update(
            {
                "workspaceId": "8",
                "tagManagerUrl": "https://tagmanager.google.com/variable/21",
                "notes": "Later copy",
                "parentFolderId": "101",
            }
        )
        path = self.root / "metadata-different-duplicates.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        finding = next(
            row
            for row in audit_export(path)["findings"]
            if row["module_name"] == "duplicate_variable_logic"
            and {"20", "21"}.issubset(row["object_ids"])
        )
        self.assertEqual("duplicate_configuration", finding["finding_type"])
        comparison = next(
            row
            for row in relationship_candidates(data["containerVersion"])
            if "exact_configuration" in row["comparison_types"]
            and {"variable:20", "variable:21"}.issubset(
                row["candidate_object_keys"]
            )
        )
        self.assertEqual(1.0, comparison["similarity_score"])
        configuration = scaffold_configuration(path)
        variable_review = next(
            row
            for row in configuration["rows"]
            if row["object_key"] == "variable:20"
        )
        metadata_suffixes = (
            ".workspaceId",
            ".tagManagerUrl",
            ".notes",
            ".parentFolderId",
        )
        self.assertFalse(
            any(
                source_path.endswith(metadata_suffixes)
                for source_path in variable_review["required_logic_anchors"]
            )
        )

    def test_relationship_run_reviews_multiple_firing_routes(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["firingTriggerId"] = ["15", "16"]
        rows = relationship_candidates(container_version(data))
        route = [
            row
            for row in rows
            if "multi_firing_route_consolidation_review" in row["comparison_types"]
            and set(row["candidate_object_keys"]) == {"trigger:15", "trigger:16"}
        ]
        self.assertEqual(1, len(route))

    def test_architecture_accepts_source_grounded_open_discovery(self) -> None:
        review = complete_architecture(self.export_path)
        review["comparisons"].append(value_discovery_row(self.export_path))
        review["open_discovery_attestation"]["discovered_comparison_ids"] = [
            "DISC-VALUE-001"
        ]
        review["open_discovery_attestation"]["zero_discovery_rationale"] = ""
        next(
            item
            for item in review["open_discovery_attestation"]["method_reviews"]
            if item["method"] == "terminal_source_formula_and_output_overlap"
        )["additional_discovery_ids"] = ["DISC-VALUE-001"]
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("open-discovery.json", review),
        )
        self.assertEqual([], errors)

        mismatched = copy.deepcopy(review)
        mismatched_discovery = next(
            row
            for row in mismatched["comparisons"]
            if row["comparison_id"] == "DISC-VALUE-001"
        )
        mismatched_discovery["comparison_types"] = [
            "semantic_name_family_candidate"
        ]
        mismatched_errors, _ = validate_architecture(
            self.export_path,
            self.write_review("mismatched-discovery-method.json", mismatched),
        )
        self.assertTrue(
            any(
                "declared comparison types require discovery methods" in error
                for error in mismatched_errors
            )
        )

    def test_configuration_scaffold_requires_every_object_branch_code_and_trace(self) -> None:
        review = scaffold_configuration(self.export_path)
        source_count = sum(
            len(sample_export()["containerVersion"].get(layer, []))
            for layer in (
                "tag",
                "trigger",
                "variable",
                "customTemplate",
                "client",
                "transformation",
            )
        )
        self.assertEqual(source_count, len(review["rows"]))
        meta = next(
            row for row in review["rows"] if row["object_id"] == "2" and row["layer"] == "tag"
        )
        self.assertTrue(meta["required_branch_reviews"])
        self.assertTrue(meta["required_code_line_hashes"])
        self.assertTrue(meta["reference_trace_requirements"])
        self.assertTrue(meta["technical_code_facts"])

    def test_configuration_gate_rejects_missing_branch_and_generic_code(self) -> None:
        review = complete_configuration(self.export_path)
        custom = next(
            row for row in review["rows"] if row["object_id"] == "2" and row["layer"] == "tag"
        )
        custom["configuration_branch_reviews"].pop()
        custom["code_behavior_blocks"][0]["purpose"] = "Code inspected and reviewed"
        path = self.write_review("bad-config.json", review)
        errors, _ = validate_configuration(self.export_path, path)
        self.assertTrue(any("branch reviews" in error for error in errors))
        self.assertTrue(any("incomplete purpose" in error for error in errors))

    def test_configuration_gate_rejects_generic_semantic_prose_with_valid_citations(self) -> None:
        review = complete_configuration(self.export_path)
        purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        purchase["purpose"] = (
            "GA4 - Purchase - All serves one concrete measurement purpose through its tag setup."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("generic-semantic-prose.json", review),
        )
        self.assertTrue(any("purpose lacks object-specific analysis" in error for error in errors))

    def test_review_context_is_content_locked_not_hash_only(self) -> None:
        review = complete_configuration(self.export_path)
        review["audit_context"]["business_model"] = "publisher"
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("tampered-context.json", review),
        )
        self.assertTrue(any("audit_context differs" in error for error in errors))

    def test_unknown_external_vendor_creates_official_research_contract(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"].append(
            {
                "tagId": "90",
                "name": "Partner widget",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            '<script src="https://unknown-cdn.example/widget.js"></script>'
                        ),
                    }
                ],
                "firingTriggerId": ["10"],
            }
        )
        path = self.root / "unknown-vendor.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        review = scaffold_configuration(path)
        partner = next(row for row in review["rows"] if row["object_key"] == "tag:90")
        self.assertTrue(
            any(context["category"] == "unknown_vendor" for context in partner["vendor_contexts"])
        )
        self.assertTrue(any(topic["research_required"] for topic in partner["required_contract_topics"]))
        context = build_context_model(path)
        self.assertIn("unknown-cdn.example", context["context"]["external_hosts"])
        self.assertNotIn("unknown-cdn.example", context["context"]["server_routing_hosts"])
        self.assertEqual("Unclassified", vendor_record('{"key":"contents"}')["name"])
        self.assertEqual("Unclassified", vendor_record('{"key":"activity"}')["name"])

    def test_configuration_requires_contract_topics_and_recursive_node_meaning(self) -> None:
        scaffold = scaffold_configuration(self.export_path)
        purchase = next(row for row in scaffold["rows"] if row["object_key"] == "tag:1")
        topics = {item["topic"] for item in purchase["required_contract_topics"]}
        self.assertIn("ecommerce_event_contract", topics)
        self.assertIn("transaction_value_currency_and_quantity", topics)
        value_variable = next(row for row in scaffold["rows"] if row["object_key"] == "variable:24")
        self.assertTrue(
            any(
                context["vendor"] == "GA4 / Google tag"
                for context in value_variable["vendor_contexts"]
            )
        )
        self.assertTrue(value_variable["required_contract_topics"])

        review = complete_configuration(self.export_path)
        completed_purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        completed_purchase["contract_checks"].pop()
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("missing-contract-topic.json", review),
        )
        self.assertTrue(any("every generated topic" in error for error in errors))

        review = complete_configuration(self.export_path)
        completed_purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        completed_purchase["contract_checks"][0]["expected_rule"] = (
            "The official vendor documentation was reviewed for this configuration."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("generic-contract-rule.json", review),
        )
        self.assertTrue(any("topic-specific contract" in error for error in errors))

        review = complete_configuration(self.export_path)
        meta = next(row for row in review["rows"] if row["object_key"] == "tag:2")
        trace = next(
            item for item in meta["reference_traces"] if item["reference"] == "DLV - Items"
        )
        trace["node_reviews"].clear()
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("missing-trace-node.json", review),
        )
        self.assertTrue(any("every variable node" in error for error in errors))

    def test_complete_configuration_and_architecture_reviews_validate(self) -> None:
        configuration = complete_configuration(self.export_path)
        architecture = complete_architecture(self.export_path)
        config_errors, _ = validate_configuration(
            self.export_path, self.write_review("configuration.json", configuration)
        )
        architecture_errors, _ = validate_architecture(
            self.export_path, self.write_review("architecture.json", architecture)
        )
        self.assertEqual([], config_errors)
        self.assertEqual([], architecture_errors)
        purchase_family = next(
            row for row in architecture["families"] if "tag:1" in row["member_object_keys"]
        )
        self.assertIn("trigger:10", purchase_family["chain_object_keys"])
        self.assertIn("trigger:13", purchase_family["chain_object_keys"])
        self.assertIn("variable:24", purchase_family["chain_object_keys"])
        self.assertIn("variable:25", purchase_family["chain_object_keys"])
        self.assertIn("tag:3", purchase_family["chain_object_keys"])
        self.assertEqual(
            set(purchase_family["chain_object_keys"]),
            {item["object_key"] for item in purchase_family["chain_assessments"]},
        )

    def test_configuration_taxonomy_uses_the_defect_not_generic_consent_facts(
        self,
    ) -> None:
        from gtm_configuration_review import validate_row_outcome
        from gtm_operation_compile import configuration_taxonomy

        consent_facts = {
            "effective_consent_route_facts": {
                "consent_status": "needed",
                "detected_consent_payload_purposes": ["analytics_storage"],
            }
        }
        self.assertEqual(
            ("Event firing logic", "Broken reference"),
            configuration_taxonomy(
                {
                    **consent_facts,
                    "layer": "tag",
                    "defects": [
                        {
                            "statement": (
                                "The setup dependency resolves as missing in the export."
                            )
                        }
                    ],
                }
            ),
        )
        self.assertEqual(
            ("Custom code & templates", "Custom code risk"),
            configuration_taxonomy(
                {
                    **consent_facts,
                    "layer": "tag",
                    "required_technical_findings": [
                        {
                            "finding_key": "unsafe_dynamic_script",
                            "statement": "Custom code injects a dynamic script.",
                        }
                    ],
                }
            ),
        )

        decision_ready_issue = {
            "object_key": "tag:1",
            "object_name": "Broken setup route",
            "correctness_verdict": "Issue",
            "disposition": "owner_decision_needed",
            "owner_question": (
                "Which valid setup target should replace the missing route for tag:1?"
            ),
            "recommended_action": (
                "For tag:1, repair defect CFG-001 at "
                "$.containerVersion.tag[0].setupTag[0].tagName by replacing the "
                "missing target with the owner-approved existing setup tag."
            ),
            "defects": [
                {
                    "defect_id": "CFG-001",
                    "evidence_anchors": [
                        "$.containerVersion.tag[0].setupTag[0].tagName"
                    ],
                }
            ],
        }
        self.assertEqual(
            [],
            validate_row_outcome(decision_ready_issue, "decision-ready issue"),
        )
        vague = {
            **decision_ready_issue,
            "recommended_action": (
                "Correct every listed source-proven defect with one exact operation."
            ),
        }
        self.assertTrue(
            any(
                "must name the affected object" in error
                or "must cite a defect ID" in error
                for error in validate_row_outcome(vague, "vague issue")
            )
        )

    def test_architecture_verdicts_and_zero_discovery_are_fail_closed(self) -> None:
        review = complete_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if "exact_configuration" not in row.get("comparison_types", [])
        )
        comparison.update(
            {
                "relationship_verdict": "Owner decision needed",
                "disposition": "keep",
                "owner_question": "",
            }
        )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("incoherent-architecture-verdict.json", review),
        )
        self.assertTrue(any("requires owner_decision_needed" in error for error in errors))
        self.assertTrue(any("precise question" in error for error in errors))

        review = complete_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if row["relationship_verdict"] in {"Intentional variant", "Complementary"}
            and len(row["candidate_object_keys"]) > 1
        )
        for assessment in comparison["member_assessments"]:
            assessment["distinguishing_configuration"] = (
                "This object is kept because the general source review found a different role."
            )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("generic-retention-distinction.json", review),
        )
        self.assertTrue(
            any("configuration term unique to that member" in error for error in errors)
        )

        review = complete_architecture(self.export_path)
        review["open_discovery_attestation"]["zero_discovery_rationale"] = (
            "Every object was reviewed carefully and no additional relationship was found."
        )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("generic-zero-discovery.json", review),
        )
        self.assertTrue(any("naming every discovery method" in error for error in errors))

    def test_purchase_transaction_id_absence_cannot_be_marked_compliant(self) -> None:
        data = sample_export()
        event_parameters = data["containerVersion"]["tag"][0]["parameter"][2]["map"]
        data["containerVersion"]["tag"][0]["parameter"][2]["map"] = [
            item for item in event_parameters if item.get("key") != "transaction_id"
        ]
        export = self.root / "missing-transaction-id.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        scaffold = scaffold_configuration(export)
        purchase = next(row for row in scaffold["rows"] if row["object_key"] == "tag:1")
        topic = next(
            topic
            for topic in purchase["required_contract_topics"]
            if topic["topic"] == "purchase_transaction_id_uniqueness"
        )
        self.assertEqual("missing", topic["configuration_presence_state"])

        completed = complete_configuration(export)
        purchase_review = next(
            row for row in completed["rows"] if row["object_key"] == "tag:1"
        )
        purchase_check = next(
            check
            for check in purchase_review["contract_checks"]
            if check["contract_topic"] == topic["topic_key"]
        )
        purchase_check["verdict"] = "Compliant"
        errors, _ = validate_configuration(
            export,
            self.write_review("missing-transaction-id-review.json", completed),
        )
        self.assertTrue(
            any(
                "required exported configuration terms" in error
                and "transaction_id" in error
                for error in errors
            )
        )

    def test_architecture_requires_every_chain_object_and_server_root(self) -> None:
        review = complete_architecture(self.export_path)
        family = next(row for row in review["families"] if "tag:1" in row["member_object_keys"])
        family["chain_assessments"].pop()
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("missing-chain-assessment.json", review),
        )
        self.assertTrue(any("chain" in error and "every member" in error for error in errors))

        data = sample_export()
        data["containerVersion"]["client"] = [
            {"clientId": "50", "name": "GA4 client", "type": "gaaw_client"}
        ]
        data["containerVersion"]["transformation"] = [
            {
                "transformationId": "60",
                "name": "Redact user data",
                "type": "exclude_parameters",
            }
        ]
        server_path = self.root / "server.json"
        server_path.write_text(json.dumps(data), encoding="utf-8")
        server_review = scaffold_architecture(server_path)
        root_keys = {
            key
            for family_row in server_review["families"]
            for key in family_row["member_object_keys"]
        }
        self.assertIn("client:50", root_keys)
        self.assertIn("transformation:60", root_keys)

    def test_operational_validator_requires_all_findings(self) -> None:
        review = complete_operational(self.export_path)
        review["findings"].pop()
        errors, _ = validate_operational(
            self.export_path, self.write_review("bad-operational.json", review)
        )
        self.assertTrue(any("missing operational findings" in error for error in errors))

    def test_compiler_blocks_unaligned_basic_consolidation(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        operation = duplicate_variable_operation()
        finding = next(
            row
            for row in operational["findings"]
            if row["module_name"] == "duplicate_variable_paths"
        )
        finding.update(copy.deepcopy(operation))
        finding["disposition"] = "cleanup_operation"
        comparison = next(
            row
            for row in architecture["comparisons"]
            if set(row.get("candidate_object_keys", []))
            == {"variable:20", "variable:21"}
        )
        comparison.update(
            {
                "relationship_verdict": "Owner decision needed",
                "disposition": "owner_decision_needed",
                "owner_question": (
                    "Which variable should remain canonical for the shared items "
                    "contract?"
                ),
                "recommended_action": (
                    "Choose one canonical variable before consolidation."
                ),
                "operations": [],
            }
        )
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], payload["operations"])
        self.assertTrue(any("lacks an aligned business-architecture" in error for error in errors))

    def test_three_runs_reconcile_and_future_state_passes(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        for name, review, validator in (
            ("operational.json", operational, validate_operational),
            ("configuration.json", configuration, validate_configuration),
            ("architecture.json", architecture, validate_architecture),
        ):
            errors, _ = validator(self.export_path, self.write_review(name, review))
            self.assertEqual([], errors)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        self.assertGreaterEqual(len(payload["operations"]), 1)
        duplicate_packet = next(
            operation
            for operation in payload["operations"]
            if operation.get("canonical_object_key") == "variable:20"
            and any(
                deletion.get("object_key") == "variable:21"
                for deletion in operation.get("deletions", [])
            )
        )
        self.assertEqual(
            ["business_architecture", "operational_sanitation"],
            duplicate_packet["source_runs"],
        )
        self.assertTrue(payload["shared_facts_sha256"])
        self.assertTrue(payload["decision_ledger"])
        self.assertLess(payload["projected_object_counts"]["variable"]["delta"], 0)
        self.assertEqual(
            ["delete"],
            duplicate_packet["execution_phases"],
        )
        report, future_errors = check_future_state(self.export_path, payload)
        self.assertEqual([], future_errors)
        self.assertEqual("pass", report["status"])
        self.assertEqual(
            payload["projected_object_counts"]["variable"]["delta"],
            report["object_counts"]["variable"]["delta"],
        )

    def test_compiler_rejects_same_key_with_different_mutations(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        comparison = next(
            row
            for row in architecture["comparisons"]
            if set(row["candidate_object_keys"]) == {"variable:20", "variable:21"}
        )
        comparison["operations"][0]["canonical_object_key"] = "variable:21"
        comparison["operations"][0]["deletions"] = [
            {
                "object_key": "variable:20",
                "reason": "Use variable 21 as the conflicting canonical object in this test.",
            }
        ]
        _, errors = compile_operations(operational, configuration, architecture, "Manual")
        self.assertTrue(
            any("reused for different structured mutations" in error for error in errors)
        )

    def test_compiler_merges_wording_variants_for_the_same_mutation(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        comparison = next(
            row
            for row in architecture["comparisons"]
            if set(row["candidate_object_keys"]) == {"variable:20", "variable:21"}
        )
        comparison["operations"][0]["problem"] = (
            "DLV - Items Copy repeats the canonical ecommerce.items variable without consumers."
        )
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Manual",
        )
        self.assertEqual([], errors)
        duplicate_packet = next(
            operation
            for operation in payload["operations"]
            if operation.get("canonical_object_key") == "variable:20"
        )
        self.assertEqual(4, len(duplicate_packet["lens_rationales"]))
        self.assertEqual(
            {"operational_sanitation", "business_architecture"},
            {
                row["source_run"]
                for row in duplicate_packet["lens_rationales"]
            },
        )

    def test_compiler_folds_redundant_delete_into_broader_exact_consolidation(
        self,
    ) -> None:
        from gtm_operation_compile import (
            merge_compatible_operations,
            normalized_operation,
        )

        unused = normalized_operation(
            {
                "operation_key": "run-1-delete-unused-trigger",
                "deletions": [
                    {
                        "object_key": "trigger:10",
                        "reason": "The trigger has no surviving consumer.",
                    }
                ],
            },
            "operational_sanitation",
            "UNUSED-TRIGGER-10",
            ["trigger:10"],
        )
        consolidation = normalized_operation(
            {
                "operation_key": "run-3-consolidate-exact-triggers",
                "canonical_object_key": "trigger:30",
                "remaps": [
                    {
                        "from_object_key": "trigger:20",
                        "to_object_key": "trigger:30",
                        "consumer_object_keys": ["tag:1"],
                    }
                ],
                "deletions": [
                    {
                        "object_key": "trigger:10",
                        "reason": "It is an unused exact duplicate.",
                    },
                    {
                        "object_key": "trigger:20",
                        "reason": "Its consumer is remapped to the canonical trigger.",
                    },
                ],
            },
            "business_architecture",
            "REL-EXACT-TRIGGERS",
            ["trigger:10", "trigger:20", "trigger:30"],
        )

        errors: list[str] = []
        merged = merge_compatible_operations(
            [unused, consolidation],
            errors,
        )

        self.assertEqual([], errors)
        self.assertEqual(1, len(merged))
        self.assertEqual(
            ["business_architecture", "operational_sanitation"],
            merged[0]["source_runs"],
        )
        self.assertEqual(
            {"trigger:10", "trigger:20"},
            {
                deletion["object_key"]
                for deletion in merged[0]["deletions"]
            },
        )
        self.assertEqual("trigger:30", merged[0]["canonical_object_key"])

    def test_compiler_composes_only_changes_to_the_same_exported_field(self) -> None:
        from gtm_operation_compile import (
            merge_compatible_operations,
            normalized_operation,
        )

        def operation(key: str, object_key: str, path: str, after: str) -> dict:
            return normalized_operation(
                {
                    "operation_key": key,
                    "changes": [
                        {
                            "object_key": object_key,
                            "json_path": path,
                            "before": "alpha beta gamma",
                            "after": after,
                        }
                    ],
                },
                "operational_sanitation",
                key,
                [object_key],
            )

        unrelated_errors: list[str] = []
        unrelated = merge_compatible_operations(
            [
                operation(
                    "repair-tag-name",
                    "tag:1",
                    "$.containerVersion.tag[0].name",
                    "ALPHA beta gamma",
                ),
                operation(
                    "repair-trigger-name",
                    "trigger:10",
                    "$.containerVersion.trigger[0].name",
                    "alpha beta GAMMA",
                ),
            ],
            unrelated_errors,
        )
        self.assertEqual([], unrelated_errors)
        self.assertEqual(2, len(unrelated))

        same_field_errors: list[str] = []
        same_field = merge_compatible_operations(
            [
                operation(
                    "repair-left-token",
                    "tag:1",
                    "$.containerVersion.tag[0].parameter[0].value",
                    "ALPHA beta gamma",
                ),
                operation(
                    "repair-right-token",
                    "tag:1",
                    "$.containerVersion.tag[0].parameter[0].value",
                    "alpha beta GAMMA",
                ),
            ],
            same_field_errors,
        )
        self.assertEqual([], same_field_errors)
        self.assertEqual(1, len(same_field))
        self.assertEqual("ALPHA beta GAMMA", same_field[0]["changes"][0]["after"])

    def test_future_state_blocks_deletion_that_breaks_a_consumer(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(operational, configuration, architecture, "Manual")
        self.assertEqual([], errors)
        payload["operations"][0]["deletions"] = [
            {"object_key": "variable:20", "reason": "Intentional broken test deletion."}
        ]
        payload["operations"][0]["affected_object_keys"] = ["variable:20"]
        report, future_errors = check_future_state(self.export_path, payload)
        self.assertEqual("fail", report["status"])
        self.assertTrue(any("missing references" in error for error in future_errors))

    def test_future_state_rejects_stale_before_value(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(operational, configuration, architecture, "Manual")
        self.assertEqual([], errors)
        payload["operations"][0]["changes"] = [
            {
                "object_key": "tag:1",
                "json_path": "$.containerVersion.tag[0].name",
                "before": "A stale tag name",
                "after": "GA4 - Purchase - Global",
            }
        ]
        report, future_errors = check_future_state(self.export_path, payload)
        self.assertEqual("fail", report["status"])
        self.assertTrue(any("before value does not match" in error for error in future_errors))

    def test_future_state_rechecks_configuration_and_architecture_quality(self) -> None:
        configuration_data = sample_export()
        configuration_data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Google configuration missing its required type",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "measurementId",
                        "value": "G-TEST123",
                    }
                ],
            }
        ]
        configuration_export = self.root / "future-configuration-issue.json"
        configuration_export.write_text(json.dumps(configuration_data), encoding="utf-8")
        configuration_report, configuration_errors = check_future_state(
            configuration_export,
            {"operations": [], "plan_status": "complete"},
        )
        self.assertEqual("fail", configuration_report["projected_quality"]["status"])
        self.assertTrue(
            any(
                "retains unaccounted deterministic configuration Issues" in error
                for error in configuration_errors
            )
        )

        architecture_data = sample_export()
        architecture_export = self.root / "future-architecture-candidate.json"
        architecture_export.write_text(json.dumps(architecture_data), encoding="utf-8")
        created_variable = copy.deepcopy(architecture_data["containerVersion"]["variable"][0])
        created_variable.update(
            {"variableId": "99", "name": "DLV - Items Third Copy"}
        )
        architecture_report, architecture_errors = check_future_state(
            architecture_export,
            {
                "plan_status": "complete",
                "operations": [
                    {
                        "source_runs": ["configuration_correctness"],
                        "affected_object_keys": ["variable:99"],
                        "creations": [
                            {"layer": "variable", "object": created_variable}
                        ],
                    }
                ],
            },
        )
        self.assertTrue(
            architecture_report["projected_quality"]["architecture"][
                "unexpected_new_candidates"
            ]
        )
        self.assertTrue(
            any(
                "outside architecture-backed operations" in error
                for error in architecture_errors
            )
        )

        architecture_backed = {
            "plan_status": "complete",
            "operations": [
                {
                    "source_runs": ["business_architecture"],
                    "source_object_keys": [
                        "variable:20",
                        "variable:21",
                        "variable:99",
                    ],
                    "affected_object_keys": ["variable:99"],
                    "creations": [{"layer": "variable", "object": created_variable}],
                }
            ],
        }
        covered_report, covered_errors = check_future_state(
            architecture_export, architecture_backed
        )
        self.assertEqual(
            [],
            covered_report["projected_quality"]["architecture"][
                "unexpected_new_candidates"
            ],
        )
        self.assertFalse(
            any(
                "outside architecture-backed operations" in error
                for error in covered_errors
            )
        )

    def test_future_state_retention_coverage_requires_deletion_only_subset(self) -> None:
        from gtm_future_state_check import (
            non_deletion_mutation_keys,
            planned_deleted_keys,
            retained_architecture_comparisons,
            retention_coverage_decision,
        )

        operations = {
            "operations": [
                {
                    "affected_object_keys": ["tag:3"],
                    "deletions": [{"object_key": "tag:3"}],
                }
            ],
            "decision_ledger": [
                {
                    "decision_id": "REL-RETAINED",
                    "source_run": "business_architecture",
                    # The compiled ledger omits the deleted source member.
                    "source_object_keys": ["tag:1", "tag:2"],
                    "comparison_types": [
                        "shared_configured_destination",
                        "shared_destination_consent_inheritance_review",
                    ],
                    "verdict": "Intentional variant",
                    "disposition": "keep",
                }
            ],
        }
        source_architecture = {
            "comparisons": [
                {
                    "comparison_id": "REL-RETAINED",
                    "candidate_object_keys": ["tag:1", "tag:2", "tag:3"],
                    "comparison_types": [
                        "shared_configured_destination",
                        "shared_destination_consent_inheritance_review",
                    ],
                }
            ]
        }
        candidate = {
            "candidate_object_keys": ["tag:1", "tag:2"],
            "comparison_types": [
                "shared_configured_destination",
                "shared_destination_consent_inheritance_review",
            ],
        }
        self.assertEqual(
            "REL-RETAINED",
            retention_coverage_decision(
                candidate,
                retained_architecture_comparisons(operations, source_architecture),
                planned_deleted_keys(operations),
                non_deletion_mutation_keys(operations),
            ),
        )

        changed_survivor = copy.deepcopy(operations)
        changed_survivor["operations"].append(
            {
                "affected_object_keys": ["tag:1"],
                "changes": [
                    {
                        "object_key": "tag:1",
                        "json_path": "$.containerVersion.tag[0].parameter[0].value",
                    }
                ],
            }
        )
        self.assertEqual(
            "",
            retention_coverage_decision(
                candidate,
                retained_architecture_comparisons(
                    changed_survivor, source_architecture
                ),
                planned_deleted_keys(changed_survivor),
                non_deletion_mutation_keys(changed_survivor),
            ),
        )

    def test_future_state_allows_only_explicit_owner_blocked_configuration_issues(
        self,
    ) -> None:
        from gtm_future_state_check import projected_quality_review

        data = sample_export()
        data["containerVersion"]["tag"][0]["teardownTag"] = [{}]
        export = self.root / "owner-blocked-configuration-issue.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        operations = {
            "plan_status": "complete",
            "operations": [],
            "decision_ledger": [
                {
                    "source_run": "configuration_correctness",
                    "source_object_keys": ["tag:1"],
                    "verdict": "Issue",
                    "disposition": "owner_decision_needed",
                }
            ],
        }
        report, errors = projected_quality_review(export, data, operations)
        self.assertEqual([], errors)
        self.assertTrue(report["configuration"]["owner_blocked_issues"])
        self.assertEqual(
            [],
            report["configuration"]["unaccounted_remaining_issues"],
        )

        unaccounted = copy.deepcopy(operations)
        unaccounted["decision_ledger"] = []
        _, errors = projected_quality_review(export, data, unaccounted)
        self.assertTrue(
            any(
                "unaccounted deterministic configuration Issues" in error
                for error in errors
            ),
            errors,
        )

    def test_structured_operations_support_creation_and_missing_field_addition(self) -> None:
        operation = {
            "problem_type": "Missing tracking",
            "creations": [
                {
                    "layer": "variable",
                    "object": {
                        "variableId": "99",
                        "name": "Constant - Test Value",
                        "type": "c",
                        "parameter": [
                            {"type": "TEMPLATE", "key": "value", "value": "test"}
                        ],
                    },
                    "reason": "Create the missing constant required by the approved test tag.",
                }
            ],
            "additions": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].parameter",
                    "mode": "append",
                    "value": {
                        "type": "TEMPLATE",
                        "key": "testParameter",
                        "value": "{{Constant - Test Value}}",
                    },
                    "reason": "Add the approved parameter to the existing purchase tag.",
                }
            ],
            "changes": [],
            "remaps": [],
            "renames": [],
            "deletions": [],
        }
        validation_errors = validate_structured_actions(
            operation,
            object_keys(self.export_path),
            "creation test",
            object_consumer_map(self.export_path),
            object_source_path_map(self.export_path),
        )
        self.assertEqual([], validation_errors)
        mismatched = copy.deepcopy(operation)
        mismatched["additions"][0]["json_path"] = (
            "$.containerVersion.tag[1].parameter"
        )
        mismatch_errors = validate_structured_actions(
            mismatched,
            object_keys(self.export_path),
            "source-path binding test",
            object_consumer_map(self.export_path),
            object_source_path_map(self.export_path),
        )
        self.assertTrue(
            any("paired with another object's json_path" in error for error in mismatch_errors)
        )
        future, apply_errors = apply_operations(
            sample_export(),
            {"operations": [operation]},
        )
        self.assertEqual([], apply_errors)
        future_cv = container_version(future)
        self.assertTrue(any(row.get("variableId") == "99" for row in future_cv["variable"]))
        self.assertEqual(
            "testParameter",
            future_cv["tag"][0]["parameter"][-1]["key"],
        )

    def test_compiler_keeps_every_valid_operation_for_explicit_approval(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Manual",
        )
        self.assertEqual([], errors)
        self.assertGreaterEqual(len(payload["operations"]), 1)
        self.assertNotIn("deferred_operations", payload)
        self.assertNotIn("aggressiveness", payload)
        self.assertTrue(
            all(
                operation["resolution_status"] == "proposed"
                for operation in payload["operations"]
            )
        )
        human, human_errors = build_rows(payload)
        self.assertEqual([], human_errors)
        proposed_rows = [row for row in human if row["Status"] == "Proposed action"]
        rendered_ids = {
            operation_id
            for row in proposed_rows
            for operation_id in re.findall(r"\bOP-\d{4}\b", row["ID"])
        }
        self.assertEqual(
            {operation["operation_id"] for operation in payload["operations"]},
            rendered_ids,
        )

        deprecated = duplicate_variable_operation()
        deprecated["minimum_aggressiveness"] = "Standard"
        deprecated_errors = validate_structured_actions(
            deprecated,
            object_keys(self.export_path),
            "deprecated mode test",
            object_consumer_map(self.export_path),
            object_source_path_map(self.export_path),
        )
        self.assertTrue(any("deprecated" in error for error in deprecated_errors))

    def test_human_plan_exposes_owner_decisions_without_internal_proof_columns(self) -> None:
        payload = {
            "operations": [],
            "decision_ledger": [
                {
                    "decision_id": "CFG-001",
                    "disposition": "owner_decision_needed",
                    "area": "Governance / ownership",
                    "problem_type": "Unclear business purpose",
                    "affected_objects": "tag:1 - Legacy lead",
                    "summary": "Legacy lead sends a second conversion for the same form submit.",
                    "owner_question": (
                        "Should Legacy lead remain a separate paid-media conversion?"
                    ),
                    "recommended_action": (
                        "Retain one canonical paid-media conversion unless the owner supplies "
                        "evidence that the second conversion serves a distinct optimization goal."
                    ),
                }
            ],
        }
        rows, errors = build_rows(payload)
        self.assertEqual([], errors)
        self.assertEqual(1, len(rows))
        self.assertEqual("Owner confirmation", rows[0]["Status"])
        self.assertEqual(7, len(rows[0]))
        self.assertEqual(list(CLEANUP_PLAN_COLUMNS), list(rows[0]))
        self.assertEqual(
            "Consent & governance",
            rows[0]["General problem category"],
        )
        self.assertNotIn("related decision", rows[0]["Affected object(s)"])
        self.assertNotIn("related decision", rows[0]["Problem / evidence"])

    def test_general_problem_taxonomy_covers_every_supported_problem_type(
        self,
    ) -> None:
        self.assertEqual(
            set(PROBLEM_TYPES),
            set(GENERAL_CATEGORY_BY_PROBLEM_TYPE) - {"Incomplete action plan"},
        )
        self.assertTrue(
            all(
                general_problem_category(problem_type)
                in GENERAL_PROBLEM_CATEGORIES
                for problem_type in PROBLEM_TYPES
            )
        )

    def test_human_plan_hides_machine_source_paths_in_owner_rows(self) -> None:
        rows, errors = build_rows(
            {
                "operations": [],
                "decision_ledger": [
                    {
                        "decision_id": "CFG-001",
                        "disposition": "owner_decision_needed",
                        "area": "Custom code & templates",
                        "problem_type": "Custom code risk",
                        "affected_objects": "tag:1 - Marketing pixel",
                        "summary": "The tag needs a business-safe configuration decision.",
                        "owner_question": (
                            "Should tag:1 Marketing pixel retain the configuration at "
                            "$.containerVersion.tag[0].consentSettings, or be corrected?"
                        ),
                        "recommended_action": (
                            "For tag:1, repair or remove the exact source anchor "
                            "$.containerVersion.tag[0].consentSettings after approval."
                        ),
                    }
                ],
            }
        )
        self.assertEqual([], errors)
        visible = rows[0]["Action / priority / QA"]
        self.assertNotIn("$.containerVersion", visible)
        self.assertIn("Question:", visible)
        self.assertIn("Recommendation:", visible)

    def test_human_plan_calls_builtin_removal_disable_or_deselect(self) -> None:
        rows, errors = build_rows(
            {
                "operations": [
                    {
                        "operation_id": "OP-0001",
                        "area": "GTM hygiene",
                        "problem_type": "Unused object",
                        "affected_objects": (
                            "builtInVariable:Click Element — Click Element"
                        ),
                        "problem": "Click Element is enabled but has no consumer.",
                        "why_it_matters": "Unused built-ins add configuration noise.",
                        "exact_proposed_action": "Delete Click Element.",
                        "qa_steps": (
                            "Re-export and confirm Click Element is no longer enabled."
                        ),
                        "deletions": [
                            {
                                "object_key": "builtInVariable:Click Element",
                                "reason": "No exported consumer uses this built-in.",
                            }
                        ],
                        "priority": "Low",
                        "execution_readiness": "approval_required",
                        "approval_status": "pending_approval",
                    }
                ],
                "decision_ledger": [],
            }
        )
        self.assertEqual([], errors)
        self.assertIn(
            "Disable/deselect",
            rows[0]["Action / priority / QA"],
        )
        self.assertNotIn(
            "Target state / exact action: Delete",
            rows[0]["Action / priority / QA"],
        )

    def test_human_plan_batches_nonblocking_container_evidence_limits(self) -> None:
        evidence_decision = {
            "disposition": "container_evidence_limit",
            "area": "Consent & compliance",
            "problem_type": "Consent mismatch",
            "summary": "Live consent timing is outside the container export.",
            "owner_question": "What runtime evidence proves the live consent timing?",
            "recommended_action": "Validate the live route separately.",
        }
        rows, errors = build_rows(
            {
                "operations": [],
                "decision_ledger": [
                    {
                        **evidence_decision,
                        "decision_id": "CFG-001",
                        "affected_objects": "tag:1 - Analytics",
                    },
                    {
                        **evidence_decision,
                        "decision_id": "CFG-002",
                        "affected_objects": "variable:2 - Event name",
                    },
                ],
            }
        )
        self.assertEqual([], errors)
        self.assertEqual(1, len(rows))
        self.assertEqual("SCOPE-001", rows[0]["ID"])
        self.assertEqual("Evidence boundary", rows[0]["Status"])
        self.assertIn("2 retained review decision", rows[0]["Affected object(s)"])
        self.assertIn("do not block unrelated cleanup", rows[0]["Problem / evidence"])

    def test_remap_requires_the_exact_source_consumer_set(self) -> None:
        review = complete_configuration(self.export_path)
        row = next(item for item in review["rows"] if item["object_key"] == "variable:20")
        operation = duplicate_variable_operation()
        operation.update(
            {
                "operation_key": "remap-items-variable",
                "canonical_object_key": "variable:21",
                "changes": [],
                "deletions": [],
                "remaps": [
                    {
                        "from_object_key": "variable:20",
                        "to_object_key": "variable:21",
                        "consumer_object_keys": ["tag:1"],
                    }
                ],
                "exact_proposed_action": (
                    "Remap the items variable to the selected canonical variable."
                ),
            }
        )
        row["disposition"] = "cleanup_operation"
        row["operation"] = operation
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("bad-remap-consumers.json", review),
        )
        self.assertTrue(
            any("exactly match every source-graph consumer" in error for error in errors)
        )

    def test_package_build_contains_three_independent_review_artifacts(self) -> None:
        package_dir = self.root / "package"
        manifest = build_package(self.export_path, package_dir, pretty=True)
        self.assertEqual("pass", manifest["status"])
        for filename in (
            "context.json",
            "shared_facts.json",
            "operational_review.json",
            "configuration_review.json",
            "architecture_review.json",
            "operational_scan.json",
        ):
            self.assertTrue((package_dir / filename).is_file())
        self.assertNotIn("semantic_review", manifest["files"])
        shared = json.loads((package_dir / "shared_facts.json").read_text(encoding="utf-8"))
        context = json.loads((package_dir / "context.json").read_text(encoding="utf-8"))
        self.assertEqual(shared["shared_facts_sha256"], manifest["shared_facts_sha256"])
        self.assertEqual(context["context_sha256"], manifest["context_sha256"])
        for filename in (
            "operational_review.json",
            "configuration_review.json",
            "architecture_review.json",
        ):
            review = json.loads((package_dir / filename).read_text(encoding="utf-8"))
            self.assertEqual(shared["shared_facts_sha256"], review["shared_facts_sha256"])

    def test_package_builder_auto_shards_only_large_reviews(self) -> None:
        small_export = {
            "exportFormatVersion": 2,
            "containerVersion": {
                "accountId": "100",
                "containerId": "200",
                "containerVersionId": "1",
                "container": {"publicId": "GTM-TEST", "usageContext": ["WEB"]},
                "tag": [],
                "trigger": [],
                "variable": [],
            },
        }
        small_path = self.root / "small-container.json"
        small_path.write_text(json.dumps(small_export), encoding="utf-8")
        small_package = self.root / "small-package"
        small_manifest = build_package(small_path, small_package, pretty=True)
        self.assertTrue(
            all(
                run["strategy"] == "single_file"
                for run in small_manifest["review_work_units"]["runs"].values()
            )
        )
        self.assertFalse((small_package / "configuration-shards").exists())

        large_export = copy.deepcopy(small_export)
        large_export["containerVersion"]["variable"] = [
            {
                "variableId": str(index),
                "name": f"Constant {index}",
                "type": "c",
                "parameter": [
                    {"type": "TEMPLATE", "key": "value", "value": str(index)}
                ],
            }
            for index in range(1, 42)
        ]
        large_path = self.root / "large-container.json"
        large_path.write_text(json.dumps(large_export), encoding="utf-8")
        large_package = self.root / "large-package"
        large_manifest = build_package(large_path, large_package, pretty=True)
        configuration = large_manifest["review_work_units"]["runs"][
            "configuration_correctness"
        ]
        self.assertEqual("sharded", configuration["strategy"])
        self.assertGreater(configuration["primary_shards"], 1)
        self.assertTrue(
            (large_package / configuration["shard_manifest"]).is_file()
        )

    def test_single_oversized_configuration_obligation_group_requires_sharding(
        self,
    ) -> None:
        review = {
            "kind": "gtm_configuration_correctness_review",
            "rows": [
                {
                    "required_branch_reviews": [
                        {"json_path": f"$.parameter[{index}]"}
                        for index in range(31)
                    ]
                }
            ],
        }
        self.assertTrue(review_requires_sharding(review))
        review["rows"][0]["required_branch_reviews"].pop()
        self.assertFalse(review_requires_sharding(review))

    def test_package_gate_rejects_shared_fact_content_with_a_copied_hash(self) -> None:
        package_dir = self.root / "tampered-package"
        build_package(self.export_path, package_dir, pretty=True)
        shared_path = package_dir / "shared_facts.json"
        shared = json.loads(shared_path.read_text(encoding="utf-8"))
        shared["objects"][0]["object_name"] = "Fabricated object name"
        shared_path.write_text(json.dumps(shared), encoding="utf-8")
        report = run_gate(self.export_path, package_dir)
        self.assertEqual("fail", report["status"])
        self.assertTrue(any("recorded hash" in error for error in report["errors"]))

    def test_broken_references_remain_auditable_integrity_findings(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["firingTriggerId"] = ["999"]
        path = self.root / "broken-reference.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        model = build_model(path)
        self.assertEqual("pass_with_integrity_findings", model["coverage_gate"])
        package_dir = self.root / "broken-package"
        manifest = build_package(path, package_dir, pretty=True)
        self.assertEqual("pass", manifest["status"])
        shared = build_shared_facts(path, context=build_context_model(path))
        self.assertEqual("pass_with_integrity_findings", shared["coverage_gate"])
        findings = audit_export(path)["findings"]
        self.assertTrue(
            any(row["finding_type"] == "missing_trigger_reference" for row in findings)
        )

    def test_verdict_engines_share_only_neutral_review_helpers(self) -> None:
        configuration_source = (SCRIPTS / "gtm_configuration_review.py").read_text(encoding="utf-8")
        architecture_source = (SCRIPTS / "gtm_architecture_review.py").read_text(encoding="utf-8")
        skill_source = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        execution_contract = (
            ROOT / "references" / "03-rules" / "execution-contract.md"
        ).read_text(encoding="utf-8")
        self.assertNotIn("from gtm_operational_review import", configuration_source)
        self.assertNotIn("from gtm_operational_review import", architecture_source)
        self.assertIn("from gtm_review_common import", configuration_source)
        self.assertIn("from gtm_review_common import", architecture_source)
        self.assertIn("fresh reasoning context", skill_source)
        self.assertIn("prohibited inputs", skill_source)
        self.assertIn("exclude completed verdict artifacts", execution_contract)

        completed = complete_configuration(self.export_path)
        self.assertNotIn("related_operational_finding_ids", completed["rows"][0])
        self.assertIn(
            "operational_review",
            completed["input_contract"]["prohibited_artifact_roles"],
        )
        completed["completion_attestation"]["foreign_verdict_artifacts_used"] = [
            "operational_review.json"
        ]
        completed["completion_attestation"]["helper_modules"] = [
            "tests.test_pipeline.complete_configuration"
        ]
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("foreign-input-configuration.json", completed),
        )
        self.assertTrue(any("foreign verdict artifacts" in error for error in errors))
        self.assertTrue(any("test helpers" in error for error in errors))

    def test_large_review_shards_merge_only_with_complete_exact_coverage(self) -> None:
        completed = complete_configuration(self.export_path)
        base_path = self.write_review("configuration-complete.json", completed)
        shard_dir = self.root / "configuration-shards"
        manifest = split_review(base_path, shard_dir, max_items=3)
        self.assertGreater(len(manifest["shards"]), 1)
        output = self.root / "configuration-merged.json"
        merged = merge_review(base_path, shard_dir, output)
        self.assertEqual(completed["rows"], merged["rows"])
        self.assertEqual("complete", merged["run_status"])
        merged_in_place = merge_review(base_path, shard_dir, base_path)
        self.assertEqual(completed["rows"], merged_in_place["rows"])
        self.assertEqual("complete", merged_in_place["run_status"])

        shard_manifest_path = shard_dir / "shard_manifest.json"
        broken_manifest = json.loads(shard_manifest_path.read_text(encoding="utf-8"))
        broken_manifest["shards"][0]["filename"] = "missing-shard.json"
        shard_manifest_path.write_text(json.dumps(broken_manifest), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "missing review shard"):
            merge_review(base_path, shard_dir, self.root / "must-not-exist.json")

    def test_architecture_shards_merge_discovery_rows_and_attestation(self) -> None:
        completed = complete_architecture(self.export_path)
        completed["comparisons"] = [
            row
            for row in completed["comparisons"]
            if row.get("comparison_origin") != "analyst_discovered"
        ]
        completed["open_discovery_attestation"]["discovered_comparison_ids"] = []
        base_path = self.write_review("architecture-complete.json", completed)
        shard_dir = self.root / "architecture-shards"
        manifest = split_review(base_path, shard_dir, max_items=2)
        discovery_path = shard_dir / manifest["discovery_shard"]
        discovery = json.loads(discovery_path.read_text(encoding="utf-8"))
        discovery["discovered_comparisons"] = [value_discovery_row(self.export_path)]
        discovery["open_discovery_attestation"].update(
            {
                "review_status": "complete",
                "discovered_comparison_ids": ["DISC-VALUE-001"],
                "zero_discovery_rationale": "",
            }
        )
        next(
            item
            for item in discovery["open_discovery_attestation"]["method_reviews"]
            if item["method"] == "terminal_source_formula_and_output_overlap"
        )["additional_discovery_ids"] = ["DISC-VALUE-001"]
        discovery_path.write_text(json.dumps(discovery), encoding="utf-8")
        merged_path = self.root / "architecture-merged.json"
        merged = merge_review(base_path, shard_dir, merged_path)
        self.assertIn(
            "DISC-VALUE-001",
            {row["comparison_id"] for row in merged["comparisons"]},
        )
        errors, _ = validate_architecture(self.export_path, merged_path)
        self.assertEqual([], errors)

    def test_three_run_gate_rejects_scaffolds_and_action_incomplete_reviews(self) -> None:
        package_dir = self.root / "package"
        build_package(self.export_path, package_dir, pretty=True)
        pending = run_gate(self.export_path, package_dir)
        self.assertEqual("fail", pending["status"])
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        (package_dir / "operational_review.json").write_text(
            json.dumps(operational), encoding="utf-8"
        )
        (package_dir / "configuration_review.json").write_text(
            json.dumps(configuration), encoding="utf-8"
        )
        (package_dir / "architecture_review.json").write_text(
            json.dumps(architecture), encoding="utf-8"
        )
        payload, compile_errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], compile_errors)
        self.assertEqual("complete", payload["plan_status"])
        incomplete = copy.deepcopy(payload)
        incomplete["plan_status"] = "incomplete_actions"
        incomplete["action_completeness"] = {
            **incomplete["action_completeness"],
            "status": "incomplete",
            "errors": [
                "OPS-TEST: deterministic operational finding with a source-proven "
                "safe action has no exact cleanup operation"
            ],
        }
        operations_path = self.write_review("operations.json", incomplete)
        completed = run_gate(self.export_path, package_dir, operations_path)
        self.assertEqual("fail", completed["status"])
        self.assertTrue(
            any(
                "action completeness" in error
                or "deterministic recompilation" in error
                for error in completed["errors"]
            )
        )

        tampered = copy.deepcopy(payload)
        tampered["operations"][0]["exact_proposed_action"] = (
            "A hand-edited action that was not compiled from the reviews."
        )
        tampered_path = self.write_review("tampered-operations.json", tampered)
        rejected = run_gate(self.export_path, package_dir, tampered_path)
        self.assertEqual("fail", rejected["status"])
        self.assertTrue(any("deterministic recompilation" in error for error in rejected["errors"]))

    def test_action_completeness_accepts_exact_actions_and_genuine_owner_limits(self) -> None:
        ledger = [
            {
                "decision_id": "OPS-001",
                "source_run": "operational_sanitation",
                "disposition": "cleanup_operation",
                "verdict": "Issue",
                "compiled_operation_ids": ["OP-0001"],
            },
            {
                "decision_id": "OPS-CANDIDATE-001",
                "source_run": "operational_sanitation",
                "disposition": "keep",
                "finding_class": "review_candidate",
                "verdict": "normalized_duplicate_tag_signature",
                "recommended_action": (
                    "Retain the two routes because their exported blocker conditions prove "
                    "that they execute in distinct business scopes."
                ),
                "compiled_operation_ids": [],
            },
            {
                "decision_id": "CFG-001",
                "source_run": "configuration_correctness",
                "disposition": "owner_decision_needed",
                "verdict": "Owner decision needed",
                "recommended_action": (
                    "Obtain the named runtime evidence and correct the configuration if the "
                    "required contract cannot be demonstrated."
                ),
                "compiled_operation_ids": [],
            },
            {
                "decision_id": "ARCH-001",
                "source_run": "business_architecture",
                "disposition": "keep",
                "verdict": "Complementary",
                "compiled_operation_ids": [],
            },
        ]
        self.assertEqual("pass", action_completeness_report(ledger)["status"])

        deterministic_fallback = copy.deepcopy(ledger)
        deterministic_fallback[0].update(
            {"disposition": "owner_decision_needed", "compiled_operation_ids": []}
        )
        report = action_completeness_report(deterministic_fallback)
        self.assertEqual("incomplete", report["status"])
        self.assertTrue(any("deterministic operational finding" in error for error in report["errors"]))

    def test_reconciliation_resolves_surviving_canonical_and_rejects_its_deletion(
        self,
    ) -> None:
        from gtm_operation_compile import reconcile_ledger_resolutions

        ledger = [
            {
                "decision_id": "REL-001",
                "source_run": "business_architecture",
                "source_object_keys": ["trigger:10", "trigger:11"],
                "recommended_canonical_object_key": "trigger:10",
                "disposition": "owner_decision_needed",
                "compiled_operation_ids": [],
                "operation_keys": [],
            }
        ]
        packets = [
            {
                "operation_id": "OP-0001",
                "operation_key": "delete-noncanonical-trigger",
                "resolution_status": "proposed",
                "deletions": [
                    {
                        "object_key": "trigger:11",
                        "reason": "The exact duplicate has no surviving consumer.",
                    }
                ],
            }
        ]
        reconciled, errors = reconcile_ledger_resolutions(ledger, packets)
        self.assertEqual([], errors)
        self.assertEqual("cleanup_operation", reconciled[0]["disposition"])
        self.assertEqual(
            "resolved_by_surviving_canonical_object",
            reconciled[0]["reconciliation_status"],
        )
        self.assertEqual(["OP-0001"], reconciled[0]["compiled_operation_ids"])

        canonical_deleted = copy.deepcopy(packets)
        canonical_deleted[0]["deletions"][0]["object_key"] = "trigger:10"
        _, errors = reconcile_ledger_resolutions(ledger, canonical_deleted)
        self.assertTrue(
            any("recommended canonical object" in error for error in errors),
            errors,
        )

    def test_reconciliation_narrows_retained_relationship_after_architecture_cleanup(
        self,
    ) -> None:
        from gtm_operation_compile import reconcile_ledger_resolutions

        ledger = [
            {
                "decision_id": "REL-SECONDARY",
                "source_run": "business_architecture",
                "source_object_keys": ["trigger:10", "trigger:11"],
                "recommended_canonical_object_key": "",
                "disposition": "keep",
                "compiled_operation_ids": [],
                "operation_keys": [],
            }
        ]
        packets = [
            {
                "operation_id": "OP-ARCH-001",
                "operation_key": "consolidate-trigger",
                "source_runs": ["business_architecture"],
                "resolution_status": "proposed",
                "deletions": [
                    {
                        "object_key": "trigger:11",
                        "reason": "Source-proven duplicate is consolidated.",
                    }
                ],
            }
        ]
        reconciled, errors = reconcile_ledger_resolutions(ledger, packets)
        self.assertEqual([], errors)
        self.assertEqual("keep", reconciled[0]["disposition"])
        self.assertEqual(
            "narrowed_by_architecture_cleanup",
            reconciled[0]["reconciliation_status"],
        )
        self.assertEqual(["trigger:10"], reconciled[0]["source_object_keys"])

    def test_inactive_lifecycle_deletion_does_not_require_fabricated_architecture(self) -> None:
        finding = {
            "finding_id": "BASE-UNUSED_VARIABLES-001",
            "finding_type": "unused_object",
            "object_type": "variable",
            "object_ids": ["9"],
            "shared_fact_object_keys": ["variable:9"],
        }
        operation = {
            "source_references": [finding["finding_id"]],
            "creations": [],
            "additions": [],
            "changes": [],
            "remaps": [],
            "deletions": [
                {"object_key": "variable:9", "reason": "No active execution root reaches it."}
            ],
            "renames": [],
        }
        self.assertEqual(
            {"variable:9"},
            runtime_neutral_operational_deletions(
                operation, {finding["finding_id"]: finding}
            ),
        )
        operation["changes"] = [
            {
                "object_key": "variable:9",
                "json_path": "$.containerVersion.variable[0].parameter[0].value",
                "before": "old",
                "after": "new",
            }
        ]
        self.assertEqual(
            set(),
            runtime_neutral_operational_deletions(
                operation, {finding["finding_id"]: finding}
            ),
        )

    def test_human_rows_and_workbook_are_compact_and_separate_from_change_log(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        preservation = payload["measurement_preservation"]
        self.assertEqual(
            len(architecture["families"]),
            len(preservation["families"]),
        )
        self.assertTrue(
            all("target_state" in family for family in preservation["families"])
        )
        self.assertTrue(
            all(
                "affected_measurement_family_ids" in operation
                and "retained_behavior" in operation
                for operation in payload["operations"]
            )
        )
        human, human_errors = build_rows(payload)
        self.assertEqual([], human_errors)
        self.assertEqual(7, len(human[0]))
        self.assertEqual(list(CLEANUP_PLAN_COLUMNS), list(human[0]))
        self.assertIn(
            "same exported configuration", human[0]["Problem / evidence"]
        )
        self.assertIn("Change", human[0]["Action / priority / QA"])
        self.assertIn("Static verification:", human[0]["Action / priority / QA"])
        workbook_path = self.root / "cleanup-plan.xlsx"
        manifest = {"source_file": self.export_path.name, "source_sha256": payload["source_sha256"]}
        source = build_model(self.export_path)
        build_workbook(
            manifest,
            source,
            operational,
            configuration,
            architecture,
            payload,
            {"rows": human},
            workbook_path,
        )
        workbook = load_workbook(workbook_path)
        self.assertEqual(CANONICAL_SHEETS, workbook.sheetnames)
        self.assertEqual(
            ["01 Summary", "02 Cleanup Plan"],
            [s.title for s in workbook if s.sheet_state == "visible"],
        )
        summary_decisions = {
            str(row[0].value)
            for row in workbook["01 Summary"].iter_rows(min_row=2, max_col=1)
        }
        self.assertIn("Retained / no-change decisions", summary_decisions)
        self.assertIn("Retained business-family architecture", summary_decisions)
        self.assertIn("Measurement-family preservation", summary_decisions)
        self.assertIn("Target-state architecture", summary_decisions)
        self.assertIn("Preservation evidence boundary", summary_decisions)
        self.assertIn("Highest-impact proposed actions", summary_decisions)
        self.assertIn("Filterable problem taxonomy", summary_decisions)
        cleanup_sheet = workbook["02 Cleanup Plan"]
        self.assertEqual(
            list(CLEANUP_PLAN_COLUMNS),
            [cell.value for cell in cleanup_sheet[1]],
        )
        self.assertEqual(f"A1:G{cleanup_sheet.max_row}", cleanup_sheet.auto_filter.ref)
        self.assertEqual("A2", cleanup_sheet.freeze_panes)
        for sheet in workbook:
            column_limit = 7 if sheet.title == "02 Cleanup Plan" else 6
            self.assertLessEqual(sheet.max_column, column_limit)
            self.assertLessEqual(
                max(sheet.column_dimensions[col].width or 0 for col in sheet.column_dimensions), 92
            )
            self.assertLessEqual(
                max(sheet.row_dimensions[row].height or 0 for row in sheet.row_dimensions),
                MAX_ROW_HEIGHT,
            )
        self.assertNotIn("Change Log", workbook.sheetnames)
        gate_errors, gate_warnings = validate_workbook(
            workbook_path,
            self.write_review("workbook-operations.json", payload),
        )
        self.assertEqual([], gate_errors)
        self.assertEqual([], gate_warnings)
        invalid_workbook_path = self.root / "cleanup-plan-invalid-category.xlsx"
        cleanup_sheet["C2"] = "Invented category"
        workbook.save(invalid_workbook_path)
        invalid_errors, _invalid_warnings = validate_workbook(
            invalid_workbook_path,
            self.write_review("workbook-invalid-category-operations.json", payload),
        )
        self.assertTrue(
            any(
                "unsupported general problem category" in error
                for error in invalid_errors
            )
        )

    def test_human_rows_present_high_impact_actions_first_without_rekeying(self) -> None:
        def operation(
            operation_id: str,
            priority: str,
            area: str,
            problem_type: str,
        ) -> dict:
            return {
                "operation_id": operation_id,
                "area": area,
                "problem_type": problem_type,
                "problem": f"{priority} fixture problem",
                "why_it_matters": f"{priority} fixture business impact",
                "exact_proposed_action": f"Apply the {priority} fixture target state.",
                "qa_steps": f"Verify the {priority} fixture outcome.",
                "priority": priority,
                "execution_readiness": "approval_required",
                "execution_order": 2 if priority == "Critical" else 1,
                "affected_objects": f"tag:{operation_id}",
                "blocker": "",
            }

        rows, errors = build_rows(
            {
                "operations": [
                    operation(
                        "OP-LOW",
                        "Low",
                        "GTM hygiene",
                        "Unnecessary complexity",
                    ),
                    operation(
                        "OP-CRITICAL",
                        "Critical",
                        "Consent & compliance",
                        "Consent mismatch",
                    ),
                ],
                "decision_ledger": [],
            }
        )
        self.assertEqual([], errors)
        self.assertEqual(["OP-CRITICAL", "OP-LOW"], [row["ID"] for row in rows])
        self.assertIn("Execution order: 2", rows[0]["Action / priority / QA"])

    def test_human_rows_batch_homogeneous_hygiene_without_losing_atomic_ids(self) -> None:
        operations = [
            {
                "operation_id": f"OP-{index:04d}",
                "area": "GTM hygiene",
                "problem_type": "Exact duplicate",
                "problem": f"Duplicate object pair {index} is source-proven.",
                "why_it_matters": "Parallel copies increase maintenance and regression risk.",
                "exact_proposed_action": (
                    f"Remap every consumer to canonical variable {index} and delete its copy."
                ),
                "qa_steps": f"Verify every consumer for duplicate pair {index}.",
                "priority": "Medium",
                "execution_readiness": "approval_required",
                "execution_order": index,
                "affected_objects": f"variable:{index}; variable:{index + 10}",
                "affected_measurement_family_ids": [f"FAM-{index:05d}"],
                "retained_behavior": f"Preserve the source measurement family {index}.",
                "blocker": "",
            }
            for index in range(1, 4)
        ]
        rows, errors = build_rows(
            {"operations": operations, "decision_ledger": []}
        )
        self.assertEqual([], errors)
        self.assertEqual(1, len(rows))
        self.assertTrue(rows[0]["ID"].startswith("BATCH-"))
        for operation in operations:
            self.assertEqual(1, rows[0]["ID"].count(operation["operation_id"]))
            self.assertIn(operation["operation_id"], rows[0]["Action / priority / QA"])
        self.assertIn("Approve, reject, or amend each atomic operation ID", rows[0]["Action / priority / QA"])

    def test_hidden_proof_is_split_losslessly_and_visible_text_is_not_truncated(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        proof = "source-bound-proof-" + ("x" * (MAX_CELL_TEXT + 41))
        workbook = Workbook()
        hidden = workbook.active
        hidden.title = "Hidden"
        add_table(hidden, [{"Evidence": proof}], ["Evidence"], split_long_cells=True)
        rendered = "".join(str(hidden.cell(row, 1).value or "") for row in range(2, hidden.max_row + 1))
        self.assertEqual(proof, rendered)
        self.assertNotIn("[truncated]", rendered.lower())

        visible = workbook.create_sheet("Visible")
        with self.assertRaisesRegex(ValueError, "summarize the user-facing row"):
            add_table(visible, [{"Problem": proof}], ["Problem"])

    def test_source_model_contains_nested_facts_and_dependency_edges(self) -> None:
        model = build_model(self.export_path)
        self.assertEqual("pass", model["coverage_gate"])
        self.assertGreater(model["counts"]["field_edges"], 0)
        self.assertGreater(model["counts"]["trigger_edges"], 0)

    def test_source_integrity_blocks_ambiguous_unmodelled_and_empty_sources(self) -> None:
        duplicate = sample_export()
        duplicate["containerVersion"]["tag"].append(
            {**copy.deepcopy(duplicate["containerVersion"]["tag"][0]), "name": "Duplicate ID"}
        )
        unmodelled = sample_export()
        unmodelled["containerVersion"]["mysteryEntity"] = [{"mysteryId": "1"}]
        unmodelled_empty = sample_export()
        unmodelled_empty["containerVersion"]["futureEntity"] = []
        cases = {
            "duplicate": (duplicate, "duplicate_entity_id"),
            "unmodelled": (unmodelled, "unmodelled_entity_layer"),
            "unmodelled-empty": (unmodelled_empty, "unmodelled_entity_layer"),
            "empty": ({}, "invalid_container_version_shape"),
            "container-resource": (
                {
                    "accountId": "100",
                    "containerId": "200",
                    "name": "This is a Container resource, not a ContainerVersion",
                    "publicId": "GTM-WRONG",
                    "usageContext": ["WEB"],
                },
                "invalid_container_version_shape",
            ),
            "empty-nested-container": (
                {"container": {}},
                "invalid_container_version_shape",
            ),
        }
        for name, (payload, finding_type) in cases.items():
            with self.subTest(name=name):
                export = self.root / f"{name}.json"
                export.write_text(json.dumps(payload), encoding="utf-8")
                model = build_model(export)
                self.assertEqual("blocked_source_integrity", model["coverage_gate"])
                self.assertIn(
                    finding_type,
                    {row["finding_type"] for row in model["source_integrity_findings"]},
                )
                package = self.root / f"{name}-package"
                manifest = build_package(export, package, pretty=True)
                self.assertEqual("blocked", manifest["status"])
                self.assertFalse((package / "configuration_review.json").exists())
                self.assertFalse((package / "architecture_review.json").exists())
                operational = audit_export(export)
                self.assertEqual("blocked_source_integrity", operational["run_status"])
                for scaffold in (
                    scaffold_operational,
                    scaffold_configuration,
                    scaffold_architecture,
                ):
                    with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                        scaffold(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    source_object_catalog(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    build_context_model(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    extract_export(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    scan_relationships(export)
                future_report, future_errors = check_future_state(
                    export, {"operations": []}
                )
                self.assertEqual(
                    "blocked_source_integrity", future_report["status"]
                )
                self.assertTrue(future_errors)

    def test_duplicate_reference_names_are_never_resolved_by_fallback(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "900",
                    "name": "Click URL",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "analytics_storage"}
                    ],
                },
                {
                    "variableId": "901",
                    "name": "Click URL",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "ad_storage"}
                    ],
                },
            ]
        )
        data["containerVersion"]["tag"][0]["parameter"].append(
            {
                "type": "TEMPLATE",
                "key": "consent_payload",
                "value": "{{Click URL}}",
            }
        )
        export = self.root / "duplicate-reference-names.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        self.assertEqual("pass", build_model(export)["coverage_gate"])
        operational = audit_export(export)
        duplicate = next(
            row
            for row in operational["findings"]
            if row["module_name"] == "duplicate_variable_names"
            and row["finding_type"] != "zero_findings"
        )
        self.assertEqual({"900", "901"}, set(duplicate["object_ids"]))

        configuration = scaffold_configuration(export)
        tag_row = next(
            row for row in configuration["rows"] if row["object_key"] == "tag:1"
        )
        trace = next(
            row
            for row in tag_row["reference_trace_requirements"]
            if row["reference"] == "Click URL"
        )
        self.assertEqual(["ambiguous"], trace["terminal_states"])
        self.assertEqual(
            {"variable:900", "variable:901"}, set(trace["required_object_keys"])
        )
        self.assertIn(
            "builtInVariable:Click URL",
            trace["terminal_requirements"][0]["configured_source"],
        )

        architecture = scaffold_architecture(export)
        family = next(
            row for row in architecture["families"] if "tag:1" in row["member_object_keys"]
        )
        self.assertTrue(
            {"variable:900", "variable:901"}.issubset(family["chain_object_keys"])
        )

        consent = tag_consent_route(
            data["containerVersion"]["tag"][0],
            variables=data["containerVersion"]["variable"],
        )
        self.assertTrue(
            {"analytics_storage", "ad_storage"}.issubset(
                consent["detected_consent_payload_purposes"]
            )
        )
        self.assertEqual([], consent["forwarded_consent_purposes"])
        builtin = next(
            row
            for row in operational["lifecycle_matrix"]
            if row["object_key"] == "builtInVariable:Click URL"
        )
        self.assertEqual("used", builtin["usage_state"])

        completed = complete_configuration(export)
        errors, _ = validate_configuration(
            export,
            self.write_review("ambiguous-configuration.json", completed),
        )
        self.assertTrue(
            any("ambiguous variable identity" in error for error in errors), errors
        )

    def test_missing_javascript_parser_is_an_explicit_review_limit(self) -> None:
        with patch.dict(sys.modules, {"esprima": None}):
            review = complete_configuration(self.export_path)
            parser_requirements = [
                finding
                for row in review["rows"]
                for finding in row["required_technical_findings"]
                if finding["category"] == "parser"
            ]
            self.assertTrue(parser_requirements)
            for row in review["rows"]:
                for finding in row["technical_finding_reviews"]:
                    if finding["finding_key"] == "parser:coverage":
                        finding["verdict"] = "False positive"
            review_path = self.write_review("parser-limit.json", review)
            errors, _ = validate_configuration(self.export_path, review_path)
            self.assertTrue(any("parser coverage limit" in error for error in errors))

            for row in review["rows"]:
                for finding in row["technical_finding_reviews"]:
                    if finding["finding_key"] != "parser:coverage":
                        continue
                    finding["verdict"] = "Documented exception"
                    finding["rationale"] = (
                        "The mandatory line-by-line code blocks cover every exported line; "
                        "the parser availability boundary is recorded without claiming AST coverage."
                    )
            review_path = self.write_review("parser-limit-resolved.json", review)
            resolved_errors, _ = validate_configuration(self.export_path, review_path)
            self.assertEqual([], resolved_errors)

    def test_direct_container_version_paths_compile_and_apply_without_envelope(self) -> None:
        direct = copy.deepcopy(sample_export()["containerVersion"])
        export = self.root / "direct-container-version.json"
        export.write_text(json.dumps(direct), encoding="utf-8")
        model = build_model(export)
        self.assertEqual("pass", model["coverage_gate"])
        self.assertTrue(
            all(
                not row["json_path"].startswith("$.containerVersion")
                for row in model["field_edges"]
            )
        )
        configuration = scaffold_configuration(export)
        tag_review = next(row for row in configuration["rows"] if row["object_key"] == "tag:1")
        self.assertEqual("$.tag[0]", tag_review["source_json_path"])
        future, errors = apply_operations(
            direct,
            {
                "operations": [
                    {
                        "operation_key": "direct-path-change",
                        "creations": [],
                        "additions": [],
                        "changes": [
                            {
                                "object_key": "tag:1",
                                "json_path": "$.tag[0].parameter[0].value",
                                "before": "purchase",
                                "after": "purchase_complete",
                            }
                        ],
                        "remaps": [],
                        "deletions": [],
                        "renames": [],
                    }
                ]
            },
        )
        self.assertEqual([], errors)
        self.assertEqual("purchase_complete", future["tag"][0]["parameter"][0]["value"])

    def test_zones_and_google_tag_configs_are_reviewed_by_all_three_runs(self) -> None:
        data = sample_export()
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner FR child container",
                "childContainer": [{"publicId": "GTM-CHILD", "nickname": "Partner"}],
                "boundary": {
                    "condition": [condition("EQUALS", "partner", "partner")],
                    "customEvaluationTriggerId": ["10"],
                },
                "typeRestriction": {"enable": True, "whitelistedTypeId": ["html"]},
            }
        ]
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "71",
                "parameter": [
                    {"type": "TEMPLATE", "key": "tag_id", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "zone-and-gtag.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        source = build_model(export)
        self.assertEqual(1, source["counts"]["zones"])
        self.assertEqual(1, source["counts"]["gtagConfigs"])
        self.assertTrue(
            any(edge["relation"] == "zone_boundary_trigger" for edge in source["trigger_edges"])
        )

        operational = audit_export(export)
        self.assertEqual(1, operational["counts"]["zones"])
        self.assertEqual(1, operational["counts"]["gtagConfigs"])
        self.assertEqual(
            {"zone:70", "gtagConfig:71"},
            {
                row["object_key"]
                for row in operational["lifecycle_matrix"]
                if row["layer"] in {"zone", "gtagConfig"}
            },
        )

        configuration = scaffold_configuration(export)
        configuration_keys = {row["object_key"] for row in configuration["rows"]}
        self.assertTrue({"zone:70", "gtagConfig:71"}.issubset(configuration_keys))
        self.assertIn(
            "collect.example.test",
            configuration["audit_context"]["server_routing_hosts"],
        )
        self.assertIn("FR", configuration["audit_context"]["markets"])
        for key in ("zone:70", "gtagConfig:71"):
            row = next(item for item in configuration["rows"] if item["object_key"] == key)
            self.assertTrue(row["required_contract_topics"])
            self.assertTrue(row["required_branch_reviews"])
        gtag_row = next(
            item for item in configuration["rows"] if item["object_key"] == "gtagConfig:71"
        )
        self.assertTrue(
            any(
                fact["json_path"].endswith(".type")
                and fact["value_type"] == "missing"
                for fact in gtag_row["source_absence_facts"]
            )
        )
        self.assertNotIn(
            "Unclassified external integration (collect.example.test)",
            {item["vendor"] for item in gtag_row["vendor_contexts"]},
        )

        architecture = scaffold_architecture(export)
        root_keys = {
            key
            for family in architecture["families"]
            for key in family["member_object_keys"]
        }
        self.assertTrue({"zone:70", "gtagConfig:71"}.issubset(root_keys))
        destination_comparison = next(
            row
            for row in architecture["comparisons"]
            if "shared_configured_destination" in row["comparison_types"]
            and {"tag:1", "gtagConfig:71"}.issubset(row["candidate_object_keys"])
        )
        self.assertIn(
            "consumer_destination_and_event_overlap",
            destination_comparison["discovery_methods"],
        )

        two_zone_cv = copy.deepcopy(data["containerVersion"])
        two_zone_cv["zone"].append(
            {
                "zoneId": "72",
                "name": "Partner child container alternate boundary",
                "childContainer": [{"publicId": "GTM-CHILD", "nickname": "Partner"}],
                "boundary": {"customEvaluationTriggerId": ["11"]},
            }
        )
        zone_comparison = next(
            row
            for row in relationship_candidates(two_zone_cv)
            if "shared_zone_child_container" in row["comparison_types"]
        )
        self.assertEqual(
            {"zone:70", "zone:72"}, set(zone_comparison["candidate_object_keys"])
        )

        complete_configuration_review = complete_configuration(export)
        configuration_errors, _ = validate_configuration(
            export,
            self.write_review("zone-gtag-configuration.json", complete_configuration_review),
        )
        self.assertEqual([], configuration_errors)
        complete_architecture_review = complete_architecture(export)
        architecture_errors, _ = validate_architecture(
            export,
            self.write_review("zone-gtag-architecture.json", complete_architecture_review),
        )
        self.assertEqual([], architecture_errors)

    def test_run1_detects_same_tag_payload_across_different_execution_routes(self) -> None:
        data = sample_export()
        original = data["containerVersion"]["tag"][0]
        alternate = copy.deepcopy(original)
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase alternate control route",
                "firingTriggerId": ["11"],
                "blockingTriggerId": [],
                "setupTag": [],
                "teardownTag": [{"tagName": "Utility - Consent Defaults"}],
                "tagFiringOption": "ONCE_PER_EVENT",
                "priority": {"type": "INTEGER", "value": "7"},
                "liveOnly": True,
                "paused": True,
                "scheduleStartMs": "100",
                "scheduleEndMs": "200",
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "same-payload-different-route.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        candidates = [
            row
            for row in audit_export(export)["findings"]
            if row["finding_type"] == "normalized_duplicate_tag_signature"
        ]
        self.assertTrue(
            any({"1", "99"}.issubset(set(row["object_ids"])) for row in candidates)
        )
        candidate = next(
            row for row in candidates if {"1", "99"}.issubset(set(row["object_ids"]))
        )
        self.assertEqual("review_candidate", candidate["finding_class"])
        self.assertIn("keep", candidate["required_resolution"])

    def test_run1_queues_same_contract_different_consent_controls(self) -> None:
        data = sample_export()
        alternate = copy.deepcopy(data["containerVersion"]["tag"][0])
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase explicit consent route",
                "blockingTriggerId": [],
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "same-contract-consent-collision.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        collision = next(
            row
            for row in audit_export(export)["findings"]
            if row["finding_type"]
            == "same_contract_different_consent_control_candidate"
        )
        self.assertEqual({"1", "99"}, set(collision["object_ids"]))
        self.assertEqual(
            ["purchase"],
            collision["shared_event_destination_contract"]["events"],
        )
        self.assertEqual(2, len(collision["consent_control_comparison"]))

    def test_consumed_object_deletion_requires_complete_remap_coverage(self) -> None:
        operation = {
            "creations": [],
            "additions": [],
            "changes": [],
            "remaps": [],
            "renames": [],
            "deletions": [
                {"object_key": "trigger:10", "reason": "Remove redundant trigger logic."}
            ],
        }
        keys = {"trigger:10", "trigger:11", "tag:1"}
        consumers = {"trigger:10": {"tag:1"}, "trigger:11": set(), "tag:1": set()}
        errors = validate_structured_actions(operation, keys, "deletion test", consumers)
        self.assertTrue(any("requires remap coverage" in error for error in errors))

        operation["remaps"] = [
            {
                "from_object_key": "trigger:10",
                "to_object_key": "trigger:11",
                "consumer_object_keys": ["tag:1"],
            }
        ]
        self.assertEqual(
            [], validate_structured_actions(operation, keys, "deletion test", consumers)
        )

    def test_remaps_reject_cross_layer_deleted_targets_and_created_cycles(self) -> None:
        keys = {"trigger:12", "trigger:13", "tag:1"}
        consumers = {
            "trigger:12": {"trigger:13"},
            "trigger:13": {"trigger:12"},
            "tag:1": set(),
        }

        def operation(target: str, deletions: list[dict] | None = None) -> dict:
            return {
                "creations": [],
                "additions": [],
                "changes": [],
                "remaps": [
                    {
                        "from_object_key": "trigger:13",
                        "to_object_key": target,
                        "consumer_object_keys": ["trigger:12"],
                    }
                ],
                "renames": [],
                "deletions": deletions or [],
            }

        cross_layer = validate_structured_actions(
            operation("tag:1"), keys, "cross-layer test", consumers
        )
        self.assertTrue(any("crosses GTM layers" in error for error in cross_layer))

        deleted_target = validate_structured_actions(
            operation(
                "trigger:12",
                [
                    {
                        "object_key": "trigger:12",
                        "reason": "Delete the obsolete target after consolidation.",
                    }
                ],
            ),
            keys,
            "deleted-target test",
            consumers,
        )
        self.assertTrue(any("remap target 'trigger:12' is also deleted" in error for error in deleted_target))

        cycle = validate_structured_actions(
            operation("trigger:12"), keys, "cycle test", consumers
        )
        self.assertTrue(any("creates a dependency cycle" in error for error in cycle))

    def test_operational_operation_set_rejects_duplicate_final_names(self) -> None:
        operation = duplicate_variable_operation()
        operation["renames"] = [
            {
                "object_key": "tag:1",
                "before": "GA4 - Purchase - All",
                "after": "Meta - Purchase - All",
            }
        ]
        errors = validate_operation_set(
            [operation],
            expected_consumers=object_consumer_map(self.export_path),
            object_names=object_name_map(self.export_path),
            label="operational operation set",
        )
        self.assertTrue(any("duplicate final name" in error for error in errors))

        review = complete_operational(self.export_path)
        finding = next(
            row
            for row in review["findings"]
            if row["module_name"] == "duplicate_variable_paths"
            and set(row.get("object_ids", [])) == {"20", "21"}
        )
        finding.update(operation)
        finding["disposition"] = "cleanup_operation"
        finding["rationale"] = (
            "Variables 20 and 21 share ecommerce.items while tag 1 is proposed for an "
            "explicit final-name change in this operation."
        )
        review_errors, _ = validate_operational(
            self.export_path,
            self.write_review("duplicate-final-name-operational.json", review),
        )
        self.assertTrue(any("duplicate final name" in error for error in review_errors))

    def test_operation_set_allows_one_dead_blocker_edge_and_cross_finding_delete_chain(self) -> None:
        consumers = {
            "trigger:11": {"trigger:12"},
            "trigger:12": set(),
            "trigger:13": {"tag:1", "tag:2"},
            "tag:1": set(),
            "tag:2": set(),
        }
        empty = {
            "creations": [],
            "additions": [],
            "changes": [],
            "remaps": [],
            "deletions": [],
            "renames": [],
        }
        remove_dead_edge = {
            **empty,
            "changes": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].blockingTriggerId",
                    "before": ["13"],
                    "after": [],
                }
            ],
        }
        delete_child = {
            **empty,
            "deletions": [
                {"object_key": "trigger:11", "reason": "Delete the unused child trigger."}
            ],
        }
        delete_group = {
            **empty,
            "deletions": [
                {"object_key": "trigger:12", "reason": "Delete the unused trigger group."}
            ],
        }

        errors = validate_operation_set(
            [remove_dead_edge, delete_child, delete_group],
            expected_consumers=consumers,
            object_names={},
            label="cross-finding operation set",
        )
        self.assertEqual([], errors)

        missing_consumer_delete = validate_operation_set(
            [remove_dead_edge, delete_child],
            expected_consumers=consumers,
            object_names={},
            label="missing consumer operation set",
        )
        self.assertTrue(
            any("deleting consumed object 'trigger:11'" in error for error in missing_consumer_delete)
        )

    def test_run2_rejects_conclusions_that_contradict_decisive_source_states(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"].append(
            {
                "tagId": "5",
                "name": "Marketing bundle",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            "<script>fbq('track','Purchase',{value:{{DLV - Value}}});"
                            "ttq.track('CompletePayment',{value:{{DLV - Value}}});"
                            "var s=document.createElement('script');"
                            "s.src='https://metrics.example.test/p.js';"
                            "document.head.appendChild(s);</script>"
                        ),
                    }
                ],
                "firingTriggerId": ["10"],
            }
        )
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner boundary",
                "childContainer": [
                    {"publicId": "GTM-CHILD"},
                    {"publicId": "GTM-CHILD"},
                    {},
                ],
                "boundary": {
                    "customEvaluationTriggerId": ["10", "999"],
                    "condition": "malformed",
                },
                "typeRestriction": {"enable": True, "whitelistedTypeId": []},
            }
        ]
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        data["containerVersion"]["customTemplate"] = [
            {
                "templateId": "90",
                "name": "Opaque template",
                "templateData": '{"__wm":"TEMPLATE","permissions":[]}',
            }
        ]
        export = self.root / "decisive-run2-states.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        honest = complete_configuration(export)
        honest_errors, _ = validate_configuration(
            export,
            self.write_review("decisive-run2-honest.json", honest),
        )
        self.assertEqual([], honest_errors)
        gtag_row = next(
            item for item in honest["rows"] if item["object_key"] == "gtagConfig:80"
        )
        self.assertTrue(
            {"tag:1"}.issubset(
                {peer["object_key"] for peer in gtag_row["destination_peer_contexts"]}
            )
        )
        tag_row = next(
            item for item in honest["rows"] if item["object_key"] == "tag:1"
        )
        gtag_peer = next(
            peer
            for peer in tag_row["destination_peer_contexts"]
            if peer["object_key"] == "gtagConfig:80"
        )
        self.assertEqual(["collect.example.test"], gtag_peer["server_routing_hosts"])
        self.assertFalse(gtag_peer["type_present"])
        self.assertTrue(
            any(
                fact["json_path"].endswith(".type")
                and fact["value_type"] == "missing"
                for fact in tag_row["destination_peer_facts"]
            )
        )
        self.assertTrue(
            any(
                item["obligation_key"].startswith(
                    "peer_destination_contract_unproven:"
                )
                for item in tag_row["required_configuration_obligations"]
            )
        )
        value_row = next(
            item for item in honest["rows"] if item["object_key"] == "variable:24"
        )
        self.assertTrue(
            any(
                context["consumer_key"] == "tag:5"
                and "purchase" in context["events"]
                for context in value_row["consumer_dependency_contexts"]
            )
        )
        opaque_row = next(
            item
            for item in honest["rows"]
            if item["object_key"] == "customTemplate:90"
        )
        self.assertEqual(
            "unknown_opaque",
            opaque_row["technical_code_facts"]["returned_value_type"],
        )
        self.assertFalse(
            any(
                "no material external behavior limit" in value.lower()
                for value in opaque_row["technical_code_facts"][
                    "container_evidence_limits"
                ]
            )
        )

        poisoned = copy.deepcopy(honest)
        for key in (
            "tag:1",
            "tag:5",
            "zone:70",
            "gtagConfig:80",
            "customTemplate:90",
        ):
            row = next(item for item in poisoned["rows"] if item["object_key"] == key)
            row.update(
                {
                    "correctness_verdict": "Correct",
                    "correctness_basis": object_specific_text(
                        row,
                        "is incorrectly claimed to have no material configuration defect",
                        "correctness_basis",
                    ),
                    "defects": [],
                    "disposition": "keep",
                    "owner_question": "",
                }
            )
            for branch in row["configuration_branch_reviews"]:
                branch["correctness"] = "Correct"
            for check in row["logic_cross_checks"]:
                check["verdict"] = "Aligned"
            for check in row["contract_checks"]:
                check["verdict"] = "Compliant"
                if not check.get("source"):
                    check["source"] = "https://docs.example.test/official"
            for finding in row["technical_finding_reviews"]:
                if "no reviewable executable behavior" in finding[
                    "source_statement"
                ].lower():
                    finding["verdict"] = "False positive"

        errors, _ = validate_configuration(
            export,
            self.write_review("decisive-run2-poisoned.json", poisoned),
        )
        self.assertTrue(any("invalid_zone_boundary_field:condition" in error for error in errors))
        self.assertTrue(any("missing_required_field:type" in error for error in errors))
        self.assertTrue(any("opaque_custom_template_behavior" in error for error in errors))
        self.assertTrue(any("peer_destination_contract_unproven" in error for error in errors))
        self.assertTrue(any("reserved or non-production" in error for error in errors))

    def test_parser_fallback_requires_exact_segment_attestation(self) -> None:
        review = complete_configuration(self.export_path)
        row = next(item for item in review["rows"] if item["object_key"] == "variable:22")
        parser_review = next(
            item
            for item in row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        parser_review["fallback_line_hashes"] = parser_review["fallback_line_hashes"][:-1]
        parser_review["manual_review_method"] = (
            "A generic line-by-line parser review was performed for this code."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("weak-parser-fallback.json", review),
        )
        self.assertTrue(any("every exported code segment hash" in error for error in errors))
        self.assertTrue(any("source-specific code behavior" in error for error in errors))

        segment_poisoned = complete_configuration(self.export_path)
        segment_row = next(
            item
            for item in segment_poisoned["rows"]
            if item["object_key"] == "variable:22"
        )
        segment_parser_review = next(
            item
            for item in segment_row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        segment_parser_review["fallback_segment_reviews"][1]["behavior"] = (
            "This remaining code segment received a generic manual inspection only."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("generic-parser-segment.json", segment_poisoned),
        )
        self.assertTrue(any("parser fallback segment" in error for error in errors))

    def test_run2_rejects_code_semantic_reversal_and_risk_dismissal(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][1]["parameter"][0]["value"] = (
            "<script>fbq('track', 'Purchase');</script>"
        )
        export = self.root / "semantic-code-reversal.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        def complete_without_parser() -> dict:
            with patch.dict(sys.modules, {"esprima": None}):
                return complete_configuration(export)

        review = complete_without_parser()
        code_row = next(
            row for row in review["rows"] if row["object_key"] == "tag:2"
        )
        code_row["code_behavior_blocks"][0]["health_assessment"] = (
            "The exact fbq identifier is inspected, but Meta does not track or send the "
            "Purchase event from this segment."
        )
        parser_review = next(
            item
            for item in code_row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        parser_review["fallback_segment_reviews"][0]["behavior"] = (
            "Mandatory line-by-line review cites fbq and Purchase, but states that Meta "
            "does not track or send the Purchase event from this exact source segment."
        )
        risk_key = next(
            item["finding_key"]
            for item in code_row["required_technical_findings"]
            if item["category"] in {"health", "security"}
        )
        risk_review = next(
            item
            for item in code_row["technical_finding_reviews"]
            if item["finding_key"] == risk_key
        )
        risk_review["verdict"] = "False positive"

        errors, _ = validate_configuration(
            export,
            self.write_review("semantic-code-reversal-review.json", review),
        )
        self.assertTrue(
            any("reverses source-proven segment behavior meta_event_send" in error for error in errors)
        )
        self.assertTrue(
            any("cannot be dismissed as a false positive" in error for error in errors)
        )
        dead_path_review = complete_without_parser()
        dead_path_row = next(
            row
            for row in dead_path_review["rows"]
            if row["object_key"] == "tag:2"
        )
        dead_path_row["code_behavior_blocks"][0]["health_assessment"] = (
            "The source contains fbq track Purchase, but this is a dead code path with "
            "zero delivery."
        )
        dead_path_parser = next(
            item
            for item in dead_path_row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        dead_path_parser["fallback_segment_reviews"][0]["behavior"] = (
            "Mandatory line-by-line inspection identifies fbq track Purchase, but calls "
            "the segment unreachable with no delivery."
        )
        dead_path_errors, _ = validate_configuration(
            export,
            self.write_review("dead-path-code-reversal.json", dead_path_review),
        )
        self.assertTrue(
            any("denies the executable effect" in error for error in dead_path_errors)
        )
        alias_expectations = {
            "Cleanup opportunity": "requires one concrete proposed_action",
            "Documented exception": "requires an evidence-bound exception_basis",
            "Owner decision needed": "advisory technical review signal",
        }
        for verdict, expected_error in alias_expectations.items():
            with self.subTest(verdict=verdict):
                alias_review = complete_without_parser()
                alias_row = next(
                    row
                    for row in alias_review["rows"]
                    if row["object_key"] == "tag:2"
                )
                alias_risk = next(
                    item
                    for item in alias_row["technical_finding_reviews"]
                    if item["finding_key"] == risk_key
                )
                alias_risk.update(
                    {
                        "verdict": verdict,
                        "rationale": (
                            "The source signal receives this generic disposition for the "
                            "current container."
                        ),
                        "proposed_action": "",
                        "exception_basis": "",
                        "owner_question": "",
                    }
                )
                alias_row["owner_question"] = ""
                alias_errors, _ = validate_configuration(
                    export,
                    self.write_review(
                        f"semantic-risk-alias-{verdict.lower().replace(' ', '-')}.json",
                        alias_review,
                    ),
                )
                self.assertTrue(
                    any(expected_error in error for error in alias_errors), alias_errors
                )

        confirmed_review = complete_without_parser()
        confirmed_row = next(
            row
            for row in confirmed_review["rows"]
            if row["object_key"] == "tag:2"
        )
        confirmed_risk = next(
            item
            for item in confirmed_row["technical_finding_reviews"]
            if item["finding_key"] == risk_key
        )
        confirmed_risk["verdict"] = "Confirmed issue"
        confirmed_row.update(
            {
                "correctness_verdict": "Issue",
                "correctness_basis": object_specific_text(
                    confirmed_row,
                    "contains the confirmed custom-code risk and requires owner resolution",
                    "correctness_basis",
                ),
                "disposition": "owner_decision_needed",
                "owner_question": (
                    "Which owner will resolve the confirmed inline-script risk before this "
                    "tag remains active?"
                ),
                "defects": [
                    {
                        "defect_id": "TECH-001",
                        "statement": confirmed_risk["source_statement"],
                        "configured_effect": (
                            "The exact exported custom-code signal creates the confirmed "
                            "maintenance or security exposure."
                        ),
                        "expected_behavior": (
                            "The confirmed code risk must be corrected or accepted through a "
                            "source-bound owner decision."
                        ),
                        "evidence_anchors": [],
                        "code_line_hashes": [
                            confirmed_row["required_code_line_hashes"][0]
                        ],
                        "technical_finding_keys": [],
                    }
                ],
            }
        )
        confirmed_errors, _ = validate_configuration(
            export,
            self.write_review("unlinked-confirmed-technical-risk.json", confirmed_review),
        )
        self.assertTrue(
            any(
                f"actionable technical finding {risk_key} must link" in error
                for error in confirmed_errors
            )
        )

    def test_run2_binds_malformed_controls_and_contradictions_to_every_verdict_layer(
        self,
    ) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["teardownTag"] = [{}, {}]
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "915",
                    "name": "Impossible lead or signup conjunction",
                    "type": "CUSTOM_EVENT",
                    "filter": [
                        condition("EQUALS", "{{_event}}", "lead"),
                        condition("EQUALS", "{{_event}}", "signup"),
                    ],
                },
                {
                    "triggerId": "916",
                    "name": "Malformed trigger group members",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": ["10", {"type": "TEMPLATE", "value": "10"}, {}],
                        }
                    ],
                },
            ]
        )
        export = self.root / "malformed-controls-and-contradiction.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_configuration(export)
        honest_errors, _ = validate_configuration(
            export, self.write_review("malformed-controls-honest.json", review)
        )
        self.assertEqual([], honest_errors)

        tag = next(
            row
            for row in reversed(review["rows"])
            if row["object_key"] == "tag:1"
        )
        malformed_sequences = [
            item
            for item in tag["required_configuration_obligations"]
            if item["obligation_key"].startswith("dependency:teardownTag:")
        ]
        self.assertEqual(2, len(malformed_sequences))
        group = next(
            row for row in review["rows"] if row["object_key"] == "trigger:916"
        )
        self.assertEqual(
            2,
            sum(
                item["obligation_key"].startswith("invalid_trigger_group_member:")
                for item in group["required_configuration_obligations"]
            ),
        )
        malformed_traces = [
            trace
            for trace in group["execution_dependency_traces"]
            if trace["resolution_state"] == "malformed"
        ]
        self.assertEqual(2, len(malformed_traces))
        self.assertEqual(
            {
                f"{group['source_json_path']}.parameter[0].list[0]",
                f"{group['source_json_path']}.parameter[0].list[2]",
            },
            {
                trace["source_reference_paths"][0]
                for trace in malformed_traces
            },
        )
        valid_trace = next(
            trace
            for trace in group["execution_dependency_traces"]
            if trace["reference"] == "10" and trace["resolution_state"] == "unique"
        )
        self.assertEqual(
            [f"{group['source_json_path']}.parameter[0].list[1].value"],
            valid_trace["source_reference_paths"],
        )
        contradictory = next(
            row for row in review["rows"] if row["object_key"] == "trigger:915"
        )
        self.assertTrue(
            any(
                item["obligation_key"].startswith("contradictory_equals:")
                for item in contradictory["required_configuration_obligations"]
            )
        )

        contradictory.update(
            {
                "correctness_verdict": "Correct",
                "disposition": "keep",
                "defects": [],
                "owner_question": "",
            }
        )
        for branch in contradictory["configuration_branch_reviews"]:
            branch["correctness"] = "Correct"
        for check in contradictory["logic_cross_checks"]:
            check["verdict"] = "Aligned"
            check["conclusion"] = (
                "The exported purpose and execution values were generally reviewed and "
                "declared aligned for this object."
            )
        errors, _ = validate_configuration(
            export, self.write_review("contradiction-fail-open.json", review)
        )
        self.assertTrue(any("contradictory_equals:" in error for error in errors))
        self.assertTrue(any("does not name deterministic obligation" in error for error in errors))

    def test_run2_rejects_duplicate_rows_branches_checks_findings_and_traces(self) -> None:
        review = complete_configuration(self.export_path)
        review["rows"].append(copy.deepcopy(review["rows"][0]))
        tag = next(
            row
            for row in reversed(review["rows"])
            if row["object_key"] == "tag:1"
        )
        tag["configuration_branch_reviews"].append(
            copy.deepcopy(tag["configuration_branch_reviews"][0])
        )
        tag["logic_cross_checks"].append(copy.deepcopy(tag["logic_cross_checks"][0]))
        tag["contract_checks"].append(copy.deepcopy(tag["contract_checks"][0]))
        if tag["reference_traces"]:
            tag["reference_traces"].append(copy.deepcopy(tag["reference_traces"][0]))
        code_row = next(
            row
            for row in review["rows"]
            if row["technical_finding_reviews"]
        )
        code_row["technical_finding_reviews"].append(
            copy.deepcopy(code_row["technical_finding_reviews"][0])
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("duplicate-review-identities.json", review),
        )
        self.assertTrue(any("unique nonblank object keys" in error for error in errors))
        self.assertTrue(any("branch review paths must be unique" in error for error in errors))
        self.assertTrue(any("D3 logic check keys must be unique" in error for error in errors))
        self.assertTrue(any("unique nonblank topic keys" in error for error in errors))
        self.assertTrue(any("technical finding keys must be unique" in error for error in errors))
        if tag["reference_traces"]:
            self.assertTrue(any("trace references must be unique" in error for error in errors))

    def test_dynamic_script_load_is_network_behavior_with_full_endpoint_evidence(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][1]["parameter"][0]["value"] = (
            "<script>var s=document.createElement('script');"
            "s.src='https://metrics.example.test/p.js';"
            "document.head.appendChild(s);</script>"
        )
        export = self.root / "dynamic-script-network.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        technical = next(
            row
            for row in extract_export(export)["rows"]
            if row["layer"] == "tag" and row["object_id"] == "2"
        )
        configuration = next(
            row
            for row in scaffold_configuration(export)["rows"]
            if row["object_key"] == "tag:2"
        )
        self.assertTrue(technical["network_calls"])
        self.assertIn("https://metrics.example.test/p.js", technical["external_scripts_loaded"])
        self.assertGreater(len(configuration["code_line_facts"]), 1)
        self.assertTrue(
            any(
                "metrics.example.test" in token
                for token in configuration["specificity_tokens"]
            )
        )

    def test_run3_requires_owner_or_action_for_unsafe_retention_candidates(self) -> None:
        data = sample_export()
        alternate = copy.deepcopy(data["containerVersion"]["tag"][0])
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase alternate route",
                "firingTriggerId": ["11"],
                "blockingTriggerId": [],
                "setupTag": [],
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "unsafe-retention.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        comparison = next(
            row
            for row in review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        comparison.update(
            {
                "relationship_verdict": "Intentional variant",
                "disposition": "keep",
                "owner_question": "",
            }
        )
        family = next(
            row
            for row in review["families"]
            if {"tag:1", "tag:99"}.issubset(row["member_object_keys"])
        )
        family.update(
            {
                "relationship_verdict": "Complementary",
                "disposition": "keep",
                "owner_question": "",
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-retention-review.json", review),
        )
        self.assertTrue(any("not a source-proven intentional variant" in error for error in errors))
        self.assertTrue(any("family retention is unsupported" in error for error in errors))

    def test_run3_binds_cleanup_actions_and_discoveries_to_unsafe_candidates(self) -> None:
        data = sample_export()
        alternate = copy.deepcopy(data["containerVersion"]["tag"][0])
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase alternate route",
                "firingTriggerId": ["11"],
                "blockingTriggerId": [],
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "architecture-operation-binding.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        operation_review = complete_architecture(export)
        comparison = next(
            row
            for row in operation_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        unrelated_operation = duplicate_variable_operation()
        unrelated_operation.update(
            {
                "operation_key": "unrelated-name-only-change",
                "canonical_object_key": "",
                "changes": [],
                "remaps": [],
                "deletions": [],
                "renames": [
                    {
                        "object_key": "tag:2",
                        "before": "Meta - Purchase - All",
                        "after": "Meta - Purchase - Primary",
                    }
                ],
            }
        )
        comparison.update(
            {
                "relationship_verdict": "Functional overlap",
                "disposition": "cleanup_operation",
                "owner_question": "",
                "operations": [unrelated_operation],
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("unrelated-architecture-operation.json", operation_review),
        )
        self.assertTrue(
            any("do not change any candidate member's behavior" in error for error in errors)
        )

        no_op_review = complete_architecture(export)
        no_op_comparison = next(
            row
            for row in no_op_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        no_op_operation = copy.deepcopy(unrelated_operation)
        no_op_operation.update(
            {
                "operation_key": "candidate-no-op",
                "changes": [
                    {
                        "object_key": "tag:1",
                        "json_path": "$.containerVersion.tag[0].paused",
                        "before": False,
                        "after": False,
                    }
                ],
                "renames": [],
            }
        )
        no_op_comparison.update(
            {
                "relationship_verdict": "Functional overlap",
                "disposition": "cleanup_operation",
                "owner_question": "",
                "operations": [no_op_operation],
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("no-op-architecture-operation.json", no_op_review),
        )
        self.assertTrue(any("before and after values are identical" in error for error in errors))
        self.assertTrue(
            any("do not change any candidate member's behavior" in error for error in errors)
        )

        mismatched_path_review = complete_architecture(export)
        mismatched_comparison = next(
            row
            for row in mismatched_path_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        mismatched_operation = copy.deepcopy(unrelated_operation)
        mismatched_operation.update(
            {
                "operation_key": "candidate-key-path-mismatch",
                "changes": [
                    {
                        "object_key": "tag:1",
                        "json_path": "$.containerVersion.tag[1].paused",
                        "before": False,
                        "after": True,
                    }
                ],
                "renames": [],
            }
        )
        mismatched_comparison.update(
            {
                "relationship_verdict": "Functional overlap",
                "disposition": "cleanup_operation",
                "owner_question": "",
                "operations": [mismatched_operation],
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review(
                "mismatched-path-architecture-operation.json", mismatched_path_review
            ),
        )
        self.assertTrue(any("paired with an unrelated source path" in error for error in errors))
        self.assertTrue(
            any("do not change any candidate member's behavior" in error for error in errors)
        )

        discovery_review = complete_architecture(export)
        deterministic = next(
            row
            for row in discovery_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        discovered = copy.deepcopy(deterministic)
        discovered.update(
            {
                "comparison_id": "DISC-UNSAFE-001",
                "comparison_origin": "analyst_discovered",
                "discovery_methods": ["normalized_condition_and_route_variants"],
                "relationship_verdict": "Intentional variant",
                "disposition": "keep",
                "owner_question": "",
                "operations": [],
            }
        )
        discovery_review["comparisons"].append(discovered)
        attestation = discovery_review["open_discovery_attestation"]
        attestation["discovered_comparison_ids"] = ["DISC-UNSAFE-001"]
        attestation["zero_discovery_rationale"] = ""
        method_review = next(
            item
            for item in attestation["method_reviews"]
            if item["method"] == "normalized_condition_and_route_variants"
        )
        method_review["additional_discovery_ids"] = ["DISC-UNSAFE-001"]
        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-discovered-retention.json", discovery_review),
        )
        self.assertTrue(
            any("not a source-proven intentional variant" in error for error in errors)
        )

        method_review["additional_discovery_ids"] = []
        errors, _ = validate_architecture(
            export,
            self.write_review("unattributed-discovery.json", discovery_review),
        )
        self.assertTrue(any("do not match the discoveries attributed" in error for error in errors))

    def test_run3_unsafe_discovery_inherits_policy_for_candidate_subsets(self) -> None:
        data = sample_export()
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "type": "googtag",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "unsafe-discovery-subset.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        deterministic = next(
            row
            for row in review["comparisons"]
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        candidate_keys = [
            deterministic["candidate_object_keys"][0],
            deterministic["candidate_object_keys"][-1],
        ]
        discovered = copy.deepcopy(deterministic)
        discovered.update(
            {
                "comparison_id": "DISC-UNSAFE-SUBSET-001",
                "comparison_origin": "analyst_discovered",
                "discovery_methods": ["semantic_name_and_business_term_variants"],
                "candidate_object_keys": candidate_keys,
                "member_assessments": [
                    item
                    for item in discovered["member_assessments"]
                    if item["object_key"] in candidate_keys
                ],
                "owner_question": unsafe_owner_question(
                    {"browser_server_consent_deduplication_review"}, candidate_keys
                ),
            }
        )
        discovered.pop("comparison_types", None)
        review["comparisons"].append(discovered)
        attestation = review["open_discovery_attestation"]
        attestation["discovered_comparison_ids"] = ["DISC-UNSAFE-SUBSET-001"]
        attestation["zero_discovery_rationale"] = ""
        next(
            item
            for item in attestation["method_reviews"]
            if item["method"] == "semantic_name_and_business_term_variants"
        )["additional_discovery_ids"] = ["DISC-UNSAFE-SUBSET-001"]

        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-discovery-subset-review.json", review),
        )
        self.assertTrue(
            any(
                "browser_server_consent_deduplication_review must be attributed" in error
                and "consent_sequence_and_server_route_conflicts" in error
                for error in errors
            )
        )

    def test_run3_rejects_generic_unsafe_questions_and_positive_runtime_claims(self) -> None:
        data = sample_export()
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "type": "googtag",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "architecture-negative-runtime-facts.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        comparison = next(
            row
            for row in review["comparisons"]
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        comparison.update(
            {
                "analyst_rationale": architecture_text(
                    comparison,
                    "analyst_rationale",
                    "Runtime event ID deduplication is unproven. Browser and server event IDs "
                    "are guaranteed identical and synchronized",
                ),
                "architecture_effect": architecture_text(
                    comparison,
                    "architecture_effect",
                    "Browser and server consent is verified equivalent and consistent",
                ),
                "owner_question": (
                    f"{' '.join(comparison['candidate_object_keys'][:2])} browser server "
                    "consent route deduplication canonical. Should this relationship be reviewed?"
                ),
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("positive-runtime-overclaim.json", review),
        )
        self.assertTrue(any("overclaims a complete, guaranteed" in error for error in errors))
        self.assertTrue(any("owner question must name" in error for error in errors))
        self.assertTrue(any("inside the interrogative clause" in error for error in errors))

        separated_review = complete_architecture(export)
        separated_comparison = next(
            row
            for row in separated_review["comparisons"]
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        separated_comparison["analyst_rationale"] = architecture_text(
            separated_comparison,
            "analyst_rationale",
            "Consent and event-ID deduplication remain unproven from this export",
        )
        separated_comparison["architecture_effect"] = architecture_text(
            separated_comparison,
            "architecture_effect",
            "Consent " + ("distant context " * 30) + "is aligned and verified end to end",
        )
        separated_errors, _ = validate_architecture(
            export,
            self.write_review("separated-runtime-overclaim.json", separated_review),
        )
        self.assertTrue(
            any("overclaims a complete, guaranteed" in error for error in separated_errors)
        )

    def test_run3_generates_browser_server_consent_and_deduplication_family(self) -> None:
        data = sample_export()
        media_code = data["containerVersion"]["tag"][1]["parameter"][0]["value"]
        data["containerVersion"]["tag"][1]["parameter"][0]["value"] = media_code.replace(
            "</script>",
            "var s=document.createElement('script');"
            "s.src='https://metrics.example.test/p.js';"
            "document.head.appendChild(s);</script>",
        )
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "type": "googtag",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "browser-server-family.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        candidates = relationship_candidates(data["containerVersion"])
        family = next(
            row
            for row in candidates
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        self.assertTrue(
            {"gtagConfig:80", "tag:1", "tag:2"}.issubset(family["candidate_object_keys"])
        )
        self.assertEqual(
            {
                "consent_sequence_and_server_route_conflicts",
                "consumer_destination_and_event_overlap",
                "terminal_source_formula_and_output_overlap",
            },
            set(family["discovery_methods"]),
        )
        self.assertIn(
            "metrics.example.test",
            next(
                row
                for row in candidates
                if {"tag:1", "tag:2"}.issubset(row["candidate_object_keys"])
            )["candidate_specificity_tokens"].get("tag:2", []),
        )
        for row in candidates:
            for terms in row.get("candidate_distinguishing_terms", {}).values():
                self.assertFalse(
                    any(
                        term in {"{}", "[]", "missing", "malformed", "script", "true"}
                        or any(
                            marker in term
                            for marker in ("<script", "document.", "window.", ".src", "ttq.")
                        )
                        for term in terms
                    )
                )

    def test_run3_cannot_hide_visible_relationship_inside_evidence_limit(self) -> None:
        review = complete_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if "shared_execution_trigger" in row["comparison_types"]
        )
        comparison.update(
            {
                "relationship_verdict": "Container evidence limit",
                "disposition": "container_evidence_limit",
                "owner_question": (
                    "Which unseen runtime evidence proves that these visible routes remain separate?"
                ),
                "analyst_rationale": architecture_text(
                    comparison,
                    "analyst_rationale",
                    "The visible route is recorded while downstream runtime behavior is not visible",
                ),
                "architecture_effect": architecture_text(
                    comparison,
                    "architecture_effect",
                    "Visible configuration remains, but external evidence is unseen",
                ),
            }
        )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("deterministic-evidence-limit.json", review),
        )
        self.assertTrue(
            any("deterministic source-visible relationship" in error for error in errors)
        )

    def test_run3_rejects_keep_for_zone_overlap_and_trigger_group_cycle(self) -> None:
        data = sample_export()
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "30",
                    "name": "Cycle A",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "31"}],
                        }
                    ],
                },
                {
                    "triggerId": "31",
                    "name": "Cycle B",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "30"}],
                        }
                    ],
                },
            ]
        )
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Child A",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["10"]},
                "typeRestriction": {"enable": True, "whitelistedTypeId": ["html"]},
            },
            {
                "zoneId": "71",
                "name": "Child B",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["11"]},
                "typeRestriction": {"enable": True, "whitelistedTypeId": ["gaawe"]},
            },
        ]
        export = self.root / "unsafe-zone-cycle-retention.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        for comparison in review["comparisons"]:
            if set(comparison["comparison_types"]) & {
                "shared_zone_child_container",
                "cyclic_trigger_group_dependency",
            }:
                comparison.update(
                    {
                        "relationship_verdict": "Intentional variant",
                        "disposition": "keep",
                        "owner_question": "",
                    }
                )
        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-zone-cycle-retention-review.json", review),
        )
        self.assertTrue(any("multiple Zones governing one child" in error for error in errors))
        self.assertTrue(any("cyclic trigger-group dependency" in error for error in errors))

    def test_mandatory_operational_module_oracle_is_complete_and_fail_closed(self) -> None:
        expected = {
            "source_integrity",
            "inventory",
            "destination_inventory",
            "recognized_system_references",
            "missing_references",
            "duplicate_tag_names",
            "duplicate_trigger_names",
            "duplicate_variable_names",
            "duplicate_folder_names",
            "duplicate_zone_names",
            "duplicate_custom_template_names",
            "duplicate_client_names",
            "duplicate_transformation_names",
            "duplicate_tag_configurations",
            "normalized_duplicate_tag_signatures",
            "duplicate_trigger_logic",
            "duplicate_variable_logic",
            "duplicate_zone_configurations",
            "duplicate_google_tag_configurations",
            "duplicate_client_configurations",
            "duplicate_transformation_configurations",
            "duplicate_custom_template_configurations",
            "duplicate_variable_paths",
            "outdated_ua_styled_setup_objects",
            "unused_variables",
            "unused_triggers",
            "tags_without_firing_triggers",
            "unused_custom_templates",
            "unused_folders",
            "paused_tags",
            "used_only_by_paused_tags",
            "tag_sequence_structure",
            "tag_execution_controls",
            "single_member_trigger_groups",
            "trigger_group_structure",
            "zone_structure",
            "duplicate_custom_code",
            "variables_mirroring_builtins",
            "custom_variable_formula_logic",
            "consent_variable_logic",
            "media_tag_consent_route",
            "trigger_condition_lint",
            "ineffective_blocking_triggers",
            "unfiled_objects",
            "singleton_folders",
            "overloaded_folders",
            "name_hygiene",
            "naming_architecture_standardization",
        }
        self.assertEqual(expected, set(MANDATORY_OPERATIONAL_MODULES))
        scan = audit_export(self.export_path)
        self.assertEqual(expected, {row["module_name"] for row in scan["modules"]})
        mutated = copy.deepcopy(scan)
        mutated["modules"] = [
            row for row in mutated["modules"] if row["module_name"] != "tag_sequence_structure"
        ]
        self.assertTrue(
            any("tag_sequence_structure" in error for error in mandatory_module_errors(mutated))
        )

    def test_change_log_diff_is_field_level(self) -> None:
        before = container_version(sample_export())
        after = copy.deepcopy(before)
        after["tag"][0]["name"] = "GA4 - Purchase - Global"
        rows = diff_operations(before, after, "Direct")
        self.assertEqual(1, len(rows))
        self.assertEqual("$.name", rows[0]["field_path"])
        change_log_path = self.root / "change-log.xlsx"
        build_change_log(
            {"kind": "gtm_field_level_change_log", "execution_mode": "planned", "changes": rows},
            change_log_path,
        )
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        workbook = load_workbook(change_log_path)
        self.assertLessEqual(len(workbook.sheetnames), 3)
        for sheet in workbook:
            self.assertLessEqual(sheet.max_column, 6)

    def test_change_log_supports_zone_and_google_tag_configuration_layers(self) -> None:
        before = container_version(sample_export())
        before["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["10"]},
            }
        ]
        before["gtagConfig"] = [
            {
                "gtagConfigId": "71",
                "type": "GOOGLE_TAG",
                "parameter": [
                    {"type": "TEMPLATE", "key": "tag_id", "value": "G-TEST123"}
                ],
            }
        ]
        after = copy.deepcopy(before)
        after["zone"][0]["boundary"]["customEvaluationTriggerId"] = ["11"]
        after["gtagConfig"][0]["parameter"][0]["value"] = "G-NEW123"
        rows = diff_operations(before, after, "Direct")
        self.assertEqual(
            {"Zone", "Google tag configuration"}, {row["layer"] for row in rows}
        )

        duplicate = copy.deepcopy(before)
        duplicate["zone"].append(copy.deepcopy(duplicate["zone"][0]))
        with self.assertRaisesRegex(ValueError, "change-log source fails integrity"):
            diff_operations(duplicate, after, "Direct")

        invalid_artifact = self.root / "invalid-import-artifact.json"
        invalid_artifact.write_text(
            json.dumps({"tag": [{"name": "Missing tag identity", "type": "html"}]}),
            encoding="utf-8",
        )
        artifact_report = validate_artifact(invalid_artifact, None, "audit")
        self.assertEqual("fail", artifact_report["status"])
        self.assertIn(
            "artifact_source_integrity",
            {row["check"] for row in artifact_report["errors"]},
        )

    def test_executed_change_log_links_only_exact_approved_fields(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        approved, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        future, apply_errors = apply_operations(sample_export(), approved)
        self.assertEqual([], apply_errors)
        after = container_version(future)
        after["tag"][0]["name"] = "Unexpected unapproved rename"
        rows = diff_operations(
            container_version(sample_export()),
            after,
            "Direct GTM/MCP/API",
            approved,
            "executed",
        )
        deletion = next(row for row in rows if row["object_id"] == "21")
        unexpected = next(
            row
            for row in rows
            if row["object_id"] == "1" and row["field_path"] == "$.name"
        )
        self.assertTrue(deletion["operation_id"])
        self.assertEqual("Applied", deletion["status"])
        self.assertEqual("", unexpected["operation_id"])
        self.assertEqual("Blocked: missing approved operation link", unexpected["status"])

    def test_workbook_escapes_formula_text_and_privacy_scans_hidden_tabs(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.assertEqual("'=1+1", spreadsheet_safe_text("=1+1"))

        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        human, human_errors = build_rows(payload)
        self.assertEqual([], human_errors)
        human[0]["Problem / evidence"] = '=HYPERLINK("https://example.test","open")'
        workbook_path = self.root / "formula-safe.xlsx"
        build_workbook(
            {"source_file": self.export_path.name},
            build_model(self.export_path),
            operational,
            configuration,
            architecture,
            payload,
            {"rows": human},
            workbook_path,
        )
        rendered = load_workbook(workbook_path, data_only=False)
        formula_cell = rendered["02 Cleanup Plan"]["F2"]
        self.assertEqual("s", formula_cell.data_type)
        self.assertTrue(str(formula_cell.value).startswith("'="))
        rendered.close()

        privacy_path = self.root / "hidden-privacy.xlsx"
        raw = Workbook()
        raw.active.title = "Visible"
        hidden = raw.create_sheet("Hidden proof")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "@".join(("analyst", "example.test"))
        raw.save(privacy_path)
        self.assertTrue(
            any("Hidden proof!A1" in finding for finding in scan_xlsx(privacy_path))
        )

    def test_privacy_helpers_redact_identity_and_sensitive_urls(self) -> None:
        address = "@".join(("jane.doe", "example.test"))
        text = f"Contact {address} or +33 6 12 34 56 78 token=secret-value"
        redacted = redact_text(text)
        self.assertNotIn(address, redacted)
        self.assertNotIn("secret-value", redacted)
        url_address = "@".join(("jane", "example.test"))
        url = sanitize_url(f"https://example.test/path?email={url_address}&utm_source=test")
        self.assertNotIn(url_address, url)
        self.assertIn("utm_source=%3Cvalue%3E", url)
        self.assertTrue(privacy_findings(text))

    def test_unicode_integrity_finds_invisible_references_and_confusables_only(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "60",
                    "name": "DLV - Val\u200bue",
                    "type": "v",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "name", "value": "clean.value"}
                    ],
                },
                {
                    "variableId": "61",
                    "name": "CJS - Pаge",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "one"}
                    ],
                },
                {
                    "variableId": "62",
                    "name": "CJS - Page",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "two"}
                    ],
                },
                {
                    "variableId": "63",
                    "name": "DLV - Événement",
                    "type": "v",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "name", "value": "event.name"}
                    ],
                },
            ]
        )
        data["containerVersion"]["tag"][0]["parameter"].append(
            {
                "type": "TEMPLATE",
                "key": "invisibleReference",
                "value": "{{DLV - Val\u200bue}}",
            }
        )
        export = self.root / "unicode-integrity.json"
        export.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        findings = [
            row
            for row in audit_export(export)["findings"]
            if row["module_name"] == "name_hygiene"
            and row["finding_type"] != "zero_findings"
        ]
        self.assertTrue(
            any(
                row["finding_type"] == "unicode_name_integrity_candidate"
                and "60" in row["object_ids"]
                for row in findings
            )
        )
        self.assertTrue(
            any(
                row["finding_type"] == "unicode_reference_integrity"
                and "1" in row["object_ids"]
                for row in findings
            )
        )
        confusable = next(
            row
            for row in findings
            if row["finding_type"] == "unicode_confusable_name_candidate"
            and {"61", "62"} <= set(row["object_ids"])
        )
        self.assertEqual("review_candidate", confusable["finding_class"])
        self.assertFalse(
            any(
                row["finding_type"].startswith("unicode_")
                and "63" in row["object_ids"]
                for row in findings
            )
        )

    def test_ineffective_blocker_has_one_exact_repair_and_orphan_cleanup(self) -> None:
        finding = next(
            row
            for row in audit_export(self.export_path)["findings"]
            if row["finding_type"] == "ineffective_blocking_trigger"
        )
        repair = finding["deterministic_repair"]
        self.assertEqual("unique_ineffective_blocker_removal", repair["status"])
        self.assertEqual(
            [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].blockingTriggerId",
                    "before": ["13"],
                    "after": [],
                }
            ],
            repair["changes"],
        )
        self.assertEqual(
            ["trigger:13"],
            [item["object_key"] for item in repair["deletions"]],
        )
        self.assertEqual(
            ["tag:1", "trigger:13"],
            finding["repair_affected_object_keys"],
        )

    def test_unicode_broken_reference_keeps_normalized_target_reachable(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["parameter"][2]["map"][0]["value"] = (
            "{{DLV -\u00a0Value}}"
        )
        export = self.root / "unicode-reachable-target.json"
        export.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

        report = audit_export(export)
        findings = report["findings"]
        self.assertTrue(
            any(
                row["finding_type"] == "undefined_variable_reference"
                and "DLV -\u00a0Value" in row["object_ids"]
                for row in findings
            )
        )
        self.assertFalse(
            any(
                row["finding_type"] == "unused_object"
                and row["object_type"] == "variable"
                and "24" in row["object_ids"]
                for row in findings
            )
        )
        lifecycle = {
            row["object_key"]: row for row in report["lifecycle_matrix"]
        }
        self.assertEqual("used", lifecycle["variable:24"]["usage_state"])
        self.assertIn(
            {"from_object_key": "tag:1", "to_object_key": "variable:24"},
            report["execution_reachability"]["dependency_edges"],
        )

    def test_missing_setup_repairs_to_peer_supported_target_and_rejects_clear(
        self,
    ) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["setupTag"] = [
            {"tagName": "Google tag - Verisur"}
        ]
        data["containerVersion"]["tag"][0]["teardownTag"] = [
            {"tagName": "Utility - Consent Defaults"}
        ]
        data["containerVersion"]["tag"].extend(
            [
                {
                    "tagId": "5",
                    "name": "Google tag - Verisure",
                    "type": "googtag",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "tagId",
                            "value": "G-TEST123",
                        }
                    ],
                },
                {
                    "tagId": "6",
                    "name": "GA4 - Peer event",
                    "type": "gaawe",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "eventName",
                            "value": "peer_event",
                        },
                        {
                            "type": "TEMPLATE",
                            "key": "measurementId",
                            "value": "G-TEST123",
                        },
                    ],
                    "firingTriggerId": ["10"],
                    "setupTag": [{"tagName": "Google tag - Verisure"}],
                },
            ]
        )
        export = self.root / "peer-supported-setup.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_operational(export)
        finding = next(
            row
            for row in review["findings"]
            if row["finding_type"] == "missing_setupTag_reference"
            and "Google tag - Verisur" in row["deterministic_evidence"]
        )
        repair = finding["deterministic_repair"]
        self.assertEqual("unique_peer_supported_target", repair["status"])
        self.assertEqual("tag:5", repair["target_object_key"])
        self.assertEqual("Google tag - Verisure", repair["target_name"])
        self.assertEqual(
            "Google tag - Verisure",
            repair["changes"][0]["after"],
        )
        finding.update(
            {
                "disposition": "cleanup_operation",
                "operation_key": "repair-missing-google-tag-setup",
                "title": "Repair the missing Google tag setup reference",
                "area": "GTM hygiene",
                "problem_type": "Broken reference",
                "problem": finding["deterministic_evidence"],
                "why_it_matters": (
                    "The purchase event can execute without its intended Google tag "
                    "initialization sequence."
                ),
                "expected_clean_state": (
                    "The purchase event points to the peer-supported Google tag and "
                    "the setup edge remains explicit."
                ),
                "exact_proposed_action": (
                    "Change the missing setup name to Google tag - Verisure on tag:1."
                ),
                "canonical_object_key": "",
                "canonical_selection_rationale": "",
                "creations": [],
                "additions": [],
                "changes": copy.deepcopy(repair["changes"]),
                "remaps": [],
                "deletions": [],
                "renames": [],
                "preconditions": (
                    "Confirm tag:5 still has the peer-supported Google tag identity "
                    "before applying the field change."
                ),
                "qa_steps": (
                    "Re-export and confirm tag:1 resolves its setup edge to tag:5 "
                    "without a missing sequence reference."
                ),
                "rollback": (
                    "Restore the original setup reference from the locked source export "
                    "if the intended initializer differs."
                ),
                "priority": "High",
                "confidence": "High",
                "execution_readiness": "approval_required",
                "owner_question": "",
                "recommended_action": "",
                "challenge_review": {
                    "source_recheck": (
                        "The locked export still contains the missing setup name on tag:1 "
                        "and the existing tag:5 target."
                    ),
                    "status_and_scope_check": (
                        "Both tag:1 and the peer-supported tag:5 route remain active in "
                        "the exported execution graph."
                    ),
                    "alternative_explanation": (
                        "No second close setup target or intentional edge removal is "
                        "supported by the container evidence."
                    ),
                    "challenge_verdict": "confirmed",
                },
            }
        )
        valid_path = self.write_review("peer-supported-setup-review.json", review)
        valid_errors, _ = validate_operational(export, valid_path)
        self.assertEqual([], valid_errors)

        clearing = copy.deepcopy(review)
        clearing_finding = next(
            row
            for row in clearing["findings"]
            if row["finding_id"] == finding["finding_id"]
        )
        exact_path = repair["changes"][0]["json_path"]
        clearing_finding["changes"].append(
            {
                "object_key": "tag:1",
                "json_path": exact_path.rsplit("[", 1)[0],
                "before": [{"tagName": "Google tag - Verisur"}],
                "after": [],
            }
        )
        clearing_errors, _ = validate_operational(
            export,
            self.write_review("cleared-peer-supported-setup.json", clearing),
        )
        self.assertTrue(
            any(
                "clearing the setup/teardown edge is not the generated safe repair"
                in error
                for error in clearing_errors
            )
        )

    def test_inactive_paused_target_deletion_supersedes_stale_unique_repair(self) -> None:
        repair = {
            "status": "unique_peer_supported_target",
            "repair_kind": "setupTag",
            "changes": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].setupTag[0].tagName",
                    "before": "Stale setup tag",
                    "after": "Canonical setup tag",
                }
            ],
            "renames": [],
            "deletions": [],
        }
        expected = {"repair_affected_object_keys": ["tag:1"]}
        deleted = {
            "deterministic_repair": repair,
            "changes": [],
            "renames": [],
            "deletions": [
                {
                    "object_key": "tag:1",
                    "reason": "The tag is paused and has no active export-visible consumers.",
                }
            ],
        }
        self.assertEqual(
            [],
            validate_deterministic_repair(
                deleted,
                expected,
                {"tag:1": {"paused": True, "active_consumer_count": 0}},
                "finding paused repair",
            ),
        )
        self.assertTrue(
            validate_deterministic_repair(
                deleted,
                expected,
                {"tag:1": {"paused": False, "active_consumer_count": 0}},
                "finding active repair",
            )
        )

    def test_ua_label_only_with_ga4_consumers_gets_exact_metadata_rename(self) -> None:
        data = sample_export()
        data["containerVersion"]["trigger"].append(
            {
                "triggerId": "99",
                "name": "(UA) - Checkout complete",
                "type": "CUSTOM_EVENT",
                "customEventFilter": [
                    condition("EQUALS", "{{_event}}", "checkout_complete")
                ],
            }
        )
        data["containerVersion"]["tag"][0]["firingTriggerId"].append("99")
        export = self.root / "ua-label-ga4-consumer.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        finding = next(
            row
            for row in audit_export(export)["findings"]
            if row["module_name"] == "outdated_ua_styled_setup_objects"
            and "99" in row["object_ids"]
        )
        self.assertEqual("legacy_name_only_candidate", finding["finding_type"])
        self.assertEqual("deterministic_defect", finding["finding_class"])
        self.assertEqual(
            "unique_ga4_label_cleanup",
            finding["deterministic_repair"]["status"],
        )
        self.assertEqual(
            [
                {
                    "object_key": "trigger:99",
                    "before": "(UA) - Checkout complete",
                    "after": "Checkout complete",
                }
            ],
            finding["deterministic_repair"]["renames"],
        )
        self.assertTrue(
            any(
                item["object_key"] == "tag:1"
                and item["object_type"] == "gaawe"
                and item["destination_vendor"] == "GA4 / Google tag"
                for item in finding["consumer_route_evidence"]
            )
        )

    def test_exact_duplicate_canonical_default_is_source_ranked(self) -> None:
        review = scaffold_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if set(row["candidate_object_keys"]) == {"variable:20", "variable:21"}
            and "exact_configuration" in row["comparison_types"]
        )
        self.assertEqual(
            {"variable:20": ["tag:1", "tag:2"], "variable:21": []},
            comparison["candidate_consumer_keys"],
        )
        self.assertEqual(
            "variable:20",
            comparison["recommended_canonical_object_key"],
        )
        basis = comparison["recommended_canonical_basis"]
        self.assertIn("active over paused", basis)
        self.assertIn("more visible consumers", basis)
        self.assertIn("non-copy name", basis)
        self.assertIn("final deterministic tie-breaker", basis)

    def test_vendor_detection_ignores_monitoring_metadata_template_docs_and_comments(
        self,
    ) -> None:
        data = sample_export()
        data["containerVersion"]["customTemplate"] = [
            {
                "templateId": "901",
                "name": "Neutral utility template",
                "monitoringMetadata": {
                    "notes": "Meta Pixel and Google Ads monitoring only"
                },
                "templateData": (
                    "___INFO___\n"
                    '{"description":"Meta Pixel documentation and monitoring"}\n'
                    "___SANDBOXED_JS_FOR_WEB_TEMPLATE___\n"
                    "// Meta Pixel is mentioned only in this comment\n"
                    "data.gtmOnSuccess();\n"
                    "___TESTS___\n"
                    '[{"name":"Google Ads documentation example"}]\n'
                ),
            }
        ]
        export = self.root / "metadata-vendor-noise.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        row = next(
            item
            for item in scaffold_configuration(export)["rows"]
            if item["object_key"] == "customTemplate:901"
        )
        detected = {context["vendor"] for context in row["vendor_contexts"]}
        self.assertNotIn("Meta", detected)
        self.assertNotIn("Google Ads", detected)

    def test_identical_custom_code_requires_consistent_technical_disposition(
        self,
    ) -> None:
        data = sample_export()
        duplicate = copy.deepcopy(data["containerVersion"]["tag"][1])
        duplicate["tagId"] = "5"
        duplicate["name"] = "Meta - Purchase duplicate implementation"
        data["containerVersion"]["tag"].append(duplicate)
        export = self.root / "identical-custom-code.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_configuration(export)
        original = next(row for row in review["rows"] if row["object_key"] == "tag:2")
        copied = next(row for row in review["rows"] if row["object_key"] == "tag:5")
        self.assertEqual(
            original["technical_code_facts"]["code_hash"],
            copied["technical_code_facts"]["code_hash"],
        )
        self.assertTrue(copied["technical_finding_reviews"])
        copied["technical_finding_reviews"][0]["exception_basis"] += (
            " This copied implementation receives a different acceptance."
        )
        errors, _ = validate_configuration(
            export,
            self.write_review("inconsistent-identical-code.json", review),
        )
        self.assertTrue(
            any(
                "different technical outcome from identical code without a "
                "source-specific consumer reason" in error
                for error in errors
            )
        )

    def test_independent_attestation_rejects_bulk_completion_and_reused_context(
        self,
    ) -> None:
        operational = complete_operational(self.export_path)
        operational["completion_attestation"]["helper_modules"] = [
            "bulk_decision_writer"
        ]
        errors, _ = validate_operational(
            self.export_path,
            self.write_review("bulk-completed-operational.json", operational),
        )
        self.assertTrue(any("bulk-completion helpers" in error for error in errors))

        package_dir = self.root / "independent-package"
        build_package(self.export_path, package_dir, pretty=True)
        reviews = self.completed_reviews()
        shared_context_id = "semantic-review-context-shared"
        for filename, review in zip(
            (
                "operational_review.json",
                "configuration_review.json",
                "architecture_review.json",
            ),
            reviews,
            strict=True,
        ):
            review["completion_attestation"][
                "independent_review_context_id"
            ] = shared_context_id
            (package_dir / filename).write_text(
                json.dumps(review),
                encoding="utf-8",
            )
        gated = run_gate(self.export_path, package_dir)
        self.assertEqual("fail", gated["status"])
        self.assertTrue(
            any(
                "reuse an independent_review_context_id" in error
                for error in gated["errors"]
            )
        )

    def test_runtime_handoff_groups_one_exact_test_contract(self) -> None:
        ledger = [
            {
                "decision_id": f"CFG-DOM-{index}",
                "source_run": "configuration_correctness",
                "source_object_keys": [f"tag:{index}"],
                "disposition": "container_evidence_limit",
                "summary": (
                    "The container cannot prove whether the same checkout DOM "
                    "selector exists on the affected route."
                ),
                "recommended_action": (
                    "Verify the selector once on every affected checkout route."
                ),
            }
            for index in (1, 2)
        ]
        handoff = runtime_qa_handoff(ledger, {"rows": []})
        self.assertEqual(1, handoff["item_count"])
        item = handoff["items"][0]
        self.assertEqual(["CFG-DOM-1", "CFG-DOM-2"], item["source_references"])
        self.assertEqual(["tag:1", "tag:2"], item["affected_object_keys"])
        self.assertEqual("page_dom", item["category"])
        self.assertTrue(item["test_contract_id"].startswith("TEST-"))

    def test_runtime_handoff_targets_consuming_tags_and_omits_deleted_sources(
        self,
    ) -> None:
        ledger = [
            {
                "decision_id": "CFG-VAR-1",
                "source_run": "configuration_correctness",
                "source_object_keys": ["variable:20"],
                "consumer_object_keys": ["tag:1"],
                "disposition": "container_evidence_limit",
                "summary": (
                    "The dataLayer runtime value type is not visible in the export."
                ),
                "recommended_action": (
                    "Capture the resolved dataLayer value on the purchase tag route."
                ),
            },
            {
                "decision_id": "CFG-VAR-2",
                "source_run": "configuration_correctness",
                "source_object_keys": ["variable:21"],
                "consumer_object_keys": ["tag:2"],
                "disposition": "container_evidence_limit",
                "summary": (
                    "The dataLayer runtime value type is not visible in the export."
                ),
                "recommended_action": (
                    "Capture the resolved dataLayer value on the lead tag route."
                ),
            },
        ]
        packets = [
            {
                "deletions": [
                    {
                        "object_key": "variable:21",
                        "reason": "The variable is removed from the target state.",
                    }
                ]
            }
        ]
        handoff = runtime_qa_handoff(ledger, {"rows": []}, packets)
        self.assertEqual(1, handoff["item_count"])
        self.assertEqual(
            ["CFG-VAR-1"],
            handoff["items"][0]["source_references"],
        )
        self.assertEqual(["variable:20"], handoff["items"][0]["affected_object_keys"])
        self.assertEqual(["tag:1"], handoff["items"][0]["test_object_keys"])

    def test_action_incomplete_workbook_is_only_a_blocked_draft(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Manual",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        payload["plan_status"] = "incomplete_actions"
        payload["action_completeness"] = {
            **payload["action_completeness"],
            "status": "incomplete",
            "errors": [
                "OPS-TEST: one source-visible defect still lacks an exact operation"
            ],
        }
        rows, row_errors = build_rows(payload)
        self.assertEqual([], row_errors)
        self.assertEqual(["BLOCKED-001"], [row["ID"] for row in rows])
        workbook_path = self.root / "blocked-cleanup-plan.xlsx"
        build_workbook(
            {
                "source_file": self.export_path.name,
                "source_sha256": payload["source_sha256"],
            },
            build_model(self.export_path),
            operational,
            configuration,
            architecture,
            payload,
            {"rows": rows},
            workbook_path,
        )
        workbook = load_workbook(workbook_path, read_only=True)
        visible_ids = [
            str(row[0].value)
            for row in workbook["02 Cleanup Plan"].iter_rows(min_row=2, max_col=1)
        ]
        workbook.close()
        self.assertEqual(["BLOCKED-001"], visible_ids)
        gate_errors, gate_warnings = validate_workbook(
            workbook_path,
            self.write_review("blocked-operations.json", payload),
        )
        self.assertEqual(
            ["cleanup plan action completeness is not pass"],
            gate_errors,
        )
        self.assertEqual([], gate_warnings)

    def test_lookup_and_regex_tables_lock_bad_rows_without_flagging_valid_neighbors(
        self,
    ) -> None:
        def table_row(match: str, output: str) -> dict:
            return {
                "type": "MAP",
                "map": [
                    {"type": "TEMPLATE", "key": "key", "value": match},
                    {"type": "TEMPLATE", "key": "value", "value": output},
                ],
            }

        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "70",
                    "name": "Lookup - Duplicate",
                    "type": "smm",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "input", "value": "{{Page URL}}"},
                        {
                            "type": "LIST",
                            "key": "map",
                            "list": [table_row("a", "one"), table_row("a", "two")],
                        },
                    ],
                },
                {
                    "variableId": "71",
                    "name": "Regex - Broken",
                    "type": "remm",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "input", "value": "{{Page URL}}"},
                        {
                            "type": "LIST",
                            "key": "map",
                            "list": [
                                table_row(".*", "all"),
                                table_row("(", "broken"),
                            ],
                        },
                        {
                            "type": "BOOLEAN",
                            "key": "setDefaultValue",
                            "value": "true",
                        },
                    ],
                },
                {
                    "variableId": "72",
                    "name": "Regex - Valid",
                    "type": "remm",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "input", "value": "{{Page URL}}"},
                        {
                            "type": "LIST",
                            "key": "map",
                            "list": [
                                table_row("^/one$", "one"),
                                table_row("^/two$", "two"),
                            ],
                        },
                        {
                            "type": "BOOLEAN",
                            "key": "setDefaultValue",
                            "value": "true",
                        },
                        {
                            "type": "TEMPLATE",
                            "key": "defaultValue",
                            "value": "other",
                        },
                    ],
                },
            ]
        )
        export = self.root / "table-contracts.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = scaffold_configuration(export)
        lookup = next(row for row in review["rows"] if row["object_key"] == "variable:70")
        broken = next(row for row in review["rows"] if row["object_key"] == "variable:71")
        valid = next(row for row in review["rows"] if row["object_key"] == "variable:72")
        lookup_keys = {
            item["obligation_key"]
            for item in lookup["required_configuration_obligations"]
        }
        broken_keys = {
            item["obligation_key"]
            for item in broken["required_configuration_obligations"]
        }
        valid_keys = {
            item["obligation_key"]
            for item in valid["required_configuration_obligations"]
        }
        self.assertTrue(
            any(key.startswith("lookup_table_duplicate_match:") for key in lookup_keys)
        )
        self.assertIn("regex_table_invalid_pattern:1", broken_keys)
        self.assertIn("regex_table_permissive_pattern:0", broken_keys)
        self.assertIn("regex_table_shadowed_rows:0", broken_keys)
        self.assertIn("regex_table_enabled_default_missing", broken_keys)
        self.assertFalse(
            any(
                key.startswith(("lookup_table_", "regex_table_"))
                for key in valid_keys
            )
        )

    def test_custom_code_extension_signals_have_safe_neighboring_negatives(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"].extend(
            [
                {
                    "tagId": "80",
                    "name": "Custom - Risk signals",
                    "type": "html",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": (
                                "<script>dataLayer.reset();"
                                "var internal=google_tag_manager['GTM-TEST'];"
                                "gtag('event','purchase');debugger;"
                                "document.cookie='id=1; Path=/';"
                                "window.addEventListener('click',handler);</script>"
                            ),
                        }
                    ],
                    "firingTriggerId": ["10"],
                },
                {
                    "tagId": "81",
                    "name": "Custom - Guarded neighbor",
                    "type": "html",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": (
                                "<script>function gtag(){dataLayer.push(arguments);}"
                                "document.cookie='id=1; Secure; SameSite=Lax';"
                                "window.addEventListener('click',handler,{once:true});</script>"
                            ),
                        }
                    ],
                    "firingTriggerId": ["10"],
                },
            ]
        )
        export = self.root / "custom-code-extension-signals.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        rows = extract_export(export)["rows"]
        risky = next(row for row in rows if row["object_id"] == "80")
        safe = next(row for row in rows if row["object_id"] == "81")
        risky_text = " ".join(
            [
                *risky["technical_code_health_findings"],
                *risky["technical_code_security_findings"],
            ]
        )
        self.assertTrue(risky["dataLayer_resets"])
        self.assertTrue(risky["google_tag_manager_internal_access"])
        self.assertTrue(risky["manual_gtag_calls"])
        self.assertTrue(risky["debugger_statements"])
        self.assertIn("dataLayer.reset", risky_text)
        self.assertIn("google_tag_manager", risky_text)
        self.assertIn("Literal cookie write omits", risky_text)
        self.assertIn("without an exported remove", risky_text)
        self.assertFalse(safe["manual_gtag_calls"])
        safe_text = " ".join(
            [
                *safe["technical_code_health_findings"],
                *safe["technical_code_security_findings"],
            ]
        )
        self.assertNotIn("Literal cookie write omits", safe_text)
        self.assertNotIn("without an exported remove", safe_text)
        scaffold = scaffold_configuration(export)
        required = next(
            row
            for row in scaffold["rows"]
            if row["object_key"] == "tag:80"
        )["required_technical_findings"]
        self.assertEqual(
            "review_signal",
            next(
                item
                for item in required
                if "uses an inline script" in item["statement"]
            )["decision_class"],
        )
        self.assertEqual(
            "deterministic_defect",
            next(
                item
                for item in required
                if "internal google_tag_manager" in item["statement"]
            )["decision_class"],
        )

    def test_behavior_portability_is_source_bound_in_run2_and_compared_in_run3(
        self,
    ) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "90",
                    "name": "Constant - Staging collector",
                    "type": "c",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "value",
                            "value": (
                                "https://staging.metrics.example.com/" + "GTM-" + "AAAA1"
                            ),
                        }
                    ],
                },
                {
                    "variableId": "91",
                    "name": "Constant - Production collector",
                    "type": "c",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "value",
                            "value": (
                                "https://prod.metrics.example.com/" + "GTM-" + "BBBB2"
                            ),
                        }
                    ],
                },
                {
                    "variableId": "92",
                    "name": "Constant - Metadata only",
                    "type": "c",
                    "tagManagerUrl": (
                        "https://tagmanager.google.com/#/container/accounts/1/"
                        "containers/2/workspaces/3?publicId=" + "GTM-" + "META1"
                    ),
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "value",
                            "value": "https://collect.example.com",
                        }
                    ],
                },
            ]
        )
        export = self.root / "portability.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = scaffold_configuration(export)
        staging = next(row for row in review["rows"] if row["object_key"] == "variable:90")
        metadata = next(row for row in review["rows"] if row["object_key"] == "variable:92")
        staging_keys = {
            item["obligation_key"]
            for item in staging["required_configuration_obligations"]
        }
        metadata_keys = {
            item["obligation_key"]
            for item in metadata["required_configuration_obligations"]
        }
        self.assertTrue(
            any("hard_coded_container_identifier" in key for key in staging_keys)
        )
        self.assertTrue(
            any("environment_specific_endpoint_candidate" in key for key in staging_keys)
        )
        self.assertFalse(any(key.startswith("portability:") for key in metadata_keys))
        comparisons = relationship_candidates(data["containerVersion"])
        portability = next(
            row
            for row in comparisons
            if "behavior_portability_variant" in row["comparison_types"]
            and {"variable:90", "variable:91"} <= set(row["candidate_object_keys"])
        )
        self.assertIn(
            "normalized_condition_and_route_variants",
            portability["discovery_methods"],
        )

    def test_registry_contracts_lock_ga4_fields_but_keep_supported_tiktok_event(
        self,
    ) -> None:
        incomplete_data = sample_export()
        incomplete_parameters = incomplete_data["containerVersion"]["tag"][0]["parameter"][2][
            "map"
        ]
        incomplete_parameters[:] = [
            item for item in incomplete_parameters if item.get("key") != "items"
        ]
        incomplete_export = self.root / "registry-contract-incomplete.json"
        incomplete_export.write_text(json.dumps(incomplete_data), encoding="utf-8")
        review = scaffold_configuration(incomplete_export)
        purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        purchase_contract = next(
            topic
            for topic in purchase["required_contract_topics"]
            if topic["topic_key"] == "GA4 / Google tag:registry_contract:purchase"
        )
        self.assertEqual("known_noncompliant", purchase_contract["deterministic_contract_state"])
        self.assertIn("items", purchase_contract["required_configuration_terms"])

        data = sample_export()
        event_parameters = data["containerVersion"]["tag"][0]["parameter"][2]["map"]
        event_parameters[0]["value"] = "30.03"
        event_parameters.extend(
            [
                {
                    "type": "TEMPLATE",
                    "key": "currency",
                    "value": "EUR",
                },
                {
                    "type": "TEMPLATE",
                    "key": "items",
                    "value": "{{DLV - Items}}",
                },
            ]
        )
        safe_export = self.root / "registry-contract-safe.json"
        safe_export.write_text(json.dumps(data), encoding="utf-8")
        safe_review = scaffold_configuration(safe_export)
        safe_purchase = next(
            row for row in safe_review["rows"] if row["object_key"] == "tag:1"
        )
        safe_contract = next(
            topic
            for topic in safe_purchase["required_contract_topics"]
            if topic["topic_key"] == "GA4 / Google tag:registry_contract:purchase"
        )
        self.assertEqual("source_check_required", safe_contract["deterministic_contract_state"])

        registry = load_registry(ROOT / "references/03-rules/vendor-registry.toml")
        tiktok = next(vendor for vendor in registry["vendors"] if vendor["name"] == "TikTok")
        contract = next(
            item for item in tiktok["contracts"] if item["event"] == "CompletePayment"
        )
        self.assertEqual("supported", contract["status"])
        self.assertNotIn(
            "CompletePayment",
            tiktok.get("unsupported_standard_events", []),
        )

    def test_unicode_equivalent_missing_reference_has_one_exact_repair(self) -> None:
        from gtm_configuration_review import validate_row_outcome

        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "98",
                    "name": "DL - timerEventNumber",
                    "type": "v",
                    "parameter": [
                        {
                            "type": "INTEGER",
                            "key": "dataLayerVersion",
                            "value": "2",
                        },
                        {
                            "type": "BOOLEAN",
                            "key": "setDefaultValue",
                            "value": "false",
                        },
                        {
                            "type": "TEMPLATE",
                            "key": "name",
                            "value": "timerEventNumber",
                        },
                    ],
                },
                {
                    "variableId": "99",
                    "name": "Timer helper",
                    "type": "c",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "value",
                            "value": "{{DL -\u00a0timerEventNumber}}",
                        }
                    ],
                },
            ]
        )
        export = self.root / "unicode-reference.json"
        export.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        scaffold = scaffold_configuration(export)
        source = next(
            row for row in scaffold["rows"] if row["object_key"] == "variable:99"
        )
        trace = next(
            item
            for item in source["reference_trace_requirements"]
            if "\u00a0" in item["reference"]
        )
        terminal = trace["terminal_requirements"][0]
        self.assertEqual("unique", terminal["normalization_resolution"])
        self.assertEqual(
            ["DL - timerEventNumber"],
            terminal["normalization_candidate_names"],
        )
        self.assertEqual(1, len(trace["source_reference_paths"]))

        decision = {
            "object_key": "variable:99",
            "object_name": "Timer helper",
            "correctness_verdict": "Issue",
            "disposition": "owner_decision_needed",
            "owner_question": (
                "Which valid variable should replace the broken reference in "
                "variable:99?"
            ),
            "recommended_action": (
                "For variable:99, repair defect REF-001 at "
                f"{trace['source_reference_paths'][0]} by replacing the missing "
                "reference with the approved variable."
            ),
            "defects": [
                {
                    "defect_id": "REF-001",
                    "evidence_anchors": [trace["source_reference_paths"][0]],
                }
            ],
            "reference_trace_requirements": source[
                "reference_trace_requirements"
            ],
            "required_configuration_obligations": [],
            "contract_checks": [],
            "technical_finding_reviews": [],
            "operation": {},
        }
        errors = validate_row_outcome(decision, "unicode repair")
        self.assertTrue(
            any("Unicode/whitespace-equivalent" in error for error in errors)
        )

        decision["disposition"] = "cleanup_operation"
        decision["owner_question"] = ""
        decision["recommended_action"] = ""
        decision["operation"] = {
            "changes": [
                {
                    "object_key": "variable:99",
                    "json_path": trace["source_reference_paths"][0],
                    "before": "{{DL -\u00a0timerEventNumber}}",
                    "after": "{{DL - timerEventNumber}}",
                }
            ]
        }
        self.assertEqual([], validate_row_outcome(decision, "unicode repair"))
        human, human_errors = build_rows(
            {
                "operations": [
                    {
                        "operation_id": "OP-0001",
                        "area": "Event firing logic",
                        "problem_type": "Broken reference",
                        "problem": "The configured variable reference is missing.",
                        "why_it_matters": (
                            "The timer input can remain unresolved."
                        ),
                        "exact_proposed_action": "Repair the reference.",
                        "changes": decision["operation"]["changes"],
                        "deletions": [],
                        "remaps": [],
                        "renames": [],
                        "affected_objects": "variable:99 — Timer helper",
                        "affected_measurement_family_ids": ["FAM-0001"],
                        "priority": "High",
                        "confidence": "High",
                        "execution_readiness": "approval_required",
                        "execution_order": 1,
                        "execution_phases": ["change"],
                        "priority_basis": {
                            "active_reachability": "active",
                            "impact_classes": ["measurement_loss_or_corruption"],
                            "evidence_confidence": "High",
                        },
                        "execution_safety": {
                            "approval": {
                                "scope": "individual_operation",
                                "reasons": ["active configured reachability"],
                            },
                            "decommission": {"required": False},
                        },
                    }
                ],
                "decision_ledger": [],
                "object_catalog": {
                    "variable:99": {
                        "object_name": "Timer helper",
                    }
                },
                "measurement_preservation": {
                    "families": [
                        {
                            "family_id": "FAM-0001",
                            "family_label": "visit duration",
                        }
                    ]
                },
            }
        )
        self.assertEqual([], human_errors)
        self.assertIn(
            "invisible non-breaking space", human[0]["Problem / evidence"]
        )
        self.assertIn(
            "{{DL - timerEventNumber}}", human[0]["Problem / evidence"]
        )
        self.assertIn(
            "Keep formulas, trigger predicates", human[0]["Action / priority / QA"]
        )

    def test_exact_duplicate_projection_and_reconciliation_are_not_duplicated(
        self,
    ) -> None:
        operation = {
            "problem_type": "Exact duplicate",
            "source_runs": ["business_architecture"],
            "canonical_object_key": "variable:20",
            "creations": [],
            "additions": [],
            "changes": [],
            "remaps": [
                {
                    "from_object_key": "variable:21",
                    "to_object_key": "variable:20",
                    "consumer_object_keys": ["tag:1"],
                }
            ],
            "deletions": [{"object_key": "variable:21"}],
        }
        self.assertEqual(
            {"variable:21"},
            runtime_neutral_operational_deletions(operation, {}),
        )

        ledger = [
            {
                "decision_id": "RUN1-DUP",
                "source_run": "operational_sanitation",
                "source_object_keys": ["variable:20", "variable:21"],
                "disposition": "owner_decision_needed",
                "owner_question": "Which duplicate variable should remain?",
                "recommended_action": "Choose one canonical variable.",
            },
            {
                "decision_id": "ARCH-DUP",
                "source_run": "business_architecture",
                "source_object_keys": ["variable:20", "variable:21"],
                "comparison_types": ["exact_configuration"],
                "disposition": "owner_decision_needed",
                "owner_question": "Which duplicate variable should remain canonical?",
                "recommended_action": "Keep variable:20 and remove variable:21.",
            },
        ]
        reconciled, errors = reconcile_ledger_resolutions(ledger, [])
        self.assertEqual([], errors)
        run1 = next(row for row in reconciled if row["decision_id"] == "RUN1-DUP")
        self.assertEqual("documented_exception", run1["disposition"])
        self.assertEqual(
            "delegated_to_architecture_decision",
            run1["reconciliation_status"],
        )
        self.assertEqual(
            1,
            sum(
                row["disposition"] == "owner_decision_needed"
                for row in reconciled
            ),
        )

    def test_unused_document_write_support_requires_exact_field_action(
        self,
    ) -> None:
        from gtm_configuration_review import validate_row_outcome

        data = sample_export()
        data["containerVersion"]["tag"].append(
            {
                "tagId": "88",
                "name": "Custom HTML without document write",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": "<script>window.exampleReady=true;</script>",
                    },
                    {
                        "type": "BOOLEAN",
                        "key": "supportDocumentWrite",
                        "value": "true",
                    },
                ],
                "firingTriggerId": ["10"],
            }
        )
        export = self.root / "unused-document-write-support.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        source = next(
            row
            for row in scaffold_configuration(export)["rows"]
            if row["object_key"] == "tag:88"
        )
        obligation = next(
            item
            for item in source["required_configuration_obligations"]
            if item["obligation_key"] == "unused_document_write_support"
        )
        support_path = obligation["source_known_repair"]["json_path"]
        decision = {
            "object_key": "tag:88",
            "object_name": "Custom HTML without document write",
            "correctness_verdict": "Issue",
            "disposition": "owner_decision_needed",
            "owner_question": (
                "Should the unused document.write support setting remain enabled?"
            ),
            "recommended_action": (
                f"For tag:88, repair defect DOCWRITE-001 at {support_path} by "
                "disabling the unused support setting."
            ),
            "defects": [
                {
                    "defect_id": "DOCWRITE-001",
                    "evidence_anchors": [support_path],
                }
            ],
            "reference_trace_requirements": [],
            "required_configuration_obligations": [
                obligation
            ],
            "contract_checks": [],
            "technical_finding_reviews": [],
            "operation": {},
        }
        errors = validate_row_outcome(decision, "document write")
        self.assertTrue(
            any("source-known target" in error for error in errors)
        )
        decision.update(
            {
                "disposition": "cleanup_operation",
                "owner_question": "",
                "recommended_action": "",
                "operation": {
                    "changes": [
                        {
                            "object_key": "tag:88",
                            "json_path": support_path,
                            "before": "true",
                            "after": "false",
                        }
                    ]
                },
            }
        )
        self.assertEqual(
            [], validate_row_outcome(decision, "document write")
        )

    def test_transitive_unicode_reference_owner_question_is_resolved_once(
        self,
    ) -> None:
        broken = "DL -\u00a0etapefunnel"
        ledger = [
            {
                "decision_id": "CFG-CONSUMER",
                "source_run": "configuration_correctness",
                "source_object_keys": ["trigger:10"],
                "disposition": "owner_decision_needed",
                "owner_question": "Which funnel variable should this trigger use?",
                "recommended_action": "Repair the missing funnel reference.",
                "missing_reference_terminals": [
                    {
                        "reference": broken,
                        "source_object_key": "variable:135",
                        "normalization_candidate_names": [
                            "DL - etapefunnel"
                        ],
                        "normalization_resolution": "unique",
                    }
                ],
            }
        ]
        packets = [
            {
                "operation_id": "OP-0001",
                "operation_key": "FIX-UNICODE",
                "changes": [
                    {
                        "object_key": "variable:135",
                        "before": f"{{{{{broken}}}}}",
                        "after": "{{DL - etapefunnel}}",
                    }
                ],
                "deletions": [],
                "resolution_status": "proposed",
            }
        ]
        reconciled, errors = reconcile_ledger_resolutions(ledger, packets)
        self.assertEqual([], errors)
        self.assertEqual("cleanup_operation", reconciled[0]["disposition"])
        self.assertEqual(
            "resolved_by_upstream_reference_repair",
            reconciled[0]["reconciliation_status"],
        )
        self.assertEqual(["OP-0001"], reconciled[0]["compiled_operation_ids"])

    def test_priority_basis_and_runtime_handoff_add_utility_without_new_gates(
        self,
    ) -> None:
        active_basis = operation_priority_basis(
            {
                "affected_object_keys": ["tag:1"],
                "source_object_keys": ["tag:1"],
                "area": "Consent & compliance",
                "problem_type": "Consent mismatch",
                "problem": "An active request can run before consent.",
                "why_it_matters": "This creates privacy and measurement risk.",
                "exact_proposed_action": "Correct the consent route.",
                "priority": "Medium",
                "confidence": "High",
                "execution_readiness": "approval_required",
                "changes": [{"object_key": "tag:1"}],
            },
            {
                "tag:1": {
                    "layer": "tag",
                    "reachability": "active",
                }
            },
        )
        self.assertEqual("High", active_basis["calibrated_floor"])
        self.assertEqual("below_evidence_floor_review_recommended", active_basis["alignment"])
        folder_basis = operation_priority_basis(
            {
                "affected_object_keys": ["folder:1"],
                "source_object_keys": ["folder:1"],
                "area": "GTM hygiene",
                "problem_type": "Folder organization",
                "problem": "Folder naming is inconsistent.",
                "why_it_matters": "Maintenance is slower.",
                "exact_proposed_action": "Rename the folder.",
                "priority": "Low",
                "confidence": "High",
                "execution_readiness": "approval_required",
                "renames": [{"object_key": "folder:1"}],
            },
            {
                "folder:1": {
                    "layer": "folder",
                    "reachability": "inactive_or_unreferenced",
                }
            },
        )
        self.assertEqual("metadata_only", folder_basis["active_reachability"])
        self.assertEqual("Low", folder_basis["calibrated_floor"])

        ledger = [
            {
                "decision_id": "D-1",
                "source_run": "configuration_correctness",
                "source_object_keys": ["tag:1"],
                "disposition": "container_evidence_limit",
                "summary": "The container cannot prove whether the page DOM selector exists.",
                "owner_question": "",
                "recommended_action": "Verify the selector on each live route.",
            },
            {
                "decision_id": "D-2",
                "source_run": "business_architecture",
                "source_object_keys": ["tag:2"],
                "disposition": "owner_decision_needed",
                "summary": "The business owner must decide whether the campaign is retained.",
                "owner_question": "Should the campaign be retained?",
                "recommended_action": "Confirm ownership.",
            },
        ]
        configuration = {
            "rows": [
                {
                    "review_id": "CFG-1",
                    "object_key": "tag:3",
                    "external_evidence_status": "runtime_handoff_required",
                    "external_evidence_summary": (
                        "The container proves configured endpoints but not external "
                        "script delivery or vendor acceptance."
                    ),
                    "external_evidence_next_action": (
                        "Capture the browser request and vendor acceptance response "
                        "on the affected route."
                    ),
                    "detected_vendor": "Example vendor",
                    "effective_consent_route_facts": {
                        "server_routing_hosts": []
                    },
                    "required_contract_topics": [],
                    "technical_code_facts": {
                        "container_evidence_limits": [
                            "The container proves configured endpoints but not external "
                            "script delivery or vendor acceptance."
                        ]
                    },
                }
            ]
        }
        handoff = runtime_qa_handoff(ledger, configuration)
        self.assertEqual(2, handoff["item_count"])
        self.assertEqual(
            {"page_dom", "vendor_delivery"},
            {item["category"] for item in handoff["items"]},
        )
        self.assertTrue(
            all(
                item["blocking_scope"] == "nonblocking_for_unrelated_cleanup"
                for item in handoff["items"]
            )
        )
        rows, errors = build_rows(
            {
                "operations": [],
                "decision_ledger": ledger,
                "runtime_qa_handoff": handoff,
            }
        )
        self.assertEqual([], errors)
        scope = next(row for row in rows if row["ID"] == "SCOPE-001")
        self.assertIn("runtime-QA handoff item", scope["Affected object(s)"])
        self.assertEqual(0, runtime_qa_handoff([], {"rows": []})["item_count"])

    def test_vendor_registry_is_current_and_structurally_valid(self) -> None:
        registry_path = ROOT / "references/03-rules/vendor-registry.toml"
        errors, warnings = validate_registry(registry_path, online=False, max_age_days=365)
        self.assertEqual([], errors)
        self.assertEqual([], warnings)
        self.assertGreaterEqual(len(load_registry(registry_path)["vendors"]), 15)
        self.assertEqual(
            "Universal Analytics (legacy)",
            vendor_record('{"type": "ua", "trackingId": "UA-123"}')["name"],
        )
        self.assertEqual(
            "Google Ads",
            vendor_record('{"type": "googtag", "tagId": "AW-123"}')["name"],
        )
        self.assertEqual(
            "GA4 / Google tag",
            vendor_record('{"type": "gaawe", "eventName": "purchase"}')["name"],
        )
        tiktok = next(
            vendor
            for vendor in load_registry(registry_path)["vendors"]
            if vendor["name"] == "TikTok"
        )
        self.assertNotIn("CompletePayment", tiktok.get("unsupported_standard_events", []))
        self.assertEqual(
            "supported",
            next(
                contract
                for contract in tiktok["contracts"]
                if contract["event"] == "CompletePayment"
            )["status"],
        )

    def test_vendor_registry_rejects_malformed_event_lifecycle_metadata(self) -> None:
        registry_path = self.root / "malformed-vendor-registry.toml"
        registry_path.write_text(
            """schema_version = 2
reviewed_on = "2026-07-20"

[[vendors]]
name = "Example"
category = "media"
patterns = ["example"]
official_docs = ["http://docs.example.test"]
unsupported_standard_events = ["OldEvent", "OldEvent"]
event_replacements = ["DifferentEvent=>NewEvent", "broken"]
contract_version = "latest"
contracts = [
  { id = "broken", status = "unsupported", deprecated_endpoints = ["ftp://old.example.test"], field_rules = [{ field = "currency", exact_length = 0 }] },
]
""",
            encoding="utf-8",
        )
        errors, _ = validate_registry(registry_path, online=False)
        self.assertTrue(any("absolute HTTPS" in error for error in errors))
        self.assertTrue(any("contains duplicates" in error for error in errors))
        self.assertTrue(any("not listed" in error for error in errors))
        self.assertTrue(any("must use old=>new" in error for error in errors))
        self.assertTrue(any("contract_version" in error for error in errors))
        self.assertTrue(any("requires event" in error for error in errors))
        self.assertTrue(any("absolute HTTP(S)" in error for error in errors))
        self.assertTrue(any("positive integer" in error for error in errors))

    def test_vendor_url_check_falls_back_to_get_when_head_is_rejected(self) -> None:
        head_error = urllib.error.HTTPError(
            "https://example.invalid", 404, "HEAD rejected", {}, None
        )
        get_response = MagicMock()
        get_response.status = 200
        get_response.__enter__.return_value = get_response
        with patch(
            "gtm_vendor_registry.urllib.request.urlopen",
            side_effect=[head_error, get_response],
        ) as urlopen:
            self.assertIsNone(official_url_error("https://example.invalid"))
        self.assertEqual("HEAD", urlopen.call_args_list[0].args[0].method)
        self.assertEqual("GET", urlopen.call_args_list[1].args[0].method)

    def test_release_layout_allows_ignored_editable_install_metadata(self) -> None:
        errors = check_repository_layout(ROOT)
        self.assertEqual([], errors)
        self.assertEqual([], check_production_test_imports(ROOT))

    def test_release_check_does_not_hide_git_tracking_failure(self) -> None:
        (self.root / ".git").mkdir()
        with patch(
            "check_release.subprocess.check_output",
            side_effect=subprocess.CalledProcessError(1, ["git", "ls-files"]),
        ), self.assertRaisesRegex(RuntimeError, "tracked release resources"):
            git_ls_files(self.root)

    def test_release_tag_uses_semver_and_matches_project_version(self) -> None:
        self.assertEqual([], check_release_tag("v1.0.0"))
        self.assertEqual([], check_release_tag("v1.1.0-rc.1"))
        self.assertEqual([], check_release_tag("v1.1.0+build.7"))
        self.assertTrue(check_release_tag("v2026.07.20.1"))
        self.assertTrue(check_release_tag("v01.0.0"))
        self.assertTrue(check_release_tag("1.0.0"))
        metadata = tomllib.loads((ROOT / "pyproject.toml").read_text(encoding="utf-8"))
        current_version = str(metadata["project"]["version"])
        self.assertEqual([], check_project_version(ROOT, f"v{current_version}"))
        self.assertTrue(check_project_version(ROOT, "v0.0.0"))

    def test_runtime_bundle_is_self_installable_and_excludes_repo_only_files(self) -> None:
        bundle = self.root / "runtime-bundle"
        build_skill_bundle(ROOT, bundle)
        for filename in ("SKILL.md", "LICENSE", "pyproject.toml"):
            self.assertTrue((bundle / filename).is_file())
        self.assertFalse((bundle / "README.md").exists())
        metadata = tomllib.loads((bundle / "pyproject.toml").read_text(encoding="utf-8"))
        self.assertIsInstance(metadata["project"]["readme"], dict)
        self.assertNotIn("README.md", json.dumps(metadata["project"]["readme"]))
        for relative in (
            "scripts/gtm_operational_review.py",
            "scripts/gtm_configuration_review.py",
            "scripts/gtm_architecture_review.py",
            "scripts/gtm_review_common.py",
            "scripts/gtm_review_shards.py",
            "scripts/gtm_three_run_gate.py",
        ):
            self.assertTrue((bundle / relative).is_file())
        self.assertFalse((bundle / "tests").exists())
        self.assertFalse((bundle / ".github").exists())
        self.assertFalse((bundle / "scripts/check_release.py").exists())
        self.assertFalse((bundle / "scripts/gtm_self_test.py").exists())
        for script in (bundle / "scripts").glob("*.py"):
            source = script.read_text(encoding="utf-8")
            self.assertNotRegex(source, r"(?m)^\s*(?:from|import)\s+tests?\b")

        runtime_package = self.root / "clean-runtime-package"
        result = subprocess.run(
            [
                sys.executable,
                "-B",
                str(bundle / "scripts" / "gtm_audit_package_build.py"),
                str(self.export_path),
                "--out-dir",
                str(runtime_package),
                "--pretty",
            ],
            cwd=bundle,
            capture_output=True,
            text=True,
            check=False,
        )
        self.assertEqual(0, result.returncode, result.stderr)
        for filename in (
            "operational_review.json",
            "configuration_review.json",
            "architecture_review.json",
        ):
            review = json.loads(
                (runtime_package / filename).read_text(encoding="utf-8")
            )
            self.assertTrue(review["input_contract"]["contract_sha256"])
            self.assertEqual("pending", review["completion_attestation"]["status"])


if __name__ == "__main__":
    unittest.main()
