from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from gtm_change_log_build import build_change_log  # noqa: E402
from gtm_configuration_review import (  # noqa: E402
    source_known_configuration_repair_errors,
    validate_code_fix_efficacy,
)
from gtm_custom_code_extract import (  # noqa: E402
    code_hash,
    cookie_write_facts,
    javascript_ast_facts,
    returned_value_type,
    technical_code_review,
)
from gtm_future_state_check import configured_activation_risk, remap_trigger  # noqa: E402
from gtm_lib import (  # noqa: E402
    param_value,
    refs,
    safe_scalar_preview,
    source_integrity_findings,
)
from gtm_operation_compile import (  # noqa: E402
    _compose_text_change,
    merge_compatible_operations,
    mutation_path_errors,
    normalized_operation,
    operation_has_configured_activation_risk,
    runtime_neutral_operational_deletions,
    validate_mutation_conflicts,
)
from gtm_operational_review import finding_sets, matching_owner_exception  # noqa: E402
from gtm_review_common import (  # noqa: E402
    _reference_list_errors,
    _validate_remaps,
    object_source_path_map,
    validate_challenge,
    validate_neutral_recheck_contexts,
    validate_structured_actions,
)
from gtm_review_isolation import relocate_unexpected_bundle_artifacts  # noqa: E402


class RegressionSafetyTests(unittest.TestCase):
    def test_only_registered_system_trigger_ids_bypass_object_lookup(self) -> None:
        for trigger_id in ("2147479553", "2147479573", "2147479593"):
            self.assertEqual(
                [],
                _reference_list_errors(
                    "$.containerVersion.tag[0].firingTriggerId",
                    [trigger_id],
                    set(),
                    "test change",
                ),
            )
        errors = _reference_list_errors(
            "$.containerVersion.tag[0].firingTriggerId",
            ["2147479999"],
            set(),
            "test change",
        )
        self.assertTrue(any("exact system-trigger registry" in value for value in errors))

    def test_references_are_leaf_aware_and_exclude_object_metadata(self) -> None:
        self.assertEqual(set(), refs({"name": "{{Ghost}}", "notes": "{{Also ghost}}"}))
        self.assertEqual(set(), refs({"left": "{{Ghost", "right": "}}"}))
        self.assertEqual(
            {"Real"},
            refs(
                {
                    "tagId": "1",
                    "name": "Tag",
                    "parameter": [{"key": "value", "value": "{{Real}}"}],
                }
            ),
        )
        self.assertEqual(
            {"Executable"},
            refs(
                {
                    "templateId": "1",
                    "name": "Template",
                    "templateData": (
                        "___INFO___\nDocumentation {{Example}}\n"
                        "___SANDBOXED_JS_FOR_WEB_TEMPLATE___\n"
                        "const value = '{{Executable}}';"
                    ),
                }
            ),
        )

    def test_malformed_parameters_fail_cleanly_and_helpers_remain_defensive(self) -> None:
        source = {
            "containerVersion": {
                "tag": [{"tagId": "1", "name": "Bad", "parameter": ["not-an-object"]}]
            }
        }
        findings = source_integrity_findings(source)
        self.assertTrue(
            any(
                row["finding_type"] == "invalid_parameter_entry_shape"
                and not row["blocking"]
                for row in findings
            )
        )
        self.assertIsNone(param_value(source["containerVersion"]["tag"][0], "value"))

    def test_secret_context_redacts_values_that_contain_gtm_references(self) -> None:
        preview = safe_scalar_preview(
            "sk_live_ABC123 {{Debug variable}}",
            field_name="api_secret",
        )
        self.assertEqual("<redacted secret-like container value>", preview)

    def test_code_identity_preserves_literal_whitespace_and_return_type(self) -> None:
        self.assertNotEqual(code_hash("return 'a  b';"), code_hash("return 'a b';"))
        self.assertEqual(
            code_hash("return 1;\r\n"),
            code_hash("return 1;\n"),
        )
        self.assertEqual(
            "string_or_template_string",
            returned_value_type("function () { return 'true'; }"),
        )

    def test_parser_identity_is_explicit_when_ast_facts_are_available(self) -> None:
        facts = javascript_ast_facts("variable", "function () { return true; }")
        self.assertIn("javascript_parser_version", facts)
        if facts["javascript_parser"] == "esprima":
            self.assertTrue(str(facts["javascript_parser_version"]).strip())

    def test_activation_risk_accumulates_all_mutation_classes(self) -> None:
        operation = {
            "remaps": [
                {
                    "from_object_key": "folder:1",
                    "to_object_key": "folder:2",
                    "consumer_object_keys": ["variable:1"],
                }
            ],
            "changes": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].firingTriggerId",
                    "before": [],
                    "after": ["10"],
                }
            ],
        }
        self.assertTrue(operation_has_configured_activation_risk(operation))

    def test_impossible_blocker_and_simulation_do_not_create_false_activation_gate(self) -> None:
        operation = {
            "problem_type": "Impossible blocker",
            "problem": (
                "The removed blocker uses a disjoint custom-event condition; no firing route "
                "can match both conditions."
            ),
            "changes": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].blockingTriggerId",
                    "before": ["20"],
                    "after": [],
                }
            ],
        }
        self.assertTrue(operation_has_configured_activation_risk(operation))

        reachability = configured_activation_risk(
            {"tag": []},
            {"tag": []},
            {
                "operations": [
                    {
                        "operation_id": "OP-0001",
                        "operation_key": "heuristic-only",
                        "execution_safety": {
                            "configured_activation_risk": {"flag": True}
                        },
                    }
                ]
            },
        )
        self.assertFalse(reachability["flag"])
        self.assertEqual([], reachability["candidate_operation_ids"])
        self.assertEqual(["OP-0001"], reachability["heuristic_candidate_operation_ids"])
        self.assertTrue(reachability["simulation_overrides_heuristic"])

    def test_cookie_deletion_is_not_hardened_like_cookie_creation(self) -> None:
        deletion_code = "document.cookie='consent=; Max-Age=0; path=/';"
        deletion_facts = cookie_write_facts(deletion_code)
        self.assertEqual("delete", deletion_facts[0]["operation"])
        self.assertEqual("consent", deletion_facts[0]["name"])
        self.assertEqual("/", deletion_facts[0]["path"])
        deletion_findings = technical_code_review("tag", deletion_code, ["cookie write"])[
            "technical_code_security_findings"
        ]
        rendered_deletion = " ".join(deletion_findings)
        self.assertIn("exact name/path/domain", rendered_deletion)
        self.assertIn("not automatically added", rendered_deletion)
        self.assertNotIn("set/update omits", rendered_deletion)

        setter_findings = technical_code_review(
            "tag", "document.cookie='consent=yes; path=/';", ["cookie write"]
        )["technical_code_security_findings"]
        self.assertIn("set/update omits", " ".join(setter_findings))

    def test_listener_review_distinguishes_callback_once_from_registration_safety(self) -> None:
        fragile = technical_code_review(
            "tag",
            "window.addEventListener('load',handler,{once:true});",
            ["event listener"],
        )["technical_code_health_findings"]
        fragile_text = " ".join(fragile)
        self.assertIn("does not prevent duplicate registrations", fragile_text)
        self.assertIn("readyState", fragile_text)

        guarded = technical_code_review(
            "tag",
            (
                "if(document.readyState==='complete'){handler();}"
                "else if(!window.loadListenerBound){window.loadListenerBound=true;"
                "window.addEventListener('load',handler,{once:true});}"
            ),
            ["event listener"],
        )["technical_code_health_findings"]
        guarded_text = " ".join(guarded)
        self.assertNotIn("does not prevent duplicate registrations", guarded_text)
        self.assertNotIn("handler can be missed", guarded_text)

        one_shot = technical_code_review(
            "tag", "setTimeout(handler,250);", ["timer"]
        )["technical_code_health_findings"]
        self.assertNotIn("cancellation path", " ".join(one_shot))
        interval = technical_code_review(
            "tag", "setInterval(handler,250);", ["timer"]
        )["technical_code_health_findings"]
        self.assertIn("clearInterval lifecycle", " ".join(interval))

    def test_material_custom_code_detectors_keep_safe_neighbors(self) -> None:
        unbounded = technical_code_review(
            "tag",
            (
                "function poll(){if(!window.vendor){setTimeout(poll,50);return;}"
                "window.vendor.send();} poll();"
            ),
            ["timer"],
        )
        self.assertIn(
            "without an exported attempt",
            " ".join(unbounded["technical_code_health_findings"]),
        )
        bounded = technical_code_review(
            "tag",
            (
                "var attempts=0,maxAttempts=20; function poll(){"
                "if(window.vendor){window.vendor.send();return;} attempts++;"
                "if(attempts>=maxAttempts){return;} setTimeout(poll,50);} poll();"
            ),
            ["timer"],
        )
        self.assertNotIn(
            "without an exported attempt",
            " ".join(bounded["technical_code_health_findings"]),
        )

        weak_message = technical_code_review(
            "tag",
            (
                "window.addEventListener('message',function(event){"
                "if(event.origin.indexOf('trusted.example')===-1)return;"
                "dataLayer.push(event.data.payload);});"
            ),
            ["event listener", "dataLayer push"],
        )
        weak_text = " ".join(weak_message["technical_code_security_findings"])
        self.assertIn("origin with substring matching", weak_text)
        self.assertIn("payload directly into dataLayer", weak_text)

        guarded_message = technical_code_review(
            "tag",
            (
                "var allowedOrigins=['https://trusted.example'];"
                "window.addEventListener('message',function(event){"
                "if(allowedOrigins.indexOf(event.origin)===-1)return;"
                "if(typeof event.data.payload!=='object')return;"
                "dataLayer.push(event.data.payload);});"
            ),
            ["event listener", "dataLayer push"],
        )
        guarded_text = " ".join(guarded_message["technical_code_security_findings"])
        self.assertNotIn("origin with substring matching", guarded_text)
        self.assertNotIn("payload directly into dataLayer", guarded_text)

        guarded_direct_data = technical_code_review(
            "tag",
            (
                "window.addEventListener('message',function(event){"
                "if(event.origin!=='https://trusted.example')return;"
                "if(typeof event.data!=='object')return;dataLayer.push(event.data);});"
            ),
            ["event listener", "dataLayer push"],
        )
        self.assertNotIn(
            "payload directly into dataLayer",
            " ".join(guarded_direct_data["technical_code_security_findings"]),
        )

        wrong_duration = technical_code_review(
            "tag",
            (
                "function setCookie(name,value,days){var expiry=Date.now()+"
                "days*13*24*60*60*1000;document.cookie=name+'='+value+"
                "'; Secure; SameSite=Lax; expires='+expiry;}"
            ),
            ["cookie write"],
        )
        self.assertIn(
            "declared day count by 13",
            " ".join(wrong_duration["technical_code_security_findings"]),
        )
        correct_duration = technical_code_review(
            "tag",
            (
                "function setCookie(name,value,days){var expiry=Date.now()+"
                "days*24*60*60*1000;document.cookie=name+'='+value+"
                "'; Secure; SameSite=Lax; expires='+expiry;}"
            ),
            ["cookie write"],
        )
        self.assertNotIn(
            "declared day count",
            " ".join(correct_duration["technical_code_security_findings"]),
        )

        undefined_string = technical_code_review(
            "variable",
            (
                "function getCookie(){if(document.cookie)return 'yes';}"
                "var value=getCookie();return String(value);"
            ),
            ["cookie read"],
        )
        self.assertIn(
            "literal string 'undefined'",
            " ".join(undefined_string["technical_code_health_findings"]),
        )
        safe_string = technical_code_review(
            "variable",
            (
                "function getCookie(){if(document.cookie)return 'yes';return ''; }"
                "var value=getCookie();return String(value);"
            ),
            ["cookie read"],
        )
        self.assertNotIn(
            "literal string 'undefined'",
            " ".join(safe_string["technical_code_health_findings"]),
        )

        named_hour = technical_code_review(
            "variable",
            "function(){return Date.now();}",
            [],
            object_name="Device Local Hour",
        )
        self.assertIn(
            "name promises an hour value",
            " ".join(named_hour["technical_code_health_findings"]),
        )
        actual_hour = technical_code_review(
            "variable",
            "function(){return new Date().getHours();}",
            [],
            object_name="Device Local Hour",
        )
        self.assertNotIn(
            "name promises an hour value",
            " ".join(actual_hour["technical_code_health_findings"]),
        )

    def test_source_exact_actions_reject_stale_values_and_trigger_names(self) -> None:
        export = {
            "containerVersion": {
                "tag": [
                    {
                        "tagId": "1",
                        "name": "Tag One",
                        "type": "html",
                        "firingTriggerId": ["10"],
                    }
                ],
                "trigger": [
                    {"triggerId": "10", "name": "All Pages", "type": "PAGEVIEW"},
                    {"triggerId": "11", "name": "Second Route", "type": "CUSTOM_EVENT"},
                ],
            }
        }
        with tempfile.TemporaryDirectory() as directory:
            export_path = Path(directory) / "source.json"
            export_path.write_text(json.dumps(export), encoding="utf-8")
            source_paths = object_source_path_map(export_path)
            base_path = source_paths["tag:1"]
            valid_keys = set(source_paths)

            valid_append = {
                "additions": [
                    {
                        "object_key": "tag:1",
                        "json_path": f"{base_path}.firingTriggerId",
                        "value": "11",
                        "mode": "append",
                        "reason": "Append the exact exported trigger identifier.",
                    }
                ]
            }
            self.assertEqual(
                [],
                validate_structured_actions(
                    valid_append,
                    valid_keys,
                    "valid append",
                    source_paths_by_key=source_paths,
                ),
            )

            stale_change = {
                "changes": [
                    {
                        "object_key": "tag:1",
                        "json_path": f"{base_path}.firingTriggerId",
                        "before": ["stale"],
                        "after": ["All Pages"],
                    }
                ]
            }
            errors = validate_structured_actions(
                stale_change,
                valid_keys,
                "stale change",
                source_paths_by_key=source_paths,
            )
            self.assertTrue(any("locked source value" in error for error in errors))
            self.assertTrue(any("trigger names or unknown trigger IDs" in error for error in errors))

    def test_source_known_consent_trigger_repair_cannot_be_owner_fallback(self) -> None:
        row = {
            "object_key": "tag:7",
            "required_configuration_obligations": [
                {
                    "obligation_key": "consent_default_wrong_initialization_trigger",
                    "source_known_repair": {
                        "mode": "change",
                        "object_key": "tag:7",
                        "json_path": "$.containerVersion.tag[0].firingTriggerId",
                        "before": ["10"],
                        "after": ["2147479593"],
                    },
                }
            ],
            "disposition": "owner_decision_needed",
            "operation": {},
        }
        self.assertTrue(
            source_known_configuration_repair_errors(row, "consent default")
        )
        row.update(
            {
                "disposition": "cleanup_operation",
                "operation": {
                    "changes": [
                        {
                            "object_key": "tag:7",
                            "json_path": "$.containerVersion.tag[0].firingTriggerId",
                            "before": ["10"],
                            "after": ["2147479593"],
                        }
                    ]
                },
            }
        )
        self.assertEqual(
            [], source_known_configuration_repair_errors(row, "consent default")
        )

    def test_system_trigger_repair_does_not_require_global_detach(self) -> None:
        allowed_keys = {"tag:7", "tag:8", "trigger:10"}
        self.assertEqual(
            [],
            _reference_list_errors(
                "$.containerVersion.tag[0].firingTriggerId",
                ["2147479593"],
                allowed_keys,
                "consent default change",
            ),
        )
        self.assertEqual(
            [],
            _validate_remaps(
                {
                    "changes": [
                        {
                            "object_key": "tag:7",
                            "json_path": "$.containerVersion.tag[0].firingTriggerId",
                            "before": ["10"],
                            "after": ["2147479593"],
                        }
                    ],
                    "remaps": [],
                    "deletions": [],
                },
                allowed_keys,
                "consent default change",
                {"trigger:10": {"tag:7", "tag:8"}},
            ),
        )

    def test_code_rewrite_requires_noncosmetic_source_bound_proof(self) -> None:
        before = "<script>window.dataLayer.push({event:'legacy'});</script>"
        after = "<script>window.dataLayer.push({event:'canonical'});</script>"
        row = {
            "technical_code_facts": {"code_hash": code_hash(before)},
            "operation": {
                "changes": [
                    {
                        "json_path": "$.containerVersion.tag[0].parameter[0].value",
                        "before": before,
                        "after": after,
                    }
                ]
            },
        }
        self.assertTrue(
            any("code_fix_proof" in error for error in validate_code_fix_efficacy(row, "code"))
        )
        non_code_edit = {
            "technical_code_facts": {"code_hash": code_hash(before)},
            "operation": {
                "changes": [
                    {
                        "json_path": "$.containerVersion.tag[0].monitoringMetadata.type",
                        "before": "MAP",
                        "after": "TEMPLATE",
                    }
                ]
            },
        }
        self.assertEqual([], validate_code_fix_efficacy(non_code_edit, "non-code"))
        row["operation"]["code_fix_proof"] = {
            "defective_path": "$.containerVersion.tag[0].parameter[0].value",
            "source_code_hash": code_hash(before),
            "exact_change_summary": (
                "Replace the legacy event literal with the approved canonical event literal."
            ),
            "resolution_basis": (
                "The changed literal is the exact defective source coordinate and output field."
            ),
            "preserved_behaviors": [
                "Preserve the same dataLayer push timing and browser side effect.",
                "Preserve the same object shape and downstream event route.",
            ],
        }
        self.assertEqual([], validate_code_fix_efficacy(row, "code"))
        row["operation"]["changes"][0]["after"] = (
            "<script> window.dataLayer.push({event:'legacy'}); </script>"
        )
        self.assertTrue(
            any("cosmetic" in error for error in validate_code_fix_efficacy(row, "code"))
        )

    def test_material_challenge_requires_a_fresh_neutral_source_recheck(self) -> None:
        challenge = {
            "priority": "High",
            "challenge_review": {
                "source_recheck": "The exact source object and fields were reviewed again.",
                "status_and_scope_check": "The complete active scope was reviewed again.",
                "alternative_explanation": "An intentional distinct purpose was considered and rejected.",
                "challenge_verdict": "confirmed",
                "neutral_recheck": {
                    "recheck_context_id": "scan-context-001",
                    "source_coordinates": ["$.containerVersion.tag[0]"],
                    "neutral_question": "What disposition do these exact source facts support?",
                    "expected_outcome_disclosed": False,
                    "foreign_rationale_artifacts_used": [],
                    "recheck_verdict": "confirmed",
                },
            },
        }
        self.assertEqual([], validate_challenge(challenge, "challenge"))
        self.assertTrue(
            any(
                "reused the scan context" in error
                for error in validate_neutral_recheck_contexts(
                    {"rows": [challenge]}, "scan-context-001", "review"
                )
            )
        )
        challenge["challenge_review"]["neutral_recheck"]["source_coordinates"] = [
            "$.containerVersion.tag[999]"
        ]
        self.assertTrue(
            any(
                "do not resolve in the locked source" in error
                for error in validate_challenge(
                    challenge,
                    "challenge",
                    {"tag:1": "$.containerVersion.tag[0]"},
                )
            )
        )

    def test_unexpected_review_drafts_are_recovered_outside_the_sealed_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            package = Path(directory)
            bundle = package / "review-bundles" / "configuration_correctness"
            bundle.mkdir(parents=True)
            (bundle / "bundle_manifest.json").write_text(
                json.dumps({"input_files": []}), encoding="utf-8"
            )
            draft = bundle / "analyst-notes.txt"
            draft.write_text("temporary notes", encoding="utf-8")
            moved = relocate_unexpected_bundle_artifacts(
                bundle, package, "configuration_correctness"
            )
            self.assertEqual("analyst-notes.txt", moved[0]["from"])
            self.assertFalse(draft.exists())
            recovered = package / moved[0]["to"]
            self.assertEqual("temporary notes", recovered.read_text(encoding="utf-8"))
            self.assertTrue(
                (recovered.parent / "scratch_recovery.json").is_file()
            )

    def test_merge_preserves_high_priority_challenge_metadata(self) -> None:
        action = [{"object_key": "variable:1", "reason": "Remove exact duplicate."}]
        low = normalized_operation(
            {
                "operation_key": "low-lens",
                "priority": "Low",
                "deletions": action,
            },
            "operational_sanitation",
            "OPS-1",
            ["variable:1"],
        )
        high = normalized_operation(
            {
                "operation_key": "high-lens",
                "priority": "High",
                "deletions": action,
                "challenge_review": {
                    "source_recheck": "The source configuration was rechecked object by object.",
                    "status_and_scope_check": (
                        "The active status and complete affected scope were rechecked."
                    ),
                    "alternative_explanation": (
                        "An intentional distinct business role was considered and rejected."
                    ),
                    "challenge_verdict": "confirmed",
                    "neutral_recheck": {
                        "recheck_context_id": "neutral-recheck-context-001",
                        "source_coordinates": ["$.containerVersion.variable[0]"],
                        "neutral_question": (
                            "What cleanup disposition do these exact duplicate source "
                            "objects and their consumer facts support?"
                        ),
                        "expected_outcome_disclosed": False,
                        "foreign_rationale_artifacts_used": [],
                        "recheck_verdict": "confirmed",
                    },
                },
            },
            "business_architecture",
            "REL-1",
            ["variable:1"],
        )
        errors: list[str] = []
        merged = merge_compatible_operations([low, high], errors)
        self.assertEqual([], errors)
        self.assertEqual("High", merged[0]["priority"])
        self.assertEqual(
            "confirmed",
            merged[0]["challenge_review"]["challenge_verdict"],
        )

    def test_coincident_insertions_do_not_merge_in_arbitrary_order(self) -> None:
        self.assertIsNone(_compose_text_change("ab", ["aXb", "aYb"]))
        self.assertEqual("aXb", _compose_text_change("ab", ["aXb", "aXb"]))

    def test_redundant_deletion_reconciliation_is_input_order_independent(self) -> None:
        def deletion_only(key: str, target: str) -> dict:
            return normalized_operation(
                {
                    "operation_key": key,
                    "deletions": [
                        {"object_key": target, "reason": "Delete redundant object."}
                    ],
                },
                "operational_sanitation",
                key,
                [target],
            )

        carrier = normalized_operation(
            {
                "operation_key": "carrier",
                "canonical_object_key": "trigger:3",
                "remaps": [
                    {
                        "from_object_key": "trigger:2",
                        "to_object_key": "trigger:3",
                        "consumer_object_keys": ["tag:1"],
                    }
                ],
                "deletions": [
                    {"object_key": "trigger:1", "reason": "Delete redundant object."},
                    {"object_key": "trigger:2", "reason": "Delete redundant object."},
                ],
            },
            "business_architecture",
            "carrier",
            ["trigger:1", "trigger:2", "trigger:3"],
        )
        operations = [
            deletion_only("delete-one", "trigger:1"),
            carrier,
            deletion_only("delete-two", "trigger:2"),
        ]
        left_errors: list[str] = []
        right_errors: list[str] = []
        left = merge_compatible_operations(operations, left_errors)
        right = merge_compatible_operations(list(reversed(operations)), right_errors)
        self.assertEqual([], left_errors)
        self.assertEqual([], right_errors)
        self.assertEqual(left, right)
        self.assertEqual(1, len(left))

    def test_equal_deletions_merge_despite_display_metadata(self) -> None:
        operational = normalized_operation(
            {
                "operation_key": "operational-delete",
                "deletions": [
                    {"object_key": "variable:42", "reason": "No active consumer."}
                ],
            },
            "operational_sanitation",
            "BASE-42",
            ["variable:42"],
        )
        architecture = normalized_operation(
            {
                "operation_key": "architecture-delete",
                "deletions": [
                    {
                        "object_key": "variable:42",
                        "object_name": "Legacy variable",
                        "source_json_path": "$.containerVersion.variable[0]",
                        "reason": "Superseded canonical design.",
                    }
                ],
            },
            "business_architecture",
            "REL-42:operation:1",
            ["variable:42"],
        )
        errors: list[str] = []
        merged = merge_compatible_operations([operational, architecture], errors)
        self.assertEqual([], errors)
        self.assertEqual(1, len(merged))
        self.assertEqual(["variable:42"], merged[0]["source_object_keys"])

    def test_equal_deletions_preserve_the_only_declared_canonical_target(self) -> None:
        operational = normalized_operation(
            {
                "operation_key": "operational-delete",
                "deletions": [{"object_key": "variable:42", "reason": "Unused."}],
            },
            "operational_sanitation",
            "BASE-42",
            ["variable:42"],
        )
        architecture = normalized_operation(
            {
                "operation_key": "architecture-delete",
                "canonical_object_key": "variable:43",
                "deletions": [
                    {"object_key": "variable:42", "reason": "Superseded."}
                ],
            },
            "business_architecture",
            "REL-42",
            ["variable:42", "variable:43"],
        )
        errors: list[str] = []

        merged = merge_compatible_operations([operational, architecture], errors)

        self.assertEqual([], errors)
        self.assertEqual("variable:43", merged[0]["canonical_object_key"])

    def test_equal_deletions_block_conflicting_canonical_targets(self) -> None:
        first = normalized_operation(
            {
                "operation_key": "first",
                "canonical_object_key": "variable:43",
                "deletions": [{"object_key": "variable:42", "reason": "Duplicate."}],
            },
            "business_architecture",
            "REL-42-A",
            ["variable:42", "variable:43"],
        )
        second = normalized_operation(
            {
                "operation_key": "second",
                "canonical_object_key": "variable:44",
                "deletions": [{"object_key": "variable:42", "reason": "Duplicate."}],
            },
            "business_architecture",
            "REL-42-B",
            ["variable:42", "variable:44"],
        )
        errors: list[str] = []

        merge_compatible_operations([first, second], errors)

        self.assertTrue(any("conflicting canonical targets" in value for value in errors))

    def test_ineffective_blocker_can_retire_only_its_orphaned_dependency_chain(self) -> None:
        finding = {
            "finding_id": "BASE-INEFFECTIVE",
            "finding_type": "ineffective_blocking_trigger",
            "deterministic_repair": {
                "status": "unique_safe_repair",
                "deletions": [{"object_key": "trigger:2"}],
            },
        }
        operation = {
            "source_references": ["BASE-INEFFECTIVE"],
            "deletions": [
                {"object_key": "trigger:2"},
                {"object_key": "variable:3"},
                {"object_key": "variable:4"},
            ],
        }
        self.assertEqual(
            {"trigger:2", "variable:3"},
            runtime_neutral_operational_deletions(
                operation,
                {"BASE-INEFFECTIVE": finding},
                {
                    "trigger:2": {"tag:1"},
                    "variable:3": {"trigger:2"},
                    "variable:4": set(),
                },
            ),
        )

    def test_remap_rewrite_conflicts_with_deleting_its_consumer(self) -> None:
        errors = validate_mutation_conflicts(
            [
                {
                    "operation_key": "remap",
                    "remaps": [
                        {
                            "from_object_key": "trigger:1",
                            "to_object_key": "trigger:2",
                            "consumer_object_keys": ["tag:1"],
                        }
                    ],
                },
                {
                    "operation_key": "delete",
                    "deletions": [{"object_key": "tag:1"}],
                },
            ]
        )
        self.assertTrue(any("also changed" in error for error in errors))

    def test_mutation_paths_bind_to_the_exact_object_entry(self) -> None:
        errors = mutation_path_errors(
            [
                {
                    "operation_key": "wrong-index",
                    "changes": [
                        {
                            "object_key": "tag:1",
                            "json_path": "$.containerVersion.tag[1].name",
                        }
                    ],
                }
            ],
            {"tag:1": "$.containerVersion.tag[0]"},
        )
        self.assertTrue(any("another object's source json_path" in error for error in errors))

    def test_trigger_remap_deduplicates_existing_target_references(self) -> None:
        consumer = {
            "firingTriggerId": ["1", "2"],
            "blockingTriggerId": ["1", "2"],
            "parameter": [
                {
                    "key": "triggerIds",
                    "list": [{"value": "1"}, {"value": "2"}],
                }
            ],
        }
        remap_trigger("1", "2", consumer)
        self.assertEqual(["2"], consumer["firingTriggerId"])
        self.assertEqual(["2"], consumer["blockingTriggerId"])
        self.assertEqual([{"value": "2"}], consumer["parameter"][0]["list"])

    def test_owner_exception_must_cover_the_complete_finding_scope(self) -> None:
        finding = {
            "finding_id": "OPS-1",
            "signature_key": "scope",
            "object_names": ["Tag A", "Tag B"],
            "object_ids": [],
        }
        partial = {
            "known_owner_exceptions": [
                {
                    "object_names": ["Tag A"],
                    "reason": "Tag A is explicitly retained by the campaign owner.",
                }
            ]
        }
        complete = {
            "known_owner_exceptions": [
                {
                    "object_names": ["Tag A", "Tag B"],
                    "reason": "Both tags are explicitly retained by the campaign owner.",
                }
            ]
        }
        self.assertIsNone(matching_owner_exception(finding, partial))
        self.assertIsNotNone(matching_owner_exception(finding, complete))

    def test_operational_finding_set_rejects_malformed_and_duplicate_rows(self) -> None:
        expected = {"findings": [{"finding_id": "OPS-1"}]}
        supplied = {
            "findings": [
                {"finding_id": "OPS-1"},
                {"finding_id": "OPS-1"},
                "malformed",
            ]
        }
        _, _, errors = finding_sets(supplied, expected)
        self.assertTrue(any("malformed" in error for error in errors))
        self.assertTrue(any("unique and nonblank" in error for error in errors))

    def test_change_log_splits_long_field_proof_losslessly(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        long_value = "source-value-" + ("x" * 40_000)
        payload = {
            "execution_mode": "planned",
            "changes": [
                {
                    "change_id": "CHG-1",
                    "operation_id": "OP-1",
                    "layer": "tag",
                    "object_id": "1",
                    "before_name": "Tag",
                    "after_name": "Tag",
                    "change_category": "field_change",
                    "action": "change",
                    "field_path": "$.parameter[0].value",
                    "before_value": long_value,
                    "after_value": "new",
                    "reason": "Replace the exact source value.",
                    "functional_impact": "Preserve the configured event route.",
                    "qa_method": "Static export readback.",
                    "qa_status": "planned",
                    "rollback": "Restore the before value.",
                    "status": "planned",
                    "route": "import",
                    "blocker": "",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "change-log.xlsx"
            build_change_log(payload, output)
            workbook = load_workbook(output, read_only=True)
            details = workbook["02 Change Log Details"]
            rendered = "".join(
                str(details.cell(row, 4).value or "")
                for row in range(2, details.max_row + 1)
            )
            workbook.close()
        self.assertIn(long_value, rendered)


if __name__ == "__main__":
    unittest.main()
