from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from gtm_human_rows import operation_action_text, operation_problem_text  # noqa: E402
from gtm_workbook_readability import (  # noqa: E402
    HUMAN_SHEETS,
    ORIGINAL_SHEETS,
    decision_topics,
    sha256_file,
    visible_consequence,
    workbook_sheet_hashes,
)
from gtm_workbook_readability import (  # noqa: E402
    build as build_readability,
)
from gtm_workbook_readability_gate import validate as validate_readability  # noqa: E402

FIXTURE = ROOT / "tests" / "fixtures" / "readability_case.json"
REAL_RUN_FIXTURE = (
    ROOT / "tests" / "fixtures" / "readability_real_run_cases.json"
)


class WorkbookReadabilityTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_context = tempfile.TemporaryDirectory()
        self.temp_dir = Path(self.temp_context.name)
        self.package_dir = self.temp_dir / "audit-package"
        self.package_dir.mkdir()
        fixture = json.loads(FIXTURE.read_text(encoding="utf-8"))
        for role in (
            "audit_package_manifest",
            "context",
            "source_model",
            "operational_review",
            "configuration_review",
            "architecture_review",
            "technical_code_findings",
            "future_state_gate",
            "completion_gate",
        ):
            (self.package_dir / f"{role}.json").write_text(
                json.dumps(fixture[role], ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
        self.operations = self.package_dir / "reconciled_operations.json"
        self.operations.write_text(
            json.dumps(
                fixture["reconciled_operations"],
                ensure_ascii=False,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )
        self.canonical = self.temp_dir / "cleanup_plan.xlsx"
        self.analyst = self.temp_dir / "cleanup_plan.analyst.xlsx"
        self.manifest = self.temp_dir / "cleanup_plan.analyst.manifest.json"
        self._write_canonical_workbook()

    def tearDown(self) -> None:
        self.temp_context.cleanup()

    def test_generic_impact_is_replaced_by_problem_specific_consequence(self) -> None:
        operation = {
            "affected_objects": "tag:7 Consent default",
            "why_it_matters": (
                "Preserves 2 affected measurement families; see the evidence package."
            ),
            "changes": [
                {
                    "object_key": "tag:7",
                    "json_path": "$.containerVersion.tag[0].firingTriggerId",
                    "before": ["10"],
                    "after": ["2147479593"],
                }
            ],
        }
        consequence = visible_consequence(
            operation,
            (
                "Tag exports a default consent command but does not fire on the "
                "Consent Initialization system trigger."
            ),
        )
        self.assertIn("after other tags have already evaluated", consequence)
        self.assertNotIn("measurement families", consequence)

    def test_real_run_wording_stays_literal_and_actionable(self) -> None:
        cases = json.loads(REAL_RUN_FIXTURE.read_text(encoding="utf-8"))
        for case in cases:
            with self.subTest(case=case["case_id"]):
                operation = case["operation"]
                catalog = case["catalog"]
                problem = operation_problem_text(operation, catalog)
                consequence = visible_consequence(operation, problem)
                action = operation_action_text(operation, catalog)
                rendered = " ".join((problem, consequence, action))
                for phrase in case["required_phrases"]:
                    self.assertIn(phrase, rendered)
                for phrase in case["forbidden_phrases"]:
                    self.assertNotIn(phrase, rendered)

    def test_reconciliation_closure_is_visible_without_becoming_a_fourth_scan(
        self,
    ) -> None:
        payload = json.loads(self.operations.read_text(encoding="utf-8"))
        payload["object_catalog"]["trigger:18"] = {
            "object_name": "Legacy ecommerce trigger"
        }
        payload["decision_ledger"].append(
            {
                "decision_id": "REC-CLOSURE-0001",
                "source_run": "cross_run_reconciliation",
                "source_object_keys": ["trigger:18"],
                "disposition": "cleanup_operation",
                "area": "GTM hygiene",
                "problem_type": "Unused object",
                "affected_objects": "trigger:18 — Legacy ecommerce trigger",
                "summary": (
                    "The trigger becomes unused after its final tag consumer is removed."
                ),
                "compiled_operation_ids": ["OP-0003"],
            }
        )
        payload["operations"].append(
            {
                "operation_id": "OP-0003",
                "operation_key": "cleanup-closure-trigger-18",
                "execution_order": 3,
                "priority": "Low",
                "problem_type": "Unused object",
                "problem": (
                    "trigger:18 becomes unused after the planned cleanup removes its "
                    "last tag consumer."
                ),
                "why_it_matters": (
                    "Leaving trigger:18 after its last consumer is removed preserves "
                    "an obsolete dependency that can be selected by mistake."
                ),
                "affected_objects": "trigger:18 — Legacy ecommerce trigger",
                "affected_object_keys": ["trigger:18"],
                "source_object_keys": ["trigger:18"],
                "source_references": ["REC-CLOSURE-0001"],
                "depends_on_operation_ids": ["OP-0001"],
                "creations": [],
                "additions": [],
                "changes": [],
                "remaps": [],
                "renames": [],
                "deletions": [{"object_key": "trigger:18"}],
                "execution_phases": ["delete"],
                "priority_basis": {
                    "active_reachability": "inactive_after_prerequisites"
                },
                "execution_safety": {
                    "approval": {
                        "scope": "bulk_eligible_exact_low_risk_bundle",
                        "reasons": ["inactive after exact prerequisites"],
                    },
                    "decommission": {"required": False},
                    "configured_activation_risk": {"flag": False},
                },
                "qa_steps": "Read back the complete final workspace.",
                "rollback": "Restore trigger:18 from the locked export.",
            }
        )
        self.operations.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        self._build()
        result = self._validate()
        self.assertEqual("pass", result["status"])
        workbook = load_workbook(self.analyst)
        try:
            actions = workbook["A2 Actions"]
            row = self._row_with(actions, 1, "OP-0003")
            self.assertIn("Legacy ecommerce trigger", str(actions.cell(row, 3).value))
            audit = workbook["A4 Audit Register"]
            self._row_with(audit, 1, "REC-CLOSURE-0001")
        finally:
            workbook.close()

    def test_large_owner_register_has_no_arbitrary_grouping_gate(self) -> None:
        owners = [
            {
                "decision_id": f"OWNER-{index:02d}",
                "owner_question": f"Who owns decision {index}?",
                "recommended_action": f"Confirm decision {index} ownership.",
                "problem_type": "Ownership",
            }
            for index in range(1, 17)
        ]
        source_sha256 = "a" * 64
        default_topics = decision_topics(owners, None, source_sha256)
        self.assertEqual(16, len(default_topics))

        singleton_topics = [
            {
                "topic_id": f"D-{index:02d}",
                "source_ids": [owner["decision_id"]],
                "question": owner["owner_question"],
                "recommendation": owner["recommended_action"],
            }
            for index, owner in enumerate(owners, start=1)
        ]
        singleton_result = decision_topics(
            owners,
            {
                "kind": "gtm_readability_decision_topics",
                "source_sha256": source_sha256,
                "topics": singleton_topics,
            },
            source_sha256,
        )
        self.assertEqual(16, len(singleton_result))

        grouped_topics = [
            {
                "topic_id": "D-01",
                "title": "Shared ownership",
                "source_ids": ["OWNER-01", "OWNER-02"],
                "question": "Who owns decisions 1 and 2?",
                "recommendation": "Confirm one shared owner for decisions 1 and 2.",
            },
            *[
                {
                    "topic_id": f"D-{index - 1:02d}",
                    "source_ids": [owner["decision_id"]],
                    "question": owner["owner_question"],
                    "recommendation": owner["recommended_action"],
                }
                for index, owner in enumerate(owners[2:], start=3)
            ],
        ]
        topics = decision_topics(
            owners,
            {
                "kind": "gtm_readability_decision_topics",
                "source_sha256": source_sha256,
                "topics": grouped_topics,
            },
            source_sha256,
        )
        self.assertEqual(15, len(topics))

    def test_semantic_fallback_never_hides_two_distinct_decisions_from_one_lens(self) -> None:
        shared = {
            "problem_type": "Legacy app routing",
            "source_object_keys": ["tag:1", "variable:10"],
        }
        records = [
            {
                **shared,
                "decision_id": "CFG-0001",
                "owner_question": "Which data source should become canonical?",
                "recommended_action": "Select the canonical data source after exact comparison.",
            },
            {
                **shared,
                "decision_id": "CFG-0002",
                "owner_question": "Who owns the legacy app route after migration?",
                "recommended_action": "Confirm the route owner before retirement.",
            },
        ]
        self.assertEqual(2, len(decision_topics(records, None, "a" * 64)))

    def test_overview_uses_simulated_activation_instead_of_path_heuristic(self) -> None:
        operations = json.loads(self.operations.read_text(encoding="utf-8"))
        operations["operations"][0].setdefault("execution_safety", {})[
            "configured_activation_risk"
        ] = {"flag": True}
        self.operations.write_text(
            json.dumps(operations, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        future_path = self.package_dir / "future_state_gate.json"
        future = json.loads(future_path.read_text(encoding="utf-8"))
        future["configured_activation_risk"] = {
            "flag": False,
            "candidate_operation_ids": [],
            "heuristic_candidate_operation_ids": [
                operations["operations"][0]["operation_id"]
            ],
            "simulation_overrides_heuristic": True,
        }
        future_path.write_text(
            json.dumps(future, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        manifest = self._build()
        self.assertEqual(0, manifest["coverage"]["activation_operations"])

    def _write_canonical_workbook(self) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for index, name in enumerate(ORIGINAL_SHEETS):
            sheet = workbook.create_sheet(name)
            sheet["A1"] = "Technical ID"
            sheet["B1"] = "Value"
            sheet["A2"] = f"TECH-{index + 1:02d}"
            sheet["B2"] = name
            sheet["A3"] = "Formula"
            sheet["B3"] = "=1+1"
            sheet["B2"].comment = Comment(
                f"Canonical note for {name}",
                "Audit fixture",
            )
            sheet["B2"].hyperlink = "#'01 Summary'!A1"
            sheet.sheet_state = "visible" if index < 2 else "hidden"
        workbook.active = 0
        workbook.save(self.canonical)
        workbook.close()

    def _build(
        self,
        decision_topics: Path | None = None,
        language: str = "en",
    ) -> dict:
        _output, _manifest_path, manifest = build_readability(
            self.package_dir,
            self.operations,
            self.canonical,
            self.analyst,
            decision_topics_path=decision_topics,
            manifest_path=self.manifest,
            language=language,
        )
        return manifest

    def _validate(self, decision_topics: Path | None = None) -> dict:
        return validate_readability(
            self.package_dir,
            self.operations,
            self.canonical,
            self.analyst,
            decision_topics_path=decision_topics,
            manifest_path=self.manifest,
        )

    def _refresh_analyst_hash(self) -> None:
        manifest = json.loads(self.manifest.read_text(encoding="utf-8"))
        manifest["analyst_workbook"]["sha256"] = sha256_file(self.analyst)
        self.manifest.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    def _write_topics(self, topics: list[dict], *, include_sha: bool = True) -> Path:
        path = self.temp_dir / "decision_topics.json"
        payload = {
            "kind": "gtm_readability_decision_topics",
            "topics": topics,
        }
        if include_sha:
            payload["source_sha256"] = "a" * 64
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        return path

    @staticmethod
    def _row_with(sheet, column: int, needle: str) -> int:
        for row in range(2, sheet.max_row + 1):
            if needle in str(sheet.cell(row, column).value or ""):
                return row
        raise AssertionError(f"Could not find {needle!r} in {sheet.title}")

    def test_builder_and_gate_are_lossless_complete_and_lean(self) -> None:
        canonical_hashes = workbook_sheet_hashes(self.canonical)
        built_manifest = self._build()
        result = self._validate()

        self.assertEqual(result["status"], "pass")
        self.assertTrue(all(value == "pass" for value in result["gates"].values()))
        self.assertEqual(
            built_manifest["coverage"],
            {
                "audit_records": 6,
                "operations": 2,
                "owner_source_records": 2,
                "decision_topics": 1,
                "custom_html_tags": 2,
                "retained": 1,
                "documented_exceptions": 1,
                "evidence_limits": 0,
                "priority": {
                    "Critical": 0,
                    "High": 1,
                    "Medium": 0,
                    "Low": 1,
                },
                "bulk_operations": 0,
                "individual_operations": 2,
                "activation_operations": 0,
                "maintenance_operations": 1,
                "behavior_operations": 1,
                "remaining_records": 4,
            },
        )

        workbook = load_workbook(self.analyst, data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, HUMAN_SHEETS + ORIGINAL_SHEETS)
            self.assertEqual(
                workbook_sheet_hashes(self.analyst, ORIGINAL_SHEETS),
                canonical_hashes,
            )
            self.assertEqual(workbook["A4 Audit Register"].max_column, 6)
            self.assertEqual(workbook["A2 Actions"].max_column, 8)
            self.assertEqual(workbook["A3 Decisions"].max_column, 6)
            self.assertEqual(workbook["A5 Custom HTML"].max_column, 7)
            self.assertTrue(
                all(workbook[name].sheet_state == "hidden" for name in ORIGINAL_SHEETS)
            )

            audit = workbook["A4 Audit Register"]
            audit_ids = {
                str(audit.cell(row, 1).value)
                for row in range(2, audit.max_row + 1)
                if str(audit.cell(row, 1).value or "") != "—"
            }
            self.assertEqual(
                audit_ids,
                {
                    "OPS-0001",
                    "OPS-0002",
                    "CFG-0001",
                    "CFG-0002",
                    "ARC-0001",
                    "ARC-0002",
                },
            )

            actions = workbook["A2 Actions"]
            op1 = self._row_with(actions, 1, "OP-0001")
            op2 = self._row_with(actions, 1, "OP-0002")
            op1_text = str(actions.cell(op1, 6).value)
            self.assertLess(op1_text.index("Keep variable:20"), op1_text.index("Delete variable:10"))
            self.assertIn(
                "Repoint trigger:30 — Event - App from variable:10",
                op1_text,
            )
            self.assertIn("to variable:20 — DLV - App State", op1_text)
            action_note = actions.cell(op1, 6).comment
            self.assertIsNotNone(action_note)
            self.assertIn(
                "$.containerVersion.trigger[0].customEventFilter[0].parameter[1].value",
                action_note.text,
            )
            self.assertIn(
                "canonical-route-value-bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb",
                action_note.text,
            )
            self.assertIn("Disable/deselect", str(actions.cell(op2, 6).value))

            decisions = workbook["A3 Decisions"]
            decision_values = [
                str(decisions.cell(row, 1).value or "")
                for row in range(2, decisions.max_row + 1)
            ]
            self.assertEqual(sum("CFG-0001" in value for value in decision_values), 1)
            self.assertEqual(sum("ARC-0002" in value for value in decision_values), 1)
            self.assertTrue(
                any("source items" in str(cell.value or "") for cell in decisions["D"])
            )

            custom = workbook["A5 Custom HTML"]
            tag1 = self._row_with(custom, 1, "tag:1")
            tag2 = self._row_with(custom, 1, "tag:2")
            assessment = str(custom.cell(tag1, 5).value)
            self.assertIn("variable:10", assessment)
            self.assertIn("OP-0001", assessment)
            self.assertIn("key match does not prove live equivalence", assessment)
            self.assertEqual(custom.cell(tag2, 2).value, "Paused.")
            self.assertIn("Already reads", str(custom.cell(tag2, 5).value))
            self.assertIn("selected disposition", str(custom.cell(tag1, 4).value))
            self.assertTrue(str(custom.cell(tag1, 7).value).strip())
        finally:
            workbook.close()

    def test_owner_decisions_do_not_block_delivery(self) -> None:
        manifest = self._build()
        result = self._validate()
        self.assertEqual(manifest["coverage"]["owner_source_records"], 2)
        self.assertEqual(result["status"], "pass")
        self.assertTrue(self.canonical.is_file())
        self.assertTrue(manifest["fallback"]["deliver_when_readability_gate_fails"])

    def test_french_localizes_only_builder_owned_labels(self) -> None:
        manifest = self._build(language="fr-FR")
        result = self._validate()

        self.assertEqual(result["status"], "pass")
        self.assertEqual(manifest["language"], "fr-FR")
        workbook = load_workbook(self.analyst)
        try:
            self.assertEqual(
                workbook["A4 Audit Register"]["E1"].value,
                "Résultat / attente",
            )
            overview_values = {
                str(cell.value or "")
                for row in workbook["A1 Overview"].iter_rows()
                for cell in row
            }
            self.assertIn("2 sources / 1 sujets", overview_values)
            actions = workbook["A2 Actions"]
            row = self._row_with(actions, 1, "OP-0001")
            self.assertIn("variable:20", str(actions.cell(row, 6).value))
            self.assertIn("DLV - App State", str(actions.cell(row, 6).value))
        finally:
            workbook.close()

    def test_editorial_topics_group_equivalent_questions_without_losing_sources(
        self,
    ) -> None:
        topics = self._write_topics(
            [
                {
                    "topic_id": "D-01",
                    "title": "Future app-route ownership",
                    "question": "Which source should own the app route?",
                    "recommendation": (
                        "Compare live equivalence, then retain one canonical source."
                    ),
                    "source_ids": ["CFG-0001", "ARC-0002"],
                }
            ]
        )
        manifest = self._build(topics)
        result = self._validate(topics)

        self.assertEqual(result["status"], "pass")
        self.assertEqual(manifest["coverage"]["decision_topics"], 1)
        workbook = load_workbook(self.analyst)
        try:
            decisions = workbook["A3 Decisions"]
            values = [
                str(decisions.cell(row, 1).value or "")
                for row in range(2, decisions.max_row + 1)
            ]
            self.assertEqual(sum("CFG-0001" in value for value in values), 1)
            self.assertEqual(sum("ARC-0002" in value for value in values), 1)
            self.assertTrue(
                any(
                    "2 source items" in str(cell.value or "")
                    for cell in decisions["D"]
                )
            )
        finally:
            workbook.close()

    def test_editorial_topics_reject_duplicate_source_ids_within_one_topic(
        self,
    ) -> None:
        topics = self._write_topics(
            [
                {
                    "topic_id": "D-01",
                    "question": "Which source should own the app route?",
                    "recommendation": "Retain one canonical source.",
                    "source_ids": ["CFG-0001", "CFG-0001", "ARC-0002"],
                }
            ]
        )
        with self.assertRaisesRegex(ValueError, "duplicate"):
            self._build(topics)

    def test_editorial_single_topic_cannot_rewrite_source_question(self) -> None:
        topics = self._write_topics(
            [
                {
                    "topic_id": "D-01",
                    "question": "A rewritten question?",
                    "recommendation": "A rewritten recommendation.",
                    "source_ids": ["CFG-0001"],
                },
                {
                    "topic_id": "D-02",
                    "source_ids": ["ARC-0002"],
                },
            ]
        )
        with self.assertRaisesRegex(ValueError, "cannot rewrite"):
            self._build(topics)

    def test_editorial_topics_require_the_locked_source_hash(self) -> None:
        topics = self._write_topics(
            [
                {
                    "topic_id": "D-01",
                    "question": "Which source should own the app route?",
                    "recommendation": "Retain one canonical source.",
                    "source_ids": ["CFG-0001", "ARC-0002"],
                }
            ],
            include_sha=False,
        )
        with self.assertRaisesRegex(ValueError, "no source SHA-256"):
            self._build(topics)

    def test_gate_rejects_missing_audit_record(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        sheet = workbook["A4 Audit Register"]
        sheet.delete_rows(self._row_with(sheet, 1, "OPS-0001"))
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["audit_coverage"], "fail")

    def test_gate_rejects_reversed_action_direction(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        sheet = workbook["A2 Actions"]
        row = self._row_with(sheet, 1, "OP-0001")
        sheet.cell(row, 6).value = (
            "Keep variable:10; repoint trigger:30 to variable:10; delete variable:20."
        )
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["action_coverage_and_direction"], "fail")

    def test_gate_rejects_changed_structured_action_note(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        sheet = workbook["A2 Actions"]
        row = self._row_with(sheet, 1, "OP-0001")
        sheet.cell(row, 6).comment = Comment(
            "A shortened or altered mutation",
            "Tamper test",
        )
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["action_coverage_and_direction"], "fail")

    def test_gate_rejects_changed_original_sheet(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        workbook["03 Operational Review"]["B2"] = "Changed"
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["original_preservation"], "fail")

    def test_gate_rejects_changed_original_comment(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        workbook["03 Operational Review"]["B2"].comment = Comment(
            "Changed canonical note",
            "Tamper test",
        )
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["original_preservation"], "fail")

    def test_gate_rejects_stale_inputs(self) -> None:
        self._build()
        operations = json.loads(self.operations.read_text(encoding="utf-8"))
        operations["fixture_changed_after_build"] = True
        self.operations.write_text(
            json.dumps(operations, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["input_binding"], "fail")

    def test_builder_rejects_missing_or_mismatched_source_binding(self) -> None:
        context_path = self.package_dir / "context.json"
        context = json.loads(context_path.read_text(encoding="utf-8"))
        context.pop("source_sha256")
        context_path.write_text(
            json.dumps(context, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        with self.assertRaisesRegex(ValueError, "context has no source SHA-256"):
            self._build()

        context["source_sha256"] = "a" * 64
        context_path.write_text(
            json.dumps(context, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        completion_path = self.package_dir / "completion_gate.json"
        completion = json.loads(completion_path.read_text(encoding="utf-8"))
        completion["source_sha256"] = "b" * 64
        completion_path.write_text(
            json.dumps(completion, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        with self.assertRaisesRegex(ValueError, "completion_gate does not match"):
            self._build()

        completion_path.write_text(
            json.dumps({"status": "pass"}, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        with self.assertRaisesRegex(ValueError, "invalid kind"):
            self._build()

    def test_builder_rejects_an_unprojectable_ledger_disposition(self) -> None:
        operations = json.loads(self.operations.read_text(encoding="utf-8"))
        operations["decision_ledger"][0]["disposition"] = "unknown_state"
        self.operations.write_text(
            json.dumps(operations, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        with self.assertRaisesRegex(ValueError, "unsupported dispositions"):
            self._build()
        self.assertFalse(self.analyst.exists())

    def test_gate_rejects_hidden_sheet_link(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        sheet = workbook["A4 Audit Register"]
        row = self._row_with(sheet, 1, "OPS-0001")
        sheet.cell(row, 5).hyperlink = "#'07 Reconciled Operations'!A1"
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["visible_links"], "fail")

    def test_gate_cannot_use_or_overwrite_an_audit_input_as_manifest(self) -> None:
        self._build()
        context_path = self.package_dir / "context.json"
        before = context_path.read_bytes()

        with self.assertRaisesRegex(ValueError, "stored beside|cannot overwrite"):
            validate_readability(
                self.package_dir,
                self.operations,
                self.canonical,
                self.analyst,
                manifest_path=context_path,
            )
        self.assertEqual(context_path.read_bytes(), before)

    def test_gate_rejects_unsupported_absolute_claim(self) -> None:
        self._build()
        workbook = load_workbook(self.analyst)
        workbook["A1 Overview"]["A15"] = "Zero measurement loss"
        workbook.save(self.analyst)
        workbook.close()
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["readability"], "fail")

    def test_corrupt_analyst_file_fails_without_touching_fallback(self) -> None:
        canonical_hash = sha256_file(self.canonical)
        self._build()
        self.analyst.write_bytes(b"not an xlsx")
        self._refresh_analyst_hash()

        result = self._validate()
        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["gates"]["workbook_integrity"], "fail")
        self.assertEqual(sha256_file(self.canonical), canonical_hash)


if __name__ == "__main__":
    unittest.main()
